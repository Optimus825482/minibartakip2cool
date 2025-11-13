"""
Satın Alma Excel Servisleri
Toplu sipariş yükleme, Excel şablon oluşturma ve rapor export
"""

from datetime import datetime, date
from decimal import Decimal
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from werkzeug.utils import secure_filename
from models import (
    db, SatinAlmaSiparisi, SatinAlmaSiparisDetay, Tedarikci,
    Urun, UrunTedarikciFiyat, UrunStok, SiparisDurum
)
from utils.satin_alma_servisleri import SatinAlmaServisi
from utils.tedarikci_servisleri import TedarikciServisi
import logging

logger = logging.getLogger(__name__)


class SatinAlmaExcelServisi:
    """Satın alma Excel işlemleri servisi"""

    # Excel sütun başlıkları
    SIPARIS_SABLONU_SUTUNLAR = [
        'Ürün Kodu',
        'Ürün Adı',
        'Tedarikçi Adı',
        'Miktar',
        'Birim Fiyat',
        'Toplam Fiyat',
        'Açıklama'
    ]

    @staticmethod
    def excel_sablonu_olustur(otel_id: int, dosya_yolu: str) -> Dict:
        """
        Toplu sipariş için Excel şablonu oluştur
        
        Args:
            otel_id: Otel ID
            dosya_yolu: Şablon dosyasının kaydedileceği yol
        
        Returns:
            dict: {'success': bool, 'dosya_yolu': str, 'message': str}
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sipariş Şablonu"
            
            # Başlık satırı
            baslik_font = Font(bold=True, color="FFFFFF")
            baslik_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            baslik_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_idx, baslik in enumerate(SatinAlmaExcelServisi.SIPARIS_SABLONU_SUTUNLAR, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = baslik
                cell.font = baslik_font
                cell.fill = baslik_fill
                cell.alignment = baslik_alignment
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 30
            
            # Ürünleri getir
            urunler = db.session.query(Urun, UrunStok).join(
                UrunStok, Urun.id == UrunStok.urun_id
            ).filter(
                UrunStok.otel_id == otel_id
            ).order_by(Urun.urun_adi).all()
            
            # Örnek satırlar
            row_idx = 2
            for urun, stok in urunler[:50]:
                en_uygun = TedarikciServisi.en_uygun_tedarikci_bul(urun.id, 1)
                
                if en_uygun:
                    ws.cell(row=row_idx, column=1).value = urun.urun_kodu or f"URN-{urun.id}"
                    ws.cell(row=row_idx, column=2).value = urun.urun_adi
                    ws.cell(row=row_idx, column=3).value = en_uygun['tedarikci_adi']
                    ws.cell(row=row_idx, column=4).value = ""
                    ws.cell(row=row_idx, column=5).value = float(en_uygun['birim_fiyat'])
                    ws.cell(row=row_idx, column=6).value = f"=D{row_idx}*E{row_idx}"
                    ws.cell(row=row_idx, column=7).value = ""
                    
                    ws.cell(row=row_idx, column=4).number_format = '0'
                    ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
                    ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
                    
                    row_idx += 1
            
            # Talimatlar sayfası
            ws_talimat = wb.create_sheet("Talimatlar")
            talimatlar = [
                "TOPLU SİPARİŞ YÜKLEME TALİMATLARI",
                "",
                "1. 'Sipariş Şablonu' sekmesindeki ürünler için miktar girin",
                "2. Tedarikçi adını değiştirmek isterseniz, sistemde kayıtlı tedarikçi adını yazın",
                "3. Birim fiyat otomatik gelir, değiştirmek isterseniz yeni fiyat yazın",
                "4. Toplam fiyat otomatik hesaplanır",
                "5. Açıklama alanı opsiyoneldir",
                "6. Sipariş vermek istemediğiniz ürünlerin miktarını boş bırakın",
                "7. Dosyayı kaydedin ve sisteme yükleyin",
                "",
                f"Şablon Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                f"Otel ID: {otel_id}"
            ]
            
            for idx, talimat in enumerate(talimatlar, 1):
                cell = ws_talimat.cell(row=idx, column=1)
                cell.value = talimat
                if idx == 1:
                    cell.font = Font(bold=True, size=14)
            
            ws_talimat.column_dimensions['A'].width = 80
            
            wb.save(dosya_yolu)
            
            logger.info(f"Excel şablonu oluşturuldu: {dosya_yolu}")
            
            return {
                'success': True,
                'dosya_yolu': dosya_yolu,
                'message': 'Excel şablonu başarıyla oluşturuldu'
            }
            
        except Exception as e:
            logger.error(f"Excel şablonu oluşturma hatası: {str(e)}")
            return {
                'success': False,
                'message': f'Şablon oluşturulurken hata: {str(e)}'
            }

    @staticmethod
    def toplu_siparis_yukle(
        dosya_yolu: str,
        otel_id: int,
        kullanici_id: int,
        otomatik_onayla: bool = False
    ) -> Dict:
        """
        Excel'den toplu sipariş yükle
        
        Args:
            dosya_yolu: Excel dosya yolu
            otel_id: Otel ID
            kullanici_id: Yükleyen kullanıcı ID
            otomatik_onayla: True ise siparişler otomatik onaylanır
        
        Returns:
            dict: {
                'success': bool,
                'toplam_satir': int,
                'basarili_satir': int,
                'hatali_satir': int,
                'olusturulan_siparisler': list,
                'hatalar': list
            }
        """
        try:
            wb = openpyxl.load_workbook(dosya_yolu, data_only=True)
            ws = wb.active
            
            # Başlıkları kontrol et
            headers = [cell.value for cell in ws[1]]
            
            if not SatinAlmaExcelServisi._validate_headers(headers):
                return {
                    'success': False,
                    'message': 'Excel sütun başlıkları hatalı',
                    'toplam_satir': 0,
                    'basarili_satir': 0,
                    'hatali_satir': 0,
                    'olusturulan_siparisler': [],
                    'hatalar': ['Gerekli sütunlar bulunamadı']
                }
            
            # Satırları işle
            toplam_satir = 0
            basarili_satir = 0
            hatali_satir = 0
            hatalar = []
            
            # Tedarikçi bazında grupla
            tedarikci_siparisler = {}
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                toplam_satir += 1
                
                try:
                    # Satır verilerini çıkar
                    row_data = SatinAlmaExcelServisi._extract_siparis_row(row, row_idx)
                    
                    if not row_data:
                        continue  # Boş satır
                    
                    # Validasyon
                    is_valid, error_msg = SatinAlmaExcelServisi._validate_siparis_row(
                        row_data, otel_id
                    )
                    
                    if not is_valid:
                        hatali_satir += 1
                        hatalar.append(f"Satır {row_idx}: {error_msg}")
                        continue
                    
                    # Tedarikçiye göre grupla
                    tedarikci_id = row_data['tedarikci_id']
                    
                    if tedarikci_id not in tedarikci_siparisler:
                        tedarikci_siparisler[tedarikci_id] = {
                            'tedarikci': row_data['tedarikci'],
                            'detaylar': []
                        }
                    
                    tedarikci_siparisler[tedarikci_id]['detaylar'].append({
                        'urun_id': row_data['urun_id'],
                        'urun': row_data['urun'],
                        'miktar': row_data['miktar'],
                        'birim_fiyat': row_data['birim_fiyat'],
                        'aciklama': row_data.get('aciklama', '')
                    })
                    
                    basarili_satir += 1
                    
                except Exception as e:
                    hatali_satir += 1
                    hatalar.append(f"Satır {row_idx}: {str(e)}")
                    continue
            
            # Siparişleri oluştur
            olusturulan_siparisler = []
            
            for tedarikci_id, siparis_data in tedarikci_siparisler.items():
                try:
                    # Sipariş oluştur
                    siparis_result = SatinAlmaServisi.siparis_olustur(
                        otel_id=otel_id,
                        tedarikci_id=tedarikci_id,
                        olusturan_id=kullanici_id,
                        detaylar=siparis_data['detaylar'],
                        aciklama=f"Excel toplu yükleme - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                    )
                    
                    if siparis_result['success']:
                        siparis_id = siparis_result['siparis_id']
                        
                        # Otomatik onay
                        if otomatik_onayla:
                            onay_result = SatinAlmaServisi.siparis_onayla(
                                siparis_id=siparis_id,
                                onaylayan_id=kullanici_id
                            )
                            
                            if not onay_result['success']:
                                logger.warning(
                                    f"Sipariş oluşturuldu ama onaylanamadı: {siparis_id}"
                                )
                        
                        olusturulan_siparisler.append({
                            'siparis_id': siparis_id,
                            'siparis_no': siparis_result['siparis_no'],
                            'tedarikci_adi': siparis_data['tedarikci'].tedarikci_adi,
                            'urun_sayisi': len(siparis_data['detaylar']),
                            'durum': 'onaylandi' if otomatik_onayla else 'beklemede'
                        })
                    else:
                        hatalar.append(
                            f"Tedarikçi {siparis_data['tedarikci'].tedarikci_adi}: "
                            f"{siparis_result['message']}"
                        )
                        
                except Exception as e:
                    hatalar.append(
                        f"Tedarikçi {siparis_data['tedarikci'].tedarikci_adi}: {str(e)}"
                    )
            
            db.session.commit()
            
            logger.info(
                f"Toplu sipariş yükleme tamamlandı: "
                f"{len(olusturulan_siparisler)} sipariş oluşturuldu"
            )
            
            return {
                'success': True,
                'toplam_satir': toplam_satir,
                'basarili_satir': basarili_satir,
                'hatali_satir': hatali_satir,
                'olusturulan_siparisler': olusturulan_siparisler,
                'hatalar': hatalar[:50]
            }
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Toplu sipariş yükleme hatası: {str(e)}")
            return {
                'success': False,
                'message': f'Excel işleme hatası: {str(e)}',
                'toplam_satir': 0,
                'basarili_satir': 0,
                'hatali_satir': 0,
                'olusturulan_siparisler': [],
                'hatalar': [str(e)]
            }

    @staticmethod
    def _validate_headers(headers: List) -> bool:
        """Excel başlıklarını doğrula"""
        required = ['Ürün Kodu', 'Ürün Adı', 'Tedarikçi Adı', 'Miktar', 'Birim Fiyat']
        
        for req in required:
            if req not in headers:
                return False
        
        return True

    @staticmethod
    def _extract_siparis_row(row: tuple, row_idx: int) -> Optional[Dict]:
        """Satırdan sipariş verilerini çıkar"""
        try:
            urun_kodu = str(row[0]).strip() if row[0] else None
            urun_adi = str(row[1]).strip() if row[1] else None
            tedarikci_adi = str(row[2]).strip() if row[2] else None
            miktar = row[3]
            birim_fiyat = row[4]
            aciklama = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            
            # Boş satır kontrolü
            if not miktar or miktar == "" or miktar == 0:
                return None
            
            return {
                'urun_kodu': urun_kodu,
                'urun_adi': urun_adi,
                'tedarikci_adi': tedarikci_adi,
                'miktar': miktar,
                'birim_fiyat': birim_fiyat,
                'aciklama': aciklama
            }
            
        except Exception as e:
            logger.error(f"Satır {row_idx} veri çıkarma hatası: {str(e)}")
            return None

    @staticmethod
    def _validate_siparis_row(row_data: Dict, otel_id: int) -> tuple:
        """Sipariş satırını doğrula"""
        try:
            # Ürün kontrolü
            urun = Urun.query.filter(
                (Urun.urun_kodu == row_data['urun_kodu']) |
                (Urun.urun_adi == row_data['urun_adi'])
            ).first()
            
            if not urun:
                return False, f"Ürün bulunamadı: {row_data['urun_adi']}"
            
            row_data['urun_id'] = urun.id
            row_data['urun'] = urun
            
            # Tedarikçi kontrolü
            tedarikci = Tedarikci.query.filter_by(
                tedarikci_adi=row_data['tedarikci_adi'],
                aktif=True
            ).first()
            
            if not tedarikci:
                return False, f"Tedarikçi bulunamadı: {row_data['tedarikci_adi']}"
            
            row_data['tedarikci_id'] = tedarikci.id
            row_data['tedarikci'] = tedarikci
            
            # Miktar kontrolü
            try:
                miktar = int(row_data['miktar'])
                if miktar <= 0:
                    return False, "Miktar pozitif olmalı"
                row_data['miktar'] = miktar
            except (ValueError, TypeError):
                return False, f"Geçersiz miktar: {row_data['miktar']}"
            
            # Birim fiyat kontrolü
            try:
                birim_fiyat = Decimal(str(row_data['birim_fiyat']))
                if birim_fiyat <= 0:
                    return False, "Birim fiyat pozitif olmalı"
                row_data['birim_fiyat'] = birim_fiyat
            except (ValueError, TypeError):
                return False, f"Geçersiz birim fiyat: {row_data['birim_fiyat']}"
            
            return True, None
            
        except Exception as e:
            return False, f"Validasyon hatası: {str(e)}"

    @staticmethod
    def rapor_excel_export(
        rapor_data: Dict,
        rapor_tipi: str,
        dosya_yolu: str
    ) -> Dict:
        """
        Rapor verilerini Excel'e export et
        
        Args:
            rapor_data: Rapor verileri
            rapor_tipi: 'tedarikci_performans' veya 'satin_alma_ozet'
            dosya_yolu: Export dosya yolu
        
        Returns:
            dict: {'success': bool, 'dosya_yolu': str, 'message': str}
        """
        try:
            if rapor_tipi == 'tedarikci_performans':
                return SatinAlmaExcelServisi._export_tedarikci_performans(
                    rapor_data, dosya_yolu
                )
            elif rapor_tipi == 'satin_alma_ozet':
                return SatinAlmaExcelServisi._export_satin_alma_ozet(
                    rapor_data, dosya_yolu
                )
            else:
                return {
                    'success': False,
                    'message': f'Bilinmeyen rapor tipi: {rapor_tipi}'
                }
                
        except Exception as e:
            logger.error(f"Rapor export hatası: {str(e)}")
            return {
                'success': False,
                'message': f'Export hatası: {str(e)}'
            }

    @staticmethod
    def _export_tedarikci_performans(rapor_data: Dict, dosya_yolu: str) -> Dict:
        """Tedarikçi performans raporunu Excel'e export et"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tedarikçi Performans"
            
            # Başlık
            ws['A1'] = 'TEDARİKÇİ PERFORMANS RAPORU'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')
            
            # Genel bilgiler
            ws['A3'] = 'Rapor Tarihi:'
            ws['B3'] = datetime.now().strftime('%d.%m.%Y %H:%M')
            
            if 'genel_ozet' in rapor_data:
                ozet = rapor_data['genel_ozet']
                ws['A4'] = 'Dönem:'
                ws['B4'] = f"{ozet.get('donem_baslangic', '')} - {ozet.get('donem_bitis', '')}"
            
            # Tablo başlıkları
            headers = [
                'Tedarikçi Adı',
                'Sipariş Sayısı',
                'Toplam Tutar (TL)',
                'Ort. Sipariş Tutarı (TL)',
                'Tamamlanan',
                'İptal Edilen',
                'Pay (%)'
            ]
            
            row_idx = 6
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Veriler
            if 'tedarikci_analizi' in rapor_data:
                row_idx = 7
                for item in rapor_data['tedarikci_analizi']:
                    ws.cell(row=row_idx, column=1).value = item['tedarikci_adi']
                    ws.cell(row=row_idx, column=2).value = item['siparis_sayisi']
                    ws.cell(row=row_idx, column=3).value = item['toplam_tutar']
                    ws.cell(row=row_idx, column=4).value = item['ortalama_siparis_tutari']
                    ws.cell(row=row_idx, column=5).value = item['tamamlanan_siparis']
                    ws.cell(row=row_idx, column=6).value = item['iptal_edilen_siparis']
                    ws.cell(row=row_idx, column=7).value = item['pay_yuzdesi']
                    
                    # Sayı formatları
                    ws.cell(row=row_idx, column=3).number_format = '#,##0.00'
                    ws.cell(row=row_idx, column=4).number_format = '#,##0.00'
                    ws.cell(row=row_idx, column=7).number_format = '0.00'
                    
                    row_idx += 1
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            
            wb.save(dosya_yolu)
            
            logger.info(f"Tedarikçi performans raporu export edildi: {dosya_yolu}")
            
            return {
                'success': True,
                'dosya_yolu': dosya_yolu,
                'message': 'Rapor başarıyla export edildi'
            }
            
        except Exception as e:
            logger.error(f"Tedarikçi performans export hatası: {str(e)}")
            return {
                'success': False,
                'message': f'Export hatası: {str(e)}'
            }

    @staticmethod
    def _export_satin_alma_ozet(rapor_data: Dict, dosya_yolu: str) -> Dict:
        """Satın alma özet raporunu Excel'e export et"""
        try:
            wb = openpyxl.Workbook()
            
            # Genel Özet Sayfası
            ws_ozet = wb.active
            ws_ozet.title = "Genel Özet"
            
            ws_ozet['A1'] = 'SATIN ALMA ÖZET RAPORU'
            ws_ozet['A1'].font = Font(bold=True, size=14)
            ws_ozet.merge_cells('A1:B1')
            
            if 'genel_ozet' in rapor_data:
                ozet = rapor_data['genel_ozet']
                row_idx = 3
                
                ozet_items = [
                    ('Dönem:', f"{ozet.get('donem_baslangic', '')} - {ozet.get('donem_bitis', '')}"),
                    ('Toplam Sipariş:', ozet.get('toplam_siparis', 0)),
                    ('Toplam Tutar (TL):', ozet.get('toplam_tutar', 0)),
                    ('Ortalama Sipariş Tutarı (TL):', ozet.get('ortalama_siparis_tutari', 0)),
                    ('Benzersiz Tedarikçi:', ozet.get('benzersiz_tedarikci', 0)),
                    ('Benzersiz Ürün:', ozet.get('benzersiz_urun', 0)),
                    ('Tamamlanan Sipariş:', ozet.get('tamamlanan_siparis', 0)),
                    ('İptal Edilen Sipariş:', ozet.get('iptal_edilen_siparis', 0))
                ]
                
                for label, value in ozet_items:
                    ws_ozet.cell(row=row_idx, column=1).value = label
                    ws_ozet.cell(row=row_idx, column=1).font = Font(bold=True)
                    ws_ozet.cell(row=row_idx, column=2).value = value
                    row_idx += 1
            
            ws_ozet.column_dimensions['A'].width = 30
            ws_ozet.column_dimensions['B'].width = 25
            
            # Tedarikçi Analizi Sayfası
            if 'tedarikci_analizi' in rapor_data:
                ws_ted = wb.create_sheet("Tedarikçi Analizi")
                
                headers = ['Tedarikçi', 'Sipariş', 'Toplam (TL)', 'Ortalama (TL)', 'Pay (%)']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_ted.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                row_idx = 2
                for item in rapor_data['tedarikci_analizi']:
                    ws_ted.cell(row=row_idx, column=1).value = item['tedarikci_adi']
                    ws_ted.cell(row=row_idx, column=2).value = item['siparis_sayisi']
                    ws_ted.cell(row=row_idx, column=3).value = item['toplam_tutar']
                    ws_ted.cell(row=row_idx, column=4).value = item['ortalama_siparis_tutari']
                    ws_ted.cell(row=row_idx, column=5).value = item['pay_yuzdesi']
                    
                    ws_ted.cell(row=row_idx, column=3).number_format = '#,##0.00'
                    ws_ted.cell(row=row_idx, column=4).number_format = '#,##0.00'
                    
                    row_idx += 1
                
                ws_ted.column_dimensions['A'].width = 30
                ws_ted.column_dimensions['B'].width = 12
                ws_ted.column_dimensions['C'].width = 18
                ws_ted.column_dimensions['D'].width = 18
                ws_ted.column_dimensions['E'].width = 12
            
            # Ürün Analizi Sayfası
            if 'urun_analizi' in rapor_data:
                ws_urun = wb.create_sheet("Ürün Analizi")
                
                headers = ['Ürün', 'Miktar', 'Toplam (TL)', 'Ort. Fiyat', 'Sipariş', 'Tedarikçi']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_urun.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                row_idx = 2
                for item in rapor_data['urun_analizi'][:100]:  # İlk 100 ürün
                    ws_urun.cell(row=row_idx, column=1).value = item['urun_adi']
                    ws_urun.cell(row=row_idx, column=2).value = item['toplam_miktar']
                    ws_urun.cell(row=row_idx, column=3).value = item['toplam_tutar']
                    ws_urun.cell(row=row_idx, column=4).value = item['ortalama_birim_fiyat']
                    ws_urun.cell(row=row_idx, column=5).value = item['siparis_sayisi']
                    ws_urun.cell(row=row_idx, column=6).value = item['tedarikci_sayisi']
                    
                    ws_urun.cell(row=row_idx, column=3).number_format = '#,##0.00'
                    ws_urun.cell(row=row_idx, column=4).number_format = '#,##0.00'
                    
                    row_idx += 1
                
                ws_urun.column_dimensions['A'].width = 35
                ws_urun.column_dimensions['B'].width = 12
                ws_urun.column_dimensions['C'].width = 18
                ws_urun.column_dimensions['D'].width = 15
                ws_urun.column_dimensions['E'].width = 12
                ws_urun.column_dimensions['F'].width = 12
            
            wb.save(dosya_yolu)
            
            logger.info(f"Satın alma özet raporu export edildi: {dosya_yolu}")
            
            return {
                'success': True,
                'dosya_yolu': dosya_yolu,
                'message': 'Rapor başarıyla export edildi'
            }
            
        except Exception as e:
            logger.error(f"Satın alma özet export hatası: {str(e)}")
            return {
                'success': False,
                'message': f'Export hatası: {str(e)}'
            }
