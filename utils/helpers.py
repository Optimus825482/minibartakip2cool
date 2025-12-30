from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
from flask import session, request
from models import Kullanici, StokHareket, Urun, SistemLog, HataLog, db
from sqlalchemy import case
import json
import traceback
import logging
import pytz

# Logging yapılandırması
logging.basicConfig(
    filename='minibar_errors.log',
    level=logging.ERROR,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# KKTC Timezone (Kıbrıs - Europe/Nicosia)
KKTC_TZ = pytz.timezone('Europe/Nicosia')


def get_kktc_now():
    """
    Kıbrıs saat diliminde şu anki zamanı döndürür.
    Tüm sistemde saat kaydı için bu fonksiyon kullanılmalıdır.
    
    Returns:
        datetime: KKTC timezone'unda şu anki zaman
    """
    return datetime.now(KKTC_TZ)


def get_kktc_today():
    """
    Kıbrıs saat diliminde bugünün tarihini döndürür.
    
    Returns:
        date: KKTC timezone'unda bugünün tarihi
    """
    return datetime.now(KKTC_TZ).date()


def utc_to_kktc(utc_datetime):
    """
    UTC datetime'ı KKTC timezone'una çevirir.
    
    Args:
        utc_datetime: UTC timezone'unda datetime
        
    Returns:
        datetime: KKTC timezone'unda datetime
    """
    if utc_datetime is None:
        return None
    if utc_datetime.tzinfo is None:
        utc_datetime = utc_datetime.replace(tzinfo=timezone.utc)
    return utc_datetime.astimezone(KKTC_TZ)


def kktc_to_utc(kktc_datetime):
    """
    KKTC datetime'ı UTC timezone'una çevirir.
    
    Args:
        kktc_datetime: KKTC timezone'unda datetime
        
    Returns:
        datetime: UTC timezone'unda datetime
    """
    if kktc_datetime is None:
        return None
    if kktc_datetime.tzinfo is None:
        kktc_datetime = KKTC_TZ.localize(kktc_datetime)
    return kktc_datetime.astimezone(timezone.utc)

# Logging yapılandırması
logging.basicConfig(
    filename='minibar_errors.log',
    level=logging.ERROR,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)


def get_current_user():
    """Oturumdaki kullanıcıyı getir"""
    if 'kullanici_id' in session:
        try:
            return Kullanici.query.get(session['kullanici_id'])
        except Exception as e:
            # Transaction hatası durumunda rollback yap ve tekrar dene
            logging.error(f"❌ get_current_user hatası: {str(e)}")
            db.session.rollback()
            try:
                return Kullanici.query.get(session['kullanici_id'])
            except Exception as retry_error:
                logging.error(f"❌ get_current_user retry hatası: {str(retry_error)}")
                return None
    return None


def get_stok_toplamlari(urun_ids=None):
    """Belirtilen ürünler için stok toplamlarını tek sorguda getir."""
    net_miktar = db.func.sum(
        case(
            (StokHareket.hareket_tipi.in_(['giris', 'devir', 'sayim']), StokHareket.miktar),
            (StokHareket.hareket_tipi == 'cikis', -StokHareket.miktar),
            else_=0
        )
    )

    query = db.session.query(StokHareket.urun_id, net_miktar.label('net'))

    if urun_ids:
        query = query.filter(StokHareket.urun_id.in_(urun_ids))

    stoklar = {row.urun_id: int(row.net or 0) for row in query.group_by(StokHareket.urun_id)}

    if urun_ids:
        for uid in urun_ids:
            stoklar.setdefault(uid, 0)

    return stoklar


def get_toplam_stok(urun_id):
    """Ürün için toplam stok miktarını hesapla"""
    return get_stok_toplamlari([urun_id]).get(urun_id, 0)


def get_kritik_stok_urunler():
    """Kritik stok seviyesinin altındaki ürünleri getir"""
    urunler = Urun.query.filter_by(aktif=True).all()
    stok_map = get_stok_toplamlari([urun.id for urun in urunler])
    kritik_urunler = []
    
    for urun in urunler:
        # Kritik stok seviyesi None ise atla
        if urun.kritik_stok_seviyesi is None:
            continue
            
        mevcut_stok = stok_map.get(urun.id, 0)
        if mevcut_stok <= urun.kritik_stok_seviyesi:
            kritik_urunler.append({
                'urun': urun,
                'mevcut_stok': mevcut_stok,
                'kritik_seviye': urun.kritik_stok_seviyesi
            })
    
    return kritik_urunler


def get_stok_durumu(urun_id, stok_cache=None):
    """
    Ürün stok durumunu kategorize et ve badge bilgisi döndür
    
    Returns:
        dict: {
            'durum': 'kritik' | 'dikkat' | 'normal',
            'badge_class': Tailwind CSS class,
            'badge_text': Görünecek metin,
            'icon': SVG icon HTML,
            'mevcut_stok': int,
            'kritik_seviye': int,
            'yuzde': float (stok doluluk yüzdesi)
        }
    """
    urun = Urun.query.get(urun_id)
    if not urun:
        return None
    
    if stok_cache is None:
        mevcut_stok = get_toplam_stok(urun_id)
    else:
        mevcut_stok = stok_cache.get(urun_id, 0)
    kritik_seviye = urun.kritik_stok_seviyesi
    
    # Kritik seviye None ise normal kabul et
    if kritik_seviye is None:
        kritik_seviye = 0
    
    # Kritik seviyenin %150'si dikkat eşiği
    dikkat_esigi = kritik_seviye * 1.5
    
    # Yüzde hesaplama (kritik seviyeye göre)
    if kritik_seviye > 0:
        yuzde = (mevcut_stok / kritik_seviye) * 100
    else:
        yuzde = 100 if mevcut_stok > 0 else 0
    
    if mevcut_stok <= kritik_seviye:
        # KRİTİK - Kırmızı
        return {
            'durum': 'kritik',
            'badge_class': 'bg-red-100 text-red-800 border border-red-300',
            'badge_text': 'Kritik Stok',
            'icon': '''<svg class="w-5 h-5 text-red-500" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M8.257 3.099c.765-1.36 2.722-1.36 3.486 0l5.58 9.92c.75 1.334-.213 2.98-1.742 2.98H4.42c-1.53 0-2.493-1.646-1.743-2.98l5.58-9.92zM11 13a1 1 0 11-2 0 1 1 0 012 0zm-1-8a1 1 0 00-1 1v3a1 1 0 002 0V6a1 1 0 00-1-1z" clip-rule="evenodd"/>
                      </svg>''',
            'mevcut_stok': mevcut_stok,
            'kritik_seviye': kritik_seviye,
            'yuzde': yuzde
        }
    elif mevcut_stok <= dikkat_esigi:
        # DİKKAT - Sarı
        return {
            'durum': 'dikkat',
            'badge_class': 'bg-yellow-100 text-yellow-800 border border-yellow-300',
            'badge_text': 'Dikkat',
            'icon': '''<svg class="w-5 h-5 text-yellow-500" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M18 10a8 8 0 11-16 0 8 8 0 0116 0zm-7-4a1 1 0 11-2 0 1 1 0 012 0zM9 9a1 1 0 000 2v3a1 1 0 001 1h1a1 1 0 100-2v-3a1 1 0 00-1-1H9z" clip-rule="evenodd"/>
                      </svg>''',
            'mevcut_stok': mevcut_stok,
            'kritik_seviye': kritik_seviye,
            'yuzde': yuzde
        }
    else:
        # NORMAL - Yeşil
        return {
            'durum': 'normal',
            'badge_class': 'bg-green-100 text-green-800 border border-green-300',
            'badge_text': 'Yeterli',
            'icon': '''<svg class="w-5 h-5 text-green-500" fill="currentColor" viewBox="0 0 20 20">
                        <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-9.293a1 1 0 00-1.414-1.414L9 10.586 7.707 9.293a1 1 0 00-1.414 1.414l2 2a1 1 0 001.414 0l4-4z" clip-rule="evenodd"/>
                      </svg>''',
            'mevcut_stok': mevcut_stok,
            'kritik_seviye': kritik_seviye,
            'yuzde': yuzde
        }


def get_tum_urunler_stok_durumlari():
    """
    Tüm aktif ürünlerin stok durumlarını kategorize et
    
    Returns:
        dict: {
            'kritik': [],    # Kritik stokta olan ürünler
            'dikkat': [],    # Dikkat gerektiren ürünler
            'normal': [],    # Normal stokta olan ürünler
            'istatistik': {
                'toplam': int,
                'kritik_sayi': int,
                'dikkat_sayi': int,
                'normal_sayi': int
            }
        }
    """
    urunler = Urun.query.filter_by(aktif=True).all()
    stok_map = get_stok_toplamlari([urun.id for urun in urunler])
    
    kategoriler = {
        'kritik': [],
        'dikkat': [],
        'normal': []
    }
    
    for urun in urunler:
        durum = get_stok_durumu(urun.id, stok_cache=stok_map)
        if durum:
            durum['urun'] = urun
            kategoriler[durum['durum']].append(durum)
    
    return {
        'kritik': sorted(kategoriler['kritik'], key=lambda x: x['mevcut_stok']),
        'dikkat': sorted(kategoriler['dikkat'], key=lambda x: x['mevcut_stok']),
        'normal': sorted(kategoriler['normal'], key=lambda x: x['urun'].urun_adi),
        'istatistik': {
            'toplam': len(urunler),
            'kritik_sayi': len(kategoriler['kritik']),
            'dikkat_sayi': len(kategoriler['dikkat']),
            'normal_sayi': len(kategoriler['normal'])
        }
    }


def excel_export_stok_raporu():
    """Stok durumu Excel raporu oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Raporu"
    
    # Başlık stili
    baslik_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    baslik_fill = PatternFill(start_color='475569', end_color='475569', fill_type='solid')
    baslik_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = ['Ürün Kodu', 'Ürün Adı', 'Grup', 'Birim', 'Mevcut Stok', 'Kritik Seviye', 'Durum']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = baslik_alignment
        cell.border = border
    
    # Veri satırları
    urunler = Urun.query.filter_by(aktif=True).all()
    stok_map = get_stok_toplamlari([urun.id for urun in urunler])

    for row_num, urun in enumerate(urunler, 2):
        mevcut_stok = stok_map.get(urun.id, 0)
        durum = 'KRİTİK' if mevcut_stok <= urun.kritik_stok_seviyesi else 'NORMAL'
        
        data = [
            urun.id,
            urun.urun_adi,
            urun.grup.grup_adi,
            urun.birim,
            mevcut_stok,
            urun.kritik_stok_seviyesi,
            durum
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if durum == 'KRİTİK' and col_num == 7:
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                cell.font = Font(color='DC2626', bold=True)
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    # BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def excel_export_zimmet_raporu(personel_id):
    """Personel zimmet Excel raporu oluştur"""
    from models import PersonelZimmet, PersonelZimmetDetay
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Zimmet Raporu"
    
    # Stil tanımlamaları
    baslik_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    baslik_fill = PatternFill(start_color='475569', end_color='475569', fill_type='solid')
    baslik_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıklar
    headers = ['Zimmet No', 'Ürün Adı', 'Teslim Edilen', 'Kullanılan', 'Kalan', 'Tarih', 'Durum']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = baslik_alignment
        cell.border = border
    
    # Veri satırları
    zimmetler = PersonelZimmet.query.filter_by(personel_id=personel_id).all()
    row_num = 2
    
    for zimmet in zimmetler:
        for detay in zimmet.detaylar:
            data = [
                zimmet.id,
                detay.urun.urun_adi,
                detay.miktar,
                detay.kullanilan_miktar,
                detay.kalan_miktar or (detay.miktar - detay.kullanilan_miktar),
                zimmet.zimmet_tarihi.strftime('%d.%m.%Y'),
                zimmet.durum.upper()
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            row_num += 1
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def pdf_export_stok_raporu():
    """Stok durumu PDF raporu oluştur"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Stil tanımlamaları
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=30,
        alignment=1
    )
    
    # Başlık
    title = Paragraph("STOK DURUM RAPORU", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Tablo verileri
    data = [['Ürün Kodu', 'Ürün Adı', 'Grup', 'Birim', 'Mevcut Stok', 'Kritik Seviye', 'Durum']]
    
    urunler = Urun.query.filter_by(aktif=True).all()
    stok_map = get_stok_toplamlari([urun.id for urun in urunler]) if urunler else {}
    for urun in urunler:
        mevcut_stok = stok_map.get(urun.id, 0)
        durum = 'KRİTİK' if mevcut_stok <= urun.kritik_stok_seviyesi else 'NORMAL'
        
        data.append([
            str(urun.id),
            urun.urun_adi[:30],
            urun.grup.grup_adi[:20],
            urun.birim,
            str(mevcut_stok),
            str(urun.kritik_stok_seviyesi),
            durum
        ])
    
    # Tablo oluştur
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#475569')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # PDF oluştur
    doc.build(elements)
    buffer.seek(0)
    
    return buffer


def format_tarih(tarih):
    """Tarih formatla"""
    if not tarih:
        return ''
    return tarih.strftime('%d.%m.%Y %H:%M')


def format_para(tutar):
    """Para formatla"""
    if not tutar:
        return '0,00 ₺'
    return f'{tutar:,.2f} ₺'.replace(',', 'X').replace('.', ',').replace('X', '.')


def log_islem(islem_tipi, modul, islem_detay=None):
    """
    Sistem işlemlerini logla
    
    Args:
        islem_tipi: giris, cikis, ekleme, guncelleme, silme, goruntuleme
        modul: urun, stok, zimmet, oda, kat, personel vb.
        islem_detay: İşlem detayları (dict veya string)
    """
    try:
        kullanici_id = session.get('kullanici_id')
        
        # İşlem detayını JSON'a çevir
        if isinstance(islem_detay, dict):
            detay_json = json.dumps(islem_detay, ensure_ascii=False)
        else:
            detay_json = str(islem_detay) if islem_detay else None
        
        # IP adresi ve tarayıcı bilgisi
        ip_adresi = request.remote_addr if request else None
        tarayici = request.headers.get('User-Agent', '')[:200] if request else None
        
        # Log kaydı oluştur
        log = SistemLog(
            kullanici_id=kullanici_id,
            islem_tipi=islem_tipi,
            modul=modul,
            islem_detay=detay_json,
            ip_adresi=ip_adresi,
            tarayici=tarayici
        )
        
        db.session.add(log)
        db.session.commit()
        
    except Exception as e:
        # Log hatası uygulamayı durdurmamalı
        print(f'Log hatası: {str(e)}')
        try:
            db.session.rollback()
        except Exception:
            pass


def get_son_loglar(limit=50):
    """Son log kayıtlarını getir"""
    return SistemLog.query.order_by(SistemLog.islem_tarihi.desc()).limit(limit).all()


def get_kullanici_loglari(kullanici_id, limit=50):
    """Belirli bir kullanıcının log kayıtlarını getir"""
    return SistemLog.query.filter_by(kullanici_id=kullanici_id).order_by(
        SistemLog.islem_tarihi.desc()
    ).limit(limit).all()


def get_modul_loglari(modul, limit=50):
    """Belirli bir modülün log kayıtlarını getir"""
    return SistemLog.query.filter_by(modul=modul).order_by(
        SistemLog.islem_tarihi.desc()
    ).limit(limit).all()


def log_hata(exception, modul=None, extra_info=None):
    """
    Hataları hem dosyaya hem de veritabanına logla
    
    Args:
        exception: Exception objesi
        modul: Hangi modülde oluştu (örn: 'minibar', 'stok', 'zimmet')
        extra_info: Ek bilgiler (dict)
    
    Returns:
        HataLog objesi
    """
    try:
        kullanici_id = session.get('kullanici_id')
        
        # Exception bilgileri
        hata_tipi = type(exception).__name__
        hata_mesaji = str(exception)
        hata_detay = traceback.format_exc()
        
        # Request bilgileri
        url = request.url if request else None
        method = request.method if request else None
        ip_adresi = request.remote_addr if request else None
        tarayici = request.headers.get('User-Agent', '')[:200] if request else None
        
        # Ek bilgileri JSON'a çevir
        if extra_info:
            hata_detay += f"\n\nEk Bilgiler:\n{json.dumps(extra_info, ensure_ascii=False, indent=2)}"
        
        # Dosyaya logla
        logging.error(
            f"Hata: {hata_tipi} - {hata_mesaji}\n"
            f"Modül: {modul}\n"
            f"URL: {url}\n"
            f"Kullanıcı: {kullanici_id}\n"
            f"Detay:\n{hata_detay}"
        )
        
        # Veritabanına logla
        hata_log = HataLog(
            kullanici_id=kullanici_id,
            hata_tipi=hata_tipi,
            hata_mesaji=hata_mesaji[:500],  # İlk 500 karakter
            hata_detay=hata_detay,
            modul=modul,
            url=url[:500] if url else None,
            method=method,
            ip_adresi=ip_adresi,
            tarayici=tarayici
        )
        
        db.session.add(hata_log)
        db.session.commit()
        
        return hata_log
        
    except Exception as e:
        # Hata loglama hatası uygulamayı durdurmamalı
        print(f'HATA LOGLAMA HATASI: {str(e)}')
        logging.error(f'Hata loglama hatası: {str(e)}')
        try:
            db.session.rollback()
        except Exception:
            pass
        return None


def get_son_hatalar(limit=50):
    """Son hata kayıtlarını getir"""
    return HataLog.query.order_by(HataLog.olusturma_tarihi.desc()).limit(limit).all()


def get_cozulmemis_hatalar():
    """Çözülmemiş hataları getir"""
    return HataLog.query.filter_by(cozuldu=False).order_by(HataLog.olusturma_tarihi.desc()).all()


def hata_cozuldu_isaretle(hata_id, cozum_notu=None):
    """Hatayı çözüldü olarak işaretle"""
    try:
        hata = HataLog.query.get(hata_id)
        if hata:
            hata.cozuldu = True
            hata.cozum_notu = cozum_notu
            db.session.commit()
            return True
    except Exception as e:
        logging.error(f'Hata çözüldü işaretleme hatası: {str(e)}')
        db.session.rollback()
    return False



# ============================================
# ADMIN MINIBAR YÖNETİMİ HELPER FONKSIYONLARI
# ============================================

def get_depo_stok_durumu(grup_id=None, depo_sorumlusu_id=None, otel_id=None):
    """
    Depo stok durumlarını hesaplar (Depo + Kat Sorumlusu Zimmetleri)
    
    Args:
        grup_id (int, optional): Ürün grubu filtresi
        depo_sorumlusu_id (int, optional): Depo sorumlusu filtresi
        otel_id (int, optional): Otel filtresi - Otel bazlı stok takibi için
    
    Returns:
        list: [
            {
                'urun_id': int,
                'urun_adi': str,
                'grup_adi': str,
                'birim': str,
                'depo_stok': int,
                'zimmet_stok': int,
                'toplam_stok': int,
                'kritik_stok': int,
                'durum': 'kritik'|'dikkat'|'normal',
                'badge_class': str,
                'badge_text': str
            },
            ...
        ]
    """
    try:
        from models import UrunGrup, PersonelZimmet, PersonelZimmetDetay, Kullanici, UrunStok
        
        # Ürünleri getir
        query = Urun.query.filter_by(aktif=True)
        if grup_id:
            query = query.filter_by(grup_id=grup_id)
        
        urunler = query.all()
        
        # Otel bazlı stok durumlarını hesapla (UrunStok tablosundan)
        depo_stok_map = {}
        if otel_id:
            # Tek otel bazlı stok - UrunStok tablosundan
            stok_query = UrunStok.query.filter_by(otel_id=otel_id)
            for stok in stok_query.all():
                depo_stok_map[stok.urun_id] = stok.mevcut_stok
        else:
            # Tüm otellerin toplam stoğu
            from sqlalchemy import func
            toplam_stoklar = db.session.query(
                UrunStok.urun_id,
                func.sum(UrunStok.mevcut_stok).label('toplam')
            ).group_by(UrunStok.urun_id).all()
            depo_stok_map = {row.urun_id: row.toplam or 0 for row in toplam_stoklar}
        
        # Kat sorumlusu zimmet stoklarını hesapla
        zimmet_query = db.session.query(
            PersonelZimmetDetay.urun_id,
            db.func.sum(PersonelZimmetDetay.kalan_miktar).label('toplam_zimmet')
        ).join(
            PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id
        ).join(
            Kullanici, PersonelZimmet.personel_id == Kullanici.id
        ).filter(
            PersonelZimmet.durum == 'aktif',
            Kullanici.rol == 'kat_sorumlusu',
            Kullanici.aktif == True
        )
        
        # Otel filtresi varsa, o oteldeki kat sorumlularının zimmetlerini filtrele
        if otel_id:
            zimmet_query = zimmet_query.filter(Kullanici.otel_id == otel_id)
        
        # Depo sorumlusu filtresi varsa, onun zimmet verdiği kat sorumlularını filtrele
        if depo_sorumlusu_id:
            zimmet_query = zimmet_query.filter(
                PersonelZimmet.teslim_eden_id == depo_sorumlusu_id
            )
        
        zimmet_query = zimmet_query.group_by(PersonelZimmetDetay.urun_id)
        zimmet_sonuclar = zimmet_query.all()
        
        # Zimmet map oluştur
        zimmet_map = {z.urun_id: float(z.toplam_zimmet or 0) for z in zimmet_sonuclar}
        
        sonuclar = []
        for urun in urunler:
            depo_stok = depo_stok_map.get(urun.id, 0)
            zimmet_stok = zimmet_map.get(urun.id, 0)
            toplam_stok = depo_stok + zimmet_stok
            
            # Toplam stoku 0 olan ürünleri atla
            if toplam_stok == 0:
                continue
            
            kritik_seviye = urun.kritik_stok_seviyesi
            dikkat_esigi = kritik_seviye * 1.5
            
            # Durum belirleme (toplam stoka göre)
            if toplam_stok <= kritik_seviye:
                durum = 'kritik'
                badge_class = 'bg-red-100 text-red-800 border border-red-300 dark:bg-red-900/20 dark:text-red-400'
                badge_text = 'Kritik'
            elif toplam_stok <= dikkat_esigi:
                durum = 'dikkat'
                badge_class = 'bg-yellow-100 text-yellow-800 border border-yellow-300 dark:bg-yellow-900/20 dark:text-yellow-400'
                badge_text = 'Dikkat'
            else:
                durum = 'normal'
                badge_class = 'bg-green-100 text-green-800 border border-green-300 dark:bg-green-900/20 dark:text-green-400'
                badge_text = 'Normal'
            
            sonuclar.append({
                'urun_id': urun.id,
                'urun_adi': urun.urun_adi,
                'grup_adi': urun.grup.grup_adi,
                'birim': urun.birim,
                'depo_stok': depo_stok,
                'zimmet_stok': zimmet_stok,
                'toplam_stok': toplam_stok,
                'kritik_stok': kritik_seviye,
                'durum': durum,
                'badge_class': badge_class,
                'badge_text': badge_text
            })
        
        # Kritik olanları başa al
        sonuclar.sort(key=lambda x: (0 if x['durum'] == 'kritik' else 1 if x['durum'] == 'dikkat' else 2, x['urun_adi']))
        
        return sonuclar
        
    except Exception as e:
        log_hata(e, modul='admin_minibar', extra_info={'function': 'get_depo_stok_durumu'})
        return []


def get_oda_minibar_stoklari(kat_id=None):
    """
    Tüm odaların minibar stok durumlarını getirir
    
    Args:
        kat_id (int, optional): Kat filtresi
    
    Returns:
        list: [
            {
                'oda_id': int,
                'oda_no': str,
                'kat_adi': str,
                'kat_no': int,
                'son_islem_tarihi': datetime,
                'toplam_urun_sayisi': int,
                'bos_mu': bool
            },
            ...
        ]
    """
    try:
        from models import Oda, Kat, MinibarIslem
        
        # Odaları getir
        query = Oda.query.filter_by(aktif=True).join(Kat)
        if kat_id:
            query = query.filter(Oda.kat_id == kat_id)
        
        odalar = query.order_by(Kat.kat_no, Oda.oda_no).all()
        
        sonuclar = []
        for oda in odalar:
            # Son minibar işlemini bul
            son_islem = MinibarIslem.query.filter_by(oda_id=oda.id).order_by(
                MinibarIslem.islem_tarihi.desc()
            ).first()
            
            if son_islem:
                # Ürün sayısını hesapla
                toplam_urun = len(son_islem.detaylar)
                bos_mu = False
                son_islem_tarihi = son_islem.islem_tarihi
            else:
                toplam_urun = 0
                bos_mu = True
                son_islem_tarihi = None
            
            sonuclar.append({
                'oda_id': oda.id,
                'oda_no': oda.oda_no,
                'kat_adi': oda.kat.kat_adi,
                'kat_no': oda.kat.kat_no,
                'son_islem_tarihi': son_islem_tarihi,
                'toplam_urun_sayisi': toplam_urun,
                'bos_mu': bos_mu
            })
        
        return sonuclar
        
    except Exception as e:
        log_hata(e, modul='admin_minibar', extra_info={'function': 'get_oda_minibar_stoklari'})
        return []


def get_oda_minibar_detay(oda_id):
    """
    Belirli bir odanın minibar detaylarını getirir
    
    Args:
        oda_id (int): Oda ID
    
    Returns:
        dict: {
            'oda': Oda object,
            'son_islem': MinibarIslem object,
            'urunler': [
                {
                    'urun_adi': str,
                    'baslangic_stok': int,
                    'bitis_stok': int,
                    'tuketim': int,
                    'eklenen_miktar': int
                },
                ...
            ]
        } veya None
    """
    try:
        from models import Oda, MinibarIslem
        
        # Odayı getir
        oda = Oda.query.get(oda_id)
        if not oda:
            return None
        
        # Son minibar işlemini bul
        son_islem = MinibarIslem.query.filter_by(oda_id=oda_id).order_by(
            MinibarIslem.islem_tarihi.desc()
        ).first()
        
        if not son_islem:
            return {
                'oda': oda,
                'son_islem': None,
                'urunler': []
            }
        
        # Ürün detaylarını hazırla
        urunler = []
        for detay in son_islem.detaylar:
            urunler.append({
                'urun_adi': detay.urun.urun_adi,
                'baslangic_stok': detay.baslangic_stok,
                'bitis_stok': detay.bitis_stok,
                'tuketim': detay.tuketim,
                'eklenen_miktar': detay.eklenen_miktar
            })
        
        return {
            'oda': oda,
            'son_islem': son_islem,
            'urunler': urunler
        }
        
    except Exception as e:
        log_hata(e, modul='admin_minibar', extra_info={'function': 'get_oda_minibar_detay', 'oda_id': oda_id})
        return None


def get_minibar_sifirlama_ozeti():
    """
    Sıfırlama öncesi özet bilgileri getirir
    
    Returns:
        dict: {
            'toplam_oda_sayisi': int,
            'dolu_oda_sayisi': int,
            'toplam_urun_adedi': int,
            'urun_dagilimi': [
                {'urun_adi': str, 'toplam': int},
                ...
            ]
        }
    """
    try:
        from models import Oda, MinibarIslem, MinibarIslemDetay
        
        # Toplam oda sayısı
        toplam_oda = Oda.query.filter_by(aktif=True).count()
        
        # Dolu oda sayısı (en az bir minibar işlemi olan)
        dolu_oda = db.session.query(MinibarIslem.oda_id).distinct().count()
        
        # Toplam ürün adedi (tüm odalardaki bitiş stokları toplamı)
        toplam_urun = db.session.query(
            db.func.sum(MinibarIslemDetay.bitis_stok)
        ).join(MinibarIslem).scalar() or 0
        
        # Ürün dağılımı
        urun_dagilimi = db.session.query(
            Urun.urun_adi,
            db.func.sum(MinibarIslemDetay.bitis_stok).label('toplam')
        ).join(
            MinibarIslemDetay, MinibarIslemDetay.urun_id == Urun.id
        ).group_by(
            Urun.id, Urun.urun_adi
        ).order_by(
            db.desc('toplam')
        ).limit(10).all()
        
        return {
            'toplam_oda_sayisi': toplam_oda,
            'dolu_oda_sayisi': dolu_oda,
            'toplam_urun_adedi': int(toplam_urun),
            'urun_dagilimi': [
                {'urun_adi': u[0], 'toplam': int(u[1] or 0)}
                for u in urun_dagilimi
            ]
        }
        
    except Exception as e:
        log_hata(e, modul='admin_minibar', extra_info={'function': 'get_minibar_sifirlama_ozeti'})
        return {
            'toplam_oda_sayisi': 0,
            'dolu_oda_sayisi': 0,
            'toplam_urun_adedi': 0,
            'urun_dagilimi': []
        }


def sifirla_minibar_stoklari(kullanici_id):
    """
    Tüm minibar işlemlerini ve detaylarını tamamen siler (TRUNCATE)
    
    Args:
        kullanici_id (int): İşlemi yapan kullanıcı ID
    
    Returns:
        dict: {
            'success': bool,
            'etkilenen_oda_sayisi': int,
            'toplam_sifirlanan_stok': int,
            'message': str
        }
    """
    try:
        from models import MinibarIslemDetay, MinibarIslem
        from utils.audit import audit_delete
        
        # İşlem öncesi özet bilgileri al
        ozet = get_minibar_sifirlama_ozeti()
        toplam_sifirlanan = ozet['toplam_urun_adedi']
        etkilenen_oda_sayisi = ozet['dolu_oda_sayisi']
        
        # Önce tüm detayları sil (foreign key constraint için)
        silinen_detay = MinibarIslemDetay.query.delete()
        
        # Sonra tüm işlemleri sil
        silinen_islem = MinibarIslem.query.delete()
        
        # Transaction commit
        db.session.commit()
        
        # Audit log
        audit_delete(
            tablo_adi='minibar_islem',
            kayit_id=None,
            aciklama=f'Tüm minibar işlemleri silindi (TRUNCATE). {etkilenen_oda_sayisi} oda etkilendi. {silinen_islem} işlem, {silinen_detay} detay kaydı silindi.',
            eski_deger={
                'silinen_islem': silinen_islem,
                'silinen_detay': silinen_detay,
                'etkilenen_oda_sayisi': etkilenen_oda_sayisi,
                'toplam_sifirlanan': toplam_sifirlanan
            }
        )
        
        # Sistem log
        log_islem(
            islem_tipi='silme',
            modul='minibar_sifirlama',
            islem_detay={
                'kullanici_id': kullanici_id,
                'etkilenen_oda_sayisi': etkilenen_oda_sayisi,
                'silinen_islem': silinen_islem,
                'silinen_detay': silinen_detay,
                'toplam_sifirlanan_stok': toplam_sifirlanan,
                'tarih': get_kktc_now().isoformat()
            }
        )
        
        return {
            'success': True,
            'etkilenen_oda_sayisi': etkilenen_oda_sayisi,
            'toplam_sifirlanan_stok': toplam_sifirlanan,
            'message': f'Tüm minibar kayıtları başarıyla silindi. {silinen_islem} işlem, {silinen_detay} detay kaydı temizlendi.'
        }
        
    except Exception as e:
        db.session.rollback()
        log_hata(e, modul='admin_minibar', extra_info={'function': 'sifirla_minibar_stoklari', 'kullanici_id': kullanici_id})
        
        return {
            'success': False,
            'etkilenen_oda_sayisi': 0,
            'toplam_sifirlanan_stok': 0,
            'message': f'Sıfırlama işlemi başarısız: {str(e)}'
        }


def export_depo_stok_excel(stok_listesi):
    """
    Depo stok listesini Excel formatında export eder
    
    Args:
        stok_listesi (list): get_depo_stok_durumu() çıktısı
    
    Returns:
        BytesIO: Excel dosyası buffer
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Depo Stokları"
        
        # Başlık stili
        baslik_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        baslik_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        baslik_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlıklar
        headers = ['Ürün Adı', 'Grup', 'Birim', 'Mevcut Stok', 'Kritik Stok', 'Durum']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = baslik_alignment
            cell.border = border
        
        # Veri satırları
        for row_num, item in enumerate(stok_listesi, 2):
            data = [
                item['urun_adi'],
                item['grup_adi'],
                item['birim'],
                item['mevcut_stok'],
                item['kritik_stok'],
                item['badge_text']
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Kritik stok vurgulama
                if item['durum'] == 'kritik':
                    cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                    if col_num == 6:  # Durum kolonu
                        cell.font = Font(color='DC2626', bold=True)
                elif item['durum'] == 'dikkat':
                    cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                    if col_num == 6:
                        cell.font = Font(color='D97706', bold=True)
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        
        # Freeze panes (İlk satır sabit)
        ws.freeze_panes = 'A2'
        
        # BytesIO'ya kaydet
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        log_hata(e, modul='admin_minibar', extra_info={'function': 'export_depo_stok_excel'})
        return None


# ============================================
# KAT SORUMLUSU STOK YÖNETİMİ HELPER FONKSIYONLARI
# ============================================

def get_kat_sorumlusu_zimmet_stoklari(personel_id):
    """
    Kat sorumlusunun bağlı olduğu OTELİN ORTAK zimmet stoklarını getirir.
    Aynı oteldeki TÜM kat sorumluları AYNI stokları görür.
    
    Args:
        personel_id (int): Kat sorumlusu kullanıcı ID
    
    Returns:
        list: [
            {
                'otel_id': int,
                'otel_adi': str,
                'urunler': [
                    {
                        'stok_id': int,
                        'urun_id': int,
                        'urun_adi': str,
                        'grup_adi': str,
                        'birim': str,
                        'toplam_miktar': int,
                        'kullanilan': int,
                        'kalan': int,
                        'kritik_seviye': int,
                        'kullanim_yuzdesi': float,
                        'durum': 'kritik'|'dikkat'|'normal'|'stokout',
                        'badge_class': str,
                        'badge_text': str
                    }
                ]
            }
        ]
    """
    try:
        from models import OtelZimmetStok, Kullanici, Otel, Urun, UrunGrup
        
        # Personelin otel bilgisini al
        personel = Kullanici.query.get(personel_id)
        if not personel or not personel.otel_id:
            return []
        
        otel = Otel.query.get(personel.otel_id)
        if not otel:
            return []
        
        # Otelin TÜM zimmet stoklarını getir (ORTAK HAVUZ)
        stoklar = OtelZimmetStok.query.filter_by(
            otel_id=personel.otel_id
        ).join(Urun).order_by(Urun.urun_adi).all()
        
        urunler = []
        
        for stok in stoklar:
            kalan = stok.kalan_miktar
            toplam = stok.toplam_miktar
            kullanilan = stok.kullanilan_miktar
            kritik_seviye = stok.kritik_stok_seviyesi or 10
            
            # Kullanım yüzdesi
            kullanim_yuzdesi = (kullanilan / toplam * 100) if toplam > 0 else 0
            
            # Stok durumu belirleme
            if kalan == 0:
                durum = 'stokout'
                badge_class = 'bg-red-100 text-red-800 border border-red-300 dark:bg-red-900/20 dark:text-red-400'
                badge_text = 'Stokout'
            elif kalan <= kritik_seviye:
                durum = 'kritik'
                badge_class = 'bg-red-100 text-red-800 border border-red-300 dark:bg-red-900/20 dark:text-red-400'
                badge_text = 'Kritik'
            elif kalan <= kritik_seviye * 1.5:
                durum = 'dikkat'
                badge_class = 'bg-yellow-100 text-yellow-800 border border-yellow-300 dark:bg-yellow-900/20 dark:text-yellow-400'
                badge_text = 'Dikkat'
            else:
                durum = 'normal'
                badge_class = 'bg-green-100 text-green-800 border border-green-300 dark:bg-green-900/20 dark:text-green-400'
                badge_text = 'Normal'
            
            urunler.append({
                'stok_id': stok.id,
                'detay_id': stok.id,  # Eski template uyumluluğu için
                'urun_id': stok.urun_id,
                'urun_adi': stok.urun.urun_adi if stok.urun else 'Bilinmiyor',
                'grup_adi': stok.urun.grup.grup_adi if stok.urun and stok.urun.grup else 'Genel',
                'birim': stok.urun.birim if stok.urun else 'Adet',
                'teslim_edilen': toplam,
                'toplam_miktar': toplam,
                'kullanilan': kullanilan,
                'kalan': kalan,
                'kritik_seviye': kritik_seviye,
                'kullanim_yuzdesi': round(kullanim_yuzdesi, 1),
                'durum': durum,
                'badge_class': badge_class,
                'badge_text': badge_text
            })
        
        # Ürünleri duruma göre sırala (stokout > kritik > dikkat > normal)
        durum_siralama = {'stokout': 0, 'kritik': 1, 'dikkat': 2, 'normal': 3}
        urunler.sort(key=lambda x: (durum_siralama.get(x['durum'], 4), x['urun_adi']))
        
        # Tek bir otel kaydı döndür (ortak havuz)
        if urunler:
            return [{
                'otel_id': otel.id,
                'otel_adi': otel.ad,
                'zimmet_id': f"OTEL-{otel.id}",  # Eski template uyumluluğu
                'zimmet_tarihi': stoklar[0].son_guncelleme if stoklar else None,
                'teslim_eden': 'Otel Ortak Zimmet Deposu',
                'durum': 'aktif',
                'urunler': urunler
            }]
        
        return []
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok', extra_info={'function': 'get_kat_sorumlusu_zimmet_stoklari', 'personel_id': personel_id})
        return []



def get_kat_sorumlusu_kritik_stoklar(personel_id):
    """
    Kat sorumlusunun kritik seviyedeki ürünlerini getirir
    
    Args:
        personel_id (int): Kat sorumlusu kullanıcı ID
    
    Returns:
        dict: {
            'stokout': [],      # Stok sıfır olan ürünler
            'kritik': [],       # Kritik seviyenin altındaki ürünler
            'dikkat': [],       # Kritik seviyenin %100-150 arasındaki ürünler
            'risk': [],         # Kritik seviyenin %50-100 arasındaki ürünler
            'istatistik': {
                'toplam_urun': int,
                'stokout_sayisi': int,
                'kritik_sayisi': int,
                'dikkat_sayisi': int,
                'risk_sayisi': int
            }
        }
    """
    try:
        from models import PersonelZimmet, PersonelZimmetDetay
        
        # Tüm zimmet stoklarını getir
        zimmet_stoklari = get_kat_sorumlusu_zimmet_stoklari(personel_id)
        
        kategoriler = {
            'stokout': [],
            'kritik': [],
            'dikkat': [],
            'risk': []
        }
        
        # Tüm ürünleri kategorilere ayır
        for zimmet in zimmet_stoklari:
            for urun in zimmet['urunler']:
                kalan = urun['kalan']
                kritik_seviye = urun['kritik_seviye']
                
                # Kritik seviye None ise atla
                if kritik_seviye is None:
                    continue
                
                # Ürün bilgilerini zenginleştir
                urun_bilgi = {
                    **urun,
                    'zimmet_id': zimmet['zimmet_id'],
                    'zimmet_tarihi': zimmet['zimmet_tarihi'],
                    'eksik_miktar': max(0, kritik_seviye - kalan)
                }
                
                # Kategorilere ayır
                if kalan == 0:
                    kategoriler['stokout'].append(urun_bilgi)
                elif kalan <= kritik_seviye:
                    kategoriler['kritik'].append(urun_bilgi)
                elif kalan <= kritik_seviye * 1.5:
                    kategoriler['dikkat'].append(urun_bilgi)
                elif kalan <= kritik_seviye * 2:
                    # Risk kategorisi - kritik seviyenin 2 katının altında
                    kategoriler['risk'].append(urun_bilgi)
        
        # Her kategoriyi eksik miktara göre sırala
        for kategori in kategoriler:
            kategoriler[kategori].sort(key=lambda x: (-x['eksik_miktar'], x['urun_adi']))
        
        # İstatistikleri hesapla
        toplam_urun = sum(len(zimmet['urunler']) for zimmet in zimmet_stoklari)
        
        return {
            'stokout': kategoriler['stokout'],
            'kritik': kategoriler['kritik'],
            'dikkat': kategoriler['dikkat'],
            'risk': kategoriler['risk'],
            'istatistik': {
                'toplam_urun': toplam_urun,
                'stokout_sayisi': len(kategoriler['stokout']),
                'kritik_sayisi': len(kategoriler['kritik']),
                'dikkat_sayisi': len(kategoriler['dikkat']),
                'risk_sayisi': len(kategoriler['risk'])
            }
        }
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok', extra_info={'function': 'get_kat_sorumlusu_kritik_stoklar', 'personel_id': personel_id})
        return {
            'stokout': [],
            'kritik': [],
            'dikkat': [],
            'risk': [],
            'istatistik': {
                'toplam_urun': 0,
                'stokout_sayisi': 0,
                'kritik_sayisi': 0,
                'dikkat_sayisi': 0,
                'risk_sayisi': 0
            }
        }



def olustur_otomatik_siparis(personel_id, guvenlik_marji=1.5):
    """
    Kritik seviyedeki ürünler için otomatik sipariş listesi oluşturur
    
    Args:
        personel_id (int): Kat sorumlusu kullanıcı ID
        guvenlik_marji (float): Kritik seviyenin kaç katı sipariş edilsin (default: 1.5)
    
    Returns:
        dict: {
            'siparis_listesi': [
                {
                    'detay_id': int,
                    'urun_id': int,
                    'urun_adi': str,
                    'grup_adi': str,
                    'birim': str,
                    'mevcut_stok': int,
                    'kritik_seviye': int,
                    'onerilen_miktar': int,
                    'aciliyet': 'acil'|'normal'
                }
            ],
            'toplam_urun_sayisi': int,
            'toplam_miktar': int
        }
    """
    try:
        # Kritik stokları getir
        kritik_stoklar = get_kat_sorumlusu_kritik_stoklar(personel_id)
        
        siparis_listesi = []
        toplam_miktar = 0
        
        # Stokout ürünler - Acil
        for urun in kritik_stoklar['stokout']:
            onerilen_miktar = int(urun['kritik_seviye'] * guvenlik_marji)
            
            siparis_listesi.append({
                'detay_id': urun['detay_id'],
                'urun_id': urun['urun_id'],
                'urun_adi': urun['urun_adi'],
                'grup_adi': urun['grup_adi'],
                'birim': urun['birim'],
                'mevcut_stok': urun['kalan'],
                'kritik_seviye': urun['kritik_seviye'],
                'onerilen_miktar': onerilen_miktar,
                'aciliyet': 'acil'
            })
            toplam_miktar += onerilen_miktar
        
        # Kritik ürünler - Acil
        for urun in kritik_stoklar['kritik']:
            eksik = urun['kritik_seviye'] - urun['kalan']
            onerilen_miktar = int(eksik + (urun['kritik_seviye'] * (guvenlik_marji - 1)))
            
            siparis_listesi.append({
                'detay_id': urun['detay_id'],
                'urun_id': urun['urun_id'],
                'urun_adi': urun['urun_adi'],
                'grup_adi': urun['grup_adi'],
                'birim': urun['birim'],
                'mevcut_stok': urun['kalan'],
                'kritik_seviye': urun['kritik_seviye'],
                'onerilen_miktar': onerilen_miktar,
                'aciliyet': 'acil'
            })
            toplam_miktar += onerilen_miktar
        
        # Dikkat ürünler - Normal
        for urun in kritik_stoklar['dikkat']:
            eksik = int((urun['kritik_seviye'] * guvenlik_marji) - urun['kalan'])
            if eksik > 0:
                siparis_listesi.append({
                    'detay_id': urun['detay_id'],
                    'urun_id': urun['urun_id'],
                    'urun_adi': urun['urun_adi'],
                    'grup_adi': urun['grup_adi'],
                    'birim': urun['birim'],
                    'mevcut_stok': urun['kalan'],
                    'kritik_seviye': urun['kritik_seviye'],
                    'onerilen_miktar': eksik,
                    'aciliyet': 'normal'
                })
                toplam_miktar += eksik
        
        # Acil olanları başa al
        siparis_listesi.sort(key=lambda x: (0 if x['aciliyet'] == 'acil' else 1, x['urun_adi']))
        
        return {
            'siparis_listesi': siparis_listesi,
            'toplam_urun_sayisi': len(siparis_listesi),
            'toplam_miktar': toplam_miktar
        }
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok', extra_info={'function': 'olustur_otomatik_siparis', 'personel_id': personel_id})
        return {
            'siparis_listesi': [],
            'toplam_urun_sayisi': 0,
            'toplam_miktar': 0
        }



def kaydet_siparis_talebi(personel_id, siparis_listesi, aciklama=None):
    """
    Kat sorumlusu sipariş talebini veritabanına kaydeder
    
    Args:
        personel_id (int): Kat sorumlusu kullanıcı ID
        siparis_listesi (list): Sipariş edilecek ürünler [{'detay_id': urun_id, 'onerilen_miktar': miktar, 'aciliyet': 'normal/acil'}]
        aciklama (str, optional): Ek açıklama
    
    Returns:
        dict: {'success': bool, 'talep_id': int or None, 'talep_no': str, 'message': str}
    """
    try:
        from models import db, KatSorumlusuSiparisTalebi, KatSorumlusuSiparisTalepDetay
        from utils.audit import audit_create
        
        if not siparis_listesi or len(siparis_listesi) == 0:
            return {
                'success': False,
                'talep_id': None,
                'talep_no': None,
                'message': 'Sipariş listesi boş olamaz'
            }
        
        # Talep numarası oluştur: KST-YYYYMMDD-XXXX
        tarih_str = get_kktc_now().strftime('%Y%m%d')
        son_talep = KatSorumlusuSiparisTalebi.query.filter(
            KatSorumlusuSiparisTalebi.talep_no.like(f'KST-{tarih_str}-%')
        ).order_by(KatSorumlusuSiparisTalebi.id.desc()).first()
        
        if son_talep:
            son_sira = int(son_talep.talep_no.split('-')[-1])
            yeni_sira = son_sira + 1
        else:
            yeni_sira = 1
        
        talep_no = f'KST-{tarih_str}-{yeni_sira:04d}'
        
        # Kat sorumlusunun bağlı olduğu depo sorumlusunu bul
        from models import Kullanici
        kat_sorumlusu = db.session.get(Kullanici, personel_id)
        depo_sorumlusu_id = kat_sorumlusu.depo_sorumlusu_id if kat_sorumlusu else None
        
        # Ana talep kaydı oluştur
        yeni_talep = KatSorumlusuSiparisTalebi(
            talep_no=talep_no,
            kat_sorumlusu_id=personel_id,
            depo_sorumlusu_id=depo_sorumlusu_id,  # Bağlı depo sorumlusu
            durum='beklemede',
            aciklama=aciklama
        )
        db.session.add(yeni_talep)
        db.session.flush()  # ID'yi al
        
        # Detay kayıtlarını oluştur
        for urun in siparis_listesi:
            detay = KatSorumlusuSiparisTalepDetay(
                talep_id=yeni_talep.id,
                urun_id=urun['detay_id'],
                talep_miktari=urun['onerilen_miktar'],
                aciliyet=urun.get('aciliyet', 'normal')
            )
            db.session.add(detay)
        
        db.session.commit()
        
        # Audit log
        audit_create(
            tablo_adi='kat_sorumlusu_siparis_talepleri',
            kayit_id=yeni_talep.id,
            yeni_deger={
                'talep_no': talep_no,
                'urun_sayisi': len(siparis_listesi),
                'toplam_miktar': sum(u['onerilen_miktar'] for u in siparis_listesi)
            },
            aciklama=f'Kat sorumlusu sipariş talebi oluşturdu: {talep_no}'
        )
        
        return {
            'success': True,
            'talep_id': yeni_talep.id,
            'talep_no': talep_no,
            'message': f'Sipariş talebi başarıyla oluşturuldu. Talep No: {talep_no}'
        }
        
    except Exception as e:
        db.session.rollback()
        log_hata(e, modul='kat_sorumlusu_stok', extra_info={
            'function': 'kaydet_siparis_talebi',
            'personel_id': personel_id,
            'urun_sayisi': len(siparis_listesi) if siparis_listesi else 0
        })
        return {
            'success': False,
            'talep_id': None,
            'talep_no': None,
            'message': f'Sipariş kaydedilemedi: {str(e)}'
        }



def get_zimmet_urun_gecmisi(personel_id, urun_id, gun_sayisi=30):
    """
    Belirli bir ürünün kullanım geçmişini getirir
    
    Args:
        personel_id (int): Kat sorumlusu kullanıcı ID
        urun_id (int): Ürün ID
        gun_sayisi (int): Kaç günlük geçmiş (default: 30)
    
    Returns:
        dict: {
            'urun': Urun object,
            'hareketler': [
                {
                    'tarih': datetime,
                    'islem_tipi': str,
                    'oda_no': str,
                    'miktar': int,
                    'aciklama': str
                }
            ],
            'istatistik': {
                'toplam_kullanim': int,
                'gunluk_ortalama': float,
                'en_cok_kullanilan_gun': datetime or None,
                'en_az_kullanilan_gun': datetime or None
            }
        }
    """
    try:
        from models import Urun, MinibarIslem, MinibarIslemDetay, Oda
        from datetime import timedelta
        
        # Ürünü getir
        urun = Urun.query.get(urun_id)
        if not urun:
            return None
        
        # Tarih aralığı
        bugun = get_kktc_now()
        baslangic_tarihi = bugun - timedelta(days=gun_sayisi)
        
        # Minibar işlemlerinden ürün hareketlerini getir
        hareketler_query = db.session.query(
            MinibarIslem.islem_tarihi,
            MinibarIslem.islem_tipi,
            Oda.oda_no,
            MinibarIslemDetay.eklenen_miktar,
            MinibarIslem.aciklama
        ).join(
            MinibarIslemDetay, MinibarIslemDetay.islem_id == MinibarIslem.id
        ).join(
            Oda, Oda.id == MinibarIslem.oda_id
        ).filter(
            MinibarIslem.personel_id == personel_id,
            MinibarIslemDetay.urun_id == urun_id,
            MinibarIslem.islem_tarihi >= baslangic_tarihi,
            MinibarIslemDetay.eklenen_miktar > 0  # Sadece ekleme yapılan hareketler
        ).order_by(
            MinibarIslem.islem_tarihi.desc()
        ).all()
        
        hareketler = []
        toplam_kullanim = 0
        gunluk_kullanim = {}
        
        for hareket in hareketler_query:
            tarih, islem_tipi, oda_no, miktar, aciklama = hareket
            
            hareketler.append({
                'tarih': tarih,
                'islem_tipi': islem_tipi,
                'oda_no': oda_no,
                'miktar': miktar,
                'aciklama': aciklama or ''
            })
            
            toplam_kullanim += miktar
            
            # Günlük kullanım hesapla
            gun_key = tarih.date()
            gunluk_kullanim[gun_key] = gunluk_kullanim.get(gun_key, 0) + miktar
        
        # İstatistikler
        gunluk_ortalama = toplam_kullanim / gun_sayisi if gun_sayisi > 0 else 0
        
        en_cok_gun = None
        en_az_gun = None
        if gunluk_kullanim:
            en_cok_gun = max(gunluk_kullanim, key=gunluk_kullanim.get)
            en_az_gun = min(gunluk_kullanim, key=gunluk_kullanim.get)
        
        return {
            'urun': urun,
            'hareketler': hareketler,
            'istatistik': {
                'toplam_kullanim': toplam_kullanim,
                'gunluk_ortalama': round(gunluk_ortalama, 2),
                'en_cok_kullanilan_gun': en_cok_gun,
                'en_az_kullanilan_gun': en_az_gun,
                'gunluk_kullanim': gunluk_kullanim  # Grafik için
            }
        }
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok', extra_info={
            'function': 'get_zimmet_urun_gecmisi',
            'personel_id': personel_id,
            'urun_id': urun_id
        })
        return None



def guncelle_kritik_seviye(zimmet_detay_id, kritik_seviye):
    """
    Zimmet detayındaki ürün için kritik stok seviyesi günceller
    
    Args:
        zimmet_detay_id (int): PersonelZimmetDetay ID
        kritik_seviye (int): Yeni kritik seviye
    
    Returns:
        dict: {
            'success': bool,
            'message': str
        }
    """
    try:
        from models import PersonelZimmetDetay
        from utils.audit import audit_update, serialize_model
        
        # Input validasyonu
        if not isinstance(kritik_seviye, int) or kritik_seviye <= 0:
            return {
                'success': False,
                'message': 'Kritik seviye pozitif bir tam sayı olmalıdır'
            }
        
        # Zimmet detayını getir
        detay = PersonelZimmetDetay.query.get(zimmet_detay_id)
        if not detay:
            return {
                'success': False,
                'message': 'Zimmet detayı bulunamadı'
            }
        
        # Eski değeri sakla
        eski_deger = serialize_model(detay)
        eski_seviye = detay.kritik_stok_seviyesi
        
        # Güncelle
        detay.kritik_stok_seviyesi = kritik_seviye
        db.session.commit()
        
        # Audit trail
        audit_update(
            tablo_adi='personel_zimmet_detay',
            kayit_id=detay.id,
            eski_deger=eski_deger,
            yeni_deger=detay,
            aciklama=f'Kritik stok seviyesi güncellendi: {eski_seviye} -> {kritik_seviye}'
        )
        
        # Sistem log
        log_islem(
            islem_tipi='guncelleme',
            modul='kritik_seviye',
            islem_detay={
                'zimmet_detay_id': zimmet_detay_id,
                'urun_id': detay.urun_id,
                'urun_adi': detay.urun.urun_adi,
                'eski_seviye': eski_seviye,
                'yeni_seviye': kritik_seviye
            }
        )
        
        return {
            'success': True,
            'message': f'Kritik seviye başarıyla güncellendi: {kritik_seviye}'
        }
        
    except Exception as e:
        db.session.rollback()
        log_hata(e, modul='kat_sorumlusu_stok', extra_info={
            'function': 'guncelle_kritik_seviye',
            'zimmet_detay_id': zimmet_detay_id,
            'kritik_seviye': kritik_seviye
        })
        return {
            'success': False,
            'message': f'Kritik seviye güncellenemedi: {str(e)}'
        }



def export_zimmet_stok_excel(personel_id):
    """
    Kat sorumlusunun zimmet stoklarını Excel'e export eder
    
    Args:
        personel_id (int): Kat sorumlusu kullanıcı ID
    
    Returns:
        BytesIO: Excel dosyası buffer veya None
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Zimmet stoklarını getir
        zimmet_stoklari = get_kat_sorumlusu_zimmet_stoklari(personel_id)
        
        if not zimmet_stoklari:
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Zimmet Stokları"
        
        # Başlık stili
        baslik_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        baslik_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        baslik_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlıklar
        headers = ['Zimmet No', 'Zimmet Tarihi', 'Ürün Adı', 'Grup', 'Birim', 
                   'Teslim Edilen', 'Kullanılan', 'Kalan', 'Kritik Seviye', 'Durum']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = baslik_alignment
            cell.border = border
        
        # Veri satırları
        row_num = 2
        for zimmet in zimmet_stoklari:
            for urun in zimmet['urunler']:
                data = [
                    zimmet['zimmet_id'],
                    zimmet['zimmet_tarihi'].strftime('%d.%m.%Y'),
                    urun['urun_adi'],
                    urun['grup_adi'],
                    urun['birim'],
                    urun['teslim_edilen'],
                    urun['kullanilan'],
                    urun['kalan'],
                    urun['kritik_seviye'],
                    urun['badge_text']
                ]
                
                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Renk kodlaması
                    if urun['durum'] == 'stokout' or urun['durum'] == 'kritik':
                        cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                        if col_num == 10:  # Durum kolonu
                            cell.font = Font(color='DC2626', bold=True)
                    elif urun['durum'] == 'dikkat':
                        cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                        if col_num == 10:
                            cell.font = Font(color='D97706', bold=True)
                
                row_num += 1
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 12
        
        # Freeze panes (İlk satır sabit)
        ws.freeze_panes = 'A2'
        
        # BytesIO'ya kaydet
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok', extra_info={
            'function': 'export_zimmet_stok_excel',
            'personel_id': personel_id
        })
        return None
