"""
Rapor Export Routes Modülü

Bu modül depo raporları sayfası, kullanım kılavuzu ve Excel/PDF export
endpoint'lerini içerir.

Endpoint'ler:
- /depo-raporlar - Depo raporları sayfası
- /kullanim-kilavuzu/personel-zimmet - Kullanım kılavuzu sayfası
- /excel-export/<rapor_tipi> - Excel export
- /pdf-export/<rapor_tipi> - PDF export

Roller:
- depo_sorumlusu (depo-raporlar)
- login_required (diğerleri)
"""

from flask import render_template, request, make_response, flash, redirect, url_for
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from sqlalchemy import func, distinct
import pytz

from models import (
    db, Otel, Kullanici, Kat, Oda, UrunGrup, Urun, StokHareket,
    PersonelZimmet, PersonelZimmetDetay, MinibarIslem, MinibarIslemDetay
)
from utils.decorators import login_required, role_required
from utils.helpers import (
    get_current_user, get_kritik_stok_urunler, get_stok_toplamlari
)

KKTC_TZ = pytz.timezone('Europe/Nicosia')

def get_kktc_now():
    """Kıbrıs saat diliminde şu anki zamanı döndürür."""
    return datetime.now(KKTC_TZ)


def register_rapor_export_routes(app):
    """Rapor export route'larını register et"""

    @app.route('/depo-raporlar')
    @login_required
    @role_required('depo_sorumlusu')
    def depo_raporlar():
        baslangic_tarihi = request.args.get('baslangic_tarihi')
        bitis_tarihi = request.args.get('bitis_tarihi')
        rapor_tipi = request.args.get('rapor_tipi', '')
        urun_grup_id = request.args.get('urun_grup_id', '')
        urun_id = request.args.get('urun_id', '')
        personel_id = request.args.get('personel_id', '')
        hareket_tipi = request.args.get('hareket_tipi', '')
        
        rapor_verisi = None
        rapor_baslik = ""
        
        urun_gruplari = UrunGrup.query.filter_by(aktif=True).order_by(UrunGrup.grup_adi).all()
        urunler = Urun.query.filter_by(aktif=True).order_by(Urun.urun_adi).all()
        personeller = Kullanici.query.filter_by(rol='kat_sorumlusu', aktif=True).order_by(Kullanici.ad, Kullanici.soyad).all()
        
        if rapor_tipi:
            if rapor_tipi == 'stok_durum':
                rapor_baslik = "Stok Durum Raporu"
                
                query = Urun.query.filter_by(aktif=True)
                
                if urun_grup_id:
                    query = query.filter_by(grup_id=urun_grup_id)
                
                urunler_liste = query.order_by(Urun.urun_adi).all()
                stok_map = get_stok_toplamlari([urun.id for urun in urunler_liste])
                
                rapor_verisi = []
                for urun in urunler_liste:
                    mevcut_stok = stok_map.get(urun.id, 0)
                    rapor_verisi.append({
                        'urun': urun,
                        'mevcut_stok': mevcut_stok
                    })
            
            elif rapor_tipi == 'stok_hareket':
                rapor_baslik = "Stok Hareket Raporu"
                
                query = StokHareket.query
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(StokHareket.islem_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(StokHareket.islem_tarihi < bitis)
                
                if urun_id:
                    query = query.filter_by(urun_id=urun_id)
                elif urun_grup_id:
                    query = query.join(Urun).filter(Urun.grup_id == urun_grup_id)
                
                if hareket_tipi:
                    query = query.filter_by(hareket_tipi=hareket_tipi)
                
                rapor_verisi = query.order_by(StokHareket.islem_tarihi.desc()).all()
                
                for hareket in rapor_verisi:
                    if hareket.aciklama and 'Zimmet' in hareket.aciklama:
                        try:
                            if '#' in hareket.aciklama:
                                zimmet_id = int(hareket.aciklama.split('#')[1].split()[0])
                                zimmet = db.session.get(PersonelZimmet, zimmet_id)
                                if zimmet and zimmet.personel:
                                    hareket.zimmet_personel = f"{zimmet.personel.ad} {zimmet.personel.soyad}"
                                else:
                                    hareket.zimmet_personel = None
                            else:
                                hareket.zimmet_personel = None
                        except Exception:
                            hareket.zimmet_personel = None
                    else:
                        hareket.zimmet_personel = None
            
            elif rapor_tipi == 'zimmet':
                rapor_baslik = "Zimmet Raporu"
                
                query = PersonelZimmet.query
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
                
                if personel_id:
                    query = query.filter_by(personel_id=personel_id)
                
                rapor_verisi = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
            
            elif rapor_tipi == 'zimmet_detay':
                rapor_baslik = "Ürün Bazlı Zimmet Detay Raporu"
                
                query = db.session.query(
                    PersonelZimmetDetay,
                    PersonelZimmet,
                    Kullanici,
                    Urun
                ).join(
                    PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id
                ).join(
                    Kullanici, PersonelZimmet.personel_id == Kullanici.id
                ).join(
                    Urun, PersonelZimmetDetay.urun_id == Urun.id
                )
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
                
                if personel_id:
                    query = query.filter(PersonelZimmet.personel_id == personel_id)
                
                if urun_id:
                    query = query.filter(PersonelZimmetDetay.urun_id == urun_id)
                elif urun_grup_id:
                    query = query.filter(Urun.grup_id == urun_grup_id)
                
                rapor_verisi = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
            
            elif rapor_tipi == 'minibar_tuketim':
                rapor_baslik = "Minibar Tüketim Raporu"
                
                query = db.session.query(
                    Urun.urun_adi,
                    Urun.birim,
                    UrunGrup.grup_adi,
                    Oda.oda_no,
                    Kat.kat_adi,
                    MinibarIslem.islem_tarihi,
                    MinibarIslem.islem_tipi,
                    MinibarIslemDetay.tuketim,
                    Kullanici.ad,
                    Kullanici.soyad
                ).select_from(MinibarIslemDetay).join(
                    MinibarIslem, MinibarIslemDetay.islem_id == MinibarIslem.id
                ).join(
                    Urun, MinibarIslemDetay.urun_id == Urun.id
                ).join(
                    UrunGrup, Urun.grup_id == UrunGrup.id
                ).join(
                    Oda, MinibarIslem.oda_id == Oda.id
                ).join(
                    Kat, Oda.kat_id == Kat.id
                ).join(
                    Kullanici, MinibarIslem.personel_id == Kullanici.id
                ).filter(
                    MinibarIslem.islem_tipi.in_(['setup_kontrol', 'ekstra_tuketim']),
                    MinibarIslemDetay.tuketim > 0
                )
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(MinibarIslem.islem_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(MinibarIslem.islem_tarihi < bitis)
                
                if urun_id:
                    query = query.filter(MinibarIslemDetay.urun_id == urun_id)
                elif urun_grup_id:
                    query = query.filter(Urun.grup_id == urun_grup_id)
                
                if personel_id:
                    query = query.filter(MinibarIslem.personel_id == personel_id)
                
                rapor_verisi = query.order_by(MinibarIslem.islem_tarihi.desc()).all()
            
            elif rapor_tipi == 'minibar_kontrol_odalar':
                rapor_baslik = "Günlük Minibar Kontrolü Yapılan Odalar"
                
                query = db.session.query(
                    func.max(MinibarIslem.islem_tarihi).label('son_kontrol'),
                    Oda.oda_no,
                    Kat.kat_adi,
                    Kullanici.ad,
                    Kullanici.soyad,
                    func.count(distinct(MinibarIslemDetay.urun_id)).label('urun_sayisi'),
                    func.sum(MinibarIslemDetay.tuketim).label('toplam_tuketim')
                ).select_from(MinibarIslem).join(
                    Oda, MinibarIslem.oda_id == Oda.id
                ).join(
                    Kat, Oda.kat_id == Kat.id
                ).join(
                    Kullanici, MinibarIslem.personel_id == Kullanici.id
                ).join(
                    MinibarIslemDetay, MinibarIslem.id == MinibarIslemDetay.islem_id
                ).filter(
                    MinibarIslem.islem_tipi.in_(['setup_kontrol', 'ekstra_ekleme', 'ekstra_tuketim'])
                )
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(MinibarIslem.islem_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(MinibarIslem.islem_tarihi < bitis)
                
                if personel_id:
                    query = query.filter(MinibarIslem.personel_id == personel_id)
                
                query = query.group_by(
                    Oda.oda_no,
                    Kat.kat_adi,
                    Kullanici.ad,
                    Kullanici.soyad
                )
                
                rapor_verisi = query.order_by(db.desc('son_kontrol'), Oda.oda_no).all()
            
            elif rapor_tipi == 'urun_grup':
                rapor_baslik = "Ürün Grubu Bazlı Stok Raporu"
                
                query = UrunGrup.query.filter_by(aktif=True)
                aktif_urunler = Urun.query.filter_by(aktif=True).all()
                stok_map = get_stok_toplamlari([urun.id for urun in aktif_urunler])
                urunler_by_grup = {}
                for urun in aktif_urunler:
                    urunler_by_grup.setdefault(urun.grup_id, []).append(urun)
                
                rapor_verisi = []
                for grup in query.all():
                    grup_urunleri = urunler_by_grup.get(grup.id, [])
                    toplam_urun_sayisi = len(grup_urunleri)
                    kritik_urun_sayisi = 0
                    
                    for urun in grup_urunleri:
                        mevcut_stok = stok_map.get(urun.id, 0)
                        kritik_seviye = urun.kritik_stok_seviyesi or 0
                        if mevcut_stok <= kritik_seviye:
                            kritik_urun_sayisi += 1
                    
                    rapor_verisi.append({
                        'grup': grup,
                        'toplam_urun': toplam_urun_sayisi,
                        'kritik_urun': kritik_urun_sayisi
                    })
            
            elif rapor_tipi == 'ozet':
                rapor_baslik = "Genel Sistem Özet Raporu"
                
                toplam_urun = Urun.query.filter_by(aktif=True).count()
                
                kritik_urunler = get_kritik_stok_urunler()
                
                aktif_zimmet = PersonelZimmet.query.filter_by(durum='aktif').count()
                
                bugun = get_kktc_now().date()
                bugun_baslangic = datetime.combine(bugun, datetime.min.time())
                bugun_bitis = datetime.combine(bugun, datetime.max.time())
                
                bugun_giris = StokHareket.query.filter(
                    StokHareket.hareket_tipi == 'giris',
                    StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis)
                ).count()
                
                bugun_cikis = StokHareket.query.filter(
                    StokHareket.hareket_tipi == 'cikis',
                    StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis)
                ).count()
                
                ay_baslangic = get_kktc_now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                ay_zimmet = PersonelZimmet.query.filter(PersonelZimmet.zimmet_tarihi >= ay_baslangic).count()
                
                rapor_verisi = {
                    'toplam_urun': toplam_urun,
                    'kritik_urun': len(kritik_urunler),
                    'aktif_zimmet': aktif_zimmet,
                    'bugun_giris': bugun_giris,
                    'bugun_cikis': bugun_cikis,
                    'ay_zimmet': ay_zimmet
                }
        
        return render_template('depo_sorumlusu/raporlar.html',
                             rapor_verisi=rapor_verisi,
                             rapor_baslik=rapor_baslik,
                             rapor_tipi=rapor_tipi,
                             urun_gruplari=urun_gruplari,
                             urunler=urunler,
                             personeller=personeller)

    @app.route('/kullanim-kilavuzu/personel-zimmet')
    @login_required
    def kullanim_kilavuzu_personel_zimmet():
        """Personel zimmet kullanım kılavuzu sayfası"""
        return render_template('kullanim_kilavuzu/personel_zimmet_kilavuzu.html')

    @app.route('/excel-export/<rapor_tipi>')
    @login_required
    def excel_export(rapor_tipi):
        try:
            baslangic_tarihi = request.args.get('baslangic_tarihi')
            bitis_tarihi = request.args.get('bitis_tarihi')
            urun_grup_id = request.args.get('urun_grup_id')
            urun_id = request.args.get('urun_id')
            personel_id = request.args.get('personel_id')
            hareket_tipi = request.args.get('hareket_tipi')
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            rapor_basliklari = {
                'stok_durum': 'Stok Durum Raporu',
                'stok_hareket': 'Stok Hareket Raporu',
                'zimmet': 'Zimmet Raporu',
                'zimmet_detay': 'Ürün Bazlı Zimmet Detay Raporu',
                'urun_grup': 'Ürün Grubu Bazlı Stok Raporu',
                'ozet': 'Genel Sistem Özet Raporu'
            }
            
            baslik = rapor_basliklari.get(rapor_tipi, 'Rapor')
            ws.title = baslik[:31]
            
            ws['A1'] = baslik
            ws['A1'].font = Font(size=16, bold=True)
            
            ws['A2'] = f"Rapor Tarihi: {get_kktc_now().strftime('%d.%m.%Y %H:%M')}"
            ws['A2'].font = Font(size=10)
            
            row_num = 4
            
            if rapor_tipi == 'stok_durum':
                headers = ['Ürün Adı', 'Ürün Grubu', 'Birim', 'Mevcut Stok', 'Kritik Seviye', 'Durum']
                ws.merge_cells('A1:F1')
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                query = Urun.query.filter_by(aktif=True)
                if urun_grup_id:
                    query = query.filter_by(grup_id=urun_grup_id)
                
                urunler_liste = query.order_by(Urun.urun_adi).all()
                stok_map = get_stok_toplamlari([urun.id for urun in urunler_liste])
                
                for urun in urunler_liste:
                    row_num += 1
                    mevcut_stok = stok_map.get(urun.id, 0)
                    kritik_seviye = urun.kritik_stok_seviyesi or 0
                    durum = 'KRİTİK' if mevcut_stok <= kritik_seviye else 'NORMAL'
                    
                    ws.cell(row=row_num, column=1, value=urun.urun_adi)
                    ws.cell(row=row_num, column=2, value=urun.grup.grup_adi)
                    ws.cell(row=row_num, column=3, value=urun.birim)
                    ws.cell(row=row_num, column=4, value=mevcut_stok)
                    ws.cell(row=row_num, column=5, value=urun.kritik_stok_seviyesi)
                    ws.cell(row=row_num, column=6, value=durum)
            
            elif rapor_tipi == 'stok_hareket':
                headers = ['Tarih', 'Ürün Adı', 'Hareket Tipi', 'Miktar', 'Açıklama', 'İşlem Yapan']
                ws.merge_cells('A1:F1')
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                
                query = StokHareket.query
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(StokHareket.islem_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(StokHareket.islem_tarihi < bitis)
                
                if urun_id:
                    query = query.filter_by(urun_id=urun_id)
                elif urun_grup_id:
                    query = query.join(Urun).filter(Urun.grup_id == urun_grup_id)
                
                if hareket_tipi:
                    query = query.filter_by(hareket_tipi=hareket_tipi)
                
                hareketler = query.order_by(StokHareket.islem_tarihi.desc()).all()
                
                for hareket in hareketler:
                    row_num += 1
                    islem_yapan = f"{hareket.islem_yapan.ad} {hareket.islem_yapan.soyad}" if hareket.islem_yapan else '-'
                    
                    aciklama = hareket.aciklama or '-'
                    if hareket.aciklama and 'Zimmet' in hareket.aciklama:
                        try:
                            if '#' in hareket.aciklama:
                                zimmet_id = int(hareket.aciklama.split('#')[1].split()[0])
                                zimmet = db.session.get(PersonelZimmet, zimmet_id)
                                if zimmet and zimmet.personel:
                                    aciklama += f" → {zimmet.personel.ad} {zimmet.personel.soyad}"
                        except Exception:
                            pass
                    
                    ws.cell(row=row_num, column=1, value=hareket.islem_tarihi.strftime('%d.%m.%Y %H:%M'))
                    ws.cell(row=row_num, column=2, value=hareket.urun.urun_adi)
                    ws.cell(row=row_num, column=3, value=hareket.hareket_tipi.upper())
                    ws.cell(row=row_num, column=4, value=hareket.miktar)
                    ws.cell(row=row_num, column=5, value=aciklama)
                    ws.cell(row=row_num, column=6, value=islem_yapan)
            
            elif rapor_tipi == 'zimmet':
                headers = ['Zimmet No', 'Personel', 'Zimmet Tarihi', 'Ürün Sayısı', 'Toplam Miktar', 'Durum']
                ws.merge_cells('A1:F1')
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                
                query = PersonelZimmet.query
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
                
                if personel_id:
                    query = query.filter_by(personel_id=personel_id)
                
                zimmetler = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
                
                for zimmet in zimmetler:
                    row_num += 1
                    toplam_miktar = sum(d.miktar for d in zimmet.detaylar)
                    
                    ws.cell(row=row_num, column=1, value=f"#{zimmet.id}")
                    ws.cell(row=row_num, column=2, value=f"{zimmet.personel.ad} {zimmet.personel.soyad}")
                    ws.cell(row=row_num, column=3, value=zimmet.zimmet_tarihi.strftime('%d.%m.%Y %H:%M'))
                    ws.cell(row=row_num, column=4, value=len(zimmet.detaylar))
                    ws.cell(row=row_num, column=5, value=toplam_miktar)
                    ws.cell(row=row_num, column=6, value=zimmet.durum.upper())
            
            elif rapor_tipi == 'zimmet_detay':
                headers = ['Zimmet No', 'Personel', 'Zimmet Tarihi', 'Ürün Adı', 'Grup', 'Miktar', 'Durum']
                ws.merge_cells('A1:G1')
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
                
                query = db.session.query(
                    PersonelZimmetDetay,
                    PersonelZimmet,
                    Kullanici,
                    Urun
                ).join(
                    PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id
                ).join(
                    Kullanici, PersonelZimmet.personel_id == Kullanici.id
                ).join(
                    Urun, PersonelZimmetDetay.urun_id == Urun.id
                )
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
                
                if personel_id:
                    query = query.filter(PersonelZimmet.personel_id == personel_id)
                
                if urun_id:
                    query = query.filter(PersonelZimmetDetay.urun_id == urun_id)
                elif urun_grup_id:
                    query = query.filter(Urun.grup_id == urun_grup_id)
                
                detaylar = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
                
                for detay, zimmet, kullanici, urun in detaylar:
                    row_num += 1
                    
                    ws.cell(row=row_num, column=1, value=f"#{zimmet.id}")
                    ws.cell(row=row_num, column=2, value=f"{kullanici.ad} {kullanici.soyad}")
                    ws.cell(row=row_num, column=3, value=zimmet.zimmet_tarihi.strftime('%d.%m.%Y %H:%M'))
                    ws.cell(row=row_num, column=4, value=urun.urun_adi)
                    ws.cell(row=row_num, column=5, value=urun.grup.grup_adi)
                    ws.cell(row=row_num, column=6, value=detay.miktar)
                    ws.cell(row=row_num, column=7, value=zimmet.durum.upper())
            
            elif rapor_tipi == 'urun_grup':
                headers = ['Ürün Grubu', 'Toplam Ürün', 'Kritik Stoklu Ürün']
                ws.merge_cells('A1:C1')
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                
                gruplar = UrunGrup.query.filter_by(aktif=True).all()
                aktif_urunler = Urun.query.filter_by(aktif=True).all()
                stok_map = get_stok_toplamlari([urun.id for urun in aktif_urunler])
                urun_gruplari_map = {}
                for urun in aktif_urunler:
                    urun_gruplari_map.setdefault(urun.grup_id, []).append(urun)
                
                for grup in gruplar:
                    row_num += 1
                    grup_urunleri = urun_gruplari_map.get(grup.id, [])
                    toplam_urun_sayisi = len(grup_urunleri)
                    kritik_urun_sayisi = 0
                    
                    for urun in grup_urunleri:
                        mevcut_stok = stok_map.get(urun.id, 0)
                        kritik_seviye = urun.kritik_stok_seviyesi or 0
                        if mevcut_stok <= kritik_seviye:
                            kritik_urun_sayisi += 1
                    
                    ws.cell(row=row_num, column=1, value=grup.grup_adi)
                    ws.cell(row=row_num, column=2, value=toplam_urun_sayisi)
                    ws.cell(row=row_num, column=3, value=kritik_urun_sayisi)
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename={rapor_tipi}_raporu_{get_kktc_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return response
            
        except Exception as e:
            flash(f'Excel export hatası: {str(e)}', 'danger')
            return redirect(url_for('depo_raporlar'))

    @app.route('/pdf-export/<rapor_tipi>')
    @login_required
    def pdf_export(rapor_tipi):
        try:
            def turkce_ascii(text):
                """Türkçe karakterleri ASCII'ye dönüştür"""
                if not text:
                    return ''
                char_map = {
                    'ç': 'c', 'Ç': 'C',
                    'ğ': 'g', 'Ğ': 'G',
                    'ı': 'i', 'İ': 'I',
                    'ö': 'o', 'Ö': 'O',
                    'ş': 's', 'Ş': 'S',
                    'ü': 'u', 'Ü': 'U'
                }
                result = str(text)
                for turkish, ascii_char in char_map.items():
                    result = result.replace(turkish, ascii_char)
                return result
            
            baslangic_tarihi = request.args.get('baslangic_tarihi')
            bitis_tarihi = request.args.get('bitis_tarihi')
            urun_grup_id = request.args.get('urun_grup_id')
            urun_id = request.args.get('urun_id')
            personel_id = request.args.get('personel_id')
            hareket_tipi = request.args.get('hareket_tipi')
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
            styles = getSampleStyleSheet()
            story = []
            
            rapor_basliklari = {
                'stok_durum': 'Stok Durum Raporu',
                'stok_hareket': 'Stok Hareket Raporu',
                'zimmet': 'Zimmet Raporu',
                'zimmet_detay': 'Urun Bazli Zimmet Detay Raporu',
                'urun_grup': 'Urun Grubu Bazli Stok Raporu',
                'ozet': 'Genel Sistem Ozet Raporu'
            }
            
            baslik = turkce_ascii(rapor_basliklari.get(rapor_tipi, 'Rapor'))
            
            title = Paragraph(baslik, styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            date_text = f"Rapor Tarihi: {get_kktc_now().strftime('%d.%m.%Y %H:%M')}"
            date_para = Paragraph(date_text, styles['Normal'])
            story.append(date_para)
            story.append(Spacer(1, 20))
            
            data = []
            
            if rapor_tipi == 'stok_durum':
                data.append([turkce_ascii(h) for h in ['Urun Adi', 'Urun Grubu', 'Birim', 'Mevcut Stok', 'Kritik Seviye', 'Durum']])
                
                query = Urun.query.filter_by(aktif=True)
                if urun_grup_id:
                    query = query.filter_by(grup_id=urun_grup_id)
                
                urunler_liste = query.order_by(Urun.urun_adi).all()
                stok_map = get_stok_toplamlari([urun.id for urun in urunler_liste]) if urunler_liste else {}
                
                for urun in urunler_liste:
                    mevcut_stok = stok_map.get(urun.id, 0)
                    kritik_seviye = urun.kritik_stok_seviyesi or 0
                    durum = 'KRITIK' if mevcut_stok <= kritik_seviye else 'NORMAL'
                    
                    data.append([
                        turkce_ascii(urun.urun_adi),
                        turkce_ascii(urun.grup.grup_adi),
                        turkce_ascii(urun.birim),
                        str(mevcut_stok),
                        str(urun.kritik_stok_seviyesi),
                        durum
                    ])
            
            elif rapor_tipi == 'stok_hareket':
                data.append([turkce_ascii(h) for h in ['Tarih', 'Urun Adi', 'Hareket', 'Miktar', 'Aciklama']])
                
                query = StokHareket.query
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(StokHareket.islem_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(StokHareket.islem_tarihi < bitis)
                
                if urun_id:
                    query = query.filter_by(urun_id=urun_id)
                elif urun_grup_id:
                    query = query.join(Urun).filter(Urun.grup_id == urun_grup_id)
                
                if hareket_tipi:
                    query = query.filter_by(hareket_tipi=hareket_tipi)
                
                hareketler = query.order_by(StokHareket.islem_tarihi.desc()).limit(100).all()
                
                for hareket in hareketler:
                    aciklama = hareket.aciklama or '-'
                    if hareket.aciklama and 'Zimmet' in hareket.aciklama:
                        try:
                            if '#' in hareket.aciklama:
                                zimmet_id = int(hareket.aciklama.split('#')[1].split()[0])
                                zimmet = db.session.get(PersonelZimmet, zimmet_id)
                                if zimmet and zimmet.personel:
                                    aciklama = f"{aciklama} → {zimmet.personel.ad} {zimmet.personel.soyad}"
                        except Exception:
                            pass
                    
                    aciklama_kisaltilmis = aciklama[:50] if len(aciklama) > 50 else aciklama
                    
                    data.append([
                        hareket.islem_tarihi.strftime('%d.%m.%Y %H:%M'),
                        turkce_ascii(hareket.urun.urun_adi),
                        turkce_ascii(hareket.hareket_tipi.upper()),
                        str(hareket.miktar),
                        turkce_ascii(aciklama_kisaltilmis)
                    ])
            
            elif rapor_tipi == 'zimmet':
                data.append([turkce_ascii(h) for h in ['Zimmet No', 'Personel', 'Tarih', 'Urun Sayisi', 'Toplam', 'Durum']])
                
                query = PersonelZimmet.query
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
                
                if personel_id:
                    query = query.filter_by(personel_id=personel_id)
                
                zimmetler = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).limit(100).all()
                
                for zimmet in zimmetler:
                    toplam_miktar = sum(d.miktar for d in zimmet.detaylar)
                    
                    data.append([
                        f"#{zimmet.id}",
                        turkce_ascii(f"{zimmet.personel.ad} {zimmet.personel.soyad}"),
                        zimmet.zimmet_tarihi.strftime('%d.%m.%Y'),
                        str(len(zimmet.detaylar)),
                        str(toplam_miktar),
                        turkce_ascii(zimmet.durum.upper())
                    ])
            
            elif rapor_tipi == 'zimmet_detay':
                data.append([turkce_ascii(h) for h in ['Zimmet', 'Personel', 'Urun', 'Grup', 'Miktar', 'Durum']])
                
                query = db.session.query(
                    PersonelZimmetDetay,
                    PersonelZimmet,
                    Kullanici,
                    Urun
                ).join(
                    PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id
                ).join(
                    Kullanici, PersonelZimmet.personel_id == Kullanici.id
                ).join(
                    Urun, PersonelZimmetDetay.urun_id == Urun.id
                )
                
                if baslangic_tarihi:
                    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                    query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
                
                if bitis_tarihi:
                    bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
                
                if personel_id:
                    query = query.filter(PersonelZimmet.personel_id == personel_id)
                
                if urun_id:
                    query = query.filter(PersonelZimmetDetay.urun_id == urun_id)
                elif urun_grup_id:
                    query = query.filter(Urun.grup_id == urun_grup_id)
                
                detaylar = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).limit(100).all()
                
                for detay, zimmet, kullanici, urun in detaylar:
                    data.append([
                        f"#{zimmet.id}",
                        turkce_ascii(f"{kullanici.ad} {kullanici.soyad}"),
                        turkce_ascii(urun.urun_adi),
                        turkce_ascii(urun.grup.grup_adi),
                        str(detay.miktar),
                        turkce_ascii(zimmet.durum.upper())
                    ])
            
            elif rapor_tipi == 'urun_grup':
                data.append([turkce_ascii(h) for h in ['Urun Grubu', 'Toplam Urun', 'Kritik Stoklu Urun']])
                
                gruplar = UrunGrup.query.filter_by(aktif=True).all()
                aktif_urunler = Urun.query.filter_by(aktif=True).all()
                stok_map = get_stok_toplamlari([urun.id for urun in aktif_urunler]) if aktif_urunler else {}

                grup_urunleri_map = {}
                for urun in aktif_urunler:
                    grup_urunleri_map.setdefault(urun.grup_id, []).append(urun)

                for grup in gruplar:
                    grup_urunleri = grup_urunleri_map.get(grup.id, [])
                    toplam_urun_sayisi = len(grup_urunleri)
                    kritik_urun_sayisi = 0
                    
                    for urun in grup_urunleri:
                        mevcut_stok = stok_map.get(urun.id, 0)
                        kritik_seviye = urun.kritik_stok_seviyesi or 0
                        if mevcut_stok <= kritik_seviye:
                            kritik_urun_sayisi += 1
                    
                    data.append([
                        turkce_ascii(grup.grup_adi),
                        str(toplam_urun_sayisi),
                        str(kritik_urun_sayisi)
                    ])
            
            elif rapor_tipi == 'ozet':
                toplam_urun = Urun.query.filter_by(aktif=True).count()
                kritik_urunler = get_kritik_stok_urunler()
                aktif_zimmet = PersonelZimmet.query.filter_by(durum='aktif').count()
                
                bugun = get_kktc_now().date()
                bugun_baslangic = datetime.combine(bugun, datetime.min.time())
                bugun_bitis = datetime.combine(bugun, datetime.max.time())
                
                bugun_giris = StokHareket.query.filter(
                    StokHareket.hareket_tipi == 'giris',
                    StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis)
                ).count()
                
                bugun_cikis = StokHareket.query.filter(
                    StokHareket.hareket_tipi == 'cikis',
                    StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis)
                ).count()
                
                ay_baslangic = get_kktc_now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                ay_zimmet = PersonelZimmet.query.filter(PersonelZimmet.zimmet_tarihi >= ay_baslangic).count()
                
                data = [
                    ['Metrik', 'Deger'],
                    [turkce_ascii('Toplam Urun Sayisi'), str(toplam_urun)],
                    [turkce_ascii('Kritik Stoklu Urun'), str(len(kritik_urunler))],
                    [turkce_ascii('Aktif Zimmet'), str(aktif_zimmet)],
                    [turkce_ascii('Bugun Stok Giris'), str(bugun_giris)],
                    [turkce_ascii('Bugun Stok Cikis'), str(bugun_cikis)],
                    [turkce_ascii('Bu Ay Zimmet'), str(ay_zimmet)]
                ]
            
            if data:
                table = Table(data, repeatRows=1)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
                ]))
                
                story.append(table)
            else:
                no_data = Paragraph(turkce_ascii("Bu filtre kriterleri icin veri bulunamadi."), styles['Normal'])
                story.append(no_data)
            
            doc.build(story)
            buffer.seek(0)
            
            response = make_response(buffer.getvalue())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename={rapor_tipi}_raporu_{get_kktc_now().strftime("%Y%m%d_%H%M%S")}.pdf'
            
            return response
            
        except Exception as e:
            flash(f'PDF export hatasi: {str(e)}', 'danger')
            return redirect(url_for('depo_raporlar'))
