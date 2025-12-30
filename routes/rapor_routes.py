"""
Rapor Routes Modülü

Bu modül tüm rapor endpoint'lerini içerir.
Excel formatında indirilebilir raporlar.

Endpoint'ler:
- /raporlar/doluluk - Doluluk raporları
- /raporlar/stok - Stok raporları
- /raporlar/minibar - Minibar raporları
- /raporlar/zimmet - Zimmet raporları
- /raporlar/gelir - Gelir raporları
- /raporlar/performans - Performans raporları

Roller:
- sistem_yoneticisi
- admin
"""

from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for, jsonify
from datetime import datetime, timedelta, date
from io import BytesIO
from sqlalchemy import func, and_, or_
from models import db, Otel, Kat, Oda, Urun, StokHareket, MinibarIslem, MinibarIslemDetay, PersonelZimmet, PersonelZimmetDetay, Kullanici
from utils.decorators import login_required, role_required
from utils.helpers import log_islem, log_hata
from utils.authorization import get_kullanici_otelleri
import pytz

# KKTC Timezone
KKTC_TZ = pytz.timezone('Europe/Nicosia')

def get_kktc_now():
    """Kıbrıs saat diliminde şu anki zamanı döndürür."""
    return datetime.now(KKTC_TZ)

# Blueprint oluştur
raporlar_bp = Blueprint('raporlar', __name__, url_prefix='/raporlar')


def register_rapor_routes(app):
    """Rapor route'larını register et"""
    app.register_blueprint(raporlar_bp)


@raporlar_bp.route('/doluluk-raporlari')
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def doluluk_raporlari():
    """Doluluk raporları sayfası"""
    try:
        oteller = get_kullanici_otelleri()
        
        # URL'den otel_id parametresini al
        otel_id = request.args.get('otel_id', type=int)
        
        return render_template('raporlar/doluluk_raporlari.html', 
                             oteller=oteller,
                             secili_otel_id=otel_id)
    except Exception as e:
        log_hata(e, modul='doluluk_raporlari')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('sistem_yoneticisi_dashboard'))


@raporlar_bp.route('/doluluk-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def doluluk_raporu_olustur():
    """Doluluk raporu oluştur ve önizle"""
    try:
        from utils.occupancy_service import OccupancyService
        
        otel_id = request.form.get('otel_id', type=int)
        baslangic = request.form.get('baslangic')
        bitis = request.form.get('bitis')
        
        if not all([otel_id, baslangic, bitis]):
            flash('Lütfen tüm alanları doldurun.', 'warning')
            return redirect(url_for('raporlar.doluluk_raporlari'))
        
        baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d').date()
        bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d').date()
        
        otel = Otel.query.get(otel_id)
        if not otel:
            flash('Otel bulunamadı.', 'danger')
            return redirect(url_for('raporlar.doluluk_raporlari'))
        
        # Rapor verilerini hazırla
        gun_adlari = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
        gunler = []
        current_date = baslangic_tarih
        toplam_doluluk = 0
        max_doluluk = 0
        min_doluluk = 100
        
        while current_date <= bitis_tarih:
            rapor = OccupancyService.get_gunluk_doluluk_raporu(current_date, otel_id)
            doluluk_orani = round((rapor['dolu_oda'] / rapor['toplam_oda'] * 100) if rapor['toplam_oda'] > 0 else 0, 1)
            
            gunler.append({
                'tarih': current_date.strftime('%d.%m.%Y'),
                'gun_adi': gun_adlari[current_date.weekday()],
                'toplam_oda': rapor['toplam_oda'],
                'dolu_oda': rapor['dolu_oda'],
                'bos_oda': rapor['bos_oda'],
                'doluluk_orani': doluluk_orani
            })
            
            toplam_doluluk += doluluk_orani
            max_doluluk = max(max_doluluk, doluluk_orani)
            min_doluluk = min(min_doluluk, doluluk_orani)
            
            current_date += timedelta(days=1)
        
        gun_sayisi = len(gunler)
        ortalama_doluluk = round(toplam_doluluk / gun_sayisi, 1) if gun_sayisi > 0 else 0
        
        rapor_verisi = {
            'otel_id': otel_id,
            'otel_ad': otel.ad,
            'otel_logo': otel.logo,
            'baslangic': baslangic,
            'bitis': bitis,
            'gun_sayisi': gun_sayisi,
            'ortalama_doluluk': ortalama_doluluk,
            'max_doluluk': max_doluluk,
            'min_doluluk': min_doluluk,
            'toplam_oda': gunler[0]['toplam_oda'] if gunler else 0,
            'gunler': gunler
        }
        
        oteller = get_kullanici_otelleri()
        
        log_islem('view', 'doluluk_raporu', {
            'otel_id': otel_id,
            'baslangic': baslangic,
            'bitis': bitis
        })
        
        return render_template('raporlar/doluluk_raporlari.html', 
                             oteller=oteller, 
                             rapor_verisi=rapor_verisi)
        
    except Exception as e:
        log_hata(e, modul='doluluk_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.doluluk_raporlari'))


@raporlar_bp.route('/stok-raporlari')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def stok_raporlari():
    """Stok raporları sayfası"""
    try:
        oteller = get_kullanici_otelleri()
        return render_template('raporlar/stok_raporlari.html', oteller=oteller)
    except Exception as e:
        log_hata(e, modul='stok_raporlari')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('sistem_yoneticisi_dashboard'))


@raporlar_bp.route('/minibar-raporlari')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def minibar_raporlari():
    """Minibar raporları sayfası"""
    try:
        oteller = get_kullanici_otelleri()
        return render_template('raporlar/minibar_raporlari.html', oteller=oteller)
    except Exception as e:
        log_hata(e, modul='minibar_raporlari')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('sistem_yoneticisi_dashboard'))


@raporlar_bp.route('/zimmet-raporlari')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def zimmet_raporlari():
    """Zimmet raporları sayfası"""
    try:
        oteller = get_kullanici_otelleri()
        personeller = Kullanici.query.filter(
            Kullanici.rol.in_(['kat_sorumlusu', 'depo_sorumlusu']),
            Kullanici.aktif.is_(True)
        ).order_by(Kullanici.ad, Kullanici.soyad).all()
        return render_template('raporlar/zimmet_raporlari.html', oteller=oteller, personeller=personeller)
    except Exception as e:
        log_hata(e, modul='zimmet_raporlari')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('sistem_yoneticisi_dashboard'))


@raporlar_bp.route('/performans-raporlari')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def performans_raporlari():
    """Performans raporları sayfası"""
    try:
        oteller = get_kullanici_otelleri()
        personeller = Kullanici.query.filter(
            Kullanici.rol == 'kat_sorumlusu',
            Kullanici.aktif.is_(True)
        ).order_by(Kullanici.ad, Kullanici.soyad).all()
        return render_template('raporlar/performans_raporlari.html', oteller=oteller, personeller=personeller)
    except Exception as e:
        log_hata(e, modul='performans_raporlari')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('sistem_yoneticisi_dashboard'))


# ============================================================================
# EXCEL EXPORT FONKSİYONLARI
# ============================================================================

@raporlar_bp.route('/doluluk-excel')
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def doluluk_excel():
    """Doluluk raporu Excel export"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from utils.occupancy_service import OccupancyService
        
        otel_id = request.args.get('otel_id', type=int)
        baslangic = request.args.get('baslangic')
        bitis = request.args.get('bitis')
        
        if not all([otel_id, baslangic, bitis]):
            flash('Lütfen tüm alanları doldurun.', 'warning')
            return redirect(url_for('raporlar.doluluk_raporlari'))
        
        baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d').date()
        bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d').date()
        
        otel = Otel.query.get(otel_id)
        if not otel:
            flash('Otel bulunamadı.', 'danger')
            return redirect(url_for('raporlar.doluluk_raporlari'))
        
        # Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Doluluk Raporu"
        
        # Başlık
        ws.merge_cells('A1:F1')
        ws['A1'] = f'{otel.ad} - Doluluk Raporu'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f'{baslangic_tarih.strftime("%d.%m.%Y")} - {bitis_tarih.strftime("%d.%m.%Y")}'
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Tablo başlıkları
        headers = ['Tarih', 'Gün', 'Toplam Oda', 'Dolu Oda', 'Boş Oda', 'Doluluk %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Veriler
        gun_adlari = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
        row = 5
        current_date = baslangic_tarih
        
        while current_date <= bitis_tarih:
            rapor = OccupancyService.get_gunluk_doluluk_raporu(current_date, otel_id)
            
            ws.cell(row=row, column=1, value=current_date.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=2, value=gun_adlari[current_date.weekday()])
            ws.cell(row=row, column=3, value=rapor['toplam_oda'])
            ws.cell(row=row, column=4, value=rapor['dolu_oda'])
            ws.cell(row=row, column=5, value=rapor['bos_oda'])
            
            doluluk_orani = round((rapor['dolu_oda'] / rapor['toplam_oda'] * 100) if rapor['toplam_oda'] > 0 else 0, 1)
            ws.cell(row=row, column=6, value=f"{doluluk_orani}%")
            
            current_date += timedelta(days=1)
            row += 1
        
        # Sütun genişlikleri
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 12
        
        # Excel'i memory'ye kaydet
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f'doluluk_raporu_{otel.ad}_{baslangic}_{bitis}.xlsx'
        
        log_islem('export', 'doluluk_raporu', {
            'otel_id': otel_id,
            'baslangic': baslangic,
            'bitis': bitis
        })
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        log_hata(e, modul='doluluk_excel')
        flash('Excel dosyası oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.doluluk_raporlari'))


@raporlar_bp.route('/mevcut-stok-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin')
def mevcut_stok_raporu_olustur():
    """Mevcut stok raporu oluştur ve önizle"""
    try:
        otel_id = request.form.get('otel_id', type=int)
        
        # Veriler - Kat sorumlusunun zimmetindeki kalan stoklar
        query = db.session.query(
            Otel.ad.label('otel_ad'),
            Urun.urun_adi.label('urun_ad'),
            Urun.barkod,
            Urun.birim,
            Kullanici.ad.label('personel_ad'),
            Kullanici.soyad.label('personel_soyad'),
            func.sum(PersonelZimmetDetay.kalan_miktar).label('toplam_stok')
        ).join(PersonelZimmetDetay, Urun.id == PersonelZimmetDetay.urun_id)\
         .join(PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id)\
         .join(Kullanici, PersonelZimmet.personel_id == Kullanici.id)\
         .join(Otel, Kullanici.otel_id == Otel.id)\
         .filter(PersonelZimmet.durum == 'aktif')
        
        if otel_id:
            query = query.filter(Kullanici.otel_id == otel_id)
        
        urunler_data = query.group_by(
            Otel.ad, Urun.urun_adi, Urun.barkod, Urun.birim, 
            Kullanici.ad, Kullanici.soyad
        ).order_by(Otel.ad, Urun.urun_adi).all()
        
        urunler = []
        for u in urunler_data:
            # Durum belirleme
            stok = u.toplam_stok or 0
            if stok <= 0:
                durum = 'Tükendi'
            elif stok < 10:
                durum = 'Kritik'
            elif stok < 50:
                durum = 'Düşük'
            else:
                durum = 'Normal'
            
            urunler.append({
                'otel': u.otel_ad,
                'urun_ad': u.urun_ad,
                'barkod': u.barkod or '-',
                'birim': u.birim,
                'personel': f'{u.personel_ad} {u.personel_soyad}',
                'stok': stok,
                'durum': durum
            })
        
        # Otel adını al
        otel_ad = None
        if otel_id:
            otel = Otel.query.get(otel_id)
            if otel:
                otel_ad = otel.ad
        
        rapor_verisi = {
            'otel_id': otel_id,
            'otel_ad': otel_ad or 'Tüm Oteller',
            'toplam_urun': len(urunler),
            'urunler': urunler
        }
        
        oteller = get_kullanici_otelleri()
        
        log_islem('view', 'mevcut_stok_raporu', {'otel_id': otel_id})
        
        return render_template('raporlar/stok_raporlari.html',
                             oteller=oteller,
                             mevcut_stok_raporu=rapor_verisi)
        
    except Exception as e:
        log_hata(e, modul='mevcut_stok_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.stok_raporlari'))


@raporlar_bp.route('/stok-hareketleri-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin')
def stok_hareketleri_raporu_olustur():
    """Stok hareketleri raporu oluştur ve önizle"""
    try:
        otel_id = request.form.get('otel_id', type=int)
        baslangic = request.form.get('baslangic')
        bitis = request.form.get('bitis')
        
        if not all([baslangic, bitis]):
            flash('Lütfen tarih aralığı seçin.', 'warning')
            return redirect(url_for('raporlar.stok_raporlari'))
        
        baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d')
        bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d')
        
        # Veriler
        query = db.session.query(
            StokHareket.islem_tarihi,
            Otel.ad.label('otel_ad'),
            Urun.urun_adi.label('urun_ad'),
            StokHareket.hareket_tipi,
            StokHareket.miktar,
            Kullanici.ad.label('kullanici_ad'),
            Kullanici.soyad.label('kullanici_soyad')
        ).join(Urun, StokHareket.urun_id == Urun.id)\
         .join(Kullanici, StokHareket.islem_yapan_id == Kullanici.id)\
         .join(Otel, Kullanici.otel_id == Otel.id)\
         .filter(
            StokHareket.islem_tarihi >= baslangic_tarih,
            StokHareket.islem_tarihi <= bitis_tarih
        )
        
        if otel_id:
            query = query.filter(Kullanici.otel_id == otel_id)
        
        hareketler_data = query.order_by(StokHareket.islem_tarihi.desc()).all()
        
        hareketler = []
        for h in hareketler_data:
            hareketler.append({
                'tarih': h.islem_tarihi.strftime('%d.%m.%Y %H:%M'),
                'otel': h.otel_ad,
                'urun': h.urun_ad,
                'islem_tipi': h.hareket_tipi.upper(),
                'miktar': h.miktar,
                'kullanici': f'{h.kullanici_ad} {h.kullanici_soyad}'
            })
        
        # Otel adını al
        otel_ad = None
        if otel_id:
            otel = Otel.query.get(otel_id)
            if otel:
                otel_ad = otel.ad
        
        rapor_verisi = {
            'otel_id': otel_id,
            'otel_ad': otel_ad or 'Tüm Oteller',
            'baslangic': baslangic,
            'bitis': bitis,
            'toplam_hareket': len(hareketler),
            'hareketler': hareketler
        }
        
        oteller = get_kullanici_otelleri()
        
        log_islem('view', 'stok_hareketleri_raporu', {'otel_id': otel_id, 'baslangic': baslangic, 'bitis': bitis})
        
        return render_template('raporlar/stok_raporlari.html',
                             oteller=oteller,
                             stok_hareketleri_raporu=rapor_verisi)
        
    except Exception as e:
        log_hata(e, modul='stok_hareketleri_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.stok_raporlari'))


@raporlar_bp.route('/stok-excel')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def stok_excel():
    """Stok raporu Excel export"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        tip = request.args.get('tip', 'mevcut')
        otel_id = request.args.get('otel_id', type=int)
        
        wb = Workbook()
        ws = wb.active
        
        if tip == 'mevcut':
            # Mevcut stok durumu
            ws.title = "Mevcut Stok"
            
            # Başlık
            ws.merge_cells('A1:F1')
            ws['A1'] = 'Mevcut Stok Durumu'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:F2')
            ws['A2'] = f'Rapor Tarihi: {get_kktc_now().strftime("%d.%m.%Y %H:%M")}'
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Tablo başlıkları
            headers = ['Otel', 'Ürün Adı', 'Barkod', 'Mevcut Stok', 'Birim', 'Durum']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Veriler - Kat sorumlusunun zimmetindeki kalan stoklar
            query = db.session.query(
                Otel.ad.label('otel_ad'),
                Urun.urun_adi.label('urun_ad'),
                Urun.barkod,
                Urun.birim,
                func.sum(PersonelZimmetDetay.kalan_miktar).label('stok_miktari')
            ).join(PersonelZimmetDetay, Urun.id == PersonelZimmetDetay.urun_id)\
             .join(PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id)\
             .join(Kullanici, PersonelZimmet.personel_id == Kullanici.id)\
             .join(Otel, Kullanici.otel_id == Otel.id)\
             .filter(PersonelZimmet.durum == 'aktif')
            
            if otel_id:
                query = query.filter(Kullanici.otel_id == otel_id)
            
            urunler = query.group_by(Otel.ad, Urun.urun_adi, Urun.barkod, Urun.birim)\
                          .order_by(Otel.ad, Urun.urun_adi).all()
            
            row = 5
            for urun in urunler:
                ws.cell(row=row, column=1, value=urun.otel_ad)
                ws.cell(row=row, column=2, value=urun.urun_ad)
                ws.cell(row=row, column=3, value=urun.barkod)
                ws.cell(row=row, column=4, value=urun.stok_miktari)
                ws.cell(row=row, column=5, value=urun.birim)
                
                # Durum
                if urun.stok_miktari <= 0:
                    durum = 'Tükendi'
                elif urun.stok_miktari < 10:
                    durum = 'Kritik'
                elif urun.stok_miktari < 50:
                    durum = 'Düşük'
                else:
                    durum = 'Normal'
                ws.cell(row=row, column=6, value=durum)
                
                row += 1
            
            filename = f'mevcut_stok_{get_kktc_now().strftime("%Y%m%d_%H%M")}.xlsx'
            
        else:
            # Stok hareketleri
            ws.title = "Stok Hareketleri"
            
            baslangic = request.args.get('baslangic')
            bitis = request.args.get('bitis')
            
            if not all([baslangic, bitis]):
                flash('Lütfen tarih aralığı seçin.', 'warning')
                return redirect(url_for('raporlar.stok_raporlari'))
            
            baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d')
            bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d')
            
            # Başlık
            ws.merge_cells('A1:G1')
            ws['A1'] = 'Stok Hareketleri Raporu'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:G2')
            ws['A2'] = f'{baslangic_tarih.strftime("%d.%m.%Y")} - {bitis_tarih.strftime("%d.%m.%Y")}'
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Tablo başlıkları
            headers = ['Tarih', 'Otel', 'Ürün', 'İşlem Tipi', 'Miktar', 'Açıklama', 'Kullanıcı']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Veriler
            query = db.session.query(
                StokHareket.islem_tarihi,
                Otel.ad.label('otel_ad'),
                Urun.urun_adi.label('urun_ad'),
                StokHareket.hareket_tipi,
                StokHareket.miktar,
                StokHareket.aciklama,
                Kullanici.ad.label('kullanici_ad'),
                Kullanici.soyad.label('kullanici_soyad')
            ).join(Urun, StokHareket.urun_id == Urun.id)\
             .join(Kullanici, StokHareket.islem_yapan_id == Kullanici.id)\
             .join(Otel, Kullanici.otel_id == Otel.id)\
             .filter(
                StokHareket.islem_tarihi >= baslangic_tarih,
                StokHareket.islem_tarihi <= bitis_tarih
            )
            
            if otel_id:
                query = query.filter(Kullanici.otel_id == otel_id)
            
            hareketler = query.order_by(StokHareket.islem_tarihi.desc()).all()
            
            row = 5
            for hareket in hareketler:
                ws.cell(row=row, column=1, value=hareket.islem_tarihi.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row, column=2, value=hareket.otel_ad)
                ws.cell(row=row, column=3, value=hareket.urun_ad)
                ws.cell(row=row, column=4, value=hareket.hareket_tipi.upper())
                ws.cell(row=row, column=5, value=hareket.miktar)
                ws.cell(row=row, column=6, value=hareket.aciklama or '')
                ws.cell(row=row, column=7, value=f'{hareket.kullanici_ad} {hareket.kullanici_soyad}')
                row += 1
            
            filename = f'stok_hareketleri_{baslangic}_{bitis}.xlsx'
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 30
        if tip == 'hareketler':
            ws.column_dimensions['G'].width = 20
        
        # Excel'i memory'ye kaydet
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        log_islem('export', 'stok_raporu', {'tip': tip, 'otel_id': otel_id})
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        log_hata(e, modul='stok_excel')
        flash('Excel dosyası oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.stok_raporlari'))


@raporlar_bp.route('/minibar-detayli-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin')
def minibar_detayli_raporu_olustur():
    """Minibar detaylı raporu oluştur ve önizle"""
    try:
        otel_id = request.form.get('otel_id', type=int)
        baslangic = request.form.get('baslangic')
        bitis = request.form.get('bitis')
        
        if not all([otel_id, baslangic, bitis]):
            flash('Lütfen tüm alanları doldurun.', 'warning')
            return redirect(url_for('raporlar.minibar_raporlari'))
        
        baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d')
        bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d')
        
        otel = Otel.query.get(otel_id)
        if not otel:
            flash('Otel bulunamadı.', 'danger')
            return redirect(url_for('raporlar.minibar_raporlari'))
        
        # Veriler
        query = db.session.query(
            MinibarIslem.islem_tarihi,
            Oda.oda_no,
            Urun.urun_adi.label('urun_ad'),
            MinibarIslemDetay.tuketim,
            Kullanici.ad.label('personel_ad'),
            Kullanici.soyad.label('personel_soyad')
        ).join(MinibarIslemDetay, MinibarIslem.id == MinibarIslemDetay.islem_id)\
         .join(Oda, MinibarIslem.oda_id == Oda.id)\
         .join(Kat, Oda.kat_id == Kat.id)\
         .join(Urun, MinibarIslemDetay.urun_id == Urun.id)\
         .join(Kullanici, MinibarIslem.personel_id == Kullanici.id)\
         .filter(
            Kat.otel_id == otel_id,
            MinibarIslem.islem_tarihi >= baslangic_tarih,
            MinibarIslem.islem_tarihi <= bitis_tarih
        ).order_by(MinibarIslem.islem_tarihi.desc())
        
        islemler_data = query.all()
        
        islemler = []
        toplam_urun = 0
        for i in islemler_data:
            islemler.append({
                'tarih': i.islem_tarihi.strftime('%d.%m.%Y %H:%M'),
                'oda': i.oda_no,
                'urun': i.urun_ad,
                'miktar': i.tuketim,
                'personel': f'{i.personel_ad} {i.personel_soyad}'
            })
            toplam_urun += i.tuketim
        
        rapor_verisi = {
            'otel_id': otel_id,
            'otel_ad': otel.ad,
            'otel_logo': otel.logo,
            'baslangic': baslangic,
            'bitis': bitis,
            'toplam_islem': len(islemler),
            'toplam_urun': toplam_urun,
            'islemler': islemler
        }
        
        oteller = get_kullanici_otelleri()
        
        log_islem('view', 'minibar_detayli_raporu', {'otel_id': otel_id, 'baslangic': baslangic, 'bitis': bitis})
        
        return render_template('raporlar/minibar_raporlari.html',
                             oteller=oteller,
                             detayli_rapor=rapor_verisi)
        
    except Exception as e:
        log_hata(e, modul='minibar_detayli_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.minibar_raporlari'))


@raporlar_bp.route('/minibar-ozet-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin')
def minibar_ozet_raporu_olustur():
    """Minibar özet raporu oluştur ve önizle"""
    try:
        otel_id = request.form.get('otel_id', type=int)
        baslangic = request.form.get('baslangic')
        bitis = request.form.get('bitis')
        
        if not all([otel_id, baslangic, bitis]):
            flash('Lütfen tüm alanları doldurun.', 'warning')
            return redirect(url_for('raporlar.minibar_raporlari'))
        
        baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d')
        bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d')
        
        otel = Otel.query.get(otel_id)
        if not otel:
            flash('Otel bulunamadı.', 'danger')
            return redirect(url_for('raporlar.minibar_raporlari'))
        
        # Veriler
        query = db.session.query(
            Urun.urun_adi.label('urun_ad'),
            func.sum(MinibarIslemDetay.tuketim).label('toplam_miktar'),
            Urun.birim
        ).join(MinibarIslem, MinibarIslemDetay.islem_id == MinibarIslem.id)\
         .join(Oda, MinibarIslem.oda_id == Oda.id)\
         .join(Kat, Oda.kat_id == Kat.id)\
         .join(Urun, MinibarIslemDetay.urun_id == Urun.id)\
         .filter(
            Kat.otel_id == otel_id,
            MinibarIslem.islem_tarihi >= baslangic_tarih,
            MinibarIslem.islem_tarihi <= bitis_tarih
        ).group_by(Urun.urun_adi, Urun.birim)\
         .order_by(func.sum(MinibarIslemDetay.tuketim).desc())
        
        urunler_data = query.all()
        
        urunler = []
        genel_toplam = 0
        for u in urunler_data:
            urunler.append({
                'urun': u.urun_ad,
                'miktar': u.toplam_miktar,
                'birim': u.birim
            })
            genel_toplam += u.toplam_miktar
        
        rapor_verisi = {
            'otel_id': otel_id,
            'otel_ad': otel.ad,
            'otel_logo': otel.logo,
            'baslangic': baslangic,
            'bitis': bitis,
            'toplam_urun_cesidi': len(urunler),
            'genel_toplam': genel_toplam,
            'urunler': urunler
        }
        
        oteller = get_kullanici_otelleri()
        
        log_islem('view', 'minibar_ozet_raporu', {'otel_id': otel_id, 'baslangic': baslangic, 'bitis': bitis})
        
        return render_template('raporlar/minibar_raporlari.html',
                             oteller=oteller,
                             ozet_rapor=rapor_verisi)
        
    except Exception as e:
        log_hata(e, modul='minibar_ozet_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.minibar_raporlari'))


@raporlar_bp.route('/minibar-excel')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def minibar_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        otel_id = request.args.get('otel_id', type=int)
        baslangic = request.args.get('baslangic')
        bitis = request.args.get('bitis')
        tip = request.args.get('tip', 'detayli')
        
        if not all([otel_id, baslangic, bitis]):
            flash('Lütfen tüm alanları doldurun.', 'warning')
            return redirect(url_for('raporlar.minibar_raporlari'))
        
        baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d')
        bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d')
        
        otel = Otel.query.get(otel_id)
        if not otel:
            flash('Otel bulunamadı.', 'danger')
            return redirect(url_for('raporlar.minibar_raporlari'))
        
        wb = Workbook()
        ws = wb.active
        
        if tip == 'detayli':
            ws.title = 'Minibar Detay'
            ws.merge_cells('A1:H1')
            ws['A1'] = f'{otel.ad} - Minibar İşlemleri Detay Raporu'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:H2')
            tarih_str = f'{baslangic_tarih.strftime("%d.%m.%Y")} - {bitis_tarih.strftime("%d.%m.%Y")}'
            ws['A2'] = tarih_str
            ws['A2'].alignment = Alignment(horizontal='center')
            
            headers = ['Tarih', 'Oda', 'Ürün', 'Tüketim', 'Personel']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            query = db.session.query(
                MinibarIslem.islem_tarihi,
                Oda.oda_no,
                Urun.urun_adi.label('urun_ad'),
                MinibarIslemDetay.tuketim,
                Kullanici.ad.label('personel_ad'),
                Kullanici.soyad.label('personel_soyad')
            ).join(MinibarIslemDetay, MinibarIslem.id == MinibarIslemDetay.islem_id)\
             .join(Oda, MinibarIslem.oda_id == Oda.id)\
             .join(Kat, Oda.kat_id == Kat.id)\
             .join(Urun, MinibarIslemDetay.urun_id == Urun.id)\
             .join(Kullanici, MinibarIslem.personel_id == Kullanici.id)\
             .filter(
                Kat.otel_id == otel_id,
                MinibarIslem.islem_tarihi >= baslangic_tarih,
                MinibarIslem.islem_tarihi <= bitis_tarih
            ).order_by(MinibarIslem.islem_tarihi.desc())
            
            islemler = query.all()
            
            row = 5
            for islem in islemler:
                ws.cell(row=row, column=1, value=islem.islem_tarihi.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row, column=2, value=islem.oda_no)
                ws.cell(row=row, column=3, value=islem.urun_ad)
                ws.cell(row=row, column=4, value=islem.tuketim)
                ws.cell(row=row, column=5, value=f'{islem.personel_ad} {islem.personel_soyad}')
                row += 1
            
            filename = f'minibar_detay_{otel.ad}_{baslangic}_{bitis}.xlsx'
            
        else:
            ws.title = 'Minibar Özet'
            ws.merge_cells('A1:E1')
            ws['A1'] = f'{otel.ad} - Minibar Özet Raporu'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:E2')
            tarih_str = f'{baslangic_tarih.strftime("%d.%m.%Y")} - {bitis_tarih.strftime("%d.%m.%Y")}'
            ws['A2'] = tarih_str
            ws['A2'].alignment = Alignment(horizontal='center')
            
            headers = ['Ürün', 'Toplam Tüketim', 'Birim']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            query = db.session.query(
                Urun.urun_adi.label('urun_ad'),
                func.sum(MinibarIslemDetay.tuketim).label('toplam_miktar'),
                Urun.birim
            ).join(MinibarIslem, MinibarIslemDetay.islem_id == MinibarIslem.id)\
             .join(Oda, MinibarIslem.oda_id == Oda.id)\
             .join(Kat, Oda.kat_id == Kat.id)\
             .join(Urun, MinibarIslemDetay.urun_id == Urun.id)\
             .filter(
                Kat.otel_id == otel_id,
                MinibarIslem.islem_tarihi >= baslangic_tarih,
                MinibarIslem.islem_tarihi <= bitis_tarih
            ).group_by(Urun.urun_adi, Urun.birim)\
             .order_by(func.sum(MinibarIslemDetay.tuketim).desc())
            
            urunler = query.all()
            
            row = 5
            for urun in urunler:
                ws.cell(row=row, column=1, value=urun.urun_ad)
                ws.cell(row=row, column=2, value=urun.toplam_miktar)
                ws.cell(row=row, column=3, value=urun.birim)
                row += 1
            
            filename = f'minibar_ozet_{otel.ad}_{baslangic}_{bitis}.xlsx'
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        log_islem('export', 'minibar_raporu', {'otel_id': otel_id, 'tip': tip})
        
        return send_file(excel_buffer, as_attachment=True, download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        log_hata(e, modul='minibar_excel')
        flash('Excel dosyası oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.minibar_raporlari'))


# İkinci register_rapor_routes tanımı kaldırıldı - duplicate temizliği (29.12.2025)

@raporlar_bp.route('/zimmet-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin')
def zimmet_raporu_olustur():
    """Zimmet raporu oluştur ve önizle"""
    try:
        personel_id = request.form.get('personel_id', type=int)
        durum = request.form.get('durum', 'aktif')
        
        # Veriler - Otel bilgisi personel üzerinden
        query = db.session.query(
            Kullanici.ad.label('personel_ad'),
            Kullanici.soyad.label('personel_soyad'),
            Otel.ad.label('otel_ad'),
            Urun.urun_adi.label('urun_ad'),
            PersonelZimmetDetay.miktar.label('verilen_miktar'),
            PersonelZimmetDetay.kalan_miktar,
            PersonelZimmetDetay.kullanilan_miktar,
            PersonelZimmet.zimmet_tarihi,
            PersonelZimmet.iade_tarihi,
            PersonelZimmet.durum
        ).join(PersonelZimmetDetay, PersonelZimmet.id == PersonelZimmetDetay.zimmet_id)\
         .join(Kullanici, PersonelZimmet.personel_id == Kullanici.id)\
         .join(Urun, PersonelZimmetDetay.urun_id == Urun.id)\
         .join(Otel, Kullanici.otel_id == Otel.id)
        
        if personel_id:
            query = query.filter(PersonelZimmet.personel_id == personel_id)
        
        if durum == 'aktif':
            query = query.filter(PersonelZimmet.durum == 'aktif')
        elif durum == 'iade':
            query = query.filter(PersonelZimmet.durum == 'iade_edildi')
        
        zimmetler_data = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
        
        zimmetler = []
        for z in zimmetler_data:
            zimmetler.append({
                'personel': f'{z.personel_ad} {z.personel_soyad}',
                'otel': z.otel_ad,
                'urun': z.urun_ad,
                'verilen_miktar': z.verilen_miktar,
                'kullanilan_miktar': z.kullanilan_miktar or 0,
                'kalan_miktar': z.kalan_miktar or 0,
                'zimmet_tarihi': z.zimmet_tarihi.strftime('%d.%m.%Y'),
                'iade_tarihi': z.iade_tarihi.strftime('%d.%m.%Y') if z.iade_tarihi else None,
                'durum': z.durum
            })
        
        rapor_verisi = {
            'personel_id': personel_id,
            'durum': durum,
            'toplam_kayit': len(zimmetler),
            'zimmetler': zimmetler
        }
        
        oteller = get_kullanici_otelleri()
        personeller = Kullanici.query.filter(
            Kullanici.rol.in_(['kat_sorumlusu', 'depo_sorumlusu']),
            Kullanici.aktif.is_(True)
        ).order_by(Kullanici.ad, Kullanici.soyad).all()
        
        log_islem('view', 'zimmet_raporu', {'personel_id': personel_id, 'durum': durum})
        
        return render_template('raporlar/zimmet_raporlari.html', 
                             oteller=oteller,
                             personeller=personeller,
                             rapor_verisi=rapor_verisi)
        
    except Exception as e:
        log_hata(e, modul='zimmet_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.zimmet_raporlari'))


@raporlar_bp.route('/zimmet-excel')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def zimmet_excel():
    """Zimmet raporu Excel export"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        personel_id = request.args.get('personel_id', type=int)
        durum = request.args.get('durum', 'aktif')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Zimmet Raporu"
        
        # Başlık
        ws.merge_cells('A1:G1')
        ws['A1'] = 'Personel Zimmet Raporu'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:G2')
        ws['A2'] = f'Rapor Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Tablo başlıkları
        headers = ['Personel', 'Otel', 'Ürün', 'Miktar', 'Zimmet Tarihi', 'İade Tarihi', 'Durum']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Veriler
        query = db.session.query(
            Kullanici.ad.label('personel_ad'),
            Kullanici.soyad.label('personel_soyad'),
            Otel.ad.label('otel_ad'),
            Urun.urun_adi.label('urun_ad'),
            PersonelZimmetDetay.miktar,
            PersonelZimmet.zimmet_tarihi,
            PersonelZimmet.iade_tarihi,
            PersonelZimmet.durum
        ).join(PersonelZimmetDetay, PersonelZimmet.id == PersonelZimmetDetay.zimmet_id)\
         .join(Kullanici, PersonelZimmet.personel_id == Kullanici.id)\
         .join(Otel, Kullanici.otel_id == Otel.id)\
         .join(Urun, PersonelZimmetDetay.urun_id == Urun.id)
        
        if personel_id:
            query = query.filter(PersonelZimmet.personel_id == personel_id)
        
        if durum == 'aktif':
            query = query.filter(PersonelZimmet.durum == 'aktif')
        elif durum == 'iade':
            query = query.filter(PersonelZimmet.durum == 'iade_edildi')
        
        zimmetler = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
        
        row = 5
        for zimmet in zimmetler:
            ws.cell(row=row, column=1, value=f'{zimmet.personel_ad} {zimmet.personel_soyad}')
            ws.cell(row=row, column=2, value=zimmet.otel_ad)
            ws.cell(row=row, column=3, value=zimmet.urun_ad)
            ws.cell(row=row, column=4, value=zimmet.miktar)
            ws.cell(row=row, column=5, value=zimmet.zimmet_tarihi.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=6, value=zimmet.iade_tarihi.strftime('%d.%m.%Y') if zimmet.iade_tarihi else '-')
            ws.cell(row=row, column=7, value=zimmet.durum.upper())
            row += 1
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        # Excel'i memory'ye kaydet
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f'zimmet_raporu_{get_kktc_now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        log_islem('export', 'zimmet_raporu', {'personel_id': personel_id, 'durum': durum})
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        log_hata(e, modul='zimmet_excel')
        flash('Excel dosyası oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.zimmet_raporlari'))


@raporlar_bp.route('/performans-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin')
def performans_raporu_olustur():
    """Performans raporu oluştur ve önizle"""
    try:
        otel_id = request.form.get('otel_id', type=int)
        baslangic = request.form.get('baslangic')
        bitis = request.form.get('bitis')
        
        if not all([otel_id, baslangic, bitis]):
            flash('Lütfen tüm alanları doldurun.', 'warning')
            return redirect(url_for('raporlar.performans_raporlari'))
        
        baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d')
        bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d')
        
        otel = Otel.query.get(otel_id)
        if not otel:
            flash('Otel bulunamadı.', 'danger')
            return redirect(url_for('raporlar.performans_raporlari'))
        
        # Veriler - MinibarIslem'de durum yok, sadece toplam işlem sayısı
        query = db.session.query(
            Kullanici.ad.label('personel_ad'),
            Kullanici.soyad.label('personel_soyad'),
            func.count(MinibarIslem.id).label('toplam_islem')
        ).join(MinibarIslem, Kullanici.id == MinibarIslem.personel_id)\
         .join(Oda, MinibarIslem.oda_id == Oda.id)\
         .join(Kat, Oda.kat_id == Kat.id)\
         .filter(
            Kat.otel_id == otel_id,
            Kullanici.rol == 'kat_sorumlusu',
            MinibarIslem.islem_tarihi >= baslangic_tarih,
            MinibarIslem.islem_tarihi <= bitis_tarih
        ).group_by(Kullanici.id, Kullanici.ad, Kullanici.soyad)\
         .order_by(func.count(MinibarIslem.id).desc()).all()
        
        performanslar = []
        for p in query:
            # MinibarIslem'de durum yok, sadece toplam işlem sayısı var
            performanslar.append({
                'personel': f'{p.personel_ad} {p.personel_soyad}',
                'toplam_islem': p.toplam_islem,
                'tamamlanan': p.toplam_islem,  # Tüm işlemler tamamlanmış sayılıyor
                'bekleyen': 0,
                'iptal': 0,
                'basari_orani': 100.0  # Tüm işlemler başarılı sayılıyor
            })
        
        rapor_verisi = {
            'otel_id': otel_id,
            'otel_ad': otel.ad,
            'otel_logo': otel.logo,
            'baslangic': baslangic,
            'bitis': bitis,
            'personel_sayisi': len(performanslar),
            'performanslar': performanslar
        }
        
        oteller = get_kullanici_otelleri()
        personeller = Kullanici.query.filter(
            Kullanici.rol == 'kat_sorumlusu',
            Kullanici.aktif.is_(True)
        ).order_by(Kullanici.ad, Kullanici.soyad).all()
        
        log_islem('view', 'performans_raporu', {'otel_id': otel_id, 'baslangic': baslangic, 'bitis': bitis})
        
        return render_template('raporlar/performans_raporlari.html',
                             oteller=oteller,
                             personeller=personeller,
                             rapor_verisi=rapor_verisi)
        
    except Exception as e:
        log_hata(e, modul='performans_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.performans_raporlari'))


@raporlar_bp.route('/performans-excel')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def performans_excel():
    """Performans raporu Excel export"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        otel_id = request.args.get('otel_id', type=int)
        baslangic = request.args.get('baslangic')
        bitis = request.args.get('bitis')
        
        if not all([otel_id, baslangic, bitis]):
            flash('Lütfen tüm alanları doldurun.', 'warning')
            return redirect(url_for('raporlar.performans_raporlari'))
        
        baslangic_tarih = datetime.strptime(baslangic, '%Y-%m-%d')
        bitis_tarih = datetime.strptime(bitis, '%Y-%m-%d')
        
        otel = Otel.query.get(otel_id)
        if not otel:
            flash('Otel bulunamadı.', 'danger')
            return redirect(url_for('raporlar.performans_raporlari'))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Performans Raporu"
        
        # Başlık
        ws.merge_cells('A1:F1')
        ws['A1'] = f'{otel.ad} - Kat Sorumlusu Performans Raporu'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f'{baslangic_tarih.strftime("%d.%m.%Y")} - {bitis_tarih.strftime("%d.%m.%Y")}'
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Tablo başlıkları
        headers = ['Personel', 'Toplam İşlem', 'Tamamlanan', 'Bekleyen', 'İptal', 'Başarı %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Veriler
        query = db.session.query(
            Kullanici.ad.label('personel_ad'),
            Kullanici.soyad.label('personel_soyad'),
            func.count(MinibarIslem.id).label('toplam_islem')
        ).join(MinibarIslem, Kullanici.id == MinibarIslem.personel_id)\
         .join(Oda, MinibarIslem.oda_id == Oda.id)\
         .join(Kat, Oda.kat_id == Kat.id)\
         .filter(
            Kat.otel_id == otel_id,
            Kullanici.rol == 'kat_sorumlusu',
            MinibarIslem.islem_tarihi >= baslangic_tarih,
            MinibarIslem.islem_tarihi <= bitis_tarih
        ).group_by(Kullanici.id, Kullanici.ad, Kullanici.soyad)\
         .order_by(func.count(MinibarIslem.id).desc())
        
        performanslar = query.all()
        
        row = 5
        for perf in performanslar:
            ws.cell(row=row, column=1, value=f'{perf.personel_ad} {perf.personel_soyad}')
            ws.cell(row=row, column=2, value=perf.toplam_islem)
            ws.cell(row=row, column=3, value=perf.toplam_islem)  # Tamamlanan = Toplam
            ws.cell(row=row, column=4, value=0)  # Bekleyen
            ws.cell(row=row, column=5, value=0)  # İptal
            ws.cell(row=row, column=6, value='100.0%')  # Başarı oranı
            row += 1
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
        
        # Excel'i memory'ye kaydet
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f'performans_raporu_{otel.ad}_{baslangic}_{bitis}.xlsx'
        
        log_islem('export', 'performans_raporu', {'otel_id': otel_id, 'baslangic': baslangic, 'bitis': bitis})
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        log_hata(e, modul='performans_excel')
        flash('Excel dosyası oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.performans_raporlari'))


# ============================================================================
# YENİ RAPORLAR - OTEL BAZLI ZİMMET, KAT SORUMLUSU KULLANIM, ODA BAZLI TÜKETİM
# ============================================================================

@raporlar_bp.route('/otel-zimmet-stok')
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def otel_zimmet_stok_raporlari():
    """Otel bazlı zimmet stok raporları sayfası"""
    try:
        oteller = get_kullanici_otelleri()
        return render_template('raporlar/otel_zimmet_stok.html', oteller=oteller)
    except Exception as e:
        log_hata(e, modul='otel_zimmet_stok_raporlari')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('raporlar.doluluk_raporlari'))


@raporlar_bp.route('/otel-zimmet-stok-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def otel_zimmet_stok_raporu_olustur():
    """Otel zimmet stok raporu oluştur"""
    try:
        from utils.rapor_servisleri import OtelZimmetRaporServisi
        
        otel_id = request.form.get('otel_id', type=int)
        export_format = request.form.get('format', '')
        
        rapor = OtelZimmetRaporServisi.get_otel_zimmet_stok_raporu(otel_id)
        
        if not rapor['success']:
            flash(rapor.get('message', 'Rapor oluşturulamadı'), 'danger')
            return redirect(url_for('raporlar.otel_zimmet_stok_raporlari'))
        
        # Excel export
        if export_format == 'excel':
            return export_otel_zimmet_stok_excel(rapor)
        
        oteller = get_kullanici_otelleri()
        
        log_islem('view', 'otel_zimmet_stok_raporu', {'otel_id': otel_id})
        
        return render_template('raporlar/otel_zimmet_stok.html',
                             oteller=oteller,
                             rapor_verisi=rapor)
        
    except Exception as e:
        log_hata(e, modul='otel_zimmet_stok_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.otel_zimmet_stok_raporlari'))


def export_otel_zimmet_stok_excel(rapor):
    """Otel zimmet stok raporunu Excel olarak export et - Logo destekli"""
    import io
    import base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from flask import send_file
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Otel Zimmet Stok Raporu"
    
    # Stiller
    title_font = Font(bold=True, size=14, color="1E3A5F")
    subtitle_font = Font(size=10, color="475569")
    date_font = Font(bold=True, size=11, color="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_border = Border(bottom=Side(style='thin', color='CBD5E1'))
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12  # Logo
    ws.column_dimensions['B'].width = 30  # Otel Adı / Ürün adı
    ws.column_dimensions['C'].width = 20  # Rapor Adı / Birim
    ws.column_dimensions['D'].width = 18  # Tarih / Toplam
    ws.column_dimensions['E'].width = 12  # Kullanılan
    ws.column_dimensions['F'].width = 12  # Kalan
    ws.column_dimensions['G'].width = 14  # Kritik Seviye
    ws.column_dimensions['H'].width = 12  # Kullanım %
    ws.column_dimensions['I'].width = 12  # Durum
    
    row = 1
    
    # Her otel için tablo
    for otel in rapor.get('oteller', []):
        # HEADER: A(row):A(row+1) Logo | B(row):C(row) Otel Adı | B(row+1):C(row+1) Rapor Adı | D(row):D(row+1) Tarih
        ws.row_dimensions[row].height = 30
        ws.row_dimensions[row + 1].height = 30
        
        # Logo (A birleşik 2 satır)
        ws.merge_cells(f'A{row}:A{row + 1}')
        if otel.get('otel_logo'):
            try:
                logo_data = otel['otel_logo']
                if ',' in logo_data:
                    logo_data = logo_data.split(',')[1]
                
                logo_bytes = base64.b64decode(logo_data)
                logo_stream = io.BytesIO(logo_bytes)
                
                img = XLImage(logo_stream)
                img.width = 85
                img.height = 80
                ws.add_image(img, f'A{row}')
            except:
                pass
        
        # Otel Adı (B:C birleşik - 1. satır)
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = otel['otel_ad']
        ws[f'B{row}'].font = title_font
        ws[f'B{row}'].alignment = center_align
        
        # Rapor Adı (B:C birleşik - 2. satır)
        ws.merge_cells(f'B{row + 1}:C{row + 1}')
        ws[f'B{row + 1}'] = "Zimmet Stok Raporu"
        ws[f'B{row + 1}'].font = subtitle_font
        ws[f'B{row + 1}'].alignment = center_align
        
        # Tarih (D birleşik 2 satır)
        ws.merge_cells(f'D{row}:D{row + 1}')
        ws[f'D{row}'] = f"📅 {rapor.get('rapor_tarihi', '')}"
        ws[f'D{row}'].font = date_font
        ws[f'D{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Header altı çizgi
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row + 1}'].border = header_border
        
        row += 3  # Header'dan sonra 1 satır boşluk
        
        # Tablo başlıkları
        headers = ['Ürün', 'Birim', 'Toplam', 'Kullanılan', 'Kalan', 'Kritik Seviye', 'Kullanım %', 'Durum']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1
        
        # Ürün verileri
        for urun in otel.get('urunler', []):
            ws.cell(row=row, column=1, value=urun['urun_adi']).border = thin_border
            ws.cell(row=row, column=2, value=urun['birim']).border = thin_border
            ws.cell(row=row, column=2).alignment = center_align
            ws.cell(row=row, column=3, value=urun['toplam_miktar']).border = thin_border
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=4, value=urun['kullanilan_miktar']).border = thin_border
            ws.cell(row=row, column=4).alignment = center_align
            ws.cell(row=row, column=5, value=urun['kalan_miktar']).border = thin_border
            ws.cell(row=row, column=5).alignment = center_align
            ws.cell(row=row, column=6, value=urun['kritik_seviye']).border = thin_border
            ws.cell(row=row, column=6).alignment = center_align
            ws.cell(row=row, column=7, value=f"{urun['kullanim_yuzdesi']}%").border = thin_border
            ws.cell(row=row, column=7).alignment = center_align
            
            # Durum
            durum_map = {'stokout': 'TÜKENDİ', 'kritik': 'KRİTİK', 'dikkat': 'DİKKAT', 'normal': 'NORMAL'}
            durum = durum_map.get(urun['stok_durumu'], 'NORMAL')
            durum_cell = ws.cell(row=row, column=8, value=durum)
            durum_cell.border = thin_border
            durum_cell.alignment = center_align
            
            # Durum renkleri
            if urun['stok_durumu'] == 'stokout':
                durum_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            elif urun['stok_durumu'] == 'kritik':
                durum_cell.fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
            elif urun['stok_durumu'] == 'dikkat':
                durum_cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
            else:
                durum_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            
            row += 1
        
        row += 2  # Oteller arası boşluk
    
    # Excel dosyasını oluştur
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"otel_zimmet_stok_raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    log_islem('export', 'otel_zimmet_stok_raporu', {'format': 'excel'})
    
    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@raporlar_bp.route('/kat-sorumlusu-kullanim')
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def kat_sorumlusu_kullanim_raporlari():
    """Kat sorumlusu gün sonu raporu sayfası"""
    try:
        oteller = get_kullanici_otelleri()
        return render_template('raporlar/kat_sorumlusu_kullanim.html', oteller=oteller)
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_kullanim_raporlari')
        flash(f'Sayfa yüklenirken hata oluştu: {str(e)}', 'danger')
        return redirect(url_for('sistem_yoneticisi_dashboard'))


@raporlar_bp.route('/api/otel-kat-sorumlulari/<int:otel_id>')
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def api_otel_kat_sorumlulari(otel_id):
    """Otele ait kat sorumlularını getir"""
    try:
        personeller = Kullanici.query.filter(
            Kullanici.rol == 'kat_sorumlusu',
            Kullanici.aktif.is_(True),
            Kullanici.otel_id == otel_id
        ).order_by(Kullanici.ad, Kullanici.soyad).all()
        
        return jsonify({
            'success': True,
            'personeller': [
                {'id': p.id, 'ad_soyad': f"{p.ad} {p.soyad}"}
                for p in personeller
            ]
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@raporlar_bp.route('/kat-sorumlusu-kullanim-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def kat_sorumlusu_kullanim_raporu_olustur():
    """Kat sorumlusu gün sonu raporu oluştur"""
    try:
        from utils.rapor_servisleri import KatSorumlusuGunSonuRaporServisi
        
        otel_id = request.form.get('otel_id', type=int)
        personel_ids = request.form.getlist('personel_ids[]', type=int)
        tarih_str = request.form.get('tarih')
        export_format = request.form.get('format', '')
        
        if not otel_id:
            flash('Lütfen otel seçin.', 'warning')
            return redirect(url_for('raporlar.kat_sorumlusu_kullanim_raporlari'))
        
        tarih = datetime.strptime(tarih_str, '%Y-%m-%d').date() if tarih_str else date.today()
        
        rapor = KatSorumlusuGunSonuRaporServisi.get_gun_sonu_raporu(
            otel_id=otel_id,
            personel_ids=personel_ids if personel_ids else None,
            tarih=tarih
        )
        
        if not rapor['success']:
            flash(rapor.get('message', 'Rapor oluşturulamadı'), 'danger')
            return redirect(url_for('raporlar.kat_sorumlusu_kullanim_raporlari'))
        
        # Excel export
        if export_format == 'excel':
            return export_gun_sonu_excel(rapor)
        
        # PDF export
        if export_format == 'pdf':
            return export_gun_sonu_pdf(rapor)
        
        oteller = get_kullanici_otelleri()
        
        log_islem('view', 'kat_sorumlusu_gun_sonu_raporu', {
            'otel_id': otel_id, 'tarih': tarih_str
        })
        
        return render_template('raporlar/kat_sorumlusu_kullanim.html',
                             oteller=oteller,
                             rapor_verisi=rapor,
                             secili_otel_id=otel_id,
                             secili_tarih=tarih_str)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_kullanim_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.kat_sorumlusu_kullanim_raporlari'))


@raporlar_bp.route('/kat-sorumlusu-kullanim-excel')
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def kat_sorumlusu_kullanim_excel():
    """Kat sorumlusu gün sonu raporu Excel export"""
    try:
        from utils.rapor_servisleri import KatSorumlusuGunSonuRaporServisi
        
        otel_id = request.args.get('otel_id', type=int)
        personel_ids_str = request.args.get('personel_ids', '')
        tarih_str = request.args.get('tarih')
        
        if not otel_id:
            flash('Lütfen otel seçin.', 'warning')
            return redirect(url_for('raporlar.kat_sorumlusu_kullanim_raporlari'))
        
        personel_ids = [int(x) for x in personel_ids_str.split(',') if x.strip().isdigit()] if personel_ids_str else None
        tarih = datetime.strptime(tarih_str, '%Y-%m-%d').date() if tarih_str else date.today()
        
        rapor = KatSorumlusuGunSonuRaporServisi.get_gun_sonu_raporu(
            otel_id=otel_id,
            personel_ids=personel_ids,
            tarih=tarih
        )
        
        if not rapor['success']:
            flash(rapor.get('message', 'Rapor oluşturulamadı'), 'danger')
            return redirect(url_for('raporlar.kat_sorumlusu_kullanim_raporlari'))
        
        return export_gun_sonu_excel(rapor)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_kullanim_excel')
        flash('Excel export başarısız.', 'danger')
        return redirect(url_for('raporlar.kat_sorumlusu_kullanim_raporlari'))


@raporlar_bp.route('/kat-sorumlusu-kullanim-pdf')
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def kat_sorumlusu_kullanim_pdf():
    """Kat sorumlusu gün sonu raporu PDF export"""
    try:
        from utils.rapor_servisleri import KatSorumlusuGunSonuRaporServisi
        
        otel_id = request.args.get('otel_id', type=int)
        personel_ids_str = request.args.get('personel_ids', '')
        tarih_str = request.args.get('tarih')
        
        if not otel_id:
            flash('Lütfen otel seçin.', 'warning')
            return redirect(url_for('raporlar.kat_sorumlusu_kullanim_raporlari'))
        
        personel_ids = [int(x) for x in personel_ids_str.split(',') if x.strip().isdigit()] if personel_ids_str else None
        tarih = datetime.strptime(tarih_str, '%Y-%m-%d').date() if tarih_str else date.today()
        
        rapor = KatSorumlusuGunSonuRaporServisi.get_gun_sonu_raporu(
            otel_id=otel_id,
            personel_ids=personel_ids,
            tarih=tarih
        )
        
        if not rapor['success']:
            flash(rapor.get('message', 'Rapor oluşturulamadı'), 'danger')
            return redirect(url_for('raporlar.kat_sorumlusu_kullanim_raporlari'))
        
        return export_gun_sonu_pdf(rapor)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_kullanim_pdf')
        flash('PDF export başarısız.', 'danger')
        return redirect(url_for('raporlar.kat_sorumlusu_kullanim_raporlari'))


def export_gun_sonu_excel(rapor):
    """Gün sonu raporunu Excel olarak export et - Yeni format + Otel Logosu"""
    import io
    import base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from flask import send_file
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Gün Sonu Raporu"
    
    # Stiller - Slate tonları
    title_font = Font(bold=True, size=14, color="1E3A5F")
    subtitle_font = Font(size=10, color="475569")
    date_font = Font(bold=True, size=11, color="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
    personel_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    toplam_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_border = Border(bottom=Side(style='thin', color='CBD5E1'))
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12  # Logo
    ws.column_dimensions['B'].width = 35  # Otel Adı / Ürün adı
    ws.column_dimensions['C'].width = 25  # Rapor Adı / Ürün devam
    ws.column_dimensions['D'].width = 20  # Tarih / Miktar
    
    # HEADER SATIRI: A1:A2 Logo | B1:C1 Otel Adı | B2:C2 Rapor Adı | D1:D2 Tarih
    # Satır yükseklikleri
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    
    # Logo (A1:A2 birleşik)
    ws.merge_cells('A1:A2')
    otel_id = rapor.get('otel_id')
    if otel_id:
        otel = Otel.query.get(otel_id)
        if otel and otel.logo:
            try:
                logo_data = otel.logo
                if ',' in logo_data:
                    logo_data = logo_data.split(',')[1]
                
                logo_bytes = base64.b64decode(logo_data)
                logo_stream = io.BytesIO(logo_bytes)
                
                img = XLImage(logo_stream)
                img.width = 85
                img.height = 80
                ws.add_image(img, 'A1')
            except:
                pass
    
    # Otel Adı (B1:C1 birleşik)
    ws.merge_cells('B1:C1')
    ws['B1'] = rapor['otel_adi']
    ws['B1'].font = title_font
    ws['B1'].alignment = center_align
    
    # Rapor Adı (B2:C2 birleşik)
    ws.merge_cells('B2:C2')
    ws['B2'] = "Kat Sorumlusu Gün Sonu Raporu"
    ws['B2'].font = subtitle_font
    ws['B2'].alignment = center_align
    
    # Tarih (D1:D2 birleşik)
    ws.merge_cells('D1:D2')
    ws['D1'] = f"📅 {rapor['rapor_tarihi']}"
    ws['D1'].font = date_font
    ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Header altı çizgi
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}2'].border = header_border
    
    row = 4  # Header'dan sonra 1 satır boşluk
    
    # Her personel için
    for personel in rapor.get('personeller', []):
        # Personel başlığı
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"👤 {personel['personel_adi']} (Toplam: {personel['toplam_eklenen']} adet)"
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = personel_fill
        cell.alignment = center_align
        row += 1
        
        # Tablo başlıkları
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'Ürün Adı'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        
        ws.merge_cells(f'C{row}:D{row}')
        ws[f'C{row}'] = 'Minibarlara Eklenen'
        ws[f'C{row}'].font = header_font
        ws[f'C{row}'].fill = header_fill
        ws[f'C{row}'].alignment = center_align
        ws[f'C{row}'].border = thin_border
        ws[f'D{row}'].border = thin_border
        row += 1
        
        # Ürünler
        for urun in personel.get('urunler', []):
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = urun['urun_adi']
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border
            
            ws.merge_cells(f'C{row}:D{row}')
            ws[f'C{row}'] = urun['toplam_eklenen']
            ws[f'C{row}'].border = thin_border
            ws[f'C{row}'].alignment = center_align
            ws[f'D{row}'].border = thin_border
            row += 1
        
        row += 1
    
    # GENEL TOPLAM
    if rapor.get('genel_toplam'):
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"📊 GENEL TOPLAM: {rapor.get('genel_toplam_adet', 0)} Adet"
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = toplam_fill
        cell.alignment = center_align
        row += 1
        
        # Başlıklar
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'Ürün Adı'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        
        ws.merge_cells(f'C{row}:D{row}')
        ws[f'C{row}'] = 'Toplam Eklenen'
        ws[f'C{row}'].font = header_font
        ws[f'C{row}'].fill = header_fill
        ws[f'C{row}'].alignment = center_align
        ws[f'C{row}'].border = thin_border
        ws[f'D{row}'].border = thin_border
        row += 1
        
        for urun in rapor['genel_toplam']:
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = urun['urun_adi']
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border
            
            ws.merge_cells(f'C{row}:D{row}')
            ws[f'C{row}'] = urun['toplam_eklenen']
            ws[f'C{row}'].border = thin_border
            ws[f'C{row}'].alignment = center_align
            ws[f'D{row}'].border = thin_border
            row += 1
    
    # Excel dosyasını oluştur
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"gun_sonu_raporu_{rapor['rapor_tarihi'].replace('.', '')}_{rapor['otel_adi'].replace(' ', '_')}.xlsx"
    
    log_islem('export', 'gun_sonu_raporu', {'format': 'excel'})
    
    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def export_gun_sonu_pdf(rapor):
    """Gün sonu raporunu PDF olarak export et - Türkçe karakter destekli + Otel Logosu"""
    import io
    import base64
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import urllib.request
    import os
    import tempfile
    
    # Türkçe karakter desteği için font
    font_name = 'Helvetica'
    
    try:
        # DejaVuSans font'u indir veya mevcut olanı kullan
        fonts_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'fonts')
        font_path = os.path.join(fonts_dir, 'DejaVuSans.ttf')
        
        if not os.path.exists(fonts_dir):
            os.makedirs(fonts_dir)
        
        if not os.path.exists(font_path):
            # Font'u indir
            font_url = "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans.ttf"
            urllib.request.urlretrieve(font_url, font_path)
        
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
            font_name = 'DejaVuSans'
    except Exception as e:
        # Font yüklenemedi, varsayılan font kullanılacak
        font_name = 'Helvetica'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm, leftMargin=1*cm, rightMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Header: Logo (sol) | Otel Adı + Rapor Adı (orta) | Tarih (sağ)
    logo_cell = ""
    otel_id = rapor.get('otel_id')
    if otel_id:
        otel = Otel.query.get(otel_id)
        if otel and otel.logo:
            try:
                logo_data = otel.logo
                if ',' in logo_data:
                    logo_data = logo_data.split(',')[1]
                logo_bytes = base64.b64decode(logo_data)
                logo_stream = io.BytesIO(logo_bytes)
                logo_cell = RLImage(logo_stream, width=1.8*cm, height=1.8*cm)
            except:
                logo_cell = ""
    
    # Orta kısım: Otel adı ve rapor adı
    center_content = f"""<para align="center">
        <font name="{font_name}" size="16" color="#1E3A5F"><b>{rapor['otel_adi']}</b></font><br/>
        <font name="{font_name}" size="11" color="#475569">Kat Sorumlusu Gün Sonu Raporu</font>
    </para>"""
    
    # Sağ kısım: Tarih
    date_content = f"""<para align="right">
        <font name="{font_name}" size="10" color="#64748B">📅</font><br/>
        <font name="{font_name}" size="12" color="#1E293B"><b>{rapor['rapor_tarihi']}</b></font>
    </para>"""
    
    header_table = Table(
        [[logo_cell, Paragraph(center_content, styles['Normal']), Paragraph(date_content, styles['Normal'])]],
        colWidths=[3*cm, 12*cm, 4*cm]
    )
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 15))
    
    # Her personel için
    for personel in rapor.get('personeller', []):
        # Personel başlık tablosu (slate arka plan)
        personel_header = Table(
            [[f"{personel['personel_adi']} - Toplam: {personel['toplam_eklenen']} adet"]],
            colWidths=[19*cm]
        )
        personel_header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#334155')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(personel_header)
        
        # Ürün tablosu
        urun_data = [['Ürün Adı', 'Minibarlara Eklenen']]
        for urun in personel.get('urunler', []):
            urun_data.append([urun['urun_adi'], str(urun['toplam_eklenen'])])
        
        urun_table = Table(urun_data, colWidths=[13*cm, 6*cm])
        urun_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#475569')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        elements.append(urun_table)
        elements.append(Spacer(1, 15))
    
    # GENEL TOPLAM
    if rapor.get('genel_toplam'):
        toplam_header = Table(
            [[f"GENEL TOPLAM: {rapor.get('genel_toplam_adet', 0)} Adet"]],
            colWidths=[19*cm]
        )
        toplam_header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(toplam_header)
        
        toplam_data = [['Ürün Adı', 'Toplam Eklenen']]
        for urun in rapor['genel_toplam']:
            toplam_data.append([urun['urun_adi'], str(urun['toplam_eklenen'])])
        
        toplam_table = Table(toplam_data, colWidths=[13*cm, 6*cm])
        toplam_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        elements.append(toplam_table)
    
    # PDF oluştur
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"gun_sonu_raporu_{rapor['rapor_tarihi'].replace('.', '')}_{rapor['otel_adi'].replace(' ', '_')}.pdf"
    
    log_islem('export', 'gun_sonu_raporu', {'format': 'pdf'})
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


@raporlar_bp.route('/oda-bazli-tuketim')
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def oda_bazli_tuketim_raporlari():
    """Oda bazlı tüketim raporları sayfası"""
    try:
        oteller = get_kullanici_otelleri()
        return render_template('raporlar/oda_bazli_tuketim.html', oteller=oteller)
    except Exception as e:
        log_hata(e, modul='oda_bazli_tuketim_raporlari')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('raporlar.doluluk_raporlari'))


@raporlar_bp.route('/oda-bazli-tuketim-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def oda_bazli_tuketim_raporu_olustur():
    """Oda bazlı tüketim raporu oluştur"""
    try:
        from utils.rapor_servisleri import OdaBazliTuketimRaporServisi
        
        otel_id = request.form.get('otel_id', type=int)
        baslangic = request.form.get('baslangic')
        bitis = request.form.get('bitis')
        
        if not otel_id:
            flash('Lütfen otel seçin.', 'warning')
            return redirect(url_for('raporlar.oda_bazli_tuketim_raporlari'))
        
        baslangic_tarihi = datetime.strptime(baslangic, '%Y-%m-%d').date() if baslangic else None
        bitis_tarihi = datetime.strptime(bitis, '%Y-%m-%d').date() if bitis else None
        
        rapor = OdaBazliTuketimRaporServisi.get_oda_bazli_tuketim_raporu(
            otel_id=otel_id,
            baslangic_tarihi=baslangic_tarihi,
            bitis_tarihi=bitis_tarihi
        )
        
        if not rapor['success']:
            flash(rapor.get('message', 'Rapor oluşturulamadı'), 'danger')
            return redirect(url_for('raporlar.oda_bazli_tuketim_raporlari'))
        
        oteller = get_kullanici_otelleri()
        
        log_islem('view', 'oda_bazli_tuketim_raporu', {
            'otel_id': otel_id, 'baslangic': baslangic, 'bitis': bitis
        })
        
        return render_template('raporlar/oda_bazli_tuketim.html',
                             oteller=oteller,
                             rapor_verisi=rapor)
        
    except Exception as e:
        log_hata(e, modul='oda_bazli_tuketim_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.oda_bazli_tuketim_raporlari'))


@raporlar_bp.route('/gunluk-gorev-detay')
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def gunluk_gorev_detay_raporlari():
    """Günlük görev detay raporları sayfası"""
    try:
        oteller = get_kullanici_otelleri()
        personeller = Kullanici.query.filter(
            Kullanici.rol == 'kat_sorumlusu',
            Kullanici.aktif.is_(True)
        ).order_by(Kullanici.ad, Kullanici.soyad).all()
        return render_template('raporlar/gunluk_gorev_detay.html', 
                             oteller=oteller, personeller=personeller)
    except Exception as e:
        log_hata(e, modul='gunluk_gorev_detay_raporlari')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('raporlar.doluluk_raporlari'))


@raporlar_bp.route('/gunluk-gorev-detay-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
def gunluk_gorev_detay_raporu_olustur():
    """Günlük görev detay raporu oluştur"""
    try:
        from utils.rapor_servisleri import GunlukGorevRaporServisi
        
        otel_id = request.form.get('otel_id', type=int)
        personel_ids_str = request.form.get('personel_ids', '')
        baslangic = request.form.get('baslangic')
        bitis = request.form.get('bitis')
        
        # Personel ID'lerini parse et
        personel_ids = [int(x) for x in personel_ids_str.split(',') if x.strip().isdigit()] if personel_ids_str else None
        
        baslangic_tarihi = datetime.strptime(baslangic, '%Y-%m-%d').date() if baslangic else None
        bitis_tarihi = datetime.strptime(bitis, '%Y-%m-%d').date() if bitis else None
        
        rapor = GunlukGorevRaporServisi.get_gunluk_gorev_raporu(
            otel_id=otel_id,
            personel_ids=personel_ids,
            baslangic_tarihi=baslangic_tarihi,
            bitis_tarihi=bitis_tarihi
        )
        
        if not rapor['success']:
            flash(rapor.get('message', 'Rapor oluşturulamadı'), 'danger')
            return redirect(url_for('raporlar.gunluk_gorev_detay_raporlari'))
        
        oteller = get_kullanici_otelleri()
        personeller = Kullanici.query.filter(
            Kullanici.rol == 'kat_sorumlusu',
            Kullanici.aktif.is_(True)
        ).order_by(Kullanici.ad, Kullanici.soyad).all()
        
        log_islem('view', 'gunluk_gorev_detay_raporu', {
            'otel_id': otel_id, 'personel_ids': personel_ids
        })
        
        return render_template('raporlar/gunluk_gorev_detay.html',
                             oteller=oteller,
                             personeller=personeller,
                             rapor_verisi=rapor,
                             secili_otel_id=otel_id,
                             secili_baslangic=baslangic,
                             secili_bitis=bitis)
        
    except Exception as e:
        log_hata(e, modul='gunluk_gorev_detay_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.gunluk_gorev_detay_raporlari'))


@raporlar_bp.route('/otel-karsilastirma')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def otel_karsilastirma_raporlari():
    """Otel karşılaştırma raporları sayfası"""
    try:
        return render_template('raporlar/otel_karsilastirma.html')
    except Exception as e:
        log_hata(e, modul='otel_karsilastirma_raporlari')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('raporlar.doluluk_raporlari'))


@raporlar_bp.route('/otel-karsilastirma-raporu-olustur', methods=['POST'])
@login_required
@role_required('sistem_yoneticisi', 'admin')
def otel_karsilastirma_raporu_olustur():
    """Otel karşılaştırma raporu oluştur"""
    try:
        from utils.rapor_servisleri import OtelKarsilastirmaRaporServisi
        
        baslangic = request.form.get('baslangic')
        bitis = request.form.get('bitis')
        
        baslangic_tarihi = datetime.strptime(baslangic, '%Y-%m-%d').date() if baslangic else None
        bitis_tarihi = datetime.strptime(bitis, '%Y-%m-%d').date() if bitis else None
        
        rapor = OtelKarsilastirmaRaporServisi.get_otel_karsilastirma_raporu(
            baslangic_tarihi=baslangic_tarihi,
            bitis_tarihi=bitis_tarihi
        )
        
        if not rapor['success']:
            flash(rapor.get('message', 'Rapor oluşturulamadı'), 'danger')
            return redirect(url_for('raporlar.otel_karsilastirma_raporlari'))
        
        log_islem('view', 'otel_karsilastirma_raporu', {
            'baslangic': baslangic, 'bitis': bitis
        })
        
        return render_template('raporlar/otel_karsilastirma.html', rapor_verisi=rapor)
        
    except Exception as e:
        log_hata(e, modul='otel_karsilastirma_raporu_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.otel_karsilastirma_raporlari'))


# Üçüncü register_rapor_routes tanımı kaldırıldı - duplicate temizliği (29.12.2025)
# Sadece dosya başındaki (satır 41) tanım kullanılıyor


# =====================================================
# KAT SORUMLUSU GÜN SONU RAPORU (Kendi Oteli İçin)
# =====================================================

@raporlar_bp.route('/kat-sorumlusu/gun-sonu-raporum')
@login_required
@role_required('kat_sorumlusu')
def kat_sorumlusu_gun_sonu_raporum():
    """Kat sorumlusu kendi gün sonu raporunu görüntüler"""
    from flask import session
    try:
        kullanici_id = session.get('kullanici_id')
        kullanici = Kullanici.query.get(kullanici_id)
        
        if not kullanici or not kullanici.otel_id:
            flash('Otel atamanız bulunamadı.', 'danger')
            return redirect(url_for('kat_sorumlusu_dashboard'))
        
        # Kat sorumlusunun oteli
        otel = Otel.query.get(kullanici.otel_id)
        
        # Aynı oteldeki diğer kat sorumluları
        personeller = Kullanici.query.filter(
            Kullanici.rol == 'kat_sorumlusu',
            Kullanici.aktif.is_(True),
            Kullanici.otel_id == kullanici.otel_id
        ).order_by(Kullanici.ad, Kullanici.soyad).all()
        
        return render_template('raporlar/kat_sorumlusu_gun_sonu_raporum.html',
                             otel=otel,
                             personeller=personeller,
                             current_user_id=kullanici_id)
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_gun_sonu_raporum')
        flash('Sayfa yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('kat_sorumlusu_dashboard'))


@raporlar_bp.route('/kat-sorumlusu/gun-sonu-raporum-olustur', methods=['POST'])
@login_required
@role_required('kat_sorumlusu')
def kat_sorumlusu_gun_sonu_raporum_olustur():
    """Kat sorumlusu gün sonu raporu oluştur"""
    from flask import session
    from utils.rapor_servisleri import KatSorumlusuGunSonuRaporServisi
    
    try:
        kullanici_id = session.get('kullanici_id')
        kullanici = Kullanici.query.get(kullanici_id)
        
        if not kullanici or not kullanici.otel_id:
            flash('Otel atamanız bulunamadı.', 'danger')
            return redirect(url_for('kat_sorumlusu_dashboard'))
        
        otel_id = kullanici.otel_id
        personel_ids_str = request.form.get('personel_ids', '')
        tarih_str = request.form.get('tarih')
        export_format = request.form.get('format', '')
        
        # Personel ID'lerini parse et
        personel_ids = [int(x) for x in personel_ids_str.split(',') if x.strip().isdigit()] if personel_ids_str else None
        
        tarih = datetime.strptime(tarih_str, '%Y-%m-%d').date() if tarih_str else date.today()
        
        rapor = KatSorumlusuGunSonuRaporServisi.get_gun_sonu_raporu(
            otel_id=otel_id,
            personel_ids=personel_ids,
            tarih=tarih
        )
        
        if not rapor['success']:
            flash(rapor.get('message', 'Rapor oluşturulamadı'), 'danger')
            return redirect(url_for('raporlar.kat_sorumlusu_gun_sonu_raporum'))
        
        # Excel export
        if export_format == 'excel':
            return export_gun_sonu_excel(rapor)
        
        # PDF export
        if export_format == 'pdf':
            return export_gun_sonu_pdf(rapor)
        
        otel = Otel.query.get(otel_id)
        personeller = Kullanici.query.filter(
            Kullanici.rol == 'kat_sorumlusu',
            Kullanici.aktif.is_(True),
            Kullanici.otel_id == otel_id
        ).order_by(Kullanici.ad, Kullanici.soyad).all()
        
        log_islem('view', 'kat_sorumlusu_gun_sonu_raporum', {
            'otel_id': otel_id, 'tarih': tarih_str
        })
        
        return render_template('raporlar/kat_sorumlusu_gun_sonu_raporum.html',
                             otel=otel,
                             personeller=personeller,
                             rapor_verisi=rapor,
                             secili_tarih=tarih_str,
                             current_user_id=kullanici_id)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_gun_sonu_raporum_olustur')
        flash('Rapor oluşturulamadı.', 'danger')
        return redirect(url_for('raporlar.kat_sorumlusu_gun_sonu_raporum'))


@raporlar_bp.route('/kat-sorumlusu/gun-sonu-excel')
@login_required
@role_required('kat_sorumlusu')
def kat_sorumlusu_gun_sonu_excel():
    """Kat sorumlusu gün sonu raporu Excel export"""
    from flask import session
    from utils.rapor_servisleri import KatSorumlusuGunSonuRaporServisi
    
    try:
        kullanici_id = session.get('kullanici_id')
        kullanici = Kullanici.query.get(kullanici_id)
        
        if not kullanici or not kullanici.otel_id:
            flash('Otel atamanız bulunamadı.', 'danger')
            return redirect(url_for('raporlar.kat_sorumlusu_gun_sonu_raporum'))
        
        otel_id = kullanici.otel_id
        personel_ids_str = request.args.get('personel_ids', '')
        tarih_str = request.args.get('tarih')
        
        personel_ids = [int(x) for x in personel_ids_str.split(',') if x.strip().isdigit()] if personel_ids_str else None
        tarih = datetime.strptime(tarih_str, '%Y-%m-%d').date() if tarih_str else date.today()
        
        rapor = KatSorumlusuGunSonuRaporServisi.get_gun_sonu_raporu(
            otel_id=otel_id,
            personel_ids=personel_ids,
            tarih=tarih
        )
        
        if not rapor['success']:
            flash('Rapor oluşturulamadı.', 'danger')
            return redirect(url_for('raporlar.kat_sorumlusu_gun_sonu_raporum'))
        
        return export_gun_sonu_excel(rapor)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_gun_sonu_excel')
        flash('Excel export başarısız.', 'danger')
        return redirect(url_for('raporlar.kat_sorumlusu_gun_sonu_raporum'))


@raporlar_bp.route('/kat-sorumlusu/gun-sonu-pdf')
@login_required
@role_required('kat_sorumlusu')
def kat_sorumlusu_gun_sonu_pdf():
    """Kat sorumlusu gün sonu raporu PDF export"""
    from flask import session
    from utils.rapor_servisleri import KatSorumlusuGunSonuRaporServisi
    
    try:
        kullanici_id = session.get('kullanici_id')
        kullanici = Kullanici.query.get(kullanici_id)
        
        if not kullanici or not kullanici.otel_id:
            flash('Otel atamanız bulunamadı.', 'danger')
            return redirect(url_for('raporlar.kat_sorumlusu_gun_sonu_raporum'))
        
        otel_id = kullanici.otel_id
        personel_ids_str = request.args.get('personel_ids', '')
        tarih_str = request.args.get('tarih')
        
        personel_ids = [int(x) for x in personel_ids_str.split(',') if x.strip().isdigit()] if personel_ids_str else None
        tarih = datetime.strptime(tarih_str, '%Y-%m-%d').date() if tarih_str else date.today()
        
        rapor = KatSorumlusuGunSonuRaporServisi.get_gun_sonu_raporu(
            otel_id=otel_id,
            personel_ids=personel_ids,
            tarih=tarih
        )
        
        if not rapor['success']:
            flash('Rapor oluşturulamadı.', 'danger')
            return redirect(url_for('raporlar.kat_sorumlusu_gun_sonu_raporum'))
        
        return export_gun_sonu_pdf(rapor)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_gun_sonu_pdf')
        flash('PDF export başarısız.', 'danger')
        return redirect(url_for('raporlar.kat_sorumlusu_gun_sonu_raporum'))
