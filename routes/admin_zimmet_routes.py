"""
Admin Zimmet Route'ları

Bu modül admin zimmet yönetimi ile ilgili endpoint'leri içerir.

Endpoint'ler:
- /admin/personel-zimmetleri - Otel bazlı personel zimmet raporu
- /admin/zimmet-detay/<int:zimmet_id> - Zimmet detaylarını görüntüleme
- /admin/zimmet-iade/<int:zimmet_id> - Zimmet iade işlemi
- /admin/zimmet-iptal/<int:zimmet_id> - Zimmet iptal işlemi

Roller:
- sistem_yoneticisi
- admin
"""

from flask import render_template, request, redirect, url_for, flash, session, jsonify, send_file
from models import db, PersonelZimmet, PersonelZimmetDetay, Kullanici, StokHareket, Otel, OtelZimmetStok, PersonelZimmetKullanim, Urun
from utils.decorators import login_required, role_required
from utils.helpers import log_islem, log_hata, get_kktc_now
from utils.audit import serialize_model, audit_update
from sqlalchemy import func
import io


def register_admin_zimmet_routes(app):
    """Admin zimmet route'larını kaydet"""
    
    # Excel export fonksiyonu
    def export_personel_zimmet_excel_func(otel, kat_sorumlulari, zimmet_stoklari, personel_kullanimlari):
        """Otel zimmet raporunu Excel'e aktar"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            
            # ===== SAYFA 1: ÖZET =====
            ws_ozet = wb.active
            ws_ozet.title = "Rapor"
            
            # Başlık
            ws_ozet['A1'] = f"{otel.ad} - Zimmet Raporu"
            ws_ozet['A1'].font = Font(bold=True, size=16)
            ws_ozet.merge_cells('A1:F1')
            
            ws_ozet['A2'] = f"Rapor Tarihi: {get_kktc_now().strftime('%d.%m.%Y %H:%M')}"
            ws_ozet['A2'].font = Font(italic=True, size=10)
            
            # Kat Sorumluları
            ws_ozet['A4'] = "Kat Sorumluları:"
            ws_ozet['A4'].font = Font(bold=True)
            
            kat_isimleri = ", ".join([f"{p.ad} {p.soyad}" for p in kat_sorumlulari]) if kat_sorumlulari else "Yok"
            ws_ozet['B4'] = kat_isimleri
            
            # ===== ZİMMET STOKLARI (Aynı sayfada) =====
            ws_ozet['A6'] = "Zimmet Stokları"
            ws_ozet['A6'].font = Font(bold=True, size=12)
            
            stok_headers = ['Ürün Adı', 'Birim', 'Toplam', 'Kullanılan', 'Kalan', 'Durum']
            for col, header in enumerate(stok_headers, 1):
                cell = ws_ozet.cell(row=7, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for i, stok in enumerate(zimmet_stoklari, start=8):
                ws_ozet.cell(row=i, column=1, value=stok.urun.urun_adi if stok.urun else '-')
                ws_ozet.cell(row=i, column=2, value=stok.urun.birim if stok.urun else '-')
                ws_ozet.cell(row=i, column=3, value=stok.toplam_miktar)
                ws_ozet.cell(row=i, column=4, value=stok.kullanilan_miktar)
                ws_ozet.cell(row=i, column=5, value=stok.kalan_miktar)
                
                durum = stok.stok_durumu
                durum_text = {'normal': 'Normal', 'dikkat': 'Dikkat', 'kritik': 'Kritik', 'stokout': 'Stok Yok'}.get(durum, durum)
                cell = ws_ozet.cell(row=i, column=6, value=durum_text)
                
                if durum == 'kritik' or durum == 'stokout':
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif durum == 'dikkat':
                    cell.fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
            
            # Sütun genişlikleri
            ws_ozet.column_dimensions['A'].width = 35
            ws_ozet.column_dimensions['B'].width = 15
            ws_ozet.column_dimensions['C'].width = 12
            ws_ozet.column_dimensions['D'].width = 12
            ws_ozet.column_dimensions['E'].width = 12
            ws_ozet.column_dimensions['F'].width = 12
            
            # Buffer'a kaydet
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            log_hata(e, modul='export_personel_zimmet_excel')
            return None
    
    @app.route('/admin/personel-zimmetleri')
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def admin_personel_zimmetleri():
        """Otel bazlı personel zimmet raporu"""
        try:
            # Filtreler
            otel_id = request.args.get('otel_id', type=int)
            export_format = request.args.get('format', '')
            
            # Oteller listesi
            oteller = Otel.query.filter_by(aktif=True).order_by(Otel.ad).all()
            
            # Seçili otel bilgisi
            secili_otel = None
            kat_sorumlulari = []
            zimmet_stoklari = []
            personel_kullanimlari = []
            ozet_bilgiler = {}
            
            if otel_id:
                secili_otel = Otel.query.get(otel_id)
                
                if secili_otel:
                    # O oteldeki kat sorumluları
                    kat_sorumlulari = Kullanici.query.filter(
                        Kullanici.otel_id == otel_id,
                        Kullanici.rol == 'kat_sorumlusu',
                        Kullanici.aktif.is_(True)
                    ).order_by(Kullanici.ad, Kullanici.soyad).all()
                    
                    # Otel zimmet stokları
                    zimmet_stoklari = OtelZimmetStok.query.options(
                        db.joinedload(OtelZimmetStok.urun)
                    ).filter(
                        OtelZimmetStok.otel_id == otel_id
                    ).order_by(OtelZimmetStok.kalan_miktar.asc()).all()
                    
                    # Personel bazlı kullanım özeti
                    kullanim_query = db.session.query(
                        PersonelZimmetKullanim.personel_id,
                        Kullanici.ad,
                        Kullanici.soyad,
                        func.sum(PersonelZimmetKullanim.kullanilan_miktar).label('toplam_kullanim'),
                        func.count(PersonelZimmetKullanim.id).label('islem_sayisi'),
                        func.max(PersonelZimmetKullanim.islem_tarihi).label('son_islem')
                    ).join(
                        Kullanici, PersonelZimmetKullanim.personel_id == Kullanici.id
                    ).join(
                        OtelZimmetStok, PersonelZimmetKullanim.otel_zimmet_stok_id == OtelZimmetStok.id
                    ).filter(
                        OtelZimmetStok.otel_id == otel_id
                    ).group_by(
                        PersonelZimmetKullanim.personel_id,
                        Kullanici.ad,
                        Kullanici.soyad
                    ).order_by(
                        func.sum(PersonelZimmetKullanim.kullanilan_miktar).desc()
                    ).all()
                    
                    personel_kullanimlari = [{
                        'personel_id': k.personel_id,
                        'ad_soyad': f"{k.ad} {k.soyad}",
                        'toplam_kullanim': k.toplam_kullanim or 0,
                        'islem_sayisi': k.islem_sayisi or 0,
                        'son_islem': k.son_islem
                    } for k in kullanim_query]
                    
                    # Özet bilgiler
                    toplam_stok = sum(z.toplam_miktar for z in zimmet_stoklari)
                    toplam_kullanilan = sum(z.kullanilan_miktar for z in zimmet_stoklari)
                    toplam_kalan = sum(z.kalan_miktar for z in zimmet_stoklari)
                    kritik_urun_sayisi = len([z for z in zimmet_stoklari if z.stok_durumu in ['kritik', 'stokout']])
                    
                    ozet_bilgiler = {
                        'toplam_stok': toplam_stok,
                        'toplam_kullanilan': toplam_kullanilan,
                        'toplam_kalan': toplam_kalan,
                        'kritik_urun_sayisi': kritik_urun_sayisi,
                        'urun_cesidi': len(zimmet_stoklari),
                        'personel_sayisi': len(kat_sorumlulari)
                    }
            
            # Excel export
            if export_format == 'excel' and otel_id and secili_otel:
                excel_buffer = export_personel_zimmet_excel_func(secili_otel, kat_sorumlulari, zimmet_stoklari, personel_kullanimlari)
                if excel_buffer:
                    safe_otel_adi = secili_otel.ad.replace(' ', '_').replace('/', '_')
                    filename = f'otel_zimmet_raporu_{safe_otel_adi}_{get_kktc_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                    return send_file(
                        excel_buffer,
                        as_attachment=True,
                        download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                else:
                    flash('Excel dosyası oluşturulamadı.', 'danger')
            
            # Log kaydı
            log_islem('goruntuleme', 'personel_zimmetleri', {
                'otel_id': otel_id,
                'personel_sayisi': len(kat_sorumlulari),
                'stok_kayit_sayisi': len(zimmet_stoklari)
            })
            
            return render_template('sistem_yoneticisi/admin_personel_zimmetleri.html',
                                 oteller=oteller,
                                 secili_otel=secili_otel,
                                 secili_otel_id=otel_id,
                                 kat_sorumlulari=kat_sorumlulari,
                                 zimmet_stoklari=zimmet_stoklari,
                                 personel_kullanimlari=personel_kullanimlari,
                                 ozet_bilgiler=ozet_bilgiler)
            
        except Exception as e:
            log_hata(e, modul='admin_personel_zimmetleri')
            flash('Zimmet raporu yüklenirken hata oluştu.', 'danger')
            return redirect(url_for('sistem_yoneticisi_dashboard'))

    @app.route('/admin/zimmet-detay/<int:zimmet_id>')
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def admin_zimmet_detay(zimmet_id):
        """Zimmet detaylarını görüntüle"""
        try:
            zimmet = db.session.get(PersonelZimmet, zimmet_id)
            if not zimmet:
                flash('Zimmet kaydı bulunamadı.', 'danger')
                return redirect(url_for('admin_personel_zimmetleri'))
            
            # Detayları eager load ile getir
            zimmet = PersonelZimmet.query.options(
                db.joinedload(PersonelZimmet.personel),
                db.joinedload(PersonelZimmet.teslim_eden),
                db.joinedload(PersonelZimmet.detaylar).joinedload(PersonelZimmetDetay.urun)
            ).get(zimmet_id)
            
            # Log kaydı
            log_islem('goruntuleme', 'zimmet_detay', {
                'zimmet_id': zimmet_id,
                'personel_id': zimmet.personel_id
            })
            
            return render_template('sistem_yoneticisi/admin_zimmet_detay.html',
                                 zimmet=zimmet)
            
        except Exception as e:
            log_hata(e, modul='admin_zimmet_detay')
            flash('Zimmet detayları yüklenirken hata oluştu.', 'danger')
            return redirect(url_for('admin_personel_zimmetleri'))

    @app.route('/admin/zimmet-iade/<int:zimmet_id>', methods=['POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def admin_zimmet_iade(zimmet_id):
        """Zimmet iade işlemi"""
        try:
            zimmet = db.session.get(PersonelZimmet, zimmet_id)
            if not zimmet:
                return jsonify({'success': False, 'message': 'Zimmet bulunamadı'}), 404
            
            if zimmet.durum != 'aktif':
                return jsonify({'success': False, 'message': 'Sadece aktif zimmetler iade edilebilir'}), 400
            
            # İade edilen miktarları al
            data = request.get_json()
            iade_miktarlari = data.get('iade_miktarlari', {})
            
            # Her detay için iade işlemi
            for detay in zimmet.detaylar:
                detay_id = str(detay.id)
                if detay_id in iade_miktarlari:
                    iade_miktar = int(iade_miktarlari[detay_id])
                    if iade_miktar > 0:
                        # İade miktarını güncelle
                        detay.iade_edilen_miktar += iade_miktar
                        detay.kalan_miktar = detay.miktar - detay.kullanilan_miktar - detay.iade_edilen_miktar
                        
                        # Stok hareketine ekle
                        hareket = StokHareket(
                            urun_id=detay.urun_id,
                            hareket_tipi='giris',
                            miktar=iade_miktar,
                            aciklama=f'Zimmet iadesi - {zimmet.personel.ad} {zimmet.personel.soyad}',
                            islem_yapan_id=session['kullanici_id']
                        )
                        db.session.add(hareket)
            
            # Zimmet durumunu güncelle
            zimmet.durum = 'tamamlandi'
            db.session.commit()
            
            # Audit log
            audit_update(
                tablo_adi='personel_zimmet',
                kayit_id=zimmet_id,
                eski_deger={'durum': 'aktif'},
                yeni_deger={'durum': 'tamamlandi'},
                aciklama='Admin zimmet iade işlemi'
            )
            
            # Log kaydı
            log_islem('guncelleme', 'zimmet_iade', {
                'zimmet_id': zimmet_id,
                'personel_id': zimmet.personel_id
            })
            
            flash('Zimmet iade işlemi başarıyla tamamlandı.', 'success')
            return jsonify({'success': True, 'message': 'İade işlemi tamamlandı'})
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, modul='admin_zimmet_iade')
            return jsonify({'success': False, 'message': 'İade işlemi başarısız'}), 500

    @app.route('/admin/zimmet-iptal/<int:zimmet_id>', methods=['POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def admin_zimmet_iptal(zimmet_id):
        """Zimmet kaydını iptal et - FIFO kayıtlarını geri al"""
        try:
            from models import UrunStok, StokFifoKayit, StokFifoKullanim
            
            zimmet = db.session.get(PersonelZimmet, zimmet_id)
            if not zimmet:
                return jsonify({'success': False, 'message': 'Zimmet bulunamadı'}), 404
            
            if zimmet.durum != 'aktif':
                return jsonify({'success': False, 'message': 'Sadece aktif zimmetler iptal edilebilir'}), 400
            
            # Eski değeri sakla
            eski_deger = serialize_model(zimmet)
            otel_id = zimmet.personel.otel_id if zimmet.personel else None
            
            # Stok hareketlerini ve FIFO kayıtlarını geri al
            for detay in zimmet.detaylar:
                # Kullanılmayan miktarı depoya geri ekle
                geri_alinacak = detay.miktar - detay.kullanilan_miktar - detay.iade_edilen_miktar
                if geri_alinacak > 0:
                    # Stok hareketi
                    hareket = StokHareket(
                        urun_id=detay.urun_id,
                        hareket_tipi='giris',
                        miktar=geri_alinacak,
                        aciklama=f'Zimmet iptali - {zimmet.personel.ad} {zimmet.personel.soyad}',
                        islem_yapan_id=session['kullanici_id']
                    )
                    db.session.add(hareket)
                    
                    # UrunStok güncelle
                    urun_stok = UrunStok.query.filter_by(otel_id=otel_id, urun_id=detay.urun_id).first()
                    if urun_stok:
                        urun_stok.mevcut_stok += geri_alinacak
                        urun_stok.son_giris_tarihi = get_kktc_now()
                        urun_stok.son_guncelleyen_id = session['kullanici_id']
                
                # FIFO kullanım kayıtlarını geri al
                fifo_kullanimlar = StokFifoKullanim.query.filter_by(
                    referans_id=detay.id,
                    islem_tipi='zimmet'
                ).all()
                
                for kullanim in fifo_kullanimlar:
                    fifo_kayit = db.session.get(StokFifoKayit, kullanim.fifo_kayit_id)
                    if fifo_kayit:
                        fifo_kayit.kalan_miktar += kullanim.miktar
                        fifo_kayit.kullanilan_miktar -= kullanim.miktar
                        fifo_kayit.tukendi = False
                    db.session.delete(kullanim)
            
            # Zimmet durumunu iptal et
            zimmet.durum = 'iptal'
            db.session.commit()
            
            # Audit log
            audit_update(
                tablo_adi='personel_zimmet',
                kayit_id=zimmet_id,
                eski_deger=eski_deger,
                yeni_deger=serialize_model(zimmet),
                aciklama='Admin zimmet iptal işlemi'
            )
            
            # Log kaydı
            log_islem('guncelleme', 'zimmet_iptal', {
                'zimmet_id': zimmet_id,
                'personel_id': zimmet.personel_id
            })
            
            flash('Zimmet kaydı başarıyla iptal edildi.', 'success')
            return jsonify({'success': True, 'message': 'Zimmet iptal edildi, stoklar geri eklendi'})
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, modul='admin_zimmet_iptal')
            return jsonify({'success': False, 'message': 'İptal işlemi başarısız'}), 500
