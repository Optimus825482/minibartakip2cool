"""
Sistem Yöneticisi Route'ları

Bu modül sistem yöneticisi ve admin rolüne ait endpoint'leri içerir.

Endpoint'ler:
- /sistem-loglari - Sistem logları
- /otel-tanimla - Otel bilgileri
- /kat-tanimla - Kat tanımlama
- /kat-duzenle/<int:kat_id> - Kat düzenleme
- /kat-sil/<int:kat_id> - Kat silme
- /oda-tanimla - Oda tanımlama
- /oda-duzenle/<int:oda_id> - Oda düzenleme
- /oda-sil/<int:oda_id> - Oda silme

Roller:
- sistem_yoneticisi
- admin
"""

from flask import render_template, request, redirect, url_for, flash, session, jsonify
from datetime import datetime, date, timezone
import pytz
import logging
from models import db, Otel, Kat, Oda, OdaTipi, Kullanici, SistemLog
from utils.decorators import login_required, role_required
from utils.helpers import log_islem, log_hata
from utils.audit import audit_create, audit_update, audit_delete, serialize_model

# Logger tanımla
logger = logging.getLogger(__name__)

# KKTC Timezone (Kıbrıs - Europe/Nicosia)
KKTC_TZ = pytz.timezone('Europe/Nicosia')

def get_kktc_now():
    """Kıbrıs saat diliminde şu anki zamanı döndürür."""
    return datetime.now(KKTC_TZ)


def register_sistem_yoneticisi_routes(app):
    """Sistem yöneticisi route'larını kaydet"""
    
    # CSRF protection instance'ını al
    csrf = app.extensions.get('csrf')
    
    @app.route('/sistem-loglari')
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def sistem_loglari():
        """Sistem logları listesi"""
        # Sayfa parametreleri
        sayfa = request.args.get('sayfa', 1, type=int)
        limit = 50
        
        # Filtreler
        islem_tipi = request.args.get('islem_tipi', '')
        modul = request.args.get('modul', '')
        kullanici_id = request.args.get('kullanici_id', type=int)
        
        # Sorgu oluştur
        query = SistemLog.query
        
        if islem_tipi:
            query = query.filter(SistemLog.islem_tipi == islem_tipi)
        if modul:
            query = query.filter(SistemLog.modul == modul)
        if kullanici_id:
            query = query.filter(SistemLog.kullanici_id == kullanici_id)
        
        # Sayfalama
        loglar = query.order_by(SistemLog.islem_tarihi.desc()).paginate(
            page=sayfa, per_page=limit, error_out=False
        )
        
        # Filtre seçenekleri
        kullanicilar = Kullanici.query.filter(Kullanici.aktif.is_(True)).order_by(Kullanici.ad, Kullanici.soyad).all()
        
        return render_template('sistem_yoneticisi/sistem_loglari.html',
                             loglar=loglar,
                             kullanicilar=kullanicilar,
                             islem_tipi=islem_tipi,
                             modul=modul,
                             kullanici_id=kullanici_id)
    
    
    @app.route('/otel-tanimla', methods=['GET', 'POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def otel_tanimla():
        """Otel bilgileri tanımlama/düzenleme"""
        from forms import OtelForm
        
        otel = Otel.query.first()
        form = OtelForm(obj=otel)
        
        if form.validate_on_submit():
            try:
                # Logo işleme
                import base64
                logo_data = None
                if form.logo.data:
                    logo_file = form.logo.data
                    logo_bytes = logo_file.read()
                    logo_data = base64.b64encode(logo_bytes).decode('utf-8')
                
                if otel:
                    # Eski değeri sakla
                    eski_deger = serialize_model(otel)
                    
                    # Güncelle
                    otel.ad = form.ad.data
                    otel.adres = form.adres.data
                    otel.telefon = form.telefon.data
                    otel.email = form.email.data
                    otel.vergi_no = form.vergi_no.data or ''
                    
                    # Logo güncelle (sadece yeni logo yüklendiyse)
                    if logo_data:
                        otel.logo = logo_data
                    
                    # Audit log
                    audit_update(
                        tablo_adi='oteller',
                        kayit_id=otel.id,
                        eski_deger=eski_deger,
                        yeni_deger=serialize_model(otel),
                        aciklama='Otel bilgileri güncellendi'
                    )
                else:
                    # Yeni oluştur
                    otel = Otel(
                        ad=form.ad.data,
                        adres=form.adres.data,
                        telefon=form.telefon.data,
                        email=form.email.data,
                        vergi_no=form.vergi_no.data or '',
                        logo=logo_data
                    )
                    db.session.add(otel)
                    db.session.flush()
                    
                    # Audit log
                    audit_create(
                        tablo_adi='oteller',
                        kayit_id=otel.id,
                        yeni_deger=serialize_model(otel),
                        aciklama='Otel bilgileri oluşturuldu'
                    )
                
                db.session.commit()
                
                # Log kaydı
                log_islem('guncelleme' if otel else 'ekleme', 'otel', {
                    'otel_id': otel.id,
                    'otel_adi': otel.ad
                })
                
                flash('Otel bilgileri başarıyla kaydedildi.', 'success')
                return redirect(url_for('otel_tanimla'))
            except Exception as e:
                db.session.rollback()
                log_hata(e, modul='otel_tanimla')
                flash('Otel bilgileri kaydedilirken hata oluştu.', 'danger')
        
        return render_template('sistem_yoneticisi/otel_tanimla.html', otel=otel, form=form)
    
    
    @app.route('/kat-tanimla', methods=['GET', 'POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def kat_tanimla():
        """Kat tanımlama"""
        try:
            from forms import KatForm
            from models import Otel
            
            form = KatForm()
            
            # Otel listesini form'a yükle
            form.otel_id.choices = [(0, 'Otel Seçin...')] + [(o.id, o.ad) for o in Otel.query.filter_by(aktif=True).order_by(Otel.ad).all()]
            
            if form.validate_on_submit():
                try:
                    # Otel seçimi kontrolü
                    if not form.otel_id.data or form.otel_id.data == 0:
                        flash('Lütfen bir otel seçin!', 'danger')
                        return render_template('sistem_yoneticisi/kat_tanimla.html', katlar=[], form=form)
                    
                    kat = Kat(
                        otel_id=form.otel_id.data,
                        kat_adi=form.kat_adi.data,
                        kat_no=form.kat_no.data,
                        aciklama=form.aciklama.data,
                        aktif=form.aktif.data
                    )
                    db.session.add(kat)
                    db.session.flush()
                    
                    # Audit log
                    audit_create(
                        tablo_adi='katlar',
                        kayit_id=kat.id,
                        yeni_deger=serialize_model(kat),
                        aciklama=f'Kat oluşturuldu - {kat.kat_adi}'
                    )
                    
                    db.session.commit()
                    
                    # Log kaydı
                    log_islem('ekleme', 'kat', {
                        'kat_id': kat.id,
                        'otel_id': kat.otel_id,
                        'kat_adi': kat.kat_adi,
                        'kat_no': kat.kat_no
                    })
                    
                    flash(f'{kat.kat_adi} başarıyla eklendi.', 'success')
                    return redirect(url_for('kat_tanimla'))
                except Exception as e:
                    db.session.rollback()
                    log_hata(e, modul='kat_tanimla')
                    flash('Kat eklenirken hata oluştu.', 'danger')
                    # Exception durumunda da katları listele
                    katlar = Kat.query.filter_by(aktif=True).order_by(Kat.kat_no).all()
                    return render_template('sistem_yoneticisi/kat_tanimla.html', katlar=katlar, form=form)
            
            # Mevcut katları listele (otel bilgisi ile)
            katlar = Kat.query.filter_by(aktif=True).order_by(Kat.kat_no).all()
            
            return render_template('sistem_yoneticisi/kat_tanimla.html', katlar=katlar, form=form)
        except Exception as e:
            import traceback
            print(f"❌ KAT TANIMLA HATASI: {e}")
            print(traceback.format_exc())
            flash('Sayfa yüklenirken hata oluştu.', 'danger')
            return redirect(url_for('dashboard'))
    
    
    @app.route('/kat-duzenle/<int:kat_id>', methods=['GET', 'POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def kat_duzenle(kat_id):
        """Kat düzenleme"""
        from forms import KatForm
        from models import Otel
        
        kat = db.session.get(Kat, kat_id)
        if not kat:
            flash('Kat bulunamadı.', 'danger')
            return redirect(url_for('kat_tanimla'))
        
        # Eski değeri sakla
        eski_deger = serialize_model(kat)
        
        form = KatForm(obj=kat)
        
        # Otel listesini form'a yükle
        form.otel_id.choices = [(o.id, o.ad) for o in Otel.query.filter_by(aktif=True).order_by(Otel.ad).all()]
        
        if form.validate_on_submit():
            try:
                kat.otel_id = form.otel_id.data
                kat.kat_adi = form.kat_adi.data
                kat.kat_no = form.kat_no.data
                kat.aciklama = form.aciklama.data
                kat.aktif = form.aktif.data
                
                db.session.commit()
                
                # Audit log
                audit_update(
                    tablo_adi='katlar',
                    kayit_id=kat.id,
                    eski_deger=eski_deger,
                    yeni_deger=serialize_model(kat),
                    aciklama=f'Kat güncellendi - {kat.kat_adi}'
                )
                
                # Log kaydı
                log_islem('guncelleme', 'kat', {
                    'kat_id': kat.id,
                    'kat_adi': kat.kat_adi
                })
                
                flash(f'{kat.kat_adi} başarıyla güncellendi.', 'success')
                return redirect(url_for('kat_tanimla'))
            except Exception as e:
                db.session.rollback()
                log_hata(e, modul='kat_duzenle')
                flash('Kat güncellenirken hata oluştu.', 'danger')
        
        return render_template('sistem_yoneticisi/kat_duzenle.html', kat=kat, form=form)
    
    
    @app.route('/kat-sil/<int:kat_id>', methods=['POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def kat_sil(kat_id):
        """Kat silme (pasif yapma)"""
        try:
            kat = db.session.get(Kat, kat_id)
            if not kat:
                flash('Kat bulunamadı.', 'danger')
                return redirect(url_for('kat_tanimla'))
            
            # Eski değeri sakla
            eski_deger = serialize_model(kat)
            
            # Pasif yap
            kat.aktif = False
            db.session.commit()
            
            # Audit log
            audit_delete(
                tablo_adi='katlar',
                kayit_id=kat.id,
                eski_deger=eski_deger,
                aciklama=f'Kat pasif yapıldı - {kat.kat_adi}'
            )
            
            # Log kaydı
            log_islem('silme', 'kat', {
                'kat_id': kat.id,
                'kat_adi': kat.kat_adi
            })
            
            flash(f'{kat.kat_adi} başarıyla silindi.', 'success')
        except Exception as e:
            db.session.rollback()
            log_hata(e, modul='kat_sil')
            flash('Kat silinirken hata oluştu.', 'danger')
        
        return redirect(url_for('kat_tanimla'))
    
    
    @app.route('/oda-tanimla', methods=['GET', 'POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def oda_tanimla():
        """Oda tanımlama"""
        from forms import OdaForm
        from models import Otel
        from flask import send_file
        from datetime import datetime
        import io
        
        # Excel export kontrolü
        export_format = request.args.get('format', '')
        if export_format == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                # Tüm odaları getir
                odalar = Oda.query.options(
                    db.joinedload(Oda.kat).joinedload(Kat.otel)
                ).filter_by(aktif=True).order_by(Oda.oda_no).all()
                
                # Excel workbook oluştur
                wb = Workbook()
                ws = wb.active
                ws.title = "Odalar"
                
                # Başlık satırı
                headers = ['Otel Adı', 'Kat No', 'Oda No', 'Oda Tipi']
                ws.append(headers)
                
                # Başlık stilini ayarla
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Veri satırları
                for oda in odalar:
                    otel_adi = oda.kat.otel.ad if oda.kat and oda.kat.otel else '-'
                    kat_no = oda.kat.kat_no if oda.kat else '-'
                    oda_tipi = oda.oda_tipi_adi if oda.oda_tipi_adi else '-'
                    
                    # Oda numarasını sayıya çevir (mümkünse)
                    try:
                        oda_no_sayi = int(oda.oda_no)
                    except (ValueError, TypeError):
                        oda_no_sayi = oda.oda_no
                    
                    ws.append([
                        otel_adi,
                        kat_no,
                        oda_no_sayi,
                        oda_tipi
                    ])
                
                # Sütun genişliklerini ayarla
                ws.column_dimensions['A'].width = 30  # Otel Adı
                ws.column_dimensions['B'].width = 15  # Kat No
                ws.column_dimensions['C'].width = 15  # Oda No
                ws.column_dimensions['D'].width = 20  # Oda Tipi
                
                # Oda No sütununu sayı formatına çevir (C sütunu, 2. satırdan itibaren)
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=3)  # C sütunu (Oda No)
                    if isinstance(cell.value, int):
                        cell.number_format = '0'  # Sayı formatı (ondalık yok)
                
                # Excel dosyasını memory'ye kaydet
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                filename = f'odalar_{get_kktc_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                
                # Log kaydı
                log_islem('export', 'odalar', {
                    'format': 'excel',
                    'kayit_sayisi': len(odalar)
                })
                
                return send_file(
                    excel_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
            except Exception as e:
                log_hata(e, modul='oda_tanimla_excel')
                flash('Excel dosyası oluşturulamadı.', 'danger')
                return redirect(url_for('oda_tanimla'))
        
        form = OdaForm()
        
        # Otel listesini form'a yükle
        oteller = Otel.query.filter_by(aktif=True).order_by(Otel.ad).all()
        form.otel_id.choices = [(0, 'Otel Seçin...')] + [(o.id, o.ad) for o in oteller]
        form.kat_id.choices = [(0, 'Önce otel seçin...')]
        
        # Oda tipi listesini form'a yükle
        from models import OdaTipi
        oda_tipleri = OdaTipi.query.filter_by(aktif=True).order_by(OdaTipi.ad).all()
        form.oda_tipi_id.choices = [(0, 'Oda Tipi Seçin...')] + [(t.id, t.ad) for t in oda_tipleri]
        
        if form.validate_on_submit():
            try:
                # Otel ve kat seçimi kontrolü
                if not form.otel_id.data or form.otel_id.data == 0:
                    flash('Lütfen bir otel seçin!', 'danger')
                    return render_template('sistem_yoneticisi/oda_tanimla.html', katlar=[], odalar=[], form=form)
                
                if not form.kat_id.data or form.kat_id.data == 0:
                    flash('Lütfen bir kat seçin!', 'danger')
                    return render_template('sistem_yoneticisi/oda_tanimla.html', katlar=[], odalar=[], form=form)
                
                # Kat'ın seçilen otele ait olduğunu kontrol et
                kat = Kat.query.get(form.kat_id.data)
                if not kat or kat.otel_id != form.otel_id.data:
                    flash('Seçilen kat, seçilen otele ait değil!', 'danger')
                    return render_template('sistem_yoneticisi/oda_tanimla.html', katlar=[], odalar=[], form=form)
                
                # Oda numarası duplikasyon kontrolü
                mevcut_oda = Oda.query.filter_by(oda_no=form.oda_no.data, aktif=True).first()
                if mevcut_oda:
                    flash(f'Oda numarası "{form.oda_no.data}" zaten kullanılıyor! Lütfen farklı bir oda numarası girin.', 'warning')
                    return render_template('sistem_yoneticisi/oda_tanimla.html', katlar=[], odalar=[], form=form)
                
                oda = Oda(
                    oda_no=form.oda_no.data,
                    kat_id=form.kat_id.data,
                    oda_tipi_id=form.oda_tipi_id.data if form.oda_tipi_id.data else None,
                    kapasite=form.kapasite.data,
                    aktif=form.aktif.data
                )
                db.session.add(oda)
                db.session.flush()
                
                # Audit log
                audit_create(
                    tablo_adi='odalar',
                    kayit_id=oda.id,
                    yeni_deger=serialize_model(oda),
                    aciklama=f'Oda oluşturuldu - {oda.oda_no}'
                )
                
                db.session.commit()
                
                # Log kaydı
                log_islem('ekleme', 'oda', {
                    'oda_id': oda.id,
                    'oda_no': oda.oda_no,
                    'kat_id': oda.kat_id
                })
                
                flash(f'Oda {oda.oda_no} başarıyla eklendi.', 'success')
                return redirect(url_for('oda_tanimla'))
            except Exception as e:
                db.session.rollback()
                log_hata(e, modul='oda_tanimla')
                flash('Oda eklenirken hata oluştu.', 'danger')
        
        # Katlar ve odalar (eager loading ile otel bilgisini de çek)
        katlar = Kat.query.filter_by(aktif=True).order_by(Kat.kat_no).all()
        odalar = Oda.query.options(
            db.joinedload(Oda.kat).joinedload(Kat.otel)
        ).filter_by(aktif=True).order_by(Oda.oda_no).all()
        
        return render_template('sistem_yoneticisi/oda_tanimla.html', katlar=katlar, odalar=odalar, form=form, oteller=oteller)
    
    
    @app.route('/oda-duzenle/<int:oda_id>', methods=['GET', 'POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def oda_duzenle(oda_id):
        """Oda düzenleme"""
        from forms import OdaForm
        from models import Otel
        
        oda = db.session.get(Oda, oda_id)
        if not oda:
            flash('Oda bulunamadı.', 'danger')
            return redirect(url_for('oda_tanimla'))
        
        # Eski değeri sakla
        eski_deger = serialize_model(oda)
        
        form = OdaForm(obj=oda)
        
        # Otel listesini yükle
        oteller = Otel.query.filter_by(aktif=True).order_by(Otel.ad).all()
        form.otel_id.choices = [(o.id, o.ad) for o in oteller]
        
        # Mevcut otel seçimini ayarla
        if oda.kat and oda.kat.otel_id:
            form.otel_id.data = oda.kat.otel_id
        
        # Kat listesini yükle (mevcut otel için)
        if oda.kat:
            katlar = Kat.query.filter_by(otel_id=oda.kat.otel_id, aktif=True).order_by(Kat.kat_no).all()
            form.kat_id.choices = [(k.id, k.kat_adi) for k in katlar]
        
        # Oda tipi listesini yükle
        from models import OdaTipi
        oda_tipleri = OdaTipi.query.filter_by(aktif=True).order_by(OdaTipi.ad).all()
        form.oda_tipi_id.choices = [(0, 'Oda Tipi Seçin...')] + [(t.id, t.ad) for t in oda_tipleri]
        
        if form.validate_on_submit():
            try:
                # Kat'ın seçilen otele ait olduğunu kontrol et
                kat = Kat.query.get(form.kat_id.data)
                if not kat or kat.otel_id != form.otel_id.data:
                    flash('Seçilen kat, seçilen otele ait değil!', 'danger')
                    return render_template('sistem_yoneticisi/oda_duzenle.html', oda=oda, form=form, oteller=oteller)
                
                oda.oda_no = form.oda_no.data
                oda.kat_id = form.kat_id.data
                oda.oda_tipi_id = form.oda_tipi_id.data if form.oda_tipi_id.data else None
                oda.kapasite = form.kapasite.data
                oda.aktif = form.aktif.data
                
                db.session.commit()
                
                # Audit log
                audit_update(
                    tablo_adi='odalar',
                    kayit_id=oda.id,
                    eski_deger=eski_deger,
                    yeni_deger=serialize_model(oda),
                    aciklama=f'Oda güncellendi - {oda.oda_no}'
                )
                
                # Log kaydı
                log_islem('guncelleme', 'oda', {
                    'oda_id': oda.id,
                    'oda_no': oda.oda_no
                })
                
                flash(f'Oda {oda.oda_no} başarıyla güncellendi.', 'success')
                return redirect(url_for('oda_tanimla'))
            except Exception as e:
                db.session.rollback()
                log_hata(e, modul='oda_duzenle')
                flash('Oda güncellenirken hata oluştu.', 'danger')
        
        return render_template('sistem_yoneticisi/oda_duzenle.html', oda=oda, form=form, oteller=oteller)
    
    
    @app.route('/oda-sil/<int:oda_id>', methods=['POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def oda_sil(oda_id):
        """Oda silme (pasif yapma)"""
        try:
            oda = db.session.get(Oda, oda_id)
            if not oda:
                flash('Oda bulunamadı.', 'danger')
                return redirect(url_for('oda_tanimla'))
            
            # Eski değeri sakla
            eski_deger = serialize_model(oda)
            
            # Pasif yap
            oda.aktif = False
            db.session.commit()
            
            # Audit log
            audit_delete(
                tablo_adi='odalar',
                kayit_id=oda.id,
                eski_deger=eski_deger,
                aciklama=f'Oda pasif yapıldı - {oda.oda_no}'
            )
            
            # Log kaydı
            log_islem('silme', 'oda', {
                'oda_id': oda.id,
                'oda_no': oda.oda_no
            })
            
            flash(f'Oda {oda.oda_no} başarıyla silindi.', 'success')
        except Exception as e:
            db.session.rollback()
            log_hata(e, modul='oda_sil')
            flash('Oda silinirken hata oluştu.', 'danger')
        
        return redirect(url_for('oda_tanimla'))
    
    
    # ============================================================================
    # API ENDPOINTS - Kat Oda Tipleri
    # ============================================================================
    
    @app.route('/api/katlar/<int:kat_id>/oda-tipleri', methods=['GET'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_kat_oda_tipleri(kat_id):
        """Bir kattaki oda tiplerini ve sayılarını getir"""
        try:
            from sqlalchemy import func
            
            # Kat kontrolü
            kat = Kat.query.get_or_404(kat_id)
            
            # Oda tiplerini grupla ve say
            oda_tipleri = db.session.query(
                OdaTipi.ad,
                func.count(Oda.id).label('sayi')
            ).join(
                Oda, Oda.oda_tipi_id == OdaTipi.id
            ).filter(
                Oda.kat_id == kat_id,
                Oda.aktif == True
            ).group_by(
                OdaTipi.ad
            ).order_by(
                func.count(Oda.id).desc()
            ).all()
            
            # Sonuçları formatla
            sonuc = []
            for oda_tipi, sayi in oda_tipleri:
                sonuc.append({
                    'oda_tipi': oda_tipi if oda_tipi else 'Belirtilmemiş',
                    'sayi': sayi
                })
            
            return {
                'success': True,
                'oda_tipleri': sonuc,
                'kat': {
                    'id': kat.id,
                    'kat_adi': kat.kat_adi,
                    'otel_adi': kat.otel.ad if kat.otel else None
                }
            }
            
        except Exception as e:
            log_hata(e, 'api_kat_oda_tipleri')
            return {
                'success': False,
                'error': str(e)
            }, 500

    # ============================================================================
    # API ENDPOINTS - Oda Tipi Yönetimi
    # ============================================================================
    
    @app.route('/api/oda-tipleri', methods=['GET'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_oda_tipleri_listele():
        """Tüm oda tiplerini listele (otel bazlı setup bilgisi ile)"""
        try:
            from models import OdaTipi, oda_tipi_setup, Setup
            from sqlalchemy import and_
            
            otel_id = request.args.get('otel_id', type=int)
            
            oda_tipleri = OdaTipi.query.filter_by(aktif=True).order_by(OdaTipi.ad).all()
            
            result = []
            for tip in oda_tipleri:
                # Otel ID verilmişse, o otele ait setup'ları getir
                if otel_id:
                    setup_rows = db.session.execute(
                        db.select(oda_tipi_setup.c.setup_id).where(
                            and_(
                                oda_tipi_setup.c.otel_id == otel_id,
                                oda_tipi_setup.c.oda_tipi_id == tip.id
                            )
                        )
                    ).fetchall()
                    setup_ids = [row[0] for row in setup_rows]
                    
                    # Setup adlarını al
                    setup_adlari = []
                    for sid in setup_ids:
                        setup = Setup.query.get(sid)
                        if setup:
                            setup_adlari.append(setup.ad)
                else:
                    # Otel ID yoksa, global ilişkiden al (geriye uyumluluk)
                    setup_ids = [s.id for s in tip.setuplar]
                    setup_adlari = [s.ad for s in tip.setuplar]
                
                result.append({
                    'id': tip.id,
                    'ad': tip.ad,
                    'dolap_sayisi': tip.dolap_sayisi,
                    'setup_ids': setup_ids,
                    'setup_adlari': setup_adlari,
                    'olusturma_tarihi': tip.olusturma_tarihi.isoformat() if tip.olusturma_tarihi else None
                })
            
            return jsonify({
                'success': True,
                'oda_tipleri': result
            })
            
        except Exception as e:
            log_hata(e, 'api_oda_tipleri_listele')
            return jsonify({
                'success': False,
                'error': str(e)
            }), 500
    
    @app.route('/api/oda-tipleri', methods=['POST'])
    @csrf.exempt
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_oda_tipi_ekle():
        """Yeni oda tipi ekle"""
        try:
            from models import OdaTipi, Setup
            
            data = request.get_json()
            ad = data.get('ad', '').strip()
            dolap_sayisi = data.get('dolap_sayisi', 0)
            setup_ids = data.get('setup_ids', [])
            
            if not ad:
                return {
                    'success': False,
                    'error': 'Oda tipi adı boş olamaz'
                }, 400
            
            if not setup_ids or len(setup_ids) == 0:
                return {
                    'success': False,
                    'error': 'En az bir setup seçimi zorunludur'
                }, 400
            
            # Aynı isimde oda tipi var mı kontrol et
            mevcut = OdaTipi.query.filter_by(ad=ad).first()
            if mevcut:
                if mevcut.aktif:
                    return {
                        'success': False,
                        'error': 'Bu oda tipi zaten mevcut'
                    }, 400
                else:
                    # Pasif oda tipini aktif et ve güncelle
                    mevcut.aktif = True
                    mevcut.dolap_sayisi = dolap_sayisi
                    
                    # Setup'ları güncelle (many-to-many)
                    mevcut.setuplar.clear()
                    for setup_id in setup_ids:
                        setup = Setup.query.get(setup_id)
                        if setup:
                            mevcut.setuplar.append(setup)
                    
                    db.session.commit()
                    
                    audit_update(
                        'OdaTipi',
                        mevcut.id,
                        {'aktif': False},
                        {'aktif': True, 'dolap_sayisi': dolap_sayisi, 'setup_ids': setup_ids},
                        session.get('kullanici_id')
                    )
                    
                    log_islem('oda_tipi_aktif', f'Oda tipi aktif edildi: {ad}')
                    
                    return {
                        'success': True,
                        'oda_tipi': {
                            'id': mevcut.id,
                            'ad': mevcut.ad,
                            'dolap_sayisi': mevcut.dolap_sayisi,
                            'setup_ids': [s.id for s in mevcut.setuplar],
                            'setup_adlari': [s.ad for s in mevcut.setuplar]
                        }
                    }
            
            # Yeni oda tipi oluştur
            yeni_tip = OdaTipi(ad=ad, dolap_sayisi=dolap_sayisi)
            
            # Setup'ları ekle (many-to-many)
            for setup_id in setup_ids:
                setup = Setup.query.get(setup_id)
                if setup:
                    yeni_tip.setuplar.append(setup)
            
            db.session.add(yeni_tip)
            db.session.commit()
            
            audit_create('OdaTipi', yeni_tip.id, serialize_model(yeni_tip), session.get('kullanici_id'))
            log_islem('oda_tipi_ekle', f'Yeni oda tipi eklendi: {ad}')
            
            return {
                'success': True,
                'oda_tipi': {
                    'id': yeni_tip.id,
                    'ad': yeni_tip.ad,
                    'dolap_sayisi': yeni_tip.dolap_sayisi,
                    'setup_ids': [s.id for s in yeni_tip.setuplar],
                    'setup_adlari': [s.ad for s in yeni_tip.setuplar]
                }
            }
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'api_oda_tipi_ekle')
            return {
                'success': False,
                'error': str(e)
            }, 500
    
    @app.route('/api/oda-tipleri/<int:tip_id>', methods=['PUT'])
    @csrf.exempt
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_oda_tipi_guncelle(tip_id):
        """Oda tipi güncelle"""
        try:
            from models import OdaTipi, Setup
            
            oda_tipi = OdaTipi.query.get_or_404(tip_id)
            data = request.get_json()
            ad = data.get('ad', '').strip()
            dolap_sayisi = data.get('dolap_sayisi', 0)
            setup_ids = data.get('setup_ids', [])
            
            if not ad:
                return {
                    'success': False,
                    'error': 'Oda tipi adı boş olamaz'
                }, 400
            
            if not setup_ids or len(setup_ids) == 0:
                return {
                    'success': False,
                    'error': 'En az bir setup seçimi zorunludur'
                }, 400
            
            # Aynı isimde başka oda tipi var mı kontrol et
            mevcut = OdaTipi.query.filter(
                OdaTipi.ad == ad,
                OdaTipi.id != tip_id
            ).first()
            
            if mevcut:
                return {
                    'success': False,
                    'error': 'Bu isimde başka bir oda tipi mevcut'
                }, 400
            
            eski_deger = serialize_model(oda_tipi)
            oda_tipi.ad = ad
            oda_tipi.dolap_sayisi = dolap_sayisi
            
            # Setup'ları güncelle (many-to-many)
            oda_tipi.setuplar.clear()
            for setup_id in setup_ids:
                setup = Setup.query.get(setup_id)
                if setup:
                    oda_tipi.setuplar.append(setup)
            
            db.session.commit()
            
            audit_update('OdaTipi', tip_id, eski_deger, serialize_model(oda_tipi), session.get('kullanici_id'))
            log_islem('oda_tipi_guncelle', f'Oda tipi güncellendi: {ad}')
            
            return {
                'success': True,
                'oda_tipi': {
                    'id': oda_tipi.id,
                    'ad': oda_tipi.ad,
                    'dolap_sayisi': oda_tipi.dolap_sayisi,
                    'setup_ids': [s.id for s in oda_tipi.setuplar],
                    'setup_adlari': [s.ad for s in oda_tipi.setuplar]
                }
            }
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'api_oda_tipi_guncelle')
            return {
                'success': False,
                'error': str(e)
            }, 500
    
    @app.route('/api/oda-tipleri/<int:tip_id>', methods=['DELETE'])
    @csrf.exempt
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_oda_tipi_sil(tip_id):
        """Oda tipi sil (soft delete)"""
        try:
            from models import OdaTipi
            
            oda_tipi = OdaTipi.query.get_or_404(tip_id)
            
            # Bu oda tipini kullanan oda var mı kontrol et
            kullanan_oda_sayisi = Oda.query.filter_by(oda_tipi_id=oda_tipi.id, aktif=True).count()
            
            if kullanan_oda_sayisi > 0:
                return {
                    'success': False,
                    'error': f'Bu oda tipi {kullanan_oda_sayisi} oda tarafından kullanılıyor. Önce odaların tipini değiştirin.'
                }, 400
            
            eski_deger = serialize_model(oda_tipi)
            oda_tipi.aktif = False
            db.session.commit()
            
            audit_delete('OdaTipi', tip_id, eski_deger, session.get('kullanici_id'))
            log_islem('oda_tipi_sil', f'Oda tipi silindi: {oda_tipi.ad}')
            
            return {
                'success': True,
                'message': 'Oda tipi başarıyla silindi'
            }
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'api_oda_tipi_sil')
            return {
                'success': False,
                'error': str(e)
            }, 500

    # ============================================================================
    # SETUP YÖNETİMİ
    # ============================================================================
    
    @app.route('/setup-yonetimi', methods=['GET'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def setup_yonetimi():
        """Setup yönetim sayfası"""
        return render_template('sistem_yoneticisi/setup_yonetimi.html')
    
    # ============================================================================
    # API ENDPOINTS - Setup Yönetimi
    # ============================================================================
    
    @app.route('/api/setuplar', methods=['GET'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_setuplar_listele():
        """Tüm setup'ları listele"""
        try:
            from models import Setup, SetupIcerik, OdaTipi, Otel, oda_tipi_setup
            from sqlalchemy import distinct
            
            # Sadece aktif setup'ları getir
            setuplar = Setup.query.filter_by(aktif=True).all()
            
            sonuc = []
            for setup in setuplar:
                # Ürün sayısını hesapla
                urun_sayisi = SetupIcerik.query.filter_by(setup_id=setup.id).count()
                
                # Bu setup'a atanmış oda tiplerini al (otel bazlı tablodan, sadece aktif oda tipleri)
                # Otel adı ile birlikte getir
                atamalar = db.session.execute(
                    db.select(
                        oda_tipi_setup.c.otel_id,
                        oda_tipi_setup.c.oda_tipi_id
                    ).where(
                        oda_tipi_setup.c.setup_id == setup.id
                    )
                ).fetchall()
                
                # Otel bazlı oda tipleri dict'i oluştur
                otel_oda_tipleri = {}  # {otel_adi: [oda_tipi_adlari]}
                oda_tipi_adlari = []  # Geriye uyumluluk için
                
                for row in atamalar:
                    otel_id = row[0]
                    oda_tipi_id = row[1]
                    
                    # Aktif oda tipini al
                    oda_tipi = OdaTipi.query.filter_by(id=oda_tipi_id, aktif=True).first()
                    if not oda_tipi:
                        continue
                    
                    # Otel adını al
                    otel = Otel.query.get(otel_id) if otel_id else None
                    otel_adi = otel.ad if otel else "Bilinmeyen Otel"
                    
                    # Dict'e ekle
                    if otel_adi not in otel_oda_tipleri:
                        otel_oda_tipleri[otel_adi] = []
                    
                    if oda_tipi.ad not in otel_oda_tipleri[otel_adi]:
                        otel_oda_tipleri[otel_adi].append(oda_tipi.ad)
                    
                    # Geriye uyumluluk için
                    if oda_tipi.ad not in oda_tipi_adlari:
                        oda_tipi_adlari.append(oda_tipi.ad)
                
                # Toplam maliyeti hesapla (ürün alış fiyatı * adet)
                toplam_maliyet = 0
                setup_icerikler = SetupIcerik.query.filter_by(setup_id=setup.id).all()
                for icerik in setup_icerikler:
                    if icerik.urun and icerik.urun.alis_fiyati:
                        toplam_maliyet += icerik.urun.alis_fiyati * icerik.adet
                
                sonuc.append({
                    'id': setup.id,
                    'ad': setup.ad,
                    'aciklama': setup.aciklama,
                    'dolap_ici': setup.dolap_ici if hasattr(setup, 'dolap_ici') else True,
                    'urun_sayisi': urun_sayisi,
                    'oda_tipleri': oda_tipi_adlari,  # Geriye uyumluluk
                    'otel_oda_tipleri': otel_oda_tipleri,  # Yeni: Otel bazlı
                    'toplam_maliyet': round(toplam_maliyet, 2)
                })
            
            return {
                'success': True,
                'setuplar': sonuc
            }
            
        except Exception as e:
            log_hata(e, 'api_setuplar_listele')
            return {
                'success': False,
                'error': str(e)
            }, 500
    
    @app.route('/api/setuplar', methods=['POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_setup_ekle():
        """Yeni setup ekle"""
        try:
            from models import Setup
            
            data = request.get_json()
            ad = data.get('ad', '').strip()
            aciklama = data.get('aciklama', '').strip()
            dolap_ici = data.get('dolap_ici', True)  # Varsayılan: Dolap İçi
            
            if not ad:
                return {
                    'success': False,
                    'error': 'Setup adı boş olamaz'
                }, 400
            
            # Aynı isimde setup var mı kontrol et
            mevcut = Setup.query.filter_by(ad=ad).first()
            if mevcut:
                return {
                    'success': False,
                    'error': 'Bu isimde setup zaten mevcut'
                }, 400
            
            yeni_setup = Setup(ad=ad, aciklama=aciklama, dolap_ici=dolap_ici)
            db.session.add(yeni_setup)
            db.session.commit()
            
            audit_create('Setup', yeni_setup.id, serialize_model(yeni_setup), session.get('kullanici_id'))
            log_islem('setup_ekle', f'Yeni setup eklendi: {ad}')
            
            return {
                'success': True,
                'setup': {
                    'id': yeni_setup.id,
                    'ad': yeni_setup.ad
                }
            }
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'api_setup_ekle')
            return {
                'success': False,
                'error': str(e)
            }, 500

    @app.route('/api/setuplar/<int:setup_id>', methods=['PUT', 'PATCH', 'POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_setup_guncelle(setup_id):
        """Setup güncelle"""
        try:
            from models import Setup, Oda
            
            data = request.get_json()
            ad = data.get('ad', '').strip()
            aciklama = data.get('aciklama', '').strip()
            dolap_ici = data.get('dolap_ici', True)
            
            # Validasyon
            if not ad:
                return {
                    'success': False,
                    'error': 'Setup adı boş olamaz'
                }, 400
            
            # Setup bul
            setup = Setup.query.get_or_404(setup_id)
            eski_ad = setup.ad
            eski_deger = serialize_model(setup)
            
            # Aynı isimde başka setup var mı kontrol et
            existing = Setup.query.filter(
                Setup.ad == ad,
                Setup.id != setup_id
            ).first()
            
            if existing:
                return {
                    'success': False,
                    'error': 'Bu isimde bir setup zaten var'
                }, 400
            
            # Güncelle
            setup.ad = ad
            setup.aciklama = aciklama
            setup.dolap_ici = dolap_ici
            db.session.commit()
            
            # NOT: Oda tipi atamaları artık Many-to-Many ilişki ile yönetiliyor
            # oda_tipi_setup ara tablosu kullanılıyor, Oda tablosunda setup kolonu yok
            
            # Audit log
            yeni_deger = serialize_model(setup)
            audit_update('Setup', setup_id, eski_deger, yeni_deger, session.get('kullanici_id'))
            log_islem('setup_guncelle', f'Setup güncellendi: {eski_ad} → {ad}')
            
            return {
                'success': True,
                'message': 'Setup başarıyla güncellendi'
            }
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'api_setup_guncelle')
            return {
                'success': False,
                'error': str(e)
            }, 500

    @app.route('/api/setuplar/<int:setup_id>', methods=['DELETE'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_setup_sil(setup_id):
        """Setup sil"""
        try:
            from models import Setup
            
            setup = Setup.query.get_or_404(setup_id)
            
            eski_deger = serialize_model(setup)
            setup.aktif = False
            db.session.commit()
            
            audit_delete('Setup', setup_id, eski_deger, session.get('kullanici_id'))
            log_islem('setup_sil', f'Setup silindi: {setup.ad}')
            
            return {
                'success': True,
                'message': 'Setup başarıyla silindi'
            }
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'api_setup_sil')
            return {
                'success': False,
                'error': str(e)
            }, 500
    
    @app.route('/api/setuplar/<int:setup_id>/icerik', methods=['GET'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_setup_icerik_listele(setup_id):
        """Setup içeriğini listele"""
        try:
            from models import SetupIcerik
            
            icerikler = SetupIcerik.query.filter_by(setup_id=setup_id).all()
            
            sonuc = []
            toplam_maliyet = 0
            
            for icerik in icerikler:
                # Ürün alış fiyatını fiyat geçmişinden çek
                alis_fiyati = 0
                try:
                    if icerik.urun and icerik.urun.alis_fiyati:
                        alis_fiyati = float(icerik.urun.alis_fiyati)
                except Exception as fiyat_hata:
                    print(f"Fiyat çekme hatası: {fiyat_hata}")
                
                tutar = alis_fiyati * icerik.adet
                toplam_maliyet += tutar
                
                sonuc.append({
                    'id': icerik.id,
                    'urun_id': icerik.urun_id,
                    'urun_ad': icerik.urun.urun_adi if icerik.urun else 'Bilinmeyen',
                    'adet': icerik.adet,
                    'alis_fiyati': float(alis_fiyati),
                    'tutar': float(tutar)
                })
            
            return {
                'success': True,
                'icerikler': sonuc,
                'toplam_maliyet': float(toplam_maliyet)
            }
            
        except Exception as e:
            log_hata(e, 'api_setup_icerik_listele')
            return {
                'success': False,
                'error': str(e)
            }, 500
    
    @app.route('/api/setuplar/<int:setup_id>/icerik', methods=['POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_setup_icerik_ekle(setup_id):
        """Setup'a ürün ekle"""
        try:
            from models import SetupIcerik
            
            data = request.get_json()
            urun_id = data.get('urun_id')
            adet = data.get('adet', 1)
            
            if not urun_id:
                return {
                    'success': False,
                    'error': 'Ürün seçimi zorunludur'
                }, 400
            
            # Aynı ürün zaten ekli mi kontrol et
            mevcut = SetupIcerik.query.filter_by(setup_id=setup_id, urun_id=urun_id).first()
            if mevcut:
                return {
                    'success': False,
                    'error': 'Bu ürün zaten ekli'
                }, 400
            
            yeni_icerik = SetupIcerik(setup_id=setup_id, urun_id=urun_id, adet=adet)
            db.session.add(yeni_icerik)
            db.session.commit()
            
            log_islem('setup_icerik_ekle', f'Setup içeriğine ürün eklendi: Setup ID {setup_id}, Ürün ID {urun_id}')
            
            return {
                'success': True,
                'icerik': {
                    'id': yeni_icerik.id
                }
            }
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'api_setup_icerik_ekle')
            return {
                'success': False,
                'error': str(e)
            }, 500

    @app.route('/api/setup-icerik/<int:icerik_id>', methods=['DELETE'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_setup_icerik_sil(icerik_id):
        """Setup içeriğinden ürün sil"""
        try:
            from models import SetupIcerik
            
            icerik = SetupIcerik.query.get_or_404(icerik_id)
            
            db.session.delete(icerik)
            db.session.commit()
            
            log_islem('setup_icerik_sil', f'Setup içeriğinden ürün silindi: İçerik ID {icerik_id}')
            
            return {
                'success': True,
                'message': 'Ürün başarıyla silindi'
            }
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'api_setup_icerik_sil')
            return {
                'success': False,
                'error': str(e)
            }, 500
    
    @app.route('/api/setup-atama', methods=['POST'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_setup_atama():
        """Setup'ları oda tiplerine ata (Otel bazlı Many-to-Many)"""
        try:
            from models import OdaTipi, Setup, Otel, oda_tipi_setup
            from sqlalchemy import and_
            
            data = request.get_json()
            
            if not data:
                return {
                    'success': False,
                    'error': 'Veri gönderilmedi'
                }, 400
            
            otel_id = data.get('otel_id')
            oda_tipi_id = data.get('oda_tipi_id')
            setup_ids = data.get('setup_ids', [])
            
            if not otel_id:
                return {
                    'success': False,
                    'error': 'Otel ID belirtilmedi'
                }, 400
            
            if not oda_tipi_id:
                return {
                    'success': False,
                    'error': 'Oda tipi ID belirtilmedi'
                }, 400
            
            # Otel kontrolü
            otel = Otel.query.get(otel_id)
            if not otel:
                return {
                    'success': False,
                    'error': f'Otel bulunamadı: {otel_id}'
                }, 404
            
            # Oda tipini bul
            oda_tipi = OdaTipi.query.get(oda_tipi_id)
            if not oda_tipi:
                return {
                    'success': False,
                    'error': f'Oda tipi bulunamadı: {oda_tipi_id}'
                }, 404
            
            # Bu otel ve oda tipi için mevcut atamaları sil
            db.session.execute(
                oda_tipi_setup.delete().where(
                    and_(
                        oda_tipi_setup.c.otel_id == otel_id,
                        oda_tipi_setup.c.oda_tipi_id == oda_tipi_id
                    )
                )
            )
            
            # Yeni setup'ları ekle
            if setup_ids:
                for setup_id in setup_ids:
                    setup = Setup.query.get(setup_id)
                    if setup:
                        db.session.execute(
                            oda_tipi_setup.insert().values(
                                otel_id=otel_id,
                                oda_tipi_id=oda_tipi_id,
                                setup_id=setup_id
                            )
                        )
            
            db.session.commit()
            
            log_islem('setup_atama', f'Otel bazlı setup ataması güncellendi: {otel.ad} - {oda_tipi.ad} -> {len(setup_ids)} setup')
            
            return {
                'success': True,
                'message': 'Atama başarıyla kaydedildi'
            }
            
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'api_setup_atama')
            return {
                'success': False,
                'error': str(e)
            }, 500
    
    @app.route('/api/urunler-liste', methods=['GET'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_urunler_liste():
        """Tüm ürünleri listele"""
        try:
            from models import Urun
            
            urunler = Urun.query.filter_by(aktif=True).order_by(Urun.urun_adi).all()
            
            return {
                'success': True,
                'urunler': [{
                    'id': urun.id,
                    'ad': urun.urun_adi,
                    'grup_id': urun.grup_id,
                    'alis_fiyati': float(urun.alis_fiyati) if urun.alis_fiyati else 0
                } for urun in urunler]
            }
            
        except Exception as e:
            log_hata(e, 'api_urunler_liste')
            return {
                'success': False,
                'error': str(e)
            }, 500
    
    @app.route('/api/urun-gruplari-liste', methods=['GET'])
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def api_urun_gruplari_liste():
        """Tüm ürün gruplarını listele"""
        try:
            from models import UrunGrup
            
            gruplar = UrunGrup.query.filter_by(aktif=True).order_by(UrunGrup.grup_adi).all()
            
            return {
                'success': True,
                'gruplar': [{
                    'id': grup.id,
                    'ad': grup.grup_adi
                } for grup in gruplar]
            }
            
        except Exception as e:
            log_hata(e, 'api_urun_gruplari_liste')
            return {
                'success': False,
                'error': str(e)
            }, 500

    # İLK STOK YÜKLEME ROUTE'LARI
    # ============================================================================
    
    @app.route('/otel-ilk-stok-yukleme', methods=['GET'])
    @login_required
    @role_required('sistem_yoneticisi')
    def otel_ilk_stok_yukleme():
        """İlk stok yükleme sayfası - Her otel için 1 kez (FIFO)"""
        from models import Urun, UrunGrup
        
        try:
            # Otelleri getir
            oteller = Otel.query.filter_by(aktif=True).order_by(Otel.ad).all()
            
            # Otel durumlarını hazırla
            otel_durumlari = []
            for otel in oteller:
                # Yükleyen kullanıcı bilgisini al
                yukleyen_adi = None
                if otel.ilk_stok_yukleyen_id:
                    yukleyen = db.session.get(Kullanici, otel.ilk_stok_yukleyen_id)
                    if yukleyen:
                        yukleyen_adi = f"{yukleyen.ad} {yukleyen.soyad}"
                
                otel_durumlari.append({
                    'id': otel.id,
                    'ad': otel.ad,
                    'ilk_stok_yuklendi': otel.ilk_stok_yuklendi,
                    'yukleme_tarihi': otel.ilk_stok_yukleme_tarihi.strftime('%d.%m.%Y %H:%M') if otel.ilk_stok_yukleme_tarihi else None,
                    'yukleyen': yukleyen_adi
                })
            
            # Ürün grupları ve ürünler
            gruplar = UrunGrup.query.filter_by(aktif=True).order_by(UrunGrup.grup_adi).all()
            urunler = Urun.query.filter_by(aktif=True).order_by(Urun.urun_adi).all()
            
            return render_template('sistem_yoneticisi/ilk_stok_yukleme.html',
                                 otel_durumlari=otel_durumlari,
                                 gruplar=gruplar,
                                 urunler=urunler)
                                 
        except Exception as e:
            log_hata(e, 'ilk_stok_yukleme')
            flash('Sayfa yüklenirken hata oluştu.', 'danger')
            return redirect(url_for('sistem_yoneticisi_dashboard'))
    
    
    @app.route('/otel-ilk-stok-yukleme-kaydet', methods=['POST'])
    @login_required
    @role_required('sistem_yoneticisi')
    def otel_ilk_stok_yukleme_kaydet():
        """İlk stok yükleme işlemini kaydet (FIFO)"""
        from utils.fifo_servisler import FifoStokServisi
        
        try:
            data = request.get_json()
            
            if not data:
                return jsonify({'success': False, 'error': 'Veri bulunamadı'}), 400
            
            otel_id = data.get('otel_id')
            stok_verileri = data.get('stok_verileri', [])
            
            if not otel_id:
                return jsonify({'success': False, 'error': 'Otel seçilmedi'}), 400
            
            if not stok_verileri:
                return jsonify({'success': False, 'error': 'Stok verisi girilmedi'}), 400
            
            # Otel kontrolü
            otel = db.session.get(Otel, otel_id)
            if not otel:
                return jsonify({'success': False, 'error': 'Otel bulunamadı'}), 404
            
            if otel.ilk_stok_yuklendi:
                return jsonify({
                    'success': False, 
                    'error': f'Bu otel için ilk stok yüklemesi zaten yapılmış ({otel.ilk_stok_yukleme_tarihi.strftime("%d.%m.%Y")})'
                }), 400
            
            # İlk stok yükle
            sonuc = FifoStokServisi.ilk_stok_yukle(
                otel_id=otel_id,
                stok_verileri=stok_verileri,
                kullanici_id=session['kullanici_id']
            )
            
            if sonuc['success']:
                return jsonify({
                    'success': True,
                    'message': sonuc['message'],
                    'yuklenen_urun_sayisi': sonuc.get('yuklenen_urun_sayisi', 0)
                })
            else:
                return jsonify({
                    'success': False,
                    'error': sonuc['message'],
                    'hatalar': sonuc.get('hatalar', [])
                }), 400
                
        except Exception as e:
            db.session.rollback()
            log_hata(e, 'ilk_stok_yukleme_kaydet')
            return jsonify({'success': False, 'error': f'Kayıt hatası: {str(e)}'}), 500
    
    
    @app.route('/api/otel-stok-durumu/<int:otel_id>')
    @login_required
    @role_required('sistem_yoneticisi', 'admin', 'depo_sorumlusu')
    def api_otel_stok_durumu(otel_id):
        """Otel bazlı FIFO stok durumunu getir"""
        from utils.fifo_servisler import FifoStokServisi
        
        try:
            otel = db.session.get(Otel, otel_id)
            if not otel:
                return jsonify({'success': False, 'error': 'Otel bulunamadı'}), 404
            
            stok_durumu = FifoStokServisi.fifo_stok_durumu(otel_id)
            
            return jsonify({
                'success': True,
                'otel_id': otel_id,
                'otel_adi': otel.ad,
                'ilk_stok_yuklendi': otel.ilk_stok_yuklendi,
                'stok_durumu': stok_durumu
            })
            
        except Exception as e:
            log_hata(e, 'api_otel_stok_durumu')
            return jsonify({'success': False, 'error': str(e)}), 500


    # ==================== ANA DEPO TEDARİK GEÇMİŞİ ====================
    
    @app.route('/admin/ana-depo-tedarik-gecmisi')
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def admin_ana_depo_tedarik_gecmisi():
        """Sistem yöneticisi için tüm otellerin ana depo tedarik geçmişi"""
        from models import AnaDepoTedarik, AnaDepoTedarikDetay, Urun
        from datetime import datetime
        
        try:
            # Filtreler
            otel_id = request.args.get('otel_id', type=int)
            ay = request.args.get('ay', type=int, default=datetime.now().month)
            yil = request.args.get('yil', type=int, default=datetime.now().year)
            durum = request.args.get('durum', '')
            
            # Tedarik sorgusu
            query = AnaDepoTedarik.query
            
            # Otel filtresi
            if otel_id:
                query = query.filter(AnaDepoTedarik.otel_id == otel_id)
            
            # Ay/Yıl filtresi
            from sqlalchemy import extract
            query = query.filter(
                extract('month', AnaDepoTedarik.islem_tarihi) == ay,
                extract('year', AnaDepoTedarik.islem_tarihi) == yil
            )
            
            # Durum filtresi
            if durum:
                query = query.filter(AnaDepoTedarik.durum == durum)
            
            # Sıralama ve sonuçlar
            tedarikler = query.order_by(AnaDepoTedarik.islem_tarihi.desc()).all()
            
            # Oteller listesi (filtre için)
            oteller = Otel.query.filter_by(aktif=True).order_by(Otel.ad).all()
            
            # İstatistikler
            toplam_tedarik = len(tedarikler)
            aktif_sayi = sum(1 for t in tedarikler if t.durum == 'aktif')
            iptal_sayi = sum(1 for t in tedarikler if t.durum == 'iptal')
            
            # Tedarik detayları
            tedarik_listesi = []
            for tedarik in tedarikler:
                detaylar = AnaDepoTedarikDetay.query.filter_by(tedarik_id=tedarik.id).all()
                toplam_urun = sum(d.miktar for d in detaylar)
                
                # Otel bilgisi
                otel = db.session.get(Otel, tedarik.otel_id)
                
                # Kullanıcı bilgisi (depo sorumlusu)
                kullanici = db.session.get(Kullanici, tedarik.depo_sorumlusu_id) if tedarik.depo_sorumlusu_id else None
                
                tedarik_listesi.append({
                    'id': tedarik.id,
                    'otel_adi': otel.ad if otel else 'Bilinmiyor',
                    'tarih': tedarik.islem_tarihi,
                    'durum': tedarik.durum,
                    'toplam_urun': toplam_urun,
                    'detay_sayisi': len(detaylar),
                    'kullanici': f"{kullanici.ad} {kullanici.soyad}" if kullanici else 'Bilinmiyor',
                    'aciklama': tedarik.aciklama,
                    'detaylar': [{
                        'urun_adi': db.session.get(Urun, d.urun_id).urun_adi if db.session.get(Urun, d.urun_id) else 'Bilinmiyor',
                        'miktar': d.miktar
                    } for d in detaylar]
                })
            
            return render_template(
                'sistem_yoneticisi/ana_depo_tedarik_gecmisi.html',
                tedarikler=tedarik_listesi,
                oteller=oteller,
                secili_otel_id=otel_id,
                secili_ay=ay,
                secili_yil=yil,
                secili_durum=durum,
                istatistikler={
                    'toplam': toplam_tedarik,
                    'aktif': aktif_sayi,
                    'iptal': iptal_sayi
                }
            )
            
        except Exception as e:
            log_hata(e, 'admin_ana_depo_tedarik_gecmisi')
            flash('Tedarik geçmişi yüklenirken hata oluştu.', 'error')
            return redirect(url_for('sistem_yoneticisi_dashboard'))
