from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify
from models import DosyaYukleme, db
from utils.decorators import login_required, role_required
from utils.helpers import log_hata
from utils.occupancy_service import OccupancyService
from datetime import date, datetime, timezone
import pytz

# KKTC Timezone (Kıbrıs - Europe/Nicosia)
KKTC_TZ = pytz.timezone('Europe/Nicosia')

def get_kktc_now():
    """Kıbrıs saat diliminde şu anki zamanı döndürür."""
    return datetime.now(KKTC_TZ)

doluluk_bp = Blueprint("doluluk", __name__)

def register_doluluk_routes(app):
    app.register_blueprint(doluluk_bp)

@doluluk_bp.route("/doluluk-yonetimi")
@login_required
@role_required("depo_sorumlusu")
def doluluk_yonetimi():
    """Doluluk yönetimi ana sayfası - Excel dosyası yükleme formu ve yükleme geçmişi"""
    try:
        from utils.authorization import get_kullanici_otelleri, get_otel_filtreleme_secenekleri
        kullanici_otelleri = get_kullanici_otelleri()
        otel_secenekleri = get_otel_filtreleme_secenekleri()
        secili_otel_id = request.args.get("otel_id", type=int)
        # Otel seçilmemişse ilk oteli seç (ID=1 Merit Royal Diamond)
        if not secili_otel_id and kullanici_otelleri:
            secili_otel_id = kullanici_otelleri[0].id
        # Silinen kayıtları gösterme
        yukleme_gecmisi = DosyaYukleme.query.filter(
            DosyaYukleme.yuklenen_kullanici_id == session["kullanici_id"],
            DosyaYukleme.durum != 'silindi'
        ).order_by(DosyaYukleme.yukleme_tarihi.desc()).limit(50).all()
        return render_template("depo_sorumlusu/doluluk_yonetimi.html",
            yukleme_gecmisi=yukleme_gecmisi, otel_secenekleri=otel_secenekleri,
            secili_otel_id=secili_otel_id)
    except Exception as e:
        log_hata("doluluk_yonetimi", str(e), session.get("kullanici_id"))
        flash(f"Hata oluştu: {str(e)}", "danger")
        return redirect(url_for("depo_dashboard"))

@doluluk_bp.route("/gunluk-doluluk")
@login_required
@role_required(["kat_sorumlusu", "depo_sorumlusu", "sistem_yoneticisi"])
def gunluk_doluluk():
    try:
        from utils.authorization import get_kullanici_otelleri, get_otel_filtreleme_secenekleri
        kullanici_otelleri = get_kullanici_otelleri()
        otel_secenekleri = get_otel_filtreleme_secenekleri()
        secili_otel_id = request.args.get("otel_id", type=int)
        if not secili_otel_id and kullanici_otelleri:
            secili_otel_id = kullanici_otelleri[0].id
        tarih_str = request.args.get("tarih")
        if tarih_str:
            try:
                secili_tarih = datetime.strptime(tarih_str, "%Y-%m-%d").date()
            except ValueError:
                secili_tarih = date.today()
        else:
            secili_tarih = date.today()
        rapor_data = OccupancyService.get_gunluk_doluluk_raporu(secili_tarih, secili_otel_id)
        
        # Template için oda listesi oluştur
        from models import Oda, Kat, MisafirKayit
        from sqlalchemy import and_
        
        oda_query = Oda.query.filter_by(aktif=True).join(Kat)
        if secili_otel_id:
            oda_query = oda_query.filter(Kat.otel_id == secili_otel_id)
        
        odalar = oda_query.order_by(Kat.kat_adi, Oda.oda_no).all()
        
        # Görev durumlarını al
        from models import GorevDetay, GunlukGorev
        gorev_durumlari = {}
        gorev_detaylari = GorevDetay.query.join(GunlukGorev).filter(
            GunlukGorev.gorev_tarihi == secili_tarih
        ).all()
        for detay in gorev_detaylari:
            gorev_durumlari[detay.oda_id] = detay.durum
        
        rapor = []
        for oda in odalar:
            # O odanın o tarihteki misafir bilgisi
            misafir = MisafirKayit.query.filter(
                and_(
                    MisafirKayit.oda_id == oda.id,
                    MisafirKayit.giris_tarihi <= secili_tarih,
                    MisafirKayit.cikis_tarihi > secili_tarih
                )
            ).first()
            
            # Görev durumunu al
            gorev_durumu = gorev_durumlari.get(oda.id, 'pending')
            
            rapor.append({
                'oda_no': oda.oda_no,
                'oda_id': oda.id,
                'kat_id': oda.kat_id,
                'kat_adi': oda.kat.kat_adi if oda.kat else '-',
                'durum': 'dolu' if misafir else 'bos',
                'gorev_durumu': gorev_durumu,
                'misafir_adi': f"{misafir.misafir_sayisi} kişi" if misafir else None,
                'giris_tarihi': misafir.giris_tarihi if misafir else None,
                'cikis_tarihi': misafir.cikis_tarihi if misafir else None
            })
        
        # Günlük özet için rapor_data kullan
        haftalik_ozet = {
            'toplam_oda': rapor_data.get('toplam_oda', 0),
            'dolu_oda': rapor_data.get('dolu_oda', 0),
            'doluluk_orani': (rapor_data.get('dolu_oda', 0) / rapor_data.get('toplam_oda', 1) * 100) if rapor_data.get('toplam_oda', 0) > 0 else 0
        }
        
        # Kat bazlı özet oluştur
        from collections import defaultdict
        kat_sayilari = defaultdict(lambda: {'toplam': 0, 'dolu': 0, 'kat_adi': '', 'kat_id': None})
        for oda_bilgi in rapor:
            kat_adi = oda_bilgi['kat_adi']
            kat_sayilari[kat_adi]['toplam'] += 1
            kat_sayilari[kat_adi]['kat_adi'] = kat_adi
            kat_sayilari[kat_adi]['kat_id'] = oda_bilgi['kat_id']
            if oda_bilgi['durum'] == 'dolu':
                kat_sayilari[kat_adi]['dolu'] += 1
        
        kat_bazli_ozet = []
        for kat_adi, veriler in sorted(kat_sayilari.items()):
            oran = (veriler['dolu'] / veriler['toplam'] * 100) if veriler['toplam'] > 0 else 0
            kat_bazli_ozet.append({
                'kat_id': veriler['kat_id'],
                'kat_adi': kat_adi,
                'toplam': veriler['toplam'],
                'dolu': veriler['dolu'],
                'oran': oran
            })
        
        # Seçili oteli bul
        from models import Otel
        secili_otel = Otel.query.get(secili_otel_id) if secili_otel_id else None
        
        return render_template("kat_sorumlusu/gunluk_doluluk.html", rapor=rapor, kat_bazli_ozet=kat_bazli_ozet,
            haftalik_ozet=haftalik_ozet, secili_tarih=secili_tarih,
            otel_secenekleri=otel_secenekleri, secili_otel_id=secili_otel_id,
            secili_otel=secili_otel)
    except Exception as e:
        log_hata("gunluk_doluluk", str(e), session.get("kullanici_id"))
        flash(f"Hata oluştu: {str(e)}", "danger")
        if session.get("rol") == "kat_sorumlusu":
            return redirect(url_for("kat_sorumlusu_dashboard"))
        elif session.get("rol") == "depo_sorumlusu":
            return redirect(url_for("depo_dashboard"))
        else:
            return redirect(url_for("sistem_yoneticisi_dashboard"))

@doluluk_bp.route("/gunluk-doluluk/yazdir")
@login_required
@role_required(["kat_sorumlusu", "depo_sorumlusu", "sistem_yoneticisi"])
def gunluk_doluluk_yazdir():
    """Yazdırma için optimize edilmiş sayfa"""
    try:
        from utils.authorization import get_kullanici_otelleri
        from models import Oda, Kat, MisafirKayit
        from sqlalchemy import and_
        
        # Parametreleri al
        tarih_str = request.args.get("tarih")
        secili_otel_id = request.args.get("otel_id", type=int)
        bos_odalar_dahil = request.args.get("bos_odalar", "0") == "1"
        
        if tarih_str:
            try:
                secili_tarih = datetime.strptime(tarih_str, "%Y-%m-%d").date()
            except ValueError:
                secili_tarih = date.today()
        else:
            secili_tarih = date.today()
        
        if not secili_otel_id:
            kullanici_otelleri = get_kullanici_otelleri()
            if kullanici_otelleri:
                secili_otel_id = kullanici_otelleri[0].id
        
        # Rapor verilerini al
        rapor = OccupancyService.get_gunluk_doluluk_raporu(secili_tarih, secili_otel_id)
        
        # Yazdırma template'ini render et
        return render_template("kat_sorumlusu/gunluk_doluluk_yazdir.html",
            rapor=rapor,
            secili_tarih=secili_tarih,
            simdi=get_kktc_now(),
            bos_odalar_dahil=bos_odalar_dahil
        )
        
    except Exception as e:
        log_hata("gunluk_doluluk_yazdir", str(e), session.get("kullanici_id"))
        flash(f"Hata oluştu: {str(e)}", "danger")
        return redirect(url_for("doluluk.gunluk_doluluk"))

@doluluk_bp.route("/kat-doluluk/<int:kat_id>")
@login_required
@role_required(["kat_sorumlusu", "depo_sorumlusu", "sistem_yoneticisi"])
def kat_doluluk_detay(kat_id):
    """Kat bazlı detaylı doluluk raporu"""
    try:
        from models import Kat, Oda, MisafirKayit, GorevDetay, GunlukGorev, MinibarIslem, DNDKontrol, OdaDNDKayit
        from sqlalchemy import and_, or_, func
        
        # Tarihi al
        tarih_str = request.args.get("tarih")
        if tarih_str:
            try:
                secili_tarih = datetime.strptime(tarih_str, "%Y-%m-%d").date()
            except ValueError:
                secili_tarih = date.today()
        else:
            secili_tarih = date.today()
        
        # Katı bul
        kat = Kat.query.get_or_404(kat_id)
        
        # Kat'a ait tüm odaları al
        odalar = Oda.query.filter_by(kat_id=kat_id, aktif=True).order_by(Oda.oda_no).all()
        oda_ids = [oda.id for oda in odalar]
        
        # Görev durumlarını ve detaylarını al (varış/çıkış saatleri için)
        gorev_durumlari = {}
        gorev_saatleri = {}  # varis_saati ve cikis_saati için
        gorev_detaylari_db = GorevDetay.query.join(GunlukGorev).filter(
            GunlukGorev.gorev_tarihi == secili_tarih
        ).all()
        for detay in gorev_detaylari_db:
            gorev_durumlari[detay.oda_id] = detay.durum
            gorev_saatleri[detay.oda_id] = {
                'varis_saati': str(detay.varis_saati) if detay.varis_saati else None,
                'cikis_saati': str(detay.cikis_saati) if detay.cikis_saati else None
            }
        
        # Bugün kontrol edilen odaları bul (MinibarIslem tablosundan)
        # islem_tipi='kontrol' veya herhangi bir minibar işlemi yapılmışsa kontrol edilmiş sayılır
        bugun_baslangic = datetime.combine(secili_tarih, datetime.min.time())
        bugun_bitis = datetime.combine(secili_tarih, datetime.max.time())
        
        kontrol_edilen_odalar = {}  # oda_id -> kontrol_tipi ('completed' veya 'dnd')
        
        # MinibarIslem tablosundan bugünkü kontrolleri al
        minibar_kontroller = db.session.query(
            MinibarIslem.oda_id,
            MinibarIslem.islem_tipi,
            MinibarIslem.aciklama
        ).filter(
            MinibarIslem.oda_id.in_(oda_ids),
            MinibarIslem.islem_tarihi >= bugun_baslangic,
            MinibarIslem.islem_tarihi <= bugun_bitis
        ).all()
        
        for kontrol in minibar_kontroller:
            # DND kontrolü mü yoksa normal kontrol mü?
            if kontrol.aciklama and 'DND' in kontrol.aciklama.upper():
                kontrol_edilen_odalar[kontrol.oda_id] = 'dnd'
            else:
                # Zaten DND olarak işaretlenmemişse completed olarak işaretle
                if kontrol.oda_id not in kontrol_edilen_odalar or kontrol_edilen_odalar[kontrol.oda_id] != 'dnd':
                    kontrol_edilen_odalar[kontrol.oda_id] = 'completed'
        
        # GorevDetay'dan da kontrol durumlarını al (dolu odalar için)
        for oda_id, durum in gorev_durumlari.items():
            if oda_id in oda_ids:
                if durum == 'dnd_pending':
                    kontrol_edilen_odalar[oda_id] = 'dnd'
                elif durum == 'completed':
                    kontrol_edilen_odalar[oda_id] = 'completed'
        
        # DND bilgilerini al (DND odaları için sayı ve son DND zamanı)
        dnd_bilgileri = {}  # oda_id -> {'dnd_sayisi': int, 'son_dnd_zamani': datetime}
        
        # 1. Önce BAĞIMSIZ DND tablosundan (oda_dnd_kayitlari) bugünkü kayıtları al
        bagimsiz_dnd_kayitlari = OdaDNDKayit.query.filter(
            OdaDNDKayit.oda_id.in_(oda_ids),
            OdaDNDKayit.kayit_tarihi == secili_tarih
        ).all()
        
        for dnd_kayit in bagimsiz_dnd_kayitlari:
            dnd_bilgileri[dnd_kayit.oda_id] = {
                'dnd_sayisi': dnd_kayit.dnd_sayisi,
                'son_dnd_zamani': dnd_kayit.son_dnd_zamani
            }
            # Kontrol durumunu da güncelle
            if dnd_kayit.durum == 'tamamlandi':
                kontrol_edilen_odalar[dnd_kayit.oda_id] = 'completed'
            else:
                kontrol_edilen_odalar[dnd_kayit.oda_id] = 'dnd'
        
        # 2. Eski görev sisteminden de DND bilgilerini al (geriye uyumluluk)
        for detay in gorev_detaylari_db:
            if detay.oda_id in oda_ids and detay.durum == 'dnd_pending':
                # Eğer bağımsız DND kaydı yoksa, eski sistemden al
                if detay.oda_id not in dnd_bilgileri:
                    # DNDKontrol tablosundan bu görev için DND kayıtlarını say
                    dnd_kayitlari = DNDKontrol.query.filter_by(gorev_detay_id=detay.id).order_by(DNDKontrol.kontrol_zamani.desc()).all()
                    
                    if dnd_kayitlari:
                        dnd_bilgileri[detay.oda_id] = {
                            'dnd_sayisi': len(dnd_kayitlari),
                            'son_dnd_zamani': dnd_kayitlari[0].kontrol_zamani
                        }
                    elif detay.dnd_sayisi and detay.dnd_sayisi > 0:
                        dnd_bilgileri[detay.oda_id] = {
                            'dnd_sayisi': detay.dnd_sayisi,
                            'son_dnd_zamani': detay.son_dnd_zamani
                        }
        
        # Her oda için doluluk durumunu kontrol et
        oda_detaylari = []
        dolu_sayisi = 0
        bos_sayisi = 0
        bugun_giris_sayisi = 0
        bugun_cikis_sayisi = 0
        
        # Arrivals ve Departures görevlerini sayaç için topla
        arrivals_gorevler = []
        departures_gorevler = []
        
        for oda in odalar:
            # Bugün için misafir kaydı var mı? (bugün çıkış yapanlar DAHİL)
            misafir = MisafirKayit.query.filter(
                MisafirKayit.oda_id == oda.id,
                MisafirKayit.giris_tarihi <= secili_tarih,
                MisafirKayit.cikis_tarihi >= secili_tarih  # >= ile bugün çıkış yapanlar da dahil
            ).first()
            
            durum = "bos"
            misafir_info = None
            kalan_gun = None
            varis_saati = None
            cikis_saati = None
            
            # Görev saatlerini al
            if oda.id in gorev_saatleri:
                varis_saati = gorev_saatleri[oda.id].get('varis_saati')
                cikis_saati = gorev_saatleri[oda.id].get('cikis_saati')
            
            if misafir:
                durum = "dolu"
                dolu_sayisi += 1
                kalan_gun = (misafir.cikis_tarihi - secili_tarih).days
                
                # Bugün giriş mi?
                if misafir.giris_tarihi == secili_tarih:
                    durum = "bugun_giris"
                    bugun_giris_sayisi += 1
                    # Arrivals listesine ekle
                    arrivals_gorevler.append({
                        'oda_no': oda.oda_no,
                        'varis_saati': varis_saati
                    })
                
                # Bugün çıkış mı?
                if misafir.cikis_tarihi == secili_tarih:
                    durum = "bugun_cikis"
                    bugun_cikis_sayisi += 1
                    # Departures listesine ekle
                    departures_gorevler.append({
                        'oda_no': oda.oda_no,
                        'cikis_saati': cikis_saati
                    })
                
                misafir_info = {
                    'misafir_sayisi': misafir.misafir_sayisi,
                    'giris_tarihi': misafir.giris_tarihi,
                    'cikis_tarihi': misafir.cikis_tarihi,
                    'kalan_gun': kalan_gun,
                    'varis_saati': varis_saati,
                    'cikis_saati': cikis_saati
                }
            else:
                bos_sayisi += 1
            
            # Kontrol durumunu belirle (hem dolu hem boş odalar için)
            kontrol_durumu = kontrol_edilen_odalar.get(oda.id, None)
            
            # DND bilgilerini al
            dnd_info = dnd_bilgileri.get(oda.id, None)
            
            # Görev durumunu belirle - bağımsız DND varsa onu kullan
            # Boş odalar için de DND kaydı olabilir
            gorev_durumu_final = gorev_durumlari.get(oda.id, 'pending') if misafir else None
            if dnd_info and dnd_info['dnd_sayisi'] > 0:
                # Bağımsız DND kaydı varsa, gorev_durumu'nu dnd_pending olarak ayarla
                gorev_durumu_final = 'dnd_pending'
                # Boş oda için de kontrol_durumu'nu dnd olarak ayarla
                if not misafir:
                    kontrol_durumu = 'dnd'
            
            oda_detaylari.append({
                'oda': oda,
                'durum': durum,
                'misafir': misafir_info,
                'gorev_durumu': gorev_durumu_final,
                'kontrol_durumu': kontrol_durumu,  # 'completed', 'dnd' veya None
                'varis_saati': varis_saati,
                'cikis_saati': cikis_saati,
                'dnd_sayisi': dnd_info['dnd_sayisi'] if dnd_info else None,
                'son_dnd_zamani': dnd_info['son_dnd_zamani'].isoformat() if dnd_info and dnd_info['son_dnd_zamani'] else None
            })
        
        return render_template("kat_sorumlusu/kat_doluluk_detay.html",
            kat=kat,
            secili_tarih=secili_tarih,
            oda_detaylari=oda_detaylari,
            toplam_oda=len(odalar),
            dolu_sayisi=dolu_sayisi,
            bos_sayisi=bos_sayisi,
            bugun_giris_sayisi=bugun_giris_sayisi,
            bugun_cikis_sayisi=bugun_cikis_sayisi,
            arrivals_gorevler=arrivals_gorevler,
            departures_gorevler=departures_gorevler
        )
        
    except Exception as e:
        log_hata("kat_doluluk_detay", str(e), session.get("kullanici_id"))
        flash(f"Hata oluştu: {str(e)}", "danger")
        return redirect(url_for("doluluk.gunluk_doluluk"))

@doluluk_bp.route("/oda-doluluk/<int:oda_id>")
@login_required
@role_required(["kat_sorumlusu", "depo_sorumlusu", "sistem_yoneticisi"])
def oda_doluluk_detay(oda_id):
    try:
        detay = OccupancyService.get_oda_detay_bilgileri(oda_id)
        if not detay:
            flash("Oda bulunamadı", "warning")
            return redirect(url_for("doluluk.gunluk_doluluk"))
        kalan_gun = None
        if detay["mevcut_misafir"]:
            kalan_gun = OccupancyService.get_kalan_gun_sayisi(detay["mevcut_misafir"].cikis_tarihi)
        gelecek_rezervasyonlar_detay = []
        for rezervasyon in detay["gelecek_rezervasyonlar"]:
            kalis_suresi = OccupancyService.get_kalis_suresi(rezervasyon.giris_tarihi, rezervasyon.cikis_tarihi)
            gelecek_rezervasyonlar_detay.append({"rezervasyon": rezervasyon, "kalis_suresi": kalis_suresi})
        return render_template("kat_sorumlusu/oda_doluluk_detay.html", oda=detay["oda"],
            mevcut_misafir=detay["mevcut_misafir"], kalan_gun=kalan_gun,
            gelecek_rezervasyonlar=gelecek_rezervasyonlar_detay, gecmis_kayitlar=detay["gecmis_kayitlar"])
    except Exception as e:
        log_hata("oda_doluluk_detay", str(e), session.get("kullanici_id"))
        flash(f"Hata oluştu: {str(e)}", "danger")
        return redirect(url_for("doluluk.gunluk_doluluk"))

@doluluk_bp.route("/doluluk-yonetimi/onizle", methods=["POST"])
@login_required
@role_required("depo_sorumlusu")
def doluluk_onizle():
    """Excel dosyasını önizle - veritabanına yazmadan özet göster ve validasyon yap"""
    try:
        from utils.excel_service import ExcelProcessingService
        from utils.file_management_service import FileManagementService
        from utils.authorization import get_kullanici_otelleri
        from models import Oda, Kat, MisafirKayit, Otel
        import tempfile
        import os
        from collections import defaultdict

        if 'excel_file' not in request.files:
            return jsonify({"success": False, "error": "Dosya bulunamadı"}), 400

        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({"success": False, "error": "Dosya seçilmedi"}), 400

        # Geçici dosya oluştur
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        file.save(temp_path)
        temp_file.close()

        # Otel ID'yi al
        otel_id = request.form.get('otel_id', type=int)
        if not otel_id:
            kullanici_otelleri = get_kullanici_otelleri()
            if kullanici_otelleri:
                otel_id = kullanici_otelleri[0].id

        # Excel'i analiz et (veritabanına yazmadan)
        import openpyxl
        import pandas as pd
        workbook = openpyxl.load_workbook(temp_path, data_only=True)
        sheet = workbook.active

        # Başlıkları al
        headers = [cell.value for cell in sheet[1]]
        header_dosya_tipi = ExcelProcessingService.detect_file_type(headers)
        
        # Tarih bazlı akıllı algılama
        try:
            df_preview = pd.read_excel(temp_path, header=0)
            if 'Arrival' in df_preview.columns and 'Departure' in df_preview.columns:
                smart_dosya_tipi = ExcelProcessingService._detect_file_type_by_dates_standard(
                    df_preview, 'Arrival', 'Departure'
                )
                if smart_dosya_tipi:
                    dosya_tipi = smart_dosya_tipi
                    print(f"✅ Önizleme - Tarih bazlı akıllı algılama: {dosya_tipi}")
                else:
                    dosya_tipi = header_dosya_tipi
                    print(f"✅ Önizleme - Header bazlı algılama: {dosya_tipi}")
            else:
                dosya_tipi = header_dosya_tipi
        except Exception as e:
            print(f"⚠️ Önizleme akıllı algılama hatası: {str(e)}")
            dosya_tipi = header_dosya_tipi

        # Sütun indekslerini bul
        col_indices = {}
        for idx, header in enumerate(headers):
            header_str = str(header).strip() if header else ''
            if header_str == 'Room no':
                col_indices['oda_no'] = idx
            elif header_str == 'Arrival':
                col_indices['giris'] = idx
            elif header_str == 'Departure':
                col_indices['cikis'] = idx
            elif header_str in ('Adult', 'Adults'):  # Her iki format da destekleniyor
                col_indices['misafir'] = idx
            elif header_str == 'Arr.Time':
                col_indices['giris_saati'] = idx
            elif header_str == 'Dep.Time':
                col_indices['cikis_saati'] = idx

        # Önce dosyadaki tüm oda numaralarını topla (otel tespiti için)
        dosyadaki_odalar = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            oda_no = str(row[col_indices['oda_no']]).strip() if col_indices.get('oda_no') is not None and row[col_indices['oda_no']] else None
            if oda_no and oda_no != 'None':
                dosyadaki_odalar.add(oda_no)

        # Otel tespiti - dosyadaki oda numaralarına bakarak hangi otele ait olduğunu bul
        tespit_edilen_otel = None
        tespit_edilen_otel_adi = None
        otel_eslesme_oranlari = {}
        
        # Tüm otelleri ve odalarını al
        tum_oteller = Otel.query.filter_by(aktif=True).all()
        for otel in tum_oteller:
            otel_odalari = {oda.oda_no for oda in Oda.query.join(Kat).filter(Kat.otel_id == otel.id).all()}
            if otel_odalari:
                eslesen_odalar = dosyadaki_odalar & otel_odalari
                eslesme_orani = len(eslesen_odalar) / len(dosyadaki_odalar) * 100 if dosyadaki_odalar else 0
                otel_eslesme_oranlari[otel.id] = {
                    'otel_adi': otel.ad,
                    'eslesen': len(eslesen_odalar),
                    'toplam_dosya': len(dosyadaki_odalar),
                    'oran': eslesme_orani
                }
                # En yüksek eşleşme oranına sahip oteli seç
                if eslesme_orani > 50:  # En az %50 eşleşme olmalı
                    if tespit_edilen_otel is None or eslesme_orani > otel_eslesme_oranlari.get(tespit_edilen_otel, {}).get('oran', 0):
                        tespit_edilen_otel = otel.id
                        tespit_edilen_otel_adi = otel.ad

        # Validasyon için veriler
        toplam_satir = 0
        gecerli_satir = 0
        hatali_satir = 0
        uyarilar = []
        tum_satirlar = []  # Tüm satırları tutacak liste
        oda_sayaci = defaultdict(int)  # Duplicate kontrolü
        bulunamayan_odalar = set()
        
        # Mevcut odaları al
        if otel_id:
            mevcut_odalar = {oda.oda_no for oda in Oda.query.join(Kat).filter(Kat.otel_id == otel_id).all()}
        else:
            mevcut_odalar = {oda.oda_no for oda in Oda.query.all()}

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Boş satır
                continue
                
            toplam_satir += 1
            
            # Veri çıkar
            oda_no = str(row[col_indices['oda_no']]).strip() if col_indices.get('oda_no') is not None and row[col_indices['oda_no']] else None
            giris = str(row[col_indices['giris']]) if col_indices.get('giris') is not None and row[col_indices['giris']] else None
            cikis = str(row[col_indices['cikis']]) if col_indices.get('cikis') is not None and row[col_indices['cikis']] else None
            misafir = str(row[col_indices['misafir']]) if col_indices.get('misafir') is not None and row[col_indices['misafir']] else None
            # ARRIVALS için giriş saati
            giris_saati = str(row[col_indices['giris_saati']]) if col_indices.get('giris_saati') is not None and row[col_indices['giris_saati']] else None
            
            # Validasyon
            satir_hatalari = []
            
            if not oda_no or oda_no == 'None':
                satir_hatalari.append("Oda numarası eksik")
                hatali_satir += 1
            else:
                # Duplicate kontrolü
                oda_sayaci[oda_no] += 1
                
                # Oda var mı kontrolü
                if oda_no not in mevcut_odalar:
                    bulunamayan_odalar.add(oda_no)
                    satir_hatalari.append(f"Oda '{oda_no}' sistemde bulunamadı")
            
            if not giris or giris == 'None':
                satir_hatalari.append("Giriş tarihi eksik")
                hatali_satir += 1
            
            if not cikis or cikis == 'None':
                satir_hatalari.append("Çıkış tarihi eksik")
                hatali_satir += 1
            
            # Misafir sayısı kontrolü - Departures için zorunlu değil
            if dosya_tipi != 'departures' and (not misafir or misafir == 'None'):
                satir_hatalari.append("Misafir sayısı eksik")
                hatali_satir += 1
            
            if not satir_hatalari:
                gecerli_satir += 1
            
            # Tüm satırları ekle (eklenecek veriler olarak gösterilecek)
            tum_satirlar.append({
                'satir_no': idx,
                'oda_no': oda_no or 'None',
                'giris_saati': giris_saati or '-',
                'giris': giris or '-',
                'cikis': cikis or '-',
                'misafir': misafir or '-',
                'hatalar': satir_hatalari,
                'gecerli': len(satir_hatalari) == 0
            })
            
            # Hataları kaydet
            if satir_hatalari:
                uyarilar.append(f"Satır {idx}: {', '.join(satir_hatalari)}")

        # Duplicate uyarıları
        duplicate_odalar = {oda: sayi for oda, sayi in oda_sayaci.items() if sayi > 1}
        if duplicate_odalar:
            for oda, sayi in duplicate_odalar.items():
                uyarilar.append(f"⚠️ Oda '{oda}' dosyada {sayi} kez tekrar ediyor (Duplicate)")

        # Bulunamayan odalar uyarısı
        if bulunamayan_odalar:
            uyarilar.append(f"⚠️ Sistemde bulunamayan odalar: {', '.join(sorted(bulunamayan_odalar))}")

        # Geçici dosyayı sil
        os.unlink(temp_path)

        return jsonify({
            "success": True,
            "dosya_tipi": dosya_tipi,
            "dosya_tipi_adi": "ARRIVALS" if dosya_tipi == "arrivals" else ("DEPARTURES" if dosya_tipi == "departures" else "IN HOUSE"),
            "toplam_satir": toplam_satir,
            "gecerli_satir": gecerli_satir,
            "hatali_satir": hatali_satir,
            "tum_satirlar": tum_satirlar,  # Tüm satırları gönder
            "uyarilar": uyarilar[:20],  # İlk 20 uyarı
            "dosya_adi": file.filename,
            "duplicate_sayisi": len(duplicate_odalar),
            "bulunamayan_oda_sayisi": len(bulunamayan_odalar),
            # Otel tespiti bilgileri
            "tespit_edilen_otel_id": tespit_edilen_otel,
            "tespit_edilen_otel_adi": tespit_edilen_otel_adi,
            "otel_eslesme_oranlari": otel_eslesme_oranlari
        })

    except Exception as e:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.unlink(temp_path)
        log_hata("doluluk_onizle", str(e), session.get("kullanici_id"))
        return jsonify({"success": False, "error": f"Önizleme hatası: {str(e)}"}), 500

@doluluk_bp.route("/doluluk-yonetimi/yukle", methods=["POST"])
@login_required
@role_required("depo_sorumlusu")
def doluluk_yukle():
    """Excel dosyası yükleme endpoint'i - Kullanıcı onaylı dosya tipi desteği"""
    try:
        from utils.excel_service import ExcelProcessingService
        from utils.file_management_service import FileManagementService
        from utils.authorization import get_kullanici_otelleri
        import json

        if 'excel_file' not in request.files:
            flash("Dosya bulunamadı", "danger")
            return redirect(url_for("doluluk.doluluk_yonetimi"))

        file = request.files['excel_file']
        if file.filename == '':
            flash("Dosya seçilmedi", "danger")
            return redirect(url_for("doluluk.doluluk_yonetimi"))

        if not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
            flash("Sadece Excel dosyaları (.xlsx, .xls, .xlsm) yüklenebilir", "danger")
            return redirect(url_for("doluluk.doluluk_yonetimi"))

        user_id = session.get("kullanici_id")
        
        # Otel ID'yi al (varsa)
        otel_id = request.form.get('otel_id', type=int)
        if not otel_id:
            kullanici_otelleri = get_kullanici_otelleri()
            if kullanici_otelleri:
                otel_id = kullanici_otelleri[0].id

        # Kullanıcının onayladığı dosya tipi (override)
        override_dosya_tipi = request.form.get('dosya_tipi')
        if override_dosya_tipi and override_dosya_tipi in ['in_house', 'arrivals', 'departures']:
            print(f"📋 Kullanıcı dosya tipini override etti: {override_dosya_tipi}")

        # 1. Dosyayı kaydet (otel_id ile birlikte)
        success, file_path, islem_kodu, error = FileManagementService.save_uploaded_file(file, user_id, otel_id)
        
        if not success:
            flash(f"Dosya kaydedilemedi: {error}", "danger")
            return redirect(url_for("doluluk.doluluk_yonetimi"))

        # 2. Excel'i işle (override_dosya_tipi varsa kullan)
        result = ExcelProcessingService.process_excel_file(file_path, islem_kodu, user_id, otel_id, override_dosya_tipi)

        # 3. Durumu güncelle
        if result['success']:
            FileManagementService.update_dosya_yukleme_status(
                islem_kodu,
                'tamamlandi',
                dosya_tipi=result['dosya_tipi'],
                toplam_satir=result['toplam_satir'],
                basarili_satir=result['basarili_satir'],
                hatali_satir=result['hatali_satir'],
                hata_detaylari=json.dumps(result['hatalar'], ensure_ascii=False) if result['hatalar'] else None
            )
            
            # YuklemeGorev tablosunu güncelle (dashboard için)
            try:
                from models import YuklemeGorev
                from datetime import date, datetime, timezone
                
                # Dosya tipini YuklemeGorev formatına çevir
                dosya_tipi_map = {
                    'in_house': 'inhouse',
                    'arrivals': 'arrivals', 
                    'departures': 'departures'
                }
                yukleme_dosya_tipi = dosya_tipi_map.get(result['dosya_tipi'], result['dosya_tipi'])
                dosya_yukleme = DosyaYukleme.query.filter_by(islem_kodu=islem_kodu).first()
                
                # Bugünkü görevi bul ve güncelle - otel_id ile ara
                yukleme_gorev = YuklemeGorev.query.filter(
                    YuklemeGorev.otel_id == otel_id,
                    YuklemeGorev.gorev_tarihi == date.today(),
                    YuklemeGorev.dosya_tipi == yukleme_dosya_tipi
                ).first()
                
                if yukleme_gorev:
                    # Mevcut görevi güncelle
                    yukleme_gorev.durum = 'completed'
                    yukleme_gorev.yukleme_zamani = get_kktc_now()
                    yukleme_gorev.dosya_yukleme_id = dosya_yukleme.id if dosya_yukleme else None
                else:
                    # Görev yoksa yeni oluştur
                    yukleme_gorev = YuklemeGorev(
                        otel_id=otel_id,
                        depo_sorumlusu_id=user_id,
                        gorev_tarihi=date.today(),
                        dosya_tipi=yukleme_dosya_tipi,
                        durum='completed',
                        yukleme_zamani=get_kktc_now(),
                        dosya_yukleme_id=dosya_yukleme.id if dosya_yukleme else None
                    )
                    db.session.add(yukleme_gorev)
                
                db.session.commit()
                
                # 🔥 KRİTİK: YuklemeGorev güncellendikten SONRA görev oluşturma kontrolü yap
                # 3 dosya da yüklendiyse günlük görevleri oluştur
                try:
                    from utils.gorev_service import GorevService
                    
                    # 3 dosya tipi için yükleme durumlarını kontrol et
                    yukleme_durumlari = {}
                    for tip in ['inhouse', 'arrivals', 'departures']:
                        yukleme = YuklemeGorev.query.filter(
                            YuklemeGorev.otel_id == otel_id,
                            YuklemeGorev.gorev_tarihi == date.today(),
                            YuklemeGorev.dosya_tipi == tip
                        ).first()
                        yukleme_durumlari[tip] = yukleme.durum if yukleme else 'pending'
                    
                    print(f"📊 Otel {otel_id} - Yükleme durumları: {yukleme_durumlari}")
                    
                    # 3 dosya da yüklendi mi kontrol et
                    tum_dosyalar_yuklendi = all(
                        durum == 'completed' for durum in yukleme_durumlari.values()
                    )
                    
                    if tum_dosyalar_yuklendi:
                        print(f"✅ Otel {otel_id} - 3 dosya da yüklendi! Görevler oluşturuluyor...")
                        gorev_result = GorevService.create_daily_tasks(otel_id, date.today())
                        print(f"✅ Görevler oluşturuldu: {gorev_result}")
                        
                        # Kat sorumlularına bildirim gönder
                        if gorev_result.get('toplam_oda_sayisi', 0) > 0:
                            try:
                                from utils.bildirim_service import BildirimService
                                from models import Kullanici
                                
                                kat_sorumluları = Kullanici.query.filter(
                                    Kullanici.otel_id == otel_id,
                                    Kullanici.rol == 'kat_sorumlusu',
                                    Kullanici.aktif == True
                                ).all()
                                
                                gorev_tipleri = [
                                    ('inhouse_kontrol', gorev_result.get('inhouse_gorev_sayisi', 0)),
                                    ('arrival_kontrol', gorev_result.get('arrival_gorev_sayisi', 0)),
                                    ('departure_kontrol', gorev_result.get('departure_gorev_sayisi', 0))
                                ]
                                
                                for ks in kat_sorumluları:
                                    for gorev_tipi, oda_sayisi in gorev_tipleri:
                                        if oda_sayisi > 0:
                                            BildirimService.send_task_created_notification(
                                                personel_id=ks.id,
                                                gorev_tipi=gorev_tipi,
                                                oda_sayisi=oda_sayisi
                                            )
                            except Exception as bildirim_err:
                                print(f"⚠️ Bildirim gönderme hatası: {bildirim_err}")
                    else:
                        eksik_dosyalar = [tip for tip, durum in yukleme_durumlari.items() if durum != 'completed']
                        print(f"⏳ Otel {otel_id} - Eksik dosyalar: {eksik_dosyalar}. Görevler henüz oluşturulmayacak.")
                        
                except Exception as gorev_olusturma_err:
                    print(f"⚠️ Görev oluşturma hatası: {gorev_olusturma_err}")
                    import traceback
                    traceback.print_exc()
                    
            except Exception as gorev_err:
                print(f"YuklemeGorev güncelleme hatası: {gorev_err}")
            
            flash(f"✓ Dosya başarıyla işlendi! İşlem Kodu: {islem_kodu} | "
                  f"Başarılı: {result['basarili_satir']}, Hatalı: {result['hatali_satir']}", "success")
        else:
            FileManagementService.update_dosya_yukleme_status(
                islem_kodu,
                'hata',
                hata_detaylari=json.dumps(result.get('hatalar', [result.get('error', 'Bilinmeyen hata')]), ensure_ascii=False)
            )
            flash(f"Dosya işlenirken hata oluştu: {result.get('error', 'Bilinmeyen hata')}", "danger")

        return redirect(url_for("doluluk.doluluk_yonetimi"))
    except Exception as e:
        log_hata("doluluk_yukle", str(e), session.get("kullanici_id"))
        flash(f"Hata oluştu: {str(e)}", "danger")
        return redirect(url_for("doluluk.doluluk_yonetimi"))

@doluluk_bp.route("/doluluk-yonetimi/sil/<string:islem_kodu>", methods=["POST"])
@login_required
@role_required("depo_sorumlusu")
def doluluk_sil(islem_kodu):
    """Yükleme işlemini ve ilgili kayıtları sil"""
    try:
        from utils.file_management_service import FileManagementService

        yukleme = DosyaYukleme.query.filter_by(islem_kodu=islem_kodu).first()
        if not yukleme:
            flash("Yükleme kaydı bulunamadı", "warning")
            return redirect(url_for("doluluk.doluluk_yonetimi"))

        # Sadece kendi yüklemelerini silebilir
        if yukleme.yuklenen_kullanici_id != session.get("kullanici_id"):
            flash("Bu işlemi silme yetkiniz yok", "danger")
            return redirect(url_for("doluluk.doluluk_yonetimi"))

        # İlişkili misafir kayıtlarını ve dosyayı sil
        # Tamamlanmış görevler korunur, bekleyen görevler silinir
        success, message, summary = FileManagementService.delete_upload_by_islem_kodu(
            islem_kodu, 
            session.get("kullanici_id")
        )

        if success:
            flash(f"✓ {message}", "success")
            # Korunan görev varsa bilgi ver
            if summary.get('preserved_completed_tasks', 0) > 0:
                flash(f"ℹ️ {summary['preserved_completed_tasks']} tamamlanmış görev kaydı korundu", "info")
        else:
            flash(f"Silme hatası: {message}", "danger")

        return redirect(url_for("doluluk.doluluk_yonetimi"))
    except Exception as e:
        log_hata("doluluk_sil", str(e), session.get("kullanici_id"))
        flash(f"Hata oluştu: {str(e)}", "danger")
        return redirect(url_for("doluluk.doluluk_yonetimi"))

@doluluk_bp.route("/doluluk-yonetimi/durum/<string:islem_kodu>", methods=["GET"])
@login_required
@role_required("depo_sorumlusu")
def doluluk_durum(islem_kodu):
    """Yükleme durumunu ve hata detaylarını getir"""
    try:
        yukleme = DosyaYukleme.query.filter_by(islem_kodu=islem_kodu).first()
        if not yukleme:
            return jsonify({"success": False, "error": "Yükleme kaydı bulunamadı"}), 404

        # Sadece kendi yüklemelerini görüntüleyebilir
        if yukleme.yuklenen_kullanici_id != session.get("kullanici_id"):
            return jsonify({"success": False, "error": "Bu işlemi görüntüleme yetkiniz yok"}), 403

        import json
        hata_detaylari = None
        if yukleme.hata_detaylari:
            try:
                # İlk parse
                parsed = json.loads(yukleme.hata_detaylari) if isinstance(yukleme.hata_detaylari, str) else yukleme.hata_detaylari
                # Eğer hala string ise (çift encode durumu) tekrar parse et
                if isinstance(parsed, str):
                    hata_detaylari = json.loads(parsed)
                else:
                    hata_detaylari = parsed
            except Exception as parse_err:
                print(f"Hata detayları parse hatası: {parse_err}")
                hata_detaylari = [yukleme.hata_detaylari]

        return jsonify({
            "success": True,
            "durum": yukleme.durum,
            "hata_detaylari": hata_detaylari,
            "basarili_satir": yukleme.basarili_satir,
            "hatali_satir": yukleme.hatali_satir,
            "toplam_satir": yukleme.toplam_satir
        })
    except Exception as e:
        log_hata("doluluk_durum", str(e), session.get("kullanici_id"))
        return jsonify({"success": False, "error": str(e)}), 500
