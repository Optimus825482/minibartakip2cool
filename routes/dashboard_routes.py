"""
Dashboard Route'ları

Bu modül tüm kullanıcı rollerine ait dashboard endpoint'lerini içerir.

Endpoint'ler:
- /dashboard - Rol bazlı yönlendirme
- /sistem-yoneticisi - Sistem yöneticisi dashboard
- /depo - Depo sorumlusu dashboard
- /kat-sorumlusu - Kat sorumlusu dashboard
- /kat-sorumlusu/dashboard - Kat sorumlusu dashboard (alternatif)
- /api/son-aktiviteler - Son aktiviteler API
- /api/bekleyen-dolum-sayisi - Bekleyen dolum sayısı API
- /api/tuketim-trendleri - Tüketim trendleri API
- /sistem-yoneticisi/audit-trail - Audit trail sayfası
- /sistem-yoneticisi/audit-trail/<id> - Audit trail detay API
- /sistem-yoneticisi/audit-trail/export - Audit trail Excel export
"""

from flask import render_template, redirect, url_for, flash, session, request, jsonify, send_file
from datetime import datetime, timedelta, timezone
from io import BytesIO
import json
import logging
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

# KKTC Timezone
KKTC_TZ = pytz.timezone('Europe/Nicosia')

def get_kktc_now():
    """Kıbrıs saat diliminde şu anki zamanı döndürür."""
    return datetime.now(KKTC_TZ)

from models import db, Kat, Oda, Kullanici, UrunGrup, Urun, StokHareket, PersonelZimmet, PersonelZimmetDetay, MinibarIslem, MinibarIslemDetay, SistemLog, MinibarDolumTalebi, AuditLog
from utils.decorators import login_required, role_required
from utils.helpers import get_kritik_stok_urunler, get_tum_urunler_stok_durumlari, get_kat_sorumlusu_kritik_stoklar

# Dashboard bildirim servisini import et
try:
    from utils.dashboard_servisleri import DashboardBildirimServisi
except ImportError:
    DashboardBildirimServisi = None

# Cache + Eager Loading servisleri (1.1.2026)
from utils.dashboard_data_service import DashboardDataService
from utils.master_data_service import MasterDataService


def register_dashboard_routes(app):
    """Dashboard route'larını kaydet"""
    
    @app.route('/dashboard')
    @login_required
    def dashboard():
        """Rol bazlı dashboard yönlendirmesi"""
        # Login sonrası cache temizleme flag'ini kaldır (bir kez çalışması için)
        if session.get('clear_cache'):
            session.pop('clear_cache', None)
        
        rol = session.get('rol')
        
        if rol == 'superadmin':
            return redirect(url_for('sistem_yoneticisi_dashboard'))
        elif rol == 'sistem_yoneticisi':
            return redirect(url_for('sistem_yoneticisi_dashboard'))
        elif rol == 'admin':
            return redirect(url_for('sistem_yoneticisi_dashboard'))  # Admin de sistem yöneticisi panelini kullanır
        elif rol == 'depo_sorumlusu':
            return redirect(url_for('depo_dashboard'))
        elif rol == 'kat_sorumlusu':
            return redirect(url_for('kat_sorumlusu_dashboard'))
        else:
            flash('Geçersiz kullanıcı rolü!', 'danger')
            return redirect(url_for('logout'))

    
    @app.route('/sistem-yoneticisi')
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def sistem_yoneticisi_dashboard():
        """Sistem yöneticisi dashboard"""
        try:
            from utils.occupancy_service import OccupancyService
            from utils.authorization import get_kullanici_otelleri
            from datetime import date
            
            # ML Alertleri (DB tabanlı toggle ile kontrol)
            ml_alerts = []
            ml_alert_count = 0
            try:
                from utils.ml_toggle import is_ml_enabled
                ml_enabled = is_ml_enabled()
            except Exception:
                ml_enabled = False
            
            if ml_enabled:
                try:
                    from models import MLAlert
                    from utils.ml.alert_manager import AlertManager
                
                    alert_manager = AlertManager(db)
                    # Son 5 kritik/yüksek alert
                    ml_alerts = alert_manager.get_active_alerts(limit=5)
                    ml_alert_count = len(alert_manager.get_active_alerts())
                except Exception as e:
                    print(f"ML alert hatası: {str(e)}")
        
            # Tüm oteller için doluluk raporları
            otel_doluluk_raporlari = []
            try:
                kullanici_otelleri = get_kullanici_otelleri()
                for otel in kullanici_otelleri:
                    doluluk = OccupancyService.get_gunluk_doluluk_raporu(date.today(), otel.id)
                    # Doluluk oranı hesapla
                    if doluluk['toplam_oda'] > 0:
                        doluluk['doluluk_orani'] = round((doluluk['dolu_oda'] / doluluk['toplam_oda']) * 100)
                    else:
                        doluluk['doluluk_orani'] = 0
                
                    # Otel bilgilerini ekle
                    doluluk['otel'] = otel
                    otel_doluluk_raporlari.append(doluluk)
            except Exception as e:
                print(f"Doluluk raporları hatası: {str(e)}")
        
            # İstatistikler - Cache + Eager Loading ile optimize edildi (1.1.2026)
            # Genel istatistikler (cached)
            genel_stats = DashboardDataService.get_genel_istatistikler()
            toplam_kat = genel_stats['toplam_kat']
            toplam_oda = genel_stats['toplam_oda']
            toplam_kullanici = genel_stats['toplam_kullanici']
            toplam_personel = genel_stats['toplam_personel']
            admin_count = genel_stats['admin_count']
            depo_count = genel_stats['depo_count']
            kat_count = genel_stats['kat_count']
            toplam_urun_grup = genel_stats['toplam_urun_grup']
            toplam_urun = genel_stats['toplam_urun']
            
            # Kat-oda dağılımı (cached)
            kat_oda = DashboardDataService.get_kat_oda_dagilimi()
            kat_labels = kat_oda['kat_labels']
            kat_oda_sayilari = kat_oda['kat_oda_sayilari']
            
            # Son eklenenler (eager loading)
            son_eklenenler = DashboardDataService.get_son_eklenenler(5)
            son_katlar = son_eklenenler['son_katlar']
            son_odalar = son_eklenenler['son_odalar']
            son_personeller = son_eklenenler['son_personeller']
            son_urunler = son_eklenenler['son_urunler']
            
            # Ürün tüketim verileri (cached)
            tuketim = DashboardDataService.get_urun_tuketim_verileri(30, 10)
            urun_labels = tuketim['urun_labels']
            urun_tuketim_miktarlari = tuketim['urun_tuketim_miktarlari']
        
            kritik_urunler = get_kritik_stok_urunler()
        
            # Gelişmiş stok durumları
            try:
                stok_durumlari = get_tum_urunler_stok_durumlari()
            except Exception as e:
                print(f"Stok durumları hatası: {str(e)}")
                flash(f'Stok durumları yüklenirken hata oluştu: {str(e)}', 'warning')
                stok_durumlari = {
                    'kritik': [],
                    'dikkat': [],
                    'normal': [],
                    'istatistik': {
                        'toplam': 0,
                        'kritik_sayi': 0,
                        'dikkat_sayi': 0,
                        'normal_sayi': 0
                    }
                }
        
            # Dashboard bildirimleri (Satın Alma Modülü)
            dashboard_bildirimleri = []
            bildirim_sayilari = {
                'kritik_stok': 0,
                'geciken_siparis': 0,
                'onay_bekleyen': 0,
                'toplam': 0
            }
            if DashboardBildirimServisi:
                try:
                    kullanici_id = session.get('kullanici_id')
                    dashboard_bildirimleri = DashboardBildirimServisi.get_dashboard_bildirimleri(
                        kullanici_id, 'sistem_yoneticisi'
                    )
                    bildirim_sayilari = DashboardBildirimServisi.get_bildirim_sayilari(
                        kullanici_id, 'sistem_yoneticisi'
                    )
                except Exception as e:
                    print(f"Dashboard bildirimleri hatası: {str(e)}")
                    db.session.rollback()  # Transaction'ı temizle
        
            # Sipariş istatistikleri (eski modül kaldırıldı)
            istatistikler = {'onaylandi': 0}
            
            # Ana Depo Tedarik bildirimleri (görülmemiş tedarikler)
            from models import AnaDepoTedarik
            ana_depo_tedarik_sayisi = 0
            try:
                ana_depo_tedarik_sayisi = AnaDepoTedarik.query.filter_by(sistem_yoneticisi_goruldu=False).count()
            except Exception as e:
                print(f"Ana depo tedarik sayısı hatası: {str(e)}")
                db.session.rollback()
        
            # bugun değişkeni tanımlı değilse tanımla
            try:
                bugun
            except NameError:
                bugun = get_kktc_now().date()
        
            # Günlük Yükleme Görev Özeti (Sistem Yöneticisi için)
            yukleme_gorev_ozeti = None
            try:
                from models import YuklemeGorev
                kullanici_otelleri = get_kullanici_otelleri()
                otel_ids = [o.id for o in kullanici_otelleri] if kullanici_otelleri else []
                
                if otel_ids:
                    yukleme_gorevler = YuklemeGorev.query.filter(
                        YuklemeGorev.otel_id.in_(otel_ids),
                        db.func.date(YuklemeGorev.gorev_tarihi) == bugun,
                        YuklemeGorev.dosya_tipi.in_(['inhouse', 'arrivals', 'departures'])
                    ).all()

                    yukleme_gorev_map = {g.dosya_tipi: g for g in yukleme_gorevler}
                    inhouse_gorev = yukleme_gorev_map.get('inhouse')
                    arrivals_gorev = yukleme_gorev_map.get('arrivals')
                    departures_gorev = yukleme_gorev_map.get('departures')
                    
                    tamamlanan = 0
                    inhouse_durum = 'pending'
                    arrivals_durum = 'pending'
                    departures_durum = 'pending'
                    
                    if inhouse_gorev and inhouse_gorev.durum == 'completed':
                        tamamlanan += 1
                        inhouse_durum = 'completed'
                    
                    if arrivals_gorev and arrivals_gorev.durum == 'completed':
                        tamamlanan += 1
                        arrivals_durum = 'completed'
                    
                    if departures_gorev and departures_gorev.durum == 'completed':
                        tamamlanan += 1
                        departures_durum = 'completed'
                    
                    yukleme_gorev_ozeti = {
                        'toplam': 3,
                        'tamamlanan': tamamlanan,
                        'bekleyen': 3 - tamamlanan,
                        'tamamlanma_orani': int((tamamlanan / 3) * 100),
                        'inhouse_durum': inhouse_durum,
                        'arrivals_durum': arrivals_durum,
                        'departures_durum': departures_durum
                    }
            except Exception as e:
                print(f"Sistem yöneticisi yükleme görev özeti hatası: {str(e)}")
            
            # Eksik Doluluk Yüklemeleri Kontrolü (Sistem Yöneticisi için - Saat 10:00 sonrası)
            eksik_doluluk_yuklemeleri = []
            try:
                import pytz
                from models import YuklemeGorev, KullaniciOtel
                
                kktc_tz = pytz.timezone('Europe/Nicosia')
                now_kktc = datetime.now(kktc_tz)
                
                # Saat 10:00'dan sonra mı kontrol et
                if now_kktc.hour >= 10:
                    kullanici_otelleri = get_kullanici_otelleri()
                    otel_ids = [o.id for o in kullanici_otelleri] if kullanici_otelleri else []
                    
                    if otel_ids:
                        for otel in kullanici_otelleri:
                            # Bu otel için bugünkü yükleme görevlerini kontrol et
                            inhouse_gorev = YuklemeGorev.query.filter(
                                YuklemeGorev.otel_id == otel.id,
                                db.func.date(YuklemeGorev.gorev_tarihi) == bugun,
                                YuklemeGorev.dosya_tipi == 'inhouse'
                            ).first()
                            
                            arrivals_gorev = YuklemeGorev.query.filter(
                                YuklemeGorev.otel_id == otel.id,
                                db.func.date(YuklemeGorev.gorev_tarihi) == bugun,
                                YuklemeGorev.dosya_tipi == 'arrivals'
                            ).first()
                            
                            departures_gorev = YuklemeGorev.query.filter(
                                YuklemeGorev.otel_id == otel.id,
                                db.func.date(YuklemeGorev.gorev_tarihi) == bugun,
                                YuklemeGorev.dosya_tipi == 'departures'
                            ).first()
                            
                            # Eksik yüklemeleri belirle (In House, Arrivals, Departures)
                            otel_eksikler = []
                            if not inhouse_gorev or inhouse_gorev.durum == 'pending':
                                otel_eksikler.append('In House')
                            if not arrivals_gorev or arrivals_gorev.durum == 'pending':
                                otel_eksikler.append('Arrivals')
                            if not departures_gorev or departures_gorev.durum == 'pending':
                                otel_eksikler.append('Departures')
                            
                            if otel_eksikler:
                                # Depo sorumlularını bul
                                depo_sorumlu_atamalari = KullaniciOtel.query.join(Kullanici).filter(
                                    KullaniciOtel.otel_id == otel.id,
                                    Kullanici.rol == 'depo_sorumlusu',
                                    Kullanici.aktif == True
                                ).all()
                                
                                depo_sorumlulari = [
                                    f"{a.kullanici.ad} {a.kullanici.soyad}" for a in depo_sorumlu_atamalari
                                ]
                                
                                eksik_doluluk_yuklemeleri.append({
                                    'otel_id': otel.id,
                                    'otel_ad': otel.ad,
                                    'eksik_dosyalar': otel_eksikler,
                                    'depo_sorumlulari': depo_sorumlulari
                                })
            except Exception as e:
                print(f"Eksik doluluk yüklemeleri kontrolü hatası: {str(e)}")
            
            # Kat Sorumlusu Görevleri Özeti (Sistem Yöneticisi için)
            kat_sorumlusu_gorev_ozeti = None
            try:
                from models import GunlukGorev
                kullanici_otelleri = get_kullanici_otelleri()
                otel_ids = [o.id for o in kullanici_otelleri] if kullanici_otelleri else []
                
                if otel_ids:
                    # Bugünkü görevleri say - GunlukGorev üzerinden
                    bugun_gorevler = GunlukGorev.query.filter(
                        GunlukGorev.otel_id.in_(otel_ids),
                        db.func.date(GunlukGorev.gorev_tarihi) == bugun
                    ).all()
                    
                    toplam_gorev = len(bugun_gorevler)
                    tamamlanan_gorev = len([g for g in bugun_gorevler if g.durum == 'completed'])
                    bekleyen_gorev = len([g for g in bugun_gorevler if g.durum == 'pending'])
                    devam_eden = len([g for g in bugun_gorevler if g.durum == 'in_progress'])
                    dnd_gorev = len([g for g in bugun_gorevler if g.durum == 'dnd_pending'])
                    
                    # Görev tiplerine göre say
                    inhouse = len([g for g in bugun_gorevler if g.gorev_tipi == 'inhouse_kontrol'])
                    arrival = len([g for g in bugun_gorevler if g.gorev_tipi == 'arrival_kontrol'])
                    departure = len([g for g in bugun_gorevler if g.gorev_tipi == 'departure_kontrol'])
                    
                    inhouse_tamamlanan = len([g for g in bugun_gorevler if g.gorev_tipi == 'inhouse_kontrol' and g.durum == 'completed'])
                    arrival_tamamlanan = len([g for g in bugun_gorevler if g.gorev_tipi == 'arrival_kontrol' and g.durum == 'completed'])
                    departure_tamamlanan = len([g for g in bugun_gorevler if g.gorev_tipi == 'departure_kontrol' and g.durum == 'completed'])
                    
                    kat_sorumlusu_gorev_ozeti = {
                        'toplam': toplam_gorev,
                        'tamamlanan': tamamlanan_gorev,
                        'bekleyen': bekleyen_gorev,
                        'devam_eden': devam_eden,
                        'dnd': dnd_gorev,
                        'tamamlanma_orani': int((tamamlanan_gorev / toplam_gorev) * 100) if toplam_gorev > 0 else 0,
                        'inhouse': {'toplam': inhouse, 'tamamlanan': inhouse_tamamlanan},
                        'arrival': {'toplam': arrival, 'tamamlanan': arrival_tamamlanan},
                        'departure': {'toplam': departure, 'tamamlanan': departure_tamamlanan}
                    }
            except Exception as e:
                print(f"Kat sorumlusu görev özeti hatası: {str(e)}")
            
            return render_template('sistem_yoneticisi/dashboard.html',
                                 otel_doluluk_raporlari=otel_doluluk_raporlari,
                                 toplam_kat=toplam_kat,
                                 toplam_oda=toplam_oda,
                                 toplam_kullanici=toplam_kullanici,
                                 toplam_personel=toplam_personel,
                                 son_katlar=son_katlar,
                                 son_odalar=son_odalar,
                                 admin_count=admin_count,
                                 depo_count=depo_count,
                                 kat_count=kat_count,
                                 kat_labels=kat_labels,
                                 kat_oda_sayilari=kat_oda_sayilari,
                                 toplam_urun_grup=toplam_urun_grup,
                                 toplam_urun=toplam_urun,
                                 kritik_urunler=kritik_urunler,
                                 stok_durumlari=stok_durumlari,
                                 son_personeller=son_personeller,
                                 son_urunler=son_urunler,
                                 urun_labels=urun_labels,
                                 urun_tuketim_miktarlari=urun_tuketim_miktarlari,
                                 ml_enabled=ml_enabled,
                                 ml_alerts=ml_alerts,
                                 ml_alert_count=ml_alert_count,
                                 dashboard_bildirimleri=dashboard_bildirimleri,
                                 bildirim_sayilari=bildirim_sayilari,
                                 istatistikler=istatistikler,
                                 yukleme_gorev_ozeti=yukleme_gorev_ozeti,
                                 kat_sorumlusu_gorev_ozeti=kat_sorumlusu_gorev_ozeti,
                                 eksik_doluluk_yuklemeleri=eksik_doluluk_yuklemeleri,
                                 ana_depo_tedarik_sayisi=ana_depo_tedarik_sayisi)
        except Exception as e:
            print(f"Sistem yöneticisi dashboard hatası: {str(e)}")
            import traceback
            traceback.print_exc()
            flash(f'Dashboard yüklenirken hata oluştu: {str(e)}', 'danger')
            # Minimal dashboard göster
            return render_template('sistem_yoneticisi/dashboard.html',
                                 otel_doluluk_raporlari=[],
                                 toplam_kat=0,
                                 toplam_oda=0,
                                 toplam_kullanici=0,
                                 toplam_personel=0,
                                 son_katlar=[],
                                 son_odalar=[],
                                 admin_count=0,
                                 depo_count=0,
                                 kat_count=0,
                                 kat_labels=[],
                                 kat_oda_sayilari=[],
                                 toplam_urun_grup=0,
                                 toplam_urun=0,
                                 kritik_urunler=[],
                                 stok_durumlari={'kritik': [], 'dikkat': [], 'normal': [], 'istatistik': {'toplam': 0, 'kritik_sayi': 0, 'dikkat_sayi': 0, 'normal_sayi': 0}},
                                 son_personeller=[],
                                 son_urunler=[],
                                 urun_labels=[],
                                 urun_tuketim_miktarlari=[],
                                 ml_enabled=False,
                                 ml_alerts=[],
                                 ml_alert_count=0,
                                 dashboard_bildirimleri=[],
                                 bildirim_sayilari={'kritik_stok': 0, 'geciken_siparis': 0, 'onay_bekleyen': 0, 'toplam': 0},
                                 istatistikler={'onaylandi': 0},
                                 yukleme_gorev_ozeti=None,
                                 kat_sorumlusu_gorev_ozeti=None,
                                 eksik_doluluk_yuklemeleri=[])

    
    @app.route('/depo')
    @login_required
    @role_required('depo_sorumlusu')
    def depo_dashboard():
        """Depo sorumlusu dashboard"""
        from utils.occupancy_service import OccupancyService
        from utils.authorization import get_kullanici_otelleri
        
        # Depo sorumlusunun atandığı otelleri al
        atanan_oteller = get_kullanici_otelleri()
        
        # Her otel için bugünkü doluluk bilgilerini al
        otel_doluluk_bilgileri = []
        bugun = get_kktc_now().date()
        
        for otel in atanan_oteller:
            doluluk_raporu = OccupancyService.get_gunluk_doluluk_raporu(bugun, otel.id)
            doluluk_orani = round((doluluk_raporu['dolu_oda'] / doluluk_raporu['toplam_oda'] * 100) if doluluk_raporu['toplam_oda'] > 0 else 0, 1)
            
            # Logo'yu kullan (boyut kontrolü kaldırıldı - template'de handle edilecek)
            logo = otel.logo
            
            otel_doluluk_bilgileri.append({
                'otel_id': otel.id,
                'otel_ad': otel.ad,
                'otel_logo': logo,
                'toplam_oda': doluluk_raporu['toplam_oda'],
                'dolu_oda': doluluk_raporu['dolu_oda'],
                'bos_oda': doluluk_raporu['bos_oda'],
                'doluluk_orani': doluluk_orani
            })
        
        # İstatistikler
        toplam_urun = Urun.query.filter_by(aktif=True).count()
        kritik_urunler = get_kritik_stok_urunler()
        aktif_zimmetler = PersonelZimmet.query.filter_by(durum='aktif').count()
        
        # Gelişmiş stok durumları
        try:
            stok_durumlari = get_tum_urunler_stok_durumlari()
        except Exception as e:
            print(f"Stok durumları hatası: {str(e)}")
            flash(f'Stok durumları yüklenirken hata oluştu: {str(e)}', 'warning')
            stok_durumlari = {
                'kritik': [],
                'dikkat': [],
                'normal': [],
                'istatistik': {
                    'toplam': 0,
                    'kritik_sayi': 0,
                    'dikkat_sayi': 0,
                    'normal_sayi': 0
                }
            }
        
        # Zimmet iade istatistikleri
        toplam_iade_edilen = db.session.query(db.func.sum(PersonelZimmetDetay.iade_edilen_miktar)).filter(
            PersonelZimmetDetay.iade_edilen_miktar > 0
        ).scalar() or 0
        
        # Bu ay yapılan iade işlemleri
        ay_basi = get_kktc_now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        bu_ay_iadeler = StokHareket.query.filter(
            StokHareket.hareket_tipi == 'giris',
            StokHareket.aciklama.like('%Zimmet iadesi%'),
            StokHareket.islem_tarihi >= ay_basi
        ).count()
        
        # İptal edilen zimmetler
        iptal_zimmetler = PersonelZimmet.query.filter_by(durum='iptal').count()
        
        # Son stok hareketleri
        son_hareketler = StokHareket.query.order_by(StokHareket.islem_tarihi.desc()).limit(10).all()
        
        # Grafik verileri — Ürün grup bazlı stok durumu — TEK SORGU ile (N×2 → 1)
        grup_stok_rows = db.session.query(
            UrunGrup.grup_adi,
            db.func.sum(
                db.case(
                    (StokHareket.hareket_tipi == 'giris', StokHareket.miktar),
                    else_=0
                )
            ).label('giris'),
            db.func.sum(
                db.case(
                    (StokHareket.hareket_tipi == 'cikis', StokHareket.miktar),
                    else_=0
                )
            ).label('cikis')
        ).join(Urun, Urun.grup_id == UrunGrup.id
        ).join(StokHareket, StokHareket.urun_id == Urun.id
        ).filter(
            UrunGrup.aktif == True,
            Urun.aktif == True
        ).group_by(UrunGrup.id, UrunGrup.grup_adi).all()
        
        grup_labels = []
        grup_stok_miktarlari = []
        for row in grup_stok_rows:
            stok = int((row.giris or 0) - (row.cikis or 0))
            if stok > 0:
                grup_labels.append(row.grup_adi)
                grup_stok_miktarlari.append(stok)
        
        # Son 7 günün stok hareket istatistikleri — TEK SORGU ile (14 → 1)
        bugun = get_kktc_now().date()
        yedi_gun_once = bugun - timedelta(days=6)
        
        stok_7gun = db.session.query(
            db.func.date(StokHareket.islem_tarihi).label('gun'),
            StokHareket.hareket_tipi,
            db.func.coalesce(db.func.sum(StokHareket.miktar), 0).label('toplam')
        ).filter(
            db.func.date(StokHareket.islem_tarihi) >= yedi_gun_once
        ).group_by(
            db.func.date(StokHareket.islem_tarihi),
            StokHareket.hareket_tipi
        ).all()
        
        stok_by_day = {}
        for row in stok_7gun:
            stok_by_day.setdefault(row.gun, {})[row.hareket_tipi] = float(row.toplam)
        
        gun_labels = []
        giris_verileri = []
        cikis_verileri = []
        for i in range(6, -1, -1):
            tarih = bugun - timedelta(days=i)
            gun_labels.append(tarih.strftime('%d.%m'))
            day_data = stok_by_day.get(tarih, {})
            giris_verileri.append(day_data.get('giris', 0))
            cikis_verileri.append(day_data.get('cikis', 0))
        
        # Dashboard bildirimleri (Satın Alma Modülü)
        dashboard_bildirimleri = []
        bildirim_sayilari = {
            'kritik_stok': 0,
            'geciken_siparis': 0,
            'onay_bekleyen': 0,
            'toplam': 0
        }
        if DashboardBildirimServisi and atanan_oteller:
            try:
                kullanici_id = session.get('kullanici_id')
                # İlk atanan otel için bildirimleri al
                otel_id = atanan_oteller[0].id
                dashboard_bildirimleri = DashboardBildirimServisi.get_dashboard_bildirimleri(
                    kullanici_id, 'depo_sorumlusu', otel_id
                )
                bildirim_sayilari = DashboardBildirimServisi.get_bildirim_sayilari(
                    kullanici_id, 'depo_sorumlusu', otel_id
                )
            except Exception as e:
                print(f"Dashboard bildirimleri hatası: {str(e)}")
        
        # Ürün bazlı tüketim verileri (Son 30 günün en çok tüketilen ürünleri)
        otuz_gun_once = bugun - timedelta(days=30)
        
        # Minibar işlemlerinden en çok tüketilen ürünleri al
        urun_tuketim = db.session.query(
            Urun.urun_adi,
            db.func.sum(MinibarIslemDetay.tuketim).label('toplam_tuketim')
        ).join(
            MinibarIslemDetay, MinibarIslemDetay.urun_id == Urun.id
        ).join(
            MinibarIslem, MinibarIslem.id == MinibarIslemDetay.islem_id
        ).filter(
            db.func.date(MinibarIslem.islem_tarihi) >= otuz_gun_once,
            MinibarIslemDetay.tuketim > 0
        ).group_by(
            Urun.id, Urun.urun_adi
        ).order_by(
            db.desc('toplam_tuketim')
        ).limit(10).all()
        
        urun_labels = [u[0] for u in urun_tuketim]
        urun_tuketim_miktarlari = [float(u[1] or 0) for u in urun_tuketim]
        
        # Yükleme görev özeti
        yukleme_gorev_ozeti = {'toplam': 3, 'tamamlanan': 0, 'bekleyen': 3, 'tamamlanma_orani': 0, 'inhouse_durum': 'pending', 'arrivals_durum': 'pending', 'departures_durum': 'pending'}
        try:
            from models import YuklemeGorev
            from datetime import date
            from utils.authorization import get_kullanici_otelleri
            
            bugun = date.today()
            kullanici_id = session.get('kullanici_id')
            
            # Kullanıcının otellerini al
            kullanici_otelleri = get_kullanici_otelleri()
            otel_ids = [o.id for o in kullanici_otelleri] if kullanici_otelleri else []
            
            if otel_ids:
                # İlk otel için görevleri kontrol et (veya seçili otel)
                secili_otel_id = otel_ids[0]
                
                # Bugünkü yükleme görevlerini otel_id ile al
                inhouse_gorev = YuklemeGorev.query.filter(
                    YuklemeGorev.otel_id == secili_otel_id,
                    YuklemeGorev.gorev_tarihi == bugun,
                    YuklemeGorev.dosya_tipi == 'inhouse'
                ).first()
                
                arrivals_gorev = YuklemeGorev.query.filter(
                    YuklemeGorev.otel_id == secili_otel_id,
                    YuklemeGorev.gorev_tarihi == bugun,
                    YuklemeGorev.dosya_tipi == 'arrivals'
                ).first()
                
                departures_gorev = YuklemeGorev.query.filter(
                    YuklemeGorev.otel_id == secili_otel_id,
                    YuklemeGorev.gorev_tarihi == bugun,
                    YuklemeGorev.dosya_tipi == 'departures'
                ).first()
                
                tamamlanan = 0
                inhouse_durum = 'pending'
                arrivals_durum = 'pending'
                departures_durum = 'pending'
                
                if inhouse_gorev and inhouse_gorev.durum == 'completed':
                    tamamlanan += 1
                    inhouse_durum = 'completed'
                
                if arrivals_gorev and arrivals_gorev.durum == 'completed':
                    tamamlanan += 1
                    arrivals_durum = 'completed'
                
                if departures_gorev and departures_gorev.durum == 'completed':
                    tamamlanan += 1
                    departures_durum = 'completed'
                
                yukleme_gorev_ozeti = {
                    'toplam': 3,
                    'tamamlanan': tamamlanan,
                    'bekleyen': 3 - tamamlanan,
                    'tamamlanma_orani': int((tamamlanan / 3) * 100),
                    'inhouse_durum': inhouse_durum,
                    'arrivals_durum': arrivals_durum,
                    'departures_durum': departures_durum
                }
        except Exception as e:
            print(f"Yükleme görev özeti hatası: {str(e)}")
        
        # Oda kontrol görevleri özeti (tüm yüklemeler tamamlandığında göster) - OTEL BAZLI
        oda_kontrol_ozeti = None
        oda_kontrol_ozeti_otel_bazli = []  # Her otel için ayrı özet
        try:
            if yukleme_gorev_ozeti and yukleme_gorev_ozeti.get('tamamlanma_orani') == 100:
                from models import GunlukGorev, GorevDetay, Otel
                from utils.authorization import get_kullanici_otelleri
                
                kullanici_otelleri = get_kullanici_otelleri()
                otel_ids = [o.id for o in kullanici_otelleri] if kullanici_otelleri else []
                
                if otel_ids:
                    # Her otel için ayrı özet oluştur
                    for otel in kullanici_otelleri:
                        # Bu otelin bugünkü görevlerini al
                        otel_gorevler = GunlukGorev.query.filter(
                            GunlukGorev.otel_id == otel.id,
                            db.func.date(GunlukGorev.gorev_tarihi) == bugun
                        ).all()
                        
                        gorev_ids = [g.id for g in otel_gorevler]
                        
                        if gorev_ids:
                            detaylar = GorevDetay.query.filter(
                                GorevDetay.gorev_id.in_(gorev_ids)
                            ).all()
                            
                            toplam = len(detaylar)
                            tamamlanan_oda = len([d for d in detaylar if d.durum == 'completed'])
                            bekleyen_oda = len([d for d in detaylar if d.durum == 'pending'])
                            devam_eden = len([d for d in detaylar if d.durum == 'in_progress'])
                            
                            # Görev tiplerine göre oda sayılarını hesapla
                            gorev_tip_map = {g.id: g.gorev_tipi for g in otel_gorevler}
                            
                            inhouse_detaylar = [d for d in detaylar if gorev_tip_map.get(d.gorev_id) == 'inhouse_kontrol']
                            arrival_detaylar = [d for d in detaylar if gorev_tip_map.get(d.gorev_id) == 'arrival_kontrol']
                            departure_detaylar = [d for d in detaylar if gorev_tip_map.get(d.gorev_id) == 'departure_kontrol']
                            
                            inhouse = len(inhouse_detaylar)
                            arrival = len(arrival_detaylar)
                            departure = len(departure_detaylar)
                            
                            inhouse_tamamlanan = len([d for d in inhouse_detaylar if d.durum == 'completed'])
                            arrival_tamamlanan = len([d for d in arrival_detaylar if d.durum == 'completed'])
                            departure_tamamlanan = len([d for d in departure_detaylar if d.durum == 'completed'])
                            
                            otel_ozet = {
                                'otel_id': otel.id,
                                'otel_ad': otel.ad,
                                'otel_logo': otel.logo if hasattr(otel, 'logo') else None,
                                'toplam': toplam,
                                'tamamlanan': tamamlanan_oda,
                                'bekleyen': bekleyen_oda,
                                'devam_eden': devam_eden,
                                'tamamlanma_orani': int((tamamlanan_oda / toplam) * 100) if toplam > 0 else 0,
                                'inhouse': {'toplam': inhouse, 'tamamlanan': inhouse_tamamlanan},
                                'arrival': {'toplam': arrival, 'tamamlanan': arrival_tamamlanan},
                                'departure': {'toplam': departure, 'tamamlanan': departure_tamamlanan}
                            }
                            oda_kontrol_ozeti_otel_bazli.append(otel_ozet)
                    
                    # Toplam özet (eski format - geriye uyumluluk için)
                    if oda_kontrol_ozeti_otel_bazli:
                        toplam_tum = sum(o['toplam'] for o in oda_kontrol_ozeti_otel_bazli)
                        tamamlanan_tum = sum(o['tamamlanan'] for o in oda_kontrol_ozeti_otel_bazli)
                        bekleyen_tum = sum(o['bekleyen'] for o in oda_kontrol_ozeti_otel_bazli)
                        devam_eden_tum = sum(o['devam_eden'] for o in oda_kontrol_ozeti_otel_bazli)
                        
                        inhouse_toplam = sum(o['inhouse']['toplam'] for o in oda_kontrol_ozeti_otel_bazli)
                        inhouse_tamamlanan = sum(o['inhouse']['tamamlanan'] for o in oda_kontrol_ozeti_otel_bazli)
                        arrival_toplam = sum(o['arrival']['toplam'] for o in oda_kontrol_ozeti_otel_bazli)
                        arrival_tamamlanan = sum(o['arrival']['tamamlanan'] for o in oda_kontrol_ozeti_otel_bazli)
                        departure_toplam = sum(o['departure']['toplam'] for o in oda_kontrol_ozeti_otel_bazli)
                        departure_tamamlanan = sum(o['departure']['tamamlanan'] for o in oda_kontrol_ozeti_otel_bazli)
                        
                        oda_kontrol_ozeti = {
                            'toplam': toplam_tum,
                            'tamamlanan': tamamlanan_tum,
                            'bekleyen': bekleyen_tum,
                            'devam_eden': devam_eden_tum,
                            'tamamlanma_orani': int((tamamlanan_tum / toplam_tum) * 100) if toplam_tum > 0 else 0,
                            'inhouse': {'toplam': inhouse_toplam, 'tamamlanan': inhouse_tamamlanan},
                            'arrival': {'toplam': arrival_toplam, 'tamamlanan': arrival_tamamlanan},
                            'departure': {'toplam': departure_toplam, 'tamamlanan': departure_tamamlanan}
                        }
        except Exception as e:
            print(f"Oda kontrol özeti hatası: {str(e)}")
        
        return render_template('depo_sorumlusu/dashboard.html',
                             otel_doluluk_bilgileri=otel_doluluk_bilgileri,
                             toplam_urun=toplam_urun,
                             kritik_urunler=kritik_urunler,
                             stok_durumlari=stok_durumlari,
                             aktif_zimmetler=aktif_zimmetler,
                             toplam_iade_edilen=toplam_iade_edilen,
                             bu_ay_iadeler=bu_ay_iadeler,
                             iptal_zimmetler=iptal_zimmetler,
                             son_hareketler=son_hareketler,
                             grup_labels=grup_labels,
                             grup_stok_miktarlari=grup_stok_miktarlari,
                             gun_labels=gun_labels,
                             giris_verileri=giris_verileri,
                             cikis_verileri=cikis_verileri,
                             urun_labels=urun_labels,
                             urun_tuketim_miktarlari=urun_tuketim_miktarlari,
                             dashboard_bildirimleri=dashboard_bildirimleri,
                             bildirim_sayilari=bildirim_sayilari,
                             yukleme_gorev_ozeti=yukleme_gorev_ozeti,
                             oda_kontrol_ozeti=oda_kontrol_ozeti,
                             oda_kontrol_ozeti_otel_bazli=oda_kontrol_ozeti_otel_bazli)

    
    @app.route('/kat-sorumlusu')
    @app.route('/kat-sorumlusu/dashboard')
    @login_required
    @role_required('kat_sorumlusu')
    def kat_sorumlusu_dashboard():
        """Kat sorumlusu dashboard"""
        from utils.occupancy_service import OccupancyService
        from utils.authorization import get_kullanici_otelleri
        from datetime import date
        
        kullanici_id = session['kullanici_id']
        
        # Kullanıcının otelini al
        kullanici_otelleri = get_kullanici_otelleri()
        otel_id = kullanici_otelleri[0].id if kullanici_otelleri else None
        
        # Görev özeti - Bugün için (OTEL BAZLI)
        gorev_ozeti = None
        try:
            from utils.gorev_service import GorevService
            if otel_id:
                gorev_ozeti = GorevService.get_task_summary(otel_id, date.today())
            else:
                gorev_ozeti = {'toplam': 0, 'tamamlanan': 0, 'bekleyen': 0, 'dnd': 0, 'tamamlanma_orani': 0}
        except Exception as e:
            print(f"Görev özeti hatası: {str(e)}")
            gorev_ozeti = {'toplam': 0, 'tamamlanan': 0, 'bekleyen': 0, 'dnd': 0, 'tamamlanma_orani': 0}
        
        # Doluluk raporu - Bugün için
        doluluk_raporu = None
        try:
            if otel_id:
                doluluk_raporu = OccupancyService.get_gunluk_doluluk_raporu(date.today(), otel_id)
                # Doluluk oranı hesapla
                if doluluk_raporu['toplam_oda'] > 0:
                    doluluk_raporu['doluluk_orani'] = round((doluluk_raporu['dolu_oda'] / doluluk_raporu['toplam_oda']) * 100)
                else:
                    doluluk_raporu['doluluk_orani'] = 0
        except Exception as e:
            print(f"Doluluk raporu hatası: {str(e)}")
        
        # İstatistikler
        aktif_zimmetler = PersonelZimmet.query.filter_by(
            personel_id=kullanici_id, 
            durum='aktif'
        ).count()
        
        # Zimmetindeki toplam ürün miktarı
        zimmet_detaylari = db.session.query(
            db.func.sum(PersonelZimmetDetay.kalan_miktar)
        ).join(PersonelZimmet).filter(
            PersonelZimmet.personel_id == kullanici_id,
            PersonelZimmet.durum == 'aktif'
        ).scalar() or 0
        
        # Kritik stok bilgileri
        kritik_stoklar = get_kat_sorumlusu_kritik_stoklar(kullanici_id)
        kritik_stok_sayisi = kritik_stoklar['istatistik']['kritik_sayisi']
        stokout_sayisi = kritik_stoklar['istatistik']['stokout_sayisi']
        
        # Bugünkü kullanım (son 24 saat)
        bugun_baslangic = get_kktc_now() - timedelta(days=1)
        bugunun_kullanimi = db.session.query(
            db.func.sum(MinibarIslemDetay.eklenen_miktar)
        ).join(MinibarIslem).filter(
            MinibarIslem.personel_id == kullanici_id,
            MinibarIslem.islem_tarihi >= bugun_baslangic
        ).scalar() or 0
        
        # Son minibar işlemleri
        son_islemler = MinibarIslem.query.filter_by(
            personel_id=kullanici_id
        ).order_by(MinibarIslem.islem_tarihi.desc()).limit(10).all()
        
        # Grafik verileri - En çok kullanılan 5 ürün (son 7 gün)
        yedi_gun_once = get_kktc_now() - timedelta(days=7)
        en_cok_kullanilan = db.session.query(
            Urun.urun_adi,
            db.func.sum(MinibarIslemDetay.eklenen_miktar).label('toplam')
        ).join(MinibarIslemDetay, MinibarIslemDetay.urun_id == Urun.id).join(
            MinibarIslem, MinibarIslem.id == MinibarIslemDetay.islem_id
        ).filter(
            MinibarIslem.personel_id == kullanici_id,
            MinibarIslem.islem_tarihi >= yedi_gun_once
        ).group_by(Urun.id, Urun.urun_adi).order_by(
            db.desc('toplam')
        ).limit(5).all()
        
        en_cok_urun_labels = [u[0] for u in en_cok_kullanilan]
        en_cok_urun_miktarlar = [float(u[1] or 0) for u in en_cok_kullanilan]
        
        # Zimmet kullanım durumu (ürün bazlı)
        zimmet_urunler = db.session.query(
            Urun.urun_adi,
            db.func.sum(PersonelZimmetDetay.miktar).label('teslim_edilen'),
            db.func.sum(PersonelZimmetDetay.kullanilan_miktar).label('kullanilan'),
            db.func.sum(PersonelZimmetDetay.kalan_miktar).label('kalan')
        ).join(Urun, PersonelZimmetDetay.urun_id == Urun.id).join(
            PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id
        ).filter(
            PersonelZimmet.personel_id == kullanici_id,
            PersonelZimmet.durum == 'aktif'
        ).group_by(Urun.id, Urun.urun_adi).all()
        
        zimmet_urun_labels = [u[0] for u in zimmet_urunler]
        zimmet_kullanilan = [float(u[2] or 0) for u in zimmet_urunler]
        zimmet_kalan = [float(u[3] or 0) for u in zimmet_urunler]
        
        # Günlük tüketim trendi (son 7 gün) — TEK SORGU ile
        yedi_gun_baslangic = (get_kktc_now() - timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)
        gunluk_tuketim_rows = db.session.query(
            db.func.date(MinibarIslem.islem_tarihi).label('gun'),
            db.func.coalesce(db.func.sum(MinibarIslemDetay.eklenen_miktar), 0).label('toplam')
        ).join(MinibarIslem).filter(
            MinibarIslem.personel_id == kullanici_id,
            MinibarIslem.islem_tarihi >= yedi_gun_baslangic
        ).group_by(db.func.date(MinibarIslem.islem_tarihi)).all()
        
        tuketim_by_date = {row.gun: float(row.toplam) for row in gunluk_tuketim_rows}
        gunluk_tuketim = []
        gunluk_labels = []
        for i in range(6, -1, -1):
            gun = (get_kktc_now() - timedelta(days=i)).date()
            gunluk_labels.append(gun.strftime('%d.%m'))
            gunluk_tuketim.append(tuketim_by_date.get(gun, 0))
        
        # Minibar işlem tipi dağılımı — TEK SORGU ile
        islem_tipi_counts = dict(
            db.session.query(
                MinibarIslem.islem_tipi,
                db.func.count(MinibarIslem.id)
            ).filter(
                MinibarIslem.personel_id == kullanici_id,
                MinibarIslem.islem_tipi.in_(['ilk_dolum', 'yeniden_dolum', 'eksik_tamamlama'])
            ).group_by(MinibarIslem.islem_tipi).all()
        )
        islem_ilk_dolum = islem_tipi_counts.get('ilk_dolum', 0)
        islem_yeniden_dolum = islem_tipi_counts.get('yeniden_dolum', 0)
        islem_eksik_tamamlama = islem_tipi_counts.get('eksik_tamamlama', 0)
        
        return render_template('kat_sorumlusu/dashboard.html',
                             gorev_ozeti=gorev_ozeti,
                             doluluk_raporu=doluluk_raporu,
                             aktif_zimmetler=aktif_zimmetler,
                             zimmet_toplam=zimmet_detaylari,
                             kritik_stok_sayisi=kritik_stok_sayisi,
                             stokout_sayisi=stokout_sayisi,
                             bugunun_kullanimi=int(bugunun_kullanimi),
                             son_islemler=son_islemler,
                             en_cok_urun_labels=en_cok_urun_labels,
                             en_cok_urun_miktarlar=en_cok_urun_miktarlar,
                             zimmet_urun_labels=zimmet_urun_labels,
                             zimmet_kullanilan=zimmet_kullanilan,
                             zimmet_kalan=zimmet_kalan,
                             gunluk_tuketim=gunluk_tuketim,
                             gunluk_labels=gunluk_labels,
                             islem_ilk_dolum=islem_ilk_dolum,
                             islem_yeniden_dolum=islem_yeniden_dolum,
                             islem_eksik_tamamlama=islem_eksik_tamamlama)

    # ========================================================================
    # API: DASHBOARD WIDGET'LARI
    # ========================================================================

    @app.route('/api/son-aktiviteler')
    @login_required
    @role_required(['sistem_yoneticisi', 'admin'])
    def api_son_aktiviteler():
        """Son kullanıcı aktivitelerini döndür"""
        try:
            limit = request.args.get('limit', 10, type=int)

            aktiviteler = SistemLog.query\
                .join(Kullanici, SistemLog.kullanici_id == Kullanici.id)\
                .filter(
                    SistemLog.islem_tipi.in_(['ekleme', 'guncelleme', 'silme']),
                    Kullanici.rol != 'superadmin'
                )\
                .order_by(SistemLog.islem_tarihi.desc())\
                .limit(limit)\
                .all()

            data = []
            for log in aktiviteler:
                try:
                    kullanici_adi = 'Sistem'
                    if log.kullanici:
                        kullanici_adi = f"{log.kullanici.ad} {log.kullanici.soyad}"

                    detay = {}
                    if log.islem_detay:
                        try:
                            detay = json.loads(log.islem_detay) if isinstance(log.islem_detay, str) else log.islem_detay
                        except Exception:
                            detay = {'aciklama': str(log.islem_detay)}

                    if isinstance(log.islem_tarihi, datetime):
                        if log.islem_tarihi.tzinfo is None:
                            islem_tarihi = log.islem_tarihi.replace(tzinfo=timezone.utc)
                        else:
                            islem_tarihi = log.islem_tarihi
                    else:
                        islem_tarihi = datetime.combine(log.islem_tarihi, datetime.min.time()).replace(tzinfo=timezone.utc)
                    
                    zaman_farki = get_kktc_now() - islem_tarihi

                    if zaman_farki < timedelta(minutes=1):
                        zaman_str = "Az önce"
                    elif zaman_farki < timedelta(hours=1):
                        dakika = int(zaman_farki.total_seconds() / 60)
                        zaman_str = f"{dakika} dakika önce"
                    elif zaman_farki < timedelta(days=1):
                        saat = int(zaman_farki.total_seconds() / 3600)
                        zaman_str = f"{saat} saat önce"
                    else:
                        gun = zaman_farki.days
                        zaman_str = f"{gun} gün önce"

                    data.append({
                        'id': log.id,
                        'kullanici': kullanici_adi,
                        'islem_tipi': log.islem_tipi,
                        'modul': log.modul,
                        'detay': detay,
                        'zaman': zaman_str,
                        'tam_tarih': islem_tarihi.strftime('%d.%m.%Y %H:%M')
                    })
                except Exception as log_error:
                    print(f"Log parse hatası (ID: {log.id}): {log_error}")
                    continue

            return jsonify({'success': True, 'aktiviteler': data})

        except Exception as e:
            print(f"Son aktiviteler hatası: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'error': str(e)}), 500


    @app.route('/api/bekleyen-dolum-sayisi')
    @login_required
    @role_required(['sistem_yoneticisi', 'admin', 'depo_sorumlusu', 'kat_sorumlusu'])
    def api_bekleyen_dolum_sayisi():
        """Bekleyen dolum talepleri sayısını döndür"""
        try:
            count = MinibarDolumTalebi.query.filter_by(durum='beklemede').count()
            
            return jsonify({
                'success': True,
                'count': count
            })
        except Exception as e:
            logger.error(f"Bekleyen dolum sayısı hatası: {e}")
            return jsonify({
                'success': False,
                'count': 0,
                'error': str(e)
            }), 500


    @app.route('/api/tuketim-trendleri')
    @login_required
    @role_required(['sistem_yoneticisi', 'admin', 'depo_sorumlusu'])
    def api_tuketim_trendleri():
        """Günlük/haftalık tüketim trendlerini döndür"""
        try:
            from sqlalchemy import func

            gun_sayisi = request.args.get('gun', 7, type=int)
            baslangic = get_kktc_now() - timedelta(days=gun_sayisi)

            gunluk_tuketim = db.session.query(
                func.date(MinibarIslem.islem_tarihi).label('tarih'),
                func.sum(MinibarIslemDetay.tuketim).label('toplam_tuketim'),
                func.count(MinibarIslemDetay.id).label('islem_sayisi')
            ).join(MinibarIslemDetay, MinibarIslemDetay.islem_id == MinibarIslem.id)\
             .filter(MinibarIslem.islem_tarihi >= baslangic)\
             .group_by(func.date(MinibarIslem.islem_tarihi))\
             .order_by(func.date(MinibarIslem.islem_tarihi))\
             .all()

            tum_gunler = {}
            for i in range(gun_sayisi):
                tarih = (get_kktc_now() - timedelta(days=gun_sayisi-i-1)).date()
                tum_gunler[str(tarih)] = {'tuketim': 0, 'islem_sayisi': 0}

            for row in gunluk_tuketim:
                tarih_str = str(row.tarih)
                tum_gunler[tarih_str] = {
                    'tuketim': int(row.toplam_tuketim or 0),
                    'islem_sayisi': int(row.islem_sayisi or 0)
                }

            labels = []
            tuketim_data = []
            islem_data = []

            for tarih_str in sorted(tum_gunler.keys()):
                tarih_obj = datetime.strptime(tarih_str, '%Y-%m-%d')
                labels.append(tarih_obj.strftime('%d/%m'))
                tuketim_data.append(tum_gunler[tarih_str]['tuketim'])
                islem_data.append(tum_gunler[tarih_str]['islem_sayisi'])

            return jsonify({
                'success': True,
                'labels': labels,
                'datasets': [
                    {
                        'label': 'Toplam Tüketim',
                        'data': tuketim_data,
                        'borderColor': 'rgb(239, 68, 68)',
                        'backgroundColor': 'rgba(239, 68, 68, 0.1)',
                        'tension': 0.3
                    },
                    {
                        'label': 'İşlem Sayısı',
                        'data': islem_data,
                        'borderColor': 'rgb(59, 130, 246)',
                        'backgroundColor': 'rgba(59, 130, 246, 0.1)',
                        'tension': 0.3
                    }
                ]
            })

        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

    # ========================================================================
    # AUDIT TRAIL ROUTE'LARI
    # ========================================================================

    @app.route('/sistem-yoneticisi/audit-trail')
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def audit_trail():
        """Audit Trail - Denetim İzi Sayfası"""
        page = request.args.get('page', 1, type=int)
        per_page = 50
        
        kullanici_id = request.args.get('kullanici_id', type=int)
        islem_tipi = request.args.get('islem_tipi')
        tablo_adi = request.args.get('tablo_adi')
        tarih_baslangic = request.args.get('tarih_baslangic')
        tarih_bitis = request.args.get('tarih_bitis')
        
        query = AuditLog.query.filter(AuditLog.kullanici_rol != 'superadmin')
        
        if kullanici_id:
            query = query.filter_by(kullanici_id=kullanici_id)
        if islem_tipi:
            query = query.filter_by(islem_tipi=islem_tipi)
        if tablo_adi:
            query = query.filter_by(tablo_adi=tablo_adi)
        if tarih_baslangic:
            tarih_baslangic_dt = datetime.strptime(tarih_baslangic, '%Y-%m-%d')
            query = query.filter(AuditLog.islem_tarihi >= tarih_baslangic_dt)
        if tarih_bitis:
            tarih_bitis_dt = datetime.strptime(tarih_bitis, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(AuditLog.islem_tarihi < tarih_bitis_dt)
        
        query = query.order_by(AuditLog.islem_tarihi.desc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        bugun = get_kktc_now().replace(hour=0, minute=0, second=0, microsecond=0)
        bu_hafta = bugun - timedelta(days=bugun.weekday())
        bu_ay = bugun.replace(day=1)
        
        stats = {
            'today': AuditLog.query.filter(AuditLog.islem_tarihi >= bugun).count(),
            'week': AuditLog.query.filter(AuditLog.islem_tarihi >= bu_hafta).count(),
            'month': AuditLog.query.filter(AuditLog.islem_tarihi >= bu_ay).count()
        }
        
        users = Kullanici.query.filter_by(aktif=True).order_by(Kullanici.kullanici_adi).all()
        
        tables = db.session.query(AuditLog.tablo_adi).distinct().order_by(AuditLog.tablo_adi).all()
        tables = [t[0] for t in tables]
        
        return render_template('sistem_yoneticisi/audit_trail.html',
                             logs=pagination.items,
                             pagination=pagination,
                             users=users,
                             tables=tables,
                             stats=stats)


    @app.route('/sistem-yoneticisi/audit-trail/<int:log_id>')
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def audit_trail_detail(log_id):
        """Audit Log Detay API"""
        log = AuditLog.query.get_or_404(log_id)
        
        return jsonify({
            'id': log.id,
            'kullanici_id': log.kullanici_id,
            'kullanici_adi': log.kullanici_adi,
            'kullanici_rol': log.kullanici_rol,
            'islem_tipi': log.islem_tipi,
            'tablo_adi': log.tablo_adi,
            'kayit_id': log.kayit_id,
            'eski_deger': log.eski_deger,
            'yeni_deger': log.yeni_deger,
            'degisiklik_ozeti': log.degisiklik_ozeti,
            'http_method': log.http_method,
            'url': log.url,
            'endpoint': log.endpoint,
            'ip_adresi': log.ip_adresi,
            'user_agent': log.user_agent,
            'islem_tarihi': log.islem_tarihi.strftime('%d.%m.%Y %H:%M:%S'),
            'aciklama': log.aciklama,
            'basarili': log.basarili,
            'hata_mesaji': log.hata_mesaji
        })


    @app.route('/sistem-yoneticisi/audit-trail/export')
    @login_required
    @role_required('sistem_yoneticisi', 'admin')
    def audit_trail_export():
        """Audit Trail Excel Export"""
        kullanici_id = request.args.get('kullanici_id', type=int)
        islem_tipi = request.args.get('islem_tipi')
        tablo_adi = request.args.get('tablo_adi')
        tarih_baslangic = request.args.get('tarih_baslangic')
        tarih_bitis = request.args.get('tarih_bitis')
        
        query = AuditLog.query.filter(AuditLog.kullanici_rol != 'superadmin')
        
        if kullanici_id:
            query = query.filter_by(kullanici_id=kullanici_id)
        if islem_tipi:
            query = query.filter_by(islem_tipi=islem_tipi)
        if tablo_adi:
            query = query.filter_by(tablo_adi=tablo_adi)
        if tarih_baslangic:
            tarih_baslangic_dt = datetime.strptime(tarih_baslangic, '%Y-%m-%d')
            query = query.filter(AuditLog.islem_tarihi >= tarih_baslangic_dt)
        if tarih_bitis:
            tarih_bitis_dt = datetime.strptime(tarih_bitis, '%Y-%m-%d')
            query = query.filter(AuditLog.islem_tarihi <= tarih_bitis_dt)
        
        logs = query.order_by(AuditLog.islem_tarihi.desc()).limit(10000).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Trail"
        
        headers = ['ID', 'Tarih', 'Kullanıcı', 'Rol', 'İşlem', 'Tablo', 'Kayıt ID', 
                   'Değişiklik', 'IP', 'URL', 'Başarılı']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        for row, log in enumerate(logs, 2):
            ws.cell(row=row, column=1, value=log.id)
            ws.cell(row=row, column=2, value=log.islem_tarihi.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row, column=3, value=log.kullanici_adi)
            ws.cell(row=row, column=4, value=log.kullanici_rol)
            ws.cell(row=row, column=5, value=log.islem_tipi)
            ws.cell(row=row, column=6, value=log.tablo_adi)
            ws.cell(row=row, column=7, value=log.kayit_id)
            ws.cell(row=row, column=8, value=log.degisiklik_ozeti or '')
            ws.cell(row=row, column=9, value=log.ip_adresi or '')
            ws.cell(row=row, column=10, value=log.url or '')
            ws.cell(row=row, column=11, value='Evet' if log.basarili else 'Hayır')
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 50
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 10
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"audit_trail_{get_kktc_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        from utils.audit import audit_export
        audit_export('audit_logs', f'Excel export: {len(logs)} kayıt')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
