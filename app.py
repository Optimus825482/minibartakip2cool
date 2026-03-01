from flask import Flask, render_template, request, redirect, url_for, flash, session, make_response, jsonify, send_file
from flask_wtf.csrf import CSRFProtect, CSRFError
from datetime import datetime, timedelta, timezone
import pytz

# KKTC Timezone (Kıbrıs - Europe/Nicosia)
KKTC_TZ = pytz.timezone('Europe/Nicosia')

def get_kktc_now():
    """Kıbrıs saat diliminde şu anki zamanı döndürür."""
    return datetime.now(KKTC_TZ)
import sentry_sdk
from sentry_sdk.integrations.flask import FlaskIntegration
# SqlAlchemy integration artık otomatik yükleniyor (Sentry 2.0+)
import os
import io
import time
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from dotenv import load_dotenv
from sqlalchemy.exc import OperationalError, TimeoutError
from sqlalchemy import inspect

# Logging ayarla - Hem console hem de dosyaya yaz
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# .env dosyasını yükle
load_dotenv()

# Flask uygulaması oluştur
app = Flask(__name__)

# Konfigürasyonu yükle
app.config.from_object('config.Config')

# Proxy arkasında çalışırken (Traefik/Nginx) doğru scheme ve IP algılama
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# CSRF Koruması Aktif
csrf = CSRFProtect(app)

# CSRF token'ı tüm template'lere ekle
@app.context_processor
def inject_csrf_token():
    """CSRF token'ı template'lere enjekte et"""
    from flask_wtf.csrf import generate_csrf
    # Hem fonksiyon hem de değişken olarak sağla (geriye dönük uyumluluk için)
    return dict(csrf_token=generate_csrf)

# Rate Limiting Devre Dışı (İhtiyaç halinde açılabilir)
# limiter = Limiter(
#     app=app,
#     key_func=get_remote_address,
#     default_limits=["200 per day", "50 per hour"],
#     storage_uri="memory://",  # Production'da Redis kullanılmalı
#     strategy="fixed-window"
# )

# Veritabanı başlat
from models import db
from flask_migrate import Migrate

db.init_app(app)
migrate = Migrate(app, db)

# Veritabanı metadata'sını yenile ve bağlantıyı test et
with app.app_context():
    try:
        db.engine.dispose()
        db.reflect()
        # Bağlantıyı test et
        with db.engine.connect() as conn:
            result = conn.execute(db.text("SELECT 1"))
            result.close()
        logger.info("✅ Database engine yenilendi, metadata reflect edildi ve bağlantı test edildi")
    except Exception as e:
        logger.warning(f"⚠️ Database başlatma hatası (görmezden geliniyor): {e}")

# Cache devre dışı (Redis sadece Celery broker olarak kullanılıyor)
cache = None
logger.info("ℹ️ Cache devre dışı - Redis sadece Celery broker olarak kullanılıyor")

# ============================================
# RATE LIMITER INITIALIZATION
# ============================================
limiter = None
if app.config.get('RATE_LIMIT_ENABLED', True):
    try:
        from utils.rate_limiter import init_rate_limiter, limiter as rate_limiter
        limiter = init_rate_limiter(app)
        logger.info("✅ Rate Limiter aktifleştirildi")
    except Exception as e:
        logger.warning(f"⚠️ Rate Limiter başlatılamadı: {str(e)}")
else:
    logger.info("ℹ️ Rate Limiter devre dışı (config)")

# ============================================
# CACHE MANAGER INITIALIZATION (Master Data Only)
# ============================================
cache_manager = None
if app.config.get('CACHE_ENABLED', True):
    try:
        from utils.cache_manager import init_cache, cache_manager as cm
        cache_manager = init_cache(app)
        logger.info("✅ Cache Manager aktifleştirildi (sadece master data)")
    except Exception as e:
        logger.warning(f"⚠️ Cache Manager başlatılamadı: {str(e)}")

# Query Logging - SQLAlchemy Event Listener
try:
    from utils.monitoring.query_analyzer import setup_query_logging
    setup_query_logging()
except Exception as e:
    logger.warning(f"Query logging setup hatası: {e}")

# İkinci dispose kaldırıldı - yukarıda zaten yapılıyor (10.02.2026)
# Engine test yukarıdaki blokta yapılıyor

# Database Connection Retry Mekanizması - Railway Timeout Fix v3 (ULTRA AGRESIF)
def init_db_with_retry(max_retries=3, retry_delay=10):
    """
    Database bağlantısını retry mekanizması ile başlat
    Railway'de cold start veya network timeout sorunlarını çözer
    v3: Daha uzun timeout, daha az deneme
    """
    for attempt in range(max_retries):
        try:
            with app.app_context():
                # Database bağlantısını test et
                connection = db.engine.connect()
                connection.close()
                logger.info(f"✅ Database bağlantısı başarılı (Deneme {attempt + 1}/{max_retries})")
                return True
        except (OperationalError, TimeoutError) as e:
            error_msg = str(e)
            logger.warning(f"⚠️ Database bağlantı hatası (Deneme {attempt + 1}/{max_retries}): {error_msg[:200]}")
            
            if attempt < max_retries - 1:
                # Sabit 30 saniye bekleme (exponential backoff yerine)
                wait_time = 30
                logger.info(f"🔄 {wait_time} saniye sonra tekrar denenecek...")
                time.sleep(wait_time)
            else:
                logger.error(f"❌ Database bağlantısı {max_retries} denemeden sonra başarısız!")
                logger.error(f"❌ Son hata: {error_msg}")
                # Production'da uygulama çalışmaya devam etsin
                return False
        except Exception as e:
            logger.error(f"❌ Beklenmeyen hata: {str(e)}")
            # Beklenmeyen hatalarda da devam et
            return False
    return False

# Uygulama başlatıldığında database bağlantısını test et
try:
    init_db_with_retry()
except Exception as e:
    logger.error(f"❌ FATAL: Database başlatılamadı: {str(e)}")
    # Production'da uygulama çalışmaya devam etsin, ilk request'te tekrar denenecek

# Yardımcı modülleri import et
from utils.decorators import login_required, role_required, setup_required, setup_not_completed
from utils.helpers import (
    get_current_user, get_kritik_stok_urunler, get_stok_toplamlari,
    log_islem, get_son_loglar, get_kullanici_loglari, get_modul_loglari,
    log_hata, get_son_hatalar, get_cozulmemis_hatalar, hata_cozuldu_isaretle,
    get_stok_durumu, get_tum_urunler_stok_durumlari
)
from utils.audit import (
    audit_create, audit_update, audit_delete,
    audit_login, audit_logout, serialize_model
)

# Modelleri import et
from models import (
    Otel, Kullanici, Kat, Oda, UrunGrup, Urun, StokHareket, 
    PersonelZimmet, PersonelZimmetDetay, MinibarIslem, MinibarIslemDetay, 
    MinibarIslemTipi, SistemAyar, SistemLog, HataLog, OtomatikRapor,
    MinibarDolumTalebi, Setup, SetupIcerik
)

# Context processor - tüm template'lere kullanıcı bilgisini gönder
@app.context_processor
def inject_user():
    return dict(current_user=get_current_user())

# Context processor - Python built-in fonksiyonları
@app.context_processor
def inject_builtins():
    return dict(min=min, max=max)

# Context processor - Cache version
@app.context_processor
def inject_cache_version():
    """Cache busting için version numarası"""
    from config import Config
    return dict(cache_version=Config.CACHE_VERSION)

# Context processor - Datetime ve tarih fonksiyonları
@app.context_processor
def inject_datetime():
    """Şablonlara datetime ve tarih yardımcı fonksiyonlarını ekle"""
    gun_adlari = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
    return dict(
        now=datetime.now,
        gun_adlari=gun_adlari
    )

# Context processor - Otel bilgisi ve logo
@app.context_processor
def inject_otel_info():
    """Kullanıcının otel bilgisini ve logosunu template'lere gönder"""
    from models import Otel, Kullanici
    from utils.authorization import get_kat_sorumlusu_otel, get_depo_sorumlusu_oteller
    
    kullanici = get_current_user()
    otel_bilgi = None
    
    if kullanici:
        try:
            # Kat sorumlusu - tek otel
            if kullanici.rol == 'kat_sorumlusu':
                otel = get_kat_sorumlusu_otel(kullanici.id)
                if otel:
                    otel_bilgi = {
                        'ad': otel.ad,
                        'logo': otel.logo
                    }
            
            # Depo sorumlusu - ilk atandığı oteli göster
            elif kullanici.rol == 'depo_sorumlusu':
                oteller = get_depo_sorumlusu_oteller(kullanici.id)
                if oteller:
                    otel = oteller[0]  # İlk oteli göster
                    otel_bilgi = {
                        'ad': otel.ad,
                        'logo': otel.logo
                    }
            
            # Admin ve sistem yöneticisi - ilk oteli göster
            elif kullanici.rol in ['admin', 'sistem_yoneticisi']:
                otel = Otel.query.filter_by(aktif=True).first()
                if otel:
                    otel_bilgi = {
                        'ad': otel.ad,
                        'logo': otel.logo
                    }
        except Exception as e:
            logger.error(f"Otel bilgisi alınırken hata: {str(e)}")
    
    return dict(kullanici_otel=otel_bilgi)

# ============================================
# CACHE CONTROL - Template ve HTML cache'ini devre dışı bırak
# ============================================
@app.after_request
def add_no_cache_headers(response):
    """HTML response'larına no-cache header'ları ekle"""
    # Sadece HTML sayfaları için cache'i devre dışı bırak
    if response.content_type and 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# ============================================
# PWA SUPPORT - Service Worker
# ============================================
@app.route('/sw.js')
def service_worker():
    """Service Worker dosyasını root'tan serve et"""
    return send_file('static/sw.js', mimetype='application/javascript')

# ============================================
# METRICS MIDDLEWARE
# ============================================
from middleware.metrics_middleware import init_metrics_middleware
init_metrics_middleware(app)

# ============================================
# ROUTE REGISTRATION - Merkezi Route Yönetimi
# ============================================
from routes import register_all_routes
register_all_routes(app)

# ============================================
# KALAN ENDPOINT'LER
# ============================================
# Not: API endpoint'leri routes/api_routes.py'ye taşındı
# Kalan endpoint'ler: Zimmet, Minibar kontrol, Raporlar, Excel/PDF export

# ============================================
# ZİMMET ENDPOINT'LERİ (Depo Sorumlusu)
# ============================================

@app.route('/zimmet-detay/<int:zimmet_id>')
@login_required
@role_required('depo_sorumlusu')
def zimmet_detay(zimmet_id):
    zimmet = PersonelZimmet.query.get_or_404(zimmet_id)
    return render_template('depo_sorumlusu/zimmet_detay.html', zimmet=zimmet)

@app.route('/zimmet-iptal/<int:zimmet_id>', methods=['POST'])
@login_required
@role_required('depo_sorumlusu')
def zimmet_iptal(zimmet_id):
    """Zimmeti tamamen iptal et ve kullanılmayan ürünleri depoya iade et"""
    try:
        zimmet = PersonelZimmet.query.get_or_404(zimmet_id)
        islem_yapan = get_current_user()
        if not islem_yapan:
            flash('Kullanıcı oturumu bulunamadı. Lütfen tekrar giriş yapın.', 'danger')
            return redirect(url_for('logout'))
        
        # Sadece aktif zimmetler iptal edilebilir
        if zimmet.durum != 'aktif':
            flash('Sadece aktif zimmetler iptal edilebilir.', 'warning')
            return redirect(url_for('personel_zimmet'))
        
        # Tüm zimmet detaylarını kontrol et ve kullanılmayan ürünleri depoya iade et
        for detay in zimmet.detaylar:
            kalan = detay.kalan_miktar or (detay.miktar - detay.kullanilan_miktar)
            
            if kalan > 0:
                # Stok hareketi oluştur (depoya giriş)
                stok_hareket = StokHareket(
                    urun_id=detay.urun_id,
                    hareket_tipi='giris',
                    miktar=kalan,
                    aciklama=f'Zimmet iptali - {zimmet.personel.ad} {zimmet.personel.soyad} - Zimmet #{zimmet.id}',
                    islem_yapan_id=islem_yapan.id
                )
                db.session.add(stok_hareket)
                
                # İade edilen miktarı kaydet
                detay.iade_edilen_miktar = (detay.iade_edilen_miktar or 0) + kalan
                detay.kalan_miktar = 0
        
        # Zimmet durumunu güncelle
        zimmet.durum = 'iptal'
        zimmet.iade_tarihi = get_kktc_now()
        
        db.session.commit()
        flash(f'{zimmet.personel.ad} {zimmet.personel.soyad} adlı personelin zimmeti iptal edildi ve kullanılmayan ürünler depoya iade edildi.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Zimmet iptal edilirken hata oluştu: {str(e)}', 'danger')
    
    return redirect(url_for('personel_zimmet'))

@app.route('/zimmet-iade/<int:detay_id>', methods=['POST'])
@login_required
@role_required('depo_sorumlusu')
def zimmet_iade(detay_id):
    """Belirli bir ürünü kısmen veya tamamen iade al"""
    try:
        detay = PersonelZimmetDetay.query.get_or_404(detay_id)
        zimmet = detay.zimmet
        islem_yapan = get_current_user()
        if not islem_yapan:
            flash('Kullanıcı oturumu bulunamadı. Lütfen tekrar giriş yapın.', 'danger')
            return redirect(url_for('logout'))
        
        # Sadece aktif zimmetlerden iade alınabilir
        if zimmet.durum != 'aktif':
            flash('Sadece aktif zimmetlerden ürün iadesi alınabilir.', 'warning')
            return redirect(url_for('zimmet_detay', zimmet_id=zimmet.id))
        
        iade_miktar = int(request.form.get('iade_miktar', 0))
        aciklama = request.form.get('aciklama', '')
        
        if iade_miktar <= 0:
            flash('İade miktarı 0\'dan büyük olmalıdır.', 'warning')
            return redirect(url_for('zimmet_detay', zimmet_id=zimmet.id))
        
        # Kalan miktarı kontrol et
        kalan = detay.kalan_miktar or (detay.miktar - detay.kullanilan_miktar)
        
        if iade_miktar > kalan:
            flash(f'İade miktarı kalan miktardan fazla olamaz. Kalan: {kalan}', 'danger')
            return redirect(url_for('zimmet_detay', zimmet_id=zimmet.id))
        
        # Stok hareketi oluştur (depoya giriş)
        stok_hareket = StokHareket(
            urun_id=detay.urun_id,
            hareket_tipi='giris',
            miktar=iade_miktar,
            aciklama=f'Zimmet iadesi - {zimmet.personel.ad} {zimmet.personel.soyad} - {aciklama}',
            islem_yapan_id=islem_yapan.id
        )
        db.session.add(stok_hareket)
        
        # Zimmet detayını güncelle
        detay.iade_edilen_miktar = (detay.iade_edilen_miktar or 0) + iade_miktar
        detay.kalan_miktar = kalan - iade_miktar
        
        db.session.commit()
        flash(f'{detay.urun.urun_adi} ürününden {iade_miktar} adet iade alındı.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'İade işlemi sırasında hata oluştu: {str(e)}', 'danger')
    
    return redirect(url_for('zimmet_detay', zimmet_id=zimmet.id))

# ============================================
# API ENDPOINT'LERİ KALDIRILDI
# ============================================
# Tüm /api/ endpoint'leri routes/api_routes.py'ye taşındı

# Eski API endpoint'leri (artık routes/api_routes.py'de):
# - /api/odalar
# - /api/odalar-by-kat/<int:kat_id>
# - /api/urun-gruplari
# - /api/urunler
# - /api/urunler-by-grup/<int:grup_id>
# - /api/stok-giris
# - /api/minibar-islem-kaydet
# - /api/minibar-ilk-dolum
# - /api/minibar-ilk-dolum-kontrol/<int:oda_id>
# - /api/urun-stok/<int:urun_id>
# - /api/zimmetim
# - /api/minibar-icerigi/<int:oda_id>
# - /api/minibar-doldur
# - /api/toplu-oda-mevcut-durum
# - /api/toplu-oda-doldur
# - /api/kat-rapor-veri

# ============================================
# MİNİBAR VE DEPO ENDPOINT'LERİ
# ============================================

@app.route('/minibar-durumlari')
@login_required
@role_required('depo_sorumlusu')
def minibar_durumlari():
    """Minibar durumları - Kat ve oda seçimine göre minibar içeriğini göster"""
    kat_id = request.args.get('kat_id', type=int)
    oda_id = request.args.get('oda_id', type=int)
    
    # Tüm katları al
    katlar = Kat.query.filter_by(aktif=True).order_by(Kat.kat_no).all()
    
    # Seçili kat varsa odaları al
    odalar = []
    if kat_id:
        odalar = Oda.query.filter_by(kat_id=kat_id, aktif=True).order_by(Oda.oda_no).all()
    
    # Seçili oda varsa minibar bilgilerini al
    minibar_bilgisi = None
    son_islem = None
    minibar_urunler = []
    
    if oda_id:
        oda = db.session.get(Oda, oda_id)
        
        # Son minibar işlemini bul
        son_islem = MinibarIslem.query.filter_by(oda_id=oda_id).order_by(
            MinibarIslem.islem_tarihi.desc()
        ).first()
        
        if son_islem:
            # Bu oda için tüm minibar işlemlerini al
            tum_islemler = MinibarIslem.query.filter_by(oda_id=oda_id).order_by(
                MinibarIslem.islem_tarihi.asc()
            ).all()
            
            # Her ürün için toplam hesaplama yap
            urun_toplam = {}
            ilk_dolum_yapildi = set()  # İlk dolum yapılan ürünleri takip et
            
            for islem in tum_islemler:
                for detay in islem.detaylar:
                    urun_id = detay.urun_id
                    if urun_id not in urun_toplam:
                        urun_toplam[urun_id] = {
                            'urun': detay.urun,
                            'toplam_eklenen_ilk_dolum': 0,  # İlk dolumda eklenen (tüketim değil)
                            'toplam_eklenen_doldurma': 0,   # Doldurmalarda eklenen (= tüketim)
                            'toplam_tuketim': 0,
                            'ilk_baslangic': detay.baslangic_stok,
                            'son_bitis': detay.bitis_stok
                        }
                    
                    # İlk dolum mu kontrol et
                    if islem.islem_tipi == 'ilk_dolum' and urun_id not in ilk_dolum_yapildi:
                        urun_toplam[urun_id]['toplam_eklenen_ilk_dolum'] += detay.eklenen_miktar
                        ilk_dolum_yapildi.add(urun_id)
                    elif islem.islem_tipi in ['doldurma', 'kontrol']:
                        # Doldurma veya kontrolde eklenen miktar = tüketim
                        urun_toplam[urun_id]['toplam_eklenen_doldurma'] += detay.eklenen_miktar
                        urun_toplam[urun_id]['toplam_tuketim'] += detay.eklenen_miktar
                    
                    urun_toplam[urun_id]['son_bitis'] = detay.bitis_stok
            
            # Son işlemdeki ürünleri listele (güncel durumda olan ürünler)
            for detay in son_islem.detaylar:
                urun_id = detay.urun_id
                urun_data = urun_toplam.get(urun_id, {})
                
                # Toplam eklenen = İlk dolum + Doldurma
                ilk_dolum_eklenen = urun_data.get('toplam_eklenen_ilk_dolum', 0)
                doldurma_eklenen = urun_data.get('toplam_eklenen_doldurma', 0)
                toplam_eklenen = ilk_dolum_eklenen + doldurma_eklenen
                toplam_tuketim = urun_data.get('toplam_tuketim', 0)
                
                # Mevcut miktar = İlk dolum + Doldurma - Tüketim
                # Ama doldurma = tüketim olduğu için: İlk dolum miktarı kadar olmalı
                mevcut_miktar = urun_data.get('son_bitis', 0)
                
                minibar_urunler.append({
                    'urun': detay.urun,
                    'baslangic_stok': urun_data.get('ilk_baslangic', 0),
                    'bitis_stok': urun_data.get('son_bitis', 0),
                    'eklenen_miktar': toplam_eklenen,
                    'tuketim': toplam_tuketim,
                    'mevcut_miktar': mevcut_miktar
                })
            
            minibar_bilgisi = {
                'oda': oda,
                'son_islem': son_islem,
                'urunler': minibar_urunler,
                'toplam_urun': len(minibar_urunler),
                'toplam_miktar': sum(u['mevcut_miktar'] for u in minibar_urunler)
            }
    
    return render_template('depo_sorumlusu/minibar_durumlari.html',
                         katlar=katlar,
                         odalar=odalar,
                         minibar_bilgisi=minibar_bilgisi,
                         kat_id=kat_id,
                         oda_id=oda_id)

@app.route('/minibar-urun-gecmis/<int:oda_id>/<int:urun_id>')
@login_required
@role_required('depo_sorumlusu')
def minibar_urun_gecmis(oda_id, urun_id):
    """Belirli bir ürünün minibar geçmişini getir"""
    oda = Oda.query.get_or_404(oda_id)
    urun = Urun.query.get_or_404(urun_id)
    
    # Bu oda için tüm minibar işlemlerini al
    gecmis = []
    minibar_islemler = MinibarIslem.query.filter_by(oda_id=oda_id).order_by(
        MinibarIslem.islem_tarihi.desc()
    ).all()
    
    for islem in minibar_islemler:
        # Bu işlemde bu ürün var mı?
        detay = MinibarIslemDetay.query.filter_by(
            islem_id=islem.id,
            urun_id=urun_id
        ).first()
        
        if detay:
            gecmis.append({
                'islem_tarihi': islem.islem_tarihi.strftime('%d.%m.%Y %H:%M'),
                'islem_tipi': islem.islem_tipi,
                'personel': f"{islem.personel.ad} {islem.personel.soyad}",
                'baslangic_stok': detay.baslangic_stok,
                'eklenen_miktar': detay.eklenen_miktar,
                'tuketim': detay.tuketim,
                'bitis_stok': detay.bitis_stok,
                'aciklama': islem.aciklama or '-'
            })
    
    return jsonify({
        'success': True,
        'oda': f"{oda.oda_no}",
        'urun': urun.urun_adi,
        'gecmis': gecmis
    })

@app.route('/depo-raporlar')
@login_required
@role_required('depo_sorumlusu')
def depo_raporlar():
    # Filtreleme parametreleri
    baslangic_tarihi = request.args.get('baslangic_tarihi')
    bitis_tarihi = request.args.get('bitis_tarihi')
    rapor_tipi = request.args.get('rapor_tipi', '')
    urun_grup_id = request.args.get('urun_grup_id', '')
    urun_id = request.args.get('urun_id', '')
    personel_id = request.args.get('personel_id', '')
    hareket_tipi = request.args.get('hareket_tipi', '')
    
    rapor_verisi = None
    rapor_baslik = ""
    
    # Ürün gruplarını ve personelleri filtre için getir
    urun_gruplari = UrunGrup.query.filter_by(aktif=True).order_by(UrunGrup.grup_adi).all()
    urunler = Urun.query.filter_by(aktif=True).order_by(Urun.urun_adi).all()
    personeller = Kullanici.query.filter_by(rol='kat_sorumlusu', aktif=True).order_by(Kullanici.ad, Kullanici.soyad).all()
    
    if rapor_tipi:
        if rapor_tipi == 'stok_durum':
            # Stok Durum Raporu - Tüm ürünler için mevcut stok durumu
            rapor_baslik = "Stok Durum Raporu"
            
            query = Urun.query.filter_by(aktif=True)
            
            if urun_grup_id:
                query = query.filter_by(grup_id=urun_grup_id)
            
            urunler_liste = query.order_by(Urun.urun_adi).all()
            stok_map = get_stok_toplamlari([urun.id for urun in urunler_liste])
            
            rapor_verisi = []
            for urun in urunler_liste:
                mevcut_stok = stok_map.get(urun.id, 0)
                rapor_verisi.append({
                    'urun': urun,
                    'mevcut_stok': mevcut_stok
                })
        
        elif rapor_tipi == 'stok_hareket':
            # Stok Hareket Raporu - Detaylı stok hareketleri
            rapor_baslik = "Stok Hareket Raporu"
            
            query = StokHareket.query
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(StokHareket.islem_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(StokHareket.islem_tarihi < bitis)
            
            if urun_id:
                query = query.filter_by(urun_id=urun_id)
            elif urun_grup_id:
                query = query.join(Urun).filter(Urun.grup_id == urun_grup_id)
            
            if hareket_tipi:
                query = query.filter_by(hareket_tipi=hareket_tipi)
            
            rapor_verisi = query.order_by(StokHareket.islem_tarihi.desc()).all()
            
            # Her hareket için zimmet bilgisini ekleyelim
            for hareket in rapor_verisi:
                # Eğer açıklamada zimmet bilgisi varsa, zimmet personelini bul
                if hareket.aciklama and 'Zimmet' in hareket.aciklama:
                    try:
                        if '#' in hareket.aciklama:
                            zimmet_id = int(hareket.aciklama.split('#')[1].split()[0])
                            zimmet = db.session.get(PersonelZimmet, zimmet_id)
                            if zimmet and zimmet.personel:
                                hareket.zimmet_personel = f"{zimmet.personel.ad} {zimmet.personel.soyad}"
                            else:
                                hareket.zimmet_personel = None
                        else:
                            hareket.zimmet_personel = None
                    except Exception:
                        hareket.zimmet_personel = None
                else:
                    hareket.zimmet_personel = None
        
        elif rapor_tipi == 'zimmet':
            # Zimmet Raporu - Personel zimmet durumu
            rapor_baslik = "Zimmet Raporu"
            
            query = PersonelZimmet.query
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
            
            if personel_id:
                query = query.filter_by(personel_id=personel_id)
            
            rapor_verisi = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
        
        elif rapor_tipi == 'zimmet_detay':
            # Zimmet Detay Raporu - Ürün bazlı zimmet bilgisi
            rapor_baslik = "Ürün Bazlı Zimmet Detay Raporu"
            
            query = db.session.query(
                PersonelZimmetDetay,
                PersonelZimmet,
                Kullanici,
                Urun
            ).join(
                PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id
            ).join(
                Kullanici, PersonelZimmet.personel_id == Kullanici.id
            ).join(
                Urun, PersonelZimmetDetay.urun_id == Urun.id
            )
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
            
            if personel_id:
                query = query.filter(PersonelZimmet.personel_id == personel_id)
            
            if urun_id:
                query = query.filter(PersonelZimmetDetay.urun_id == urun_id)
            elif urun_grup_id:
                query = query.filter(Urun.grup_id == urun_grup_id)
            
            rapor_verisi = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
        
        elif rapor_tipi == 'minibar_tuketim':
            # Minibar Tüketim Raporu - SADECE minibar kontrol sırasındaki tüketim
            rapor_baslik = "Minibar Tüketim Raporu"
            
            query = db.session.query(
                Urun.urun_adi,
                Urun.birim,
                UrunGrup.grup_adi,
                Oda.oda_no,
                Kat.kat_adi,
                MinibarIslem.islem_tarihi,
                MinibarIslem.islem_tipi,
                MinibarIslemDetay.tuketim,  # Tüketim sütununu kullan
                Kullanici.ad,
                Kullanici.soyad
            ).select_from(MinibarIslemDetay).join(
                MinibarIslem, MinibarIslemDetay.islem_id == MinibarIslem.id
            ).join(
                Urun, MinibarIslemDetay.urun_id == Urun.id
            ).join(
                UrunGrup, Urun.grup_id == UrunGrup.id
            ).join(
                Oda, MinibarIslem.oda_id == Oda.id
            ).join(
                Kat, Oda.kat_id == Kat.id
            ).join(
                Kullanici, MinibarIslem.personel_id == Kullanici.id
            ).filter(
                # Setup kontrol ve ekstra tüketim işlemlerini al
                MinibarIslem.islem_tipi.in_(['setup_kontrol', 'ekstra_tuketim']),
                MinibarIslemDetay.tuketim > 0  # Sadece tüketim olan kayıtlar
            )
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(MinibarIslem.islem_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(MinibarIslem.islem_tarihi < bitis)
            
            if urun_id:
                query = query.filter(MinibarIslemDetay.urun_id == urun_id)
            elif urun_grup_id:
                query = query.filter(Urun.grup_id == urun_grup_id)
            
            if personel_id:
                query = query.filter(MinibarIslem.personel_id == personel_id)
            
            rapor_verisi = query.order_by(MinibarIslem.islem_tarihi.desc()).all()
        
        elif rapor_tipi == 'minibar_kontrol_odalar':
            # Günlük Minibar Kontrolü Yapılan Odalar Raporu
            rapor_baslik = "Günlük Minibar Kontrolü Yapılan Odalar"
            
            from sqlalchemy import func, distinct
            
            # Oda bazında son kontrol işlemini al
            query = db.session.query(
                func.max(MinibarIslem.islem_tarihi).label('son_kontrol'),
                Oda.oda_no,
                Kat.kat_adi,
                Kullanici.ad,
                Kullanici.soyad,
                func.count(distinct(MinibarIslemDetay.urun_id)).label('urun_sayisi'),
                func.sum(MinibarIslemDetay.tuketim).label('toplam_tuketim')
            ).select_from(MinibarIslem).join(
                Oda, MinibarIslem.oda_id == Oda.id
            ).join(
                Kat, Oda.kat_id == Kat.id
            ).join(
                Kullanici, MinibarIslem.personel_id == Kullanici.id
            ).join(
                MinibarIslemDetay, MinibarIslem.id == MinibarIslemDetay.islem_id
            ).filter(
                MinibarIslem.islem_tipi.in_(['setup_kontrol', 'ekstra_ekleme', 'ekstra_tuketim'])
            )
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(MinibarIslem.islem_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(MinibarIslem.islem_tarihi < bitis)
            
            if personel_id:
                query = query.filter(MinibarIslem.personel_id == personel_id)
            
            # Oda bazında grupla
            query = query.group_by(
                Oda.oda_no,
                Kat.kat_adi,
                Kullanici.ad,
                Kullanici.soyad
            )
            
            rapor_verisi = query.order_by(db.desc('son_kontrol'), Oda.oda_no).all()
        
        elif rapor_tipi == 'urun_grup':
            # Ürün Grubu Bazlı Rapor
            rapor_baslik = "Ürün Grubu Bazlı Stok Raporu"
            
            query = UrunGrup.query.filter_by(aktif=True)
            aktif_urunler = Urun.query.filter_by(aktif=True).all()
            stok_map = get_stok_toplamlari([urun.id for urun in aktif_urunler])
            urunler_by_grup = {}
            for urun in aktif_urunler:
                urunler_by_grup.setdefault(urun.grup_id, []).append(urun)
            
            rapor_verisi = []
            for grup in query.all():
                grup_urunleri = urunler_by_grup.get(grup.id, [])
                toplam_urun_sayisi = len(grup_urunleri)
                kritik_urun_sayisi = 0
                
                for urun in grup_urunleri:
                    mevcut_stok = stok_map.get(urun.id, 0)
                    kritik_seviye = urun.kritik_stok_seviyesi or 0
                    if mevcut_stok <= kritik_seviye:
                        kritik_urun_sayisi += 1
                
                rapor_verisi.append({
                    'grup': grup,
                    'toplam_urun': toplam_urun_sayisi,
                    'kritik_urun': kritik_urun_sayisi
                })
        
        elif rapor_tipi == 'ozet':
            # Özet Rapor - Genel sistem durumu
            rapor_baslik = "Genel Sistem Özet Raporu"
            
            # Toplam ürün sayısı
            toplam_urun = Urun.query.filter_by(aktif=True).count()
            
            # Kritik stok seviyesindeki ürünler
            kritik_urunler = get_kritik_stok_urunler()
            
            # Aktif zimmetler
            aktif_zimmet = PersonelZimmet.query.filter_by(durum='aktif').count()
            
            # Bugünkü stok hareketleri
            bugun = get_kktc_now().date()
            bugun_baslangic = datetime.combine(bugun, datetime.min.time())
            bugun_bitis = datetime.combine(bugun, datetime.max.time())
            
            bugun_giris = StokHareket.query.filter(
                StokHareket.hareket_tipi == 'giris',
                StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis)
            ).count()
            
            bugun_cikis = StokHareket.query.filter(
                StokHareket.hareket_tipi == 'cikis',
                StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis)
            ).count()
            
            # Bu ayki zimmet sayısı
            ay_baslangic = get_kktc_now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            ay_zimmet = PersonelZimmet.query.filter(PersonelZimmet.zimmet_tarihi >= ay_baslangic).count()
            
            rapor_verisi = {
                'toplam_urun': toplam_urun,
                'kritik_urun': len(kritik_urunler),
                'aktif_zimmet': aktif_zimmet,
                'bugun_giris': bugun_giris,
                'bugun_cikis': bugun_cikis,
                'ay_zimmet': ay_zimmet
            }
    
    return render_template('depo_sorumlusu/raporlar.html',
                         rapor_verisi=rapor_verisi,
                         rapor_baslik=rapor_baslik,
                         rapor_tipi=rapor_tipi,
                         urun_gruplari=urun_gruplari,
                         urunler=urunler,
                         personeller=personeller)

# Kullanım Kılavuzu Sayfası
@app.route('/kullanim-kilavuzu/personel-zimmet')
@login_required
def kullanim_kilavuzu_personel_zimmet():
    """Personel zimmet kullanım kılavuzu sayfası"""
    return render_template('kullanim_kilavuzu/personel_zimmet_kilavuzu.html')

# Dolum Talepleri Rotaları
# ============================================================================
# TOPLU İŞLEM ÖZELLİKLERİ
# ============================================================================

# Excel Export
@app.route('/excel-export/<rapor_tipi>')
@login_required
def excel_export(rapor_tipi):
    try:
        # Filtreleme parametrelerini al
        baslangic_tarihi = request.args.get('baslangic_tarihi')
        bitis_tarihi = request.args.get('bitis_tarihi')
        urun_grup_id = request.args.get('urun_grup_id')
        urun_id = request.args.get('urun_id')
        personel_id = request.args.get('personel_id')
        hareket_tipi = request.args.get('hareket_tipi')
        
        # Excel dosyası oluştur
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Rapor başlığı
        rapor_basliklari = {
            'stok_durum': 'Stok Durum Raporu',
            'stok_hareket': 'Stok Hareket Raporu',
            'zimmet': 'Zimmet Raporu',
            'zimmet_detay': 'Ürün Bazlı Zimmet Detay Raporu',
            'urun_grup': 'Ürün Grubu Bazlı Stok Raporu',
            'ozet': 'Genel Sistem Özet Raporu'
        }
        
        baslik = rapor_basliklari.get(rapor_tipi, 'Rapor')
        ws.title = baslik[:31]  # Excel sheet name limit
        
        # Başlık satırı
        ws['A1'] = baslik
        ws['A1'].font = Font(size=16, bold=True)
        
        # Tarih bilgisi
        ws['A2'] = f"Rapor Tarihi: {get_kktc_now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = Font(size=10)
        
        row_num = 4
        
        if rapor_tipi == 'stok_durum':
            # Başlıkları yaz
            headers = ['Ürün Adı', 'Ürün Grubu', 'Birim', 'Mevcut Stok', 'Kritik Seviye', 'Durum']
            ws.merge_cells('A1:F1')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # Verileri yaz
            query = Urun.query.filter_by(aktif=True)
            if urun_grup_id:
                query = query.filter_by(grup_id=urun_grup_id)
            
            urunler_liste = query.order_by(Urun.urun_adi).all()
            stok_map = get_stok_toplamlari([urun.id for urun in urunler_liste])
            
            for urun in urunler_liste:
                row_num += 1
                mevcut_stok = stok_map.get(urun.id, 0)
                kritik_seviye = urun.kritik_stok_seviyesi or 0
                durum = 'KRİTİK' if mevcut_stok <= kritik_seviye else 'NORMAL'
                
                ws.cell(row=row_num, column=1, value=urun.urun_adi)
                ws.cell(row=row_num, column=2, value=urun.grup.grup_adi)
                ws.cell(row=row_num, column=3, value=urun.birim)
                ws.cell(row=row_num, column=4, value=mevcut_stok)
                ws.cell(row=row_num, column=5, value=urun.kritik_stok_seviyesi)
                ws.cell(row=row_num, column=6, value=durum)
        
        elif rapor_tipi == 'stok_hareket':
            headers = ['Tarih', 'Ürün Adı', 'Hareket Tipi', 'Miktar', 'Açıklama', 'İşlem Yapan']
            ws.merge_cells('A1:F1')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            query = StokHareket.query
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(StokHareket.islem_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(StokHareket.islem_tarihi < bitis)
            
            if urun_id:
                query = query.filter_by(urun_id=urun_id)
            elif urun_grup_id:
                query = query.join(Urun).filter(Urun.grup_id == urun_grup_id)
            
            if hareket_tipi:
                query = query.filter_by(hareket_tipi=hareket_tipi)
            
            hareketler = query.order_by(StokHareket.islem_tarihi.desc()).all()
            
            for hareket in hareketler:
                row_num += 1
                islem_yapan = f"{hareket.islem_yapan.ad} {hareket.islem_yapan.soyad}" if hareket.islem_yapan else '-'
                
                # Zimmet bilgisini açıklamaya ekle
                aciklama = hareket.aciklama or '-'
                if hareket.aciklama and 'Zimmet' in hareket.aciklama:
                    try:
                        if '#' in hareket.aciklama:
                            zimmet_id = int(hareket.aciklama.split('#')[1].split()[0])
                            zimmet = db.session.get(PersonelZimmet, zimmet_id)
                            if zimmet and zimmet.personel:
                                aciklama += f" → {zimmet.personel.ad} {zimmet.personel.soyad}"
                    except Exception:
                        pass
                
                ws.cell(row=row_num, column=1, value=hareket.islem_tarihi.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row_num, column=2, value=hareket.urun.urun_adi)
                ws.cell(row=row_num, column=3, value=hareket.hareket_tipi.upper())
                ws.cell(row=row_num, column=4, value=hareket.miktar)
                ws.cell(row=row_num, column=5, value=aciklama)
                ws.cell(row=row_num, column=6, value=islem_yapan)
        
        elif rapor_tipi == 'zimmet':
            headers = ['Zimmet No', 'Personel', 'Zimmet Tarihi', 'Ürün Sayısı', 'Toplam Miktar', 'Durum']
            ws.merge_cells('A1:F1')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            query = PersonelZimmet.query
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
            
            if personel_id:
                query = query.filter_by(personel_id=personel_id)
            
            zimmetler = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
            
            for zimmet in zimmetler:
                row_num += 1
                toplam_miktar = sum(d.miktar for d in zimmet.detaylar)
                
                ws.cell(row=row_num, column=1, value=f"#{zimmet.id}")
                ws.cell(row=row_num, column=2, value=f"{zimmet.personel.ad} {zimmet.personel.soyad}")
                ws.cell(row=row_num, column=3, value=zimmet.zimmet_tarihi.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row_num, column=4, value=len(zimmet.detaylar))
                ws.cell(row=row_num, column=5, value=toplam_miktar)
                ws.cell(row=row_num, column=6, value=zimmet.durum.upper())
        
        elif rapor_tipi == 'zimmet_detay':
            headers = ['Zimmet No', 'Personel', 'Zimmet Tarihi', 'Ürün Adı', 'Grup', 'Miktar', 'Durum']
            ws.merge_cells('A1:G1')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            
            query = db.session.query(
                PersonelZimmetDetay,
                PersonelZimmet,
                Kullanici,
                Urun
            ).join(
                PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id
            ).join(
                Kullanici, PersonelZimmet.personel_id == Kullanici.id
            ).join(
                Urun, PersonelZimmetDetay.urun_id == Urun.id
            )
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
            
            if personel_id:
                query = query.filter(PersonelZimmet.personel_id == personel_id)
            
            if urun_id:
                query = query.filter(PersonelZimmetDetay.urun_id == urun_id)
            elif urun_grup_id:
                query = query.filter(Urun.grup_id == urun_grup_id)
            
            detaylar = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).all()
            
            for detay, zimmet, kullanici, urun in detaylar:
                row_num += 1
                
                ws.cell(row=row_num, column=1, value=f"#{zimmet.id}")
                ws.cell(row=row_num, column=2, value=f"{kullanici.ad} {kullanici.soyad}")
                ws.cell(row=row_num, column=3, value=zimmet.zimmet_tarihi.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row_num, column=4, value=urun.urun_adi)
                ws.cell(row=row_num, column=5, value=urun.grup.grup_adi)
                ws.cell(row=row_num, column=6, value=detay.miktar)
                ws.cell(row=row_num, column=7, value=zimmet.durum.upper())
        
        elif rapor_tipi == 'urun_grup':
            headers = ['Ürün Grubu', 'Toplam Ürün', 'Kritik Stoklu Ürün']
            ws.merge_cells('A1:C1')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            
            gruplar = UrunGrup.query.filter_by(aktif=True).all()
            aktif_urunler = Urun.query.filter_by(aktif=True).all()
            stok_map = get_stok_toplamlari([urun.id for urun in aktif_urunler])
            urun_gruplari_map = {}
            for urun in aktif_urunler:
                urun_gruplari_map.setdefault(urun.grup_id, []).append(urun)
            
            for grup in gruplar:
                row_num += 1
                grup_urunleri = urun_gruplari_map.get(grup.id, [])
                toplam_urun_sayisi = len(grup_urunleri)
                kritik_urun_sayisi = 0
                
                for urun in grup_urunleri:
                    mevcut_stok = stok_map.get(urun.id, 0)
                    kritik_seviye = urun.kritik_stok_seviyesi or 0
                    if mevcut_stok <= kritik_seviye:
                        kritik_urun_sayisi += 1
                
                ws.cell(row=row_num, column=1, value=grup.grup_adi)
                ws.cell(row=row_num, column=2, value=toplam_urun_sayisi)
                ws.cell(row=row_num, column=3, value=kritik_urun_sayisi)
        
        # Sütun genişliklerini ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Response oluştur
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={rapor_tipi}_raporu_{get_kktc_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Excel export hatası: {str(e)}', 'danger')
        return redirect(url_for('depo_raporlar'))

# PDF Export
@app.route('/pdf-export/<rapor_tipi>')
@login_required
def pdf_export(rapor_tipi):
    try:
        # Türkçe karakter dönüşüm tablosu
        def turkce_ascii(text):
            """Türkçe karakterleri ASCII'ye dönüştür"""
            if not text:
                return ''
            char_map = {
                'ç': 'c', 'Ç': 'C',
                'ğ': 'g', 'Ğ': 'G',
                'ı': 'i', 'İ': 'I',
                'ö': 'o', 'Ö': 'O',
                'ş': 's', 'Ş': 'S',
                'ü': 'u', 'Ü': 'U'
            }
            result = str(text)
            for turkish, ascii_char in char_map.items():
                result = result.replace(turkish, ascii_char)
            return result
        
        # Filtreleme parametrelerini al
        baslangic_tarihi = request.args.get('baslangic_tarihi')
        bitis_tarihi = request.args.get('bitis_tarihi')
        urun_grup_id = request.args.get('urun_grup_id')
        urun_id = request.args.get('urun_id')
        personel_id = request.args.get('personel_id')
        hareket_tipi = request.args.get('hareket_tipi')
        
        # PDF dosyası oluştur
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        styles = getSampleStyleSheet()
        story = []
        
        # Rapor başlığı
        rapor_basliklari = {
            'stok_durum': 'Stok Durum Raporu',
            'stok_hareket': 'Stok Hareket Raporu',
            'zimmet': 'Zimmet Raporu',
            'zimmet_detay': 'Urun Bazli Zimmet Detay Raporu',
            'urun_grup': 'Urun Grubu Bazli Stok Raporu',
            'ozet': 'Genel Sistem Ozet Raporu'
        }
        
        baslik = turkce_ascii(rapor_basliklari.get(rapor_tipi, 'Rapor'))
        
        # Başlık
        title = Paragraph(baslik, styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Tarih
        date_text = f"Rapor Tarihi: {get_kktc_now().strftime('%d.%m.%Y %H:%M')}"
        date_para = Paragraph(date_text, styles['Normal'])
        story.append(date_para)
        story.append(Spacer(1, 20))
        
        # Tablo verileri
        data = []
        
        if rapor_tipi == 'stok_durum':
            data.append([turkce_ascii(h) for h in ['Urun Adi', 'Urun Grubu', 'Birim', 'Mevcut Stok', 'Kritik Seviye', 'Durum']])
            
            query = Urun.query.filter_by(aktif=True)
            if urun_grup_id:
                query = query.filter_by(grup_id=urun_grup_id)
            
            urunler_liste = query.order_by(Urun.urun_adi).all()
            stok_map = get_stok_toplamlari([urun.id for urun in urunler_liste]) if urunler_liste else {}
            
            for urun in urunler_liste:
                mevcut_stok = stok_map.get(urun.id, 0)
                kritik_seviye = urun.kritik_stok_seviyesi or 0
                durum = 'KRITIK' if mevcut_stok <= kritik_seviye else 'NORMAL'
                
                data.append([
                    turkce_ascii(urun.urun_adi),
                    turkce_ascii(urun.grup.grup_adi),
                    turkce_ascii(urun.birim),
                    str(mevcut_stok),
                    str(urun.kritik_stok_seviyesi),
                    durum
                ])
        
        elif rapor_tipi == 'stok_hareket':
            data.append([turkce_ascii(h) for h in ['Tarih', 'Urun Adi', 'Hareket', 'Miktar', 'Aciklama']])
            
            query = StokHareket.query
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(StokHareket.islem_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(StokHareket.islem_tarihi < bitis)
            
            if urun_id:
                query = query.filter_by(urun_id=urun_id)
            elif urun_grup_id:
                query = query.join(Urun).filter(Urun.grup_id == urun_grup_id)
            
            if hareket_tipi:
                query = query.filter_by(hareket_tipi=hareket_tipi)
            
            hareketler = query.order_by(StokHareket.islem_tarihi.desc()).limit(100).all()
            
            for hareket in hareketler:
                # Zimmet bilgisini açıklamaya ekle
                aciklama = hareket.aciklama or '-'
                if hareket.aciklama and 'Zimmet' in hareket.aciklama:
                    try:
                        if '#' in hareket.aciklama:
                            zimmet_id = int(hareket.aciklama.split('#')[1].split()[0])
                            zimmet = db.session.get(PersonelZimmet, zimmet_id)
                            if zimmet and zimmet.personel:
                                aciklama = f"{aciklama} → {zimmet.personel.ad} {zimmet.personel.soyad}"
                    except Exception:
                        pass
                
                # Açıklamayı kısalt (PDF için)
                aciklama_kisaltilmis = aciklama[:50] if len(aciklama) > 50 else aciklama
                
                data.append([
                    hareket.islem_tarihi.strftime('%d.%m.%Y %H:%M'),
                    turkce_ascii(hareket.urun.urun_adi),
                    turkce_ascii(hareket.hareket_tipi.upper()),
                    str(hareket.miktar),
                    turkce_ascii(aciklama_kisaltilmis)
                ])
        
        elif rapor_tipi == 'zimmet':
            data.append([turkce_ascii(h) for h in ['Zimmet No', 'Personel', 'Tarih', 'Urun Sayisi', 'Toplam', 'Durum']])
            
            query = PersonelZimmet.query
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
            
            if personel_id:
                query = query.filter_by(personel_id=personel_id)
            
            zimmetler = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).limit(100).all()
            
            for zimmet in zimmetler:
                toplam_miktar = sum(d.miktar for d in zimmet.detaylar)
                
                data.append([
                    f"#{zimmet.id}",
                    turkce_ascii(f"{zimmet.personel.ad} {zimmet.personel.soyad}"),
                    zimmet.zimmet_tarihi.strftime('%d.%m.%Y'),
                    str(len(zimmet.detaylar)),
                    str(toplam_miktar),
                    turkce_ascii(zimmet.durum.upper())
                ])
        
        elif rapor_tipi == 'zimmet_detay':
            data.append([turkce_ascii(h) for h in ['Zimmet', 'Personel', 'Urun', 'Grup', 'Miktar', 'Durum']])
            
            query = db.session.query(
                PersonelZimmetDetay,
                PersonelZimmet,
                Kullanici,
                Urun
            ).join(
                PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id
            ).join(
                Kullanici, PersonelZimmet.personel_id == Kullanici.id
            ).join(
                Urun, PersonelZimmetDetay.urun_id == Urun.id
            )
            
            if baslangic_tarihi:
                baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
                query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
            
            if bitis_tarihi:
                bitis = datetime.strptime(bitis_tarihi, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
            
            if personel_id:
                query = query.filter(PersonelZimmet.personel_id == personel_id)
            
            if urun_id:
                query = query.filter(PersonelZimmetDetay.urun_id == urun_id)
            elif urun_grup_id:
                query = query.filter(Urun.grup_id == urun_grup_id)
            
            detaylar = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).limit(100).all()
            
            for detay, zimmet, kullanici, urun in detaylar:
                data.append([
                    f"#{zimmet.id}",
                    turkce_ascii(f"{kullanici.ad} {kullanici.soyad}"),
                    turkce_ascii(urun.urun_adi),
                    turkce_ascii(urun.grup.grup_adi),
                    str(detay.miktar),
                    turkce_ascii(zimmet.durum.upper())
                ])
        
        elif rapor_tipi == 'urun_grup':
            data.append([turkce_ascii(h) for h in ['Urun Grubu', 'Toplam Urun', 'Kritik Stoklu Urun']])
            
            gruplar = UrunGrup.query.filter_by(aktif=True).all()
            aktif_urunler = Urun.query.filter_by(aktif=True).all()
            stok_map = get_stok_toplamlari([urun.id for urun in aktif_urunler]) if aktif_urunler else {}

            grup_urunleri_map = {}
            for urun in aktif_urunler:
                grup_urunleri_map.setdefault(urun.grup_id, []).append(urun)

            for grup in gruplar:
                grup_urunleri = grup_urunleri_map.get(grup.id, [])
                toplam_urun_sayisi = len(grup_urunleri)
                kritik_urun_sayisi = 0
                
                for urun in grup_urunleri:
                    mevcut_stok = stok_map.get(urun.id, 0)
                    kritik_seviye = urun.kritik_stok_seviyesi or 0
                    if mevcut_stok <= kritik_seviye:
                        kritik_urun_sayisi += 1
                
                data.append([
                    turkce_ascii(grup.grup_adi),
                    str(toplam_urun_sayisi),
                    str(kritik_urun_sayisi)
                ])
        
        elif rapor_tipi == 'ozet':
            # Özet raporu için özel tablo
            toplam_urun = Urun.query.filter_by(aktif=True).count()
            kritik_urunler = get_kritik_stok_urunler()
            aktif_zimmet = PersonelZimmet.query.filter_by(durum='aktif').count()
            
            bugun = get_kktc_now().date()
            bugun_baslangic = datetime.combine(bugun, datetime.min.time())
            bugun_bitis = datetime.combine(bugun, datetime.max.time())
            
            bugun_giris = StokHareket.query.filter(
                StokHareket.hareket_tipi == 'giris',
                StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis)
            ).count()
            
            bugun_cikis = StokHareket.query.filter(
                StokHareket.hareket_tipi == 'cikis',
                StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis)
            ).count()
            
            ay_baslangic = get_kktc_now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            ay_zimmet = PersonelZimmet.query.filter(PersonelZimmet.zimmet_tarihi >= ay_baslangic).count()
            
            data = [
                ['Metrik', 'Deger'],
                [turkce_ascii('Toplam Urun Sayisi'), str(toplam_urun)],
                [turkce_ascii('Kritik Stoklu Urun'), str(len(kritik_urunler))],
                [turkce_ascii('Aktif Zimmet'), str(aktif_zimmet)],
                [turkce_ascii('Bugun Stok Giris'), str(bugun_giris)],
                [turkce_ascii('Bugun Stok Cikis'), str(bugun_cikis)],
                [turkce_ascii('Bu Ay Zimmet'), str(ay_zimmet)]
            ]
        
        # Tablo oluştur
        if data:
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
            ]))
            
            story.append(table)
        else:
            no_data = Paragraph(turkce_ascii("Bu filtre kriterleri icin veri bulunamadi."), styles['Normal'])
            story.append(no_data)
        
        # PDF'i oluştur
        doc.build(story)
        buffer.seek(0)
        
        # Response oluştur
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={rapor_tipi}_raporu_{get_kktc_now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
        
    except Exception as e:
        flash(f'PDF export hatasi: {str(e)}', 'danger')
        return redirect(url_for('depo_raporlar'))

# ============================================================================
# API: DASHBOARD WIDGET'LARI
# ============================================================================

@app.route('/api/son-aktiviteler')
@login_required
@role_required(['sistem_yoneticisi', 'admin'])
def api_son_aktiviteler():
    """Son kullanıcı aktivitelerini döndür"""
    try:
        limit = request.args.get('limit', 10, type=int)

        # Son aktiviteleri çek (sadece önemli işlemler, superadmin hariç)
        aktiviteler = SistemLog.query\
            .join(Kullanici, SistemLog.kullanici_id == Kullanici.id)\
            .filter(
                SistemLog.islem_tipi.in_(['ekleme', 'guncelleme', 'silme']),
                Kullanici.rol != 'superadmin'
            )\
            .order_by(SistemLog.islem_tarihi.desc())\
            .limit(limit)\
            .all()

        data = []
        for log in aktiviteler:
            try:
                # Kullanıcı bilgisi
                kullanici_adi = 'Sistem'
                if log.kullanici:
                    kullanici_adi = f"{log.kullanici.ad} {log.kullanici.soyad}"

                # İşlem detayını parse et
                import json
                detay = {}
                if log.islem_detay:
                    try:
                        detay = json.loads(log.islem_detay) if isinstance(log.islem_detay, str) else log.islem_detay
                    except Exception:
                        detay = {'aciklama': str(log.islem_detay)}

                # Zaman farkı hesapla
                # islem_tarihi'ni datetime'a çevir
                if isinstance(log.islem_tarihi, datetime):
                    # Datetime objesi
                    if log.islem_tarihi.tzinfo is None:
                        # Naive datetime ise, UTC olarak kabul et
                        islem_tarihi = log.islem_tarihi.replace(tzinfo=timezone.utc)
                    else:
                        islem_tarihi = log.islem_tarihi
                else:
                    # Date objesi ise datetime'a çevir
                    islem_tarihi = datetime.combine(log.islem_tarihi, datetime.min.time()).replace(tzinfo=timezone.utc)
                
                zaman_farki = get_kktc_now() - islem_tarihi

                if zaman_farki < timedelta(minutes=1):
                    zaman_str = "Az önce"
                elif zaman_farki < timedelta(hours=1):
                    dakika = int(zaman_farki.total_seconds() / 60)
                    zaman_str = f"{dakika} dakika önce"
                elif zaman_farki < timedelta(days=1):
                    saat = int(zaman_farki.total_seconds() / 3600)
                    zaman_str = f"{saat} saat önce"
                else:
                    gun = zaman_farki.days
                    zaman_str = f"{gun} gün önce"

                data.append({
                    'id': log.id,
                    'kullanici': kullanici_adi,
                    'islem_tipi': log.islem_tipi,
                    'modul': log.modul,
                    'detay': detay,
                    'zaman': zaman_str,
                    'tam_tarih': islem_tarihi.strftime('%d.%m.%Y %H:%M')
                })
            except Exception as log_error:
                # Tek bir log hatası tüm endpoint'i bozmasın
                print(f"Log parse hatası (ID: {log.id}): {log_error}")
                continue

        return jsonify({'success': True, 'aktiviteler': data})

    except Exception as e:
        print(f"Son aktiviteler hatası: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/bekleyen-dolum-sayisi')
@login_required
@role_required(['sistem_yoneticisi', 'admin', 'depo_sorumlusu', 'kat_sorumlusu'])
def api_bekleyen_dolum_sayisi():
    """Bekleyen dolum talepleri sayısını döndür"""
    try:
        # Bekleyen dolum taleplerini say
        count = MinibarDolumTalebi.query.filter_by(durum='beklemede').count()
        
        return jsonify({
            'success': True,
            'count': count
        })
    except Exception as e:
        logger.error(f"Bekleyen dolum sayısı hatası: {e}")
        return jsonify({
            'success': False,
            'count': 0,
            'error': str(e)
        }), 500


@app.route('/api/tuketim-trendleri')
@login_required
@role_required(['sistem_yoneticisi', 'admin', 'depo_sorumlusu'])
def api_tuketim_trendleri():
    """Günlük/haftalık tüketim trendlerini döndür"""
    try:
        from sqlalchemy import func

        gun_sayisi = request.args.get('gun', 7, type=int)  # Varsayılan 7 gün

        # Son N günün tüketim verilerini al
        baslangic = get_kktc_now() - timedelta(days=gun_sayisi)

        # Günlük tüketim toplamı (MinibarIslemDetay'dan)
        gunluk_tuketim = db.session.query(
            func.date(MinibarIslem.islem_tarihi).label('tarih'),
            func.sum(MinibarIslemDetay.tuketim).label('toplam_tuketim'),
            func.count(MinibarIslemDetay.id).label('islem_sayisi')
        ).join(MinibarIslemDetay, MinibarIslemDetay.islem_id == MinibarIslem.id)\
         .filter(MinibarIslem.islem_tarihi >= baslangic)\
         .group_by(func.date(MinibarIslem.islem_tarihi))\
         .order_by(func.date(MinibarIslem.islem_tarihi))\
         .all()

        # Tüm günleri doldur (veri olmayan günler için 0)
        tum_gunler = {}
        for i in range(gun_sayisi):
            tarih = (get_kktc_now() - timedelta(days=gun_sayisi-i-1)).date()
            tum_gunler[str(tarih)] = {'tuketim': 0, 'islem_sayisi': 0}

        # Veri olanları güncelle
        for row in gunluk_tuketim:
            tarih_str = str(row.tarih)
            tum_gunler[tarih_str] = {
                'tuketim': int(row.toplam_tuketim or 0),
                'islem_sayisi': int(row.islem_sayisi or 0)
            }

        # Chart.js formatına çevir
        labels = []
        tuketim_data = []
        islem_data = []

        for tarih_str in sorted(tum_gunler.keys()):
            # Tarih formatla (DD/MM)
            tarih_obj = datetime.strptime(tarih_str, '%Y-%m-%d')
            labels.append(tarih_obj.strftime('%d/%m'))
            tuketim_data.append(tum_gunler[tarih_str]['tuketim'])
            islem_data.append(tum_gunler[tarih_str]['islem_sayisi'])

        return jsonify({
            'success': True,
            'labels': labels,
            'datasets': [
                {
                    'label': 'Toplam Tüketim',
                    'data': tuketim_data,
                    'borderColor': 'rgb(239, 68, 68)',
                    'backgroundColor': 'rgba(239, 68, 68, 0.1)',
                    'tension': 0.3
                },
                {
                    'label': 'İşlem Sayısı',
                    'data': islem_data,
                    'borderColor': 'rgb(59, 130, 246)',
                    'backgroundColor': 'rgba(59, 130, 246, 0.1)',
                    'tension': 0.3
                }
            ]
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================
# AUDIT TRAIL ROUTE'LARI
# ==========================

@app.route('/sistem-yoneticisi/audit-trail')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def audit_trail():
    """Audit Trail - Denetim İzi Sayfası"""
    from models import AuditLog, Kullanici
    from datetime import datetime, timedelta
    
    # Sayfalama
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Filtreler
    kullanici_id = request.args.get('kullanici_id', type=int)
    islem_tipi = request.args.get('islem_tipi')
    tablo_adi = request.args.get('tablo_adi')
    tarih_baslangic = request.args.get('tarih_baslangic')
    tarih_bitis = request.args.get('tarih_bitis')
    
    # Base query (superadmin aktiviteleri hariç)
    query = AuditLog.query.filter(AuditLog.kullanici_rol != 'superadmin')
    
    # Filtreleme
    if kullanici_id:
        query = query.filter_by(kullanici_id=kullanici_id)
    if islem_tipi:
        query = query.filter_by(islem_tipi=islem_tipi)
    if tablo_adi:
        query = query.filter_by(tablo_adi=tablo_adi)
    if tarih_baslangic:
        tarih_baslangic_dt = datetime.strptime(tarih_baslangic, '%Y-%m-%d')
        query = query.filter(AuditLog.islem_tarihi >= tarih_baslangic_dt)
    if tarih_bitis:
        tarih_bitis_dt = datetime.strptime(tarih_bitis, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(AuditLog.islem_tarihi < tarih_bitis_dt)
    
    # Sıralama ve sayfalama
    query = query.order_by(AuditLog.islem_tarihi.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # İstatistikler
    bugun = get_kktc_now().replace(hour=0, minute=0, second=0, microsecond=0)
    bu_hafta = bugun - timedelta(days=bugun.weekday())
    bu_ay = bugun.replace(day=1)
    
    stats = {
        'today': AuditLog.query.filter(AuditLog.islem_tarihi >= bugun).count(),
        'week': AuditLog.query.filter(AuditLog.islem_tarihi >= bu_hafta).count(),
        'month': AuditLog.query.filter(AuditLog.islem_tarihi >= bu_ay).count()
    }
    
    # Filtre için kullanıcı listesi
    users = Kullanici.query.filter_by(aktif=True).order_by(Kullanici.kullanici_adi).all()
    
    # Filtre için tablo listesi (unique)
    tables = db.session.query(AuditLog.tablo_adi).distinct().order_by(AuditLog.tablo_adi).all()
    tables = [t[0] for t in tables]
    
    return render_template('sistem_yoneticisi/audit_trail.html',
                         logs=pagination.items,
                         pagination=pagination,
                         users=users,
                         tables=tables,
                         stats=stats)


@app.route('/sistem-yoneticisi/audit-trail/<int:log_id>')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def audit_trail_detail(log_id):
    """Audit Log Detay API"""
    from models import AuditLog
    
    log = AuditLog.query.get_or_404(log_id)
    
    return jsonify({
        'id': log.id,
        'kullanici_id': log.kullanici_id,
        'kullanici_adi': log.kullanici_adi,
        'kullanici_rol': log.kullanici_rol,
        'islem_tipi': log.islem_tipi,
        'tablo_adi': log.tablo_adi,
        'kayit_id': log.kayit_id,
        'eski_deger': log.eski_deger,
        'yeni_deger': log.yeni_deger,
        'degisiklik_ozeti': log.degisiklik_ozeti,
        'http_method': log.http_method,
        'url': log.url,
        'endpoint': log.endpoint,
        'ip_adresi': log.ip_adresi,
        'user_agent': log.user_agent,
        'islem_tarihi': log.islem_tarihi.strftime('%d.%m.%Y %H:%M:%S'),
        'aciklama': log.aciklama,
        'basarili': log.basarili,
        'hata_mesaji': log.hata_mesaji
    })


@app.route('/sistem-yoneticisi/audit-trail/export')
@login_required
@role_required('sistem_yoneticisi', 'admin')
def audit_trail_export():
    """Audit Trail Excel Export"""
    from models import AuditLog
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    
    # Filtreleri al
    kullanici_id = request.args.get('kullanici_id', type=int)
    islem_tipi = request.args.get('islem_tipi')
    tablo_adi = request.args.get('tablo_adi')
    tarih_baslangic = request.args.get('tarih_baslangic')
    tarih_bitis = request.args.get('tarih_bitis')
    
    # Query oluştur (superadmin aktiviteleri hariç)
    query = AuditLog.query.filter(AuditLog.kullanici_rol != 'superadmin')
    
    if kullanici_id:
        query = query.filter_by(kullanici_id=kullanici_id)
    if islem_tipi:
        query = query.filter_by(islem_tipi=islem_tipi)
    if tablo_adi:
        query = query.filter_by(tablo_adi=tablo_adi)
    if tarih_baslangic:
        tarih_baslangic_dt = datetime.strptime(tarih_baslangic, '%Y-%m-%d')
        query = query.filter(AuditLog.islem_tarihi >= tarih_baslangic_dt)
    if tarih_bitis:
        tarih_bitis_dt = datetime.strptime(tarih_bitis, '%Y-%m-%d')
        query = query.filter(AuditLog.islem_tarihi <= tarih_bitis_dt)
    
    logs = query.order_by(AuditLog.islem_tarihi.desc()).limit(10000).all()
    
    # Excel oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Trail"
    
    # Başlıklar
    headers = ['ID', 'Tarih', 'Kullanıcı', 'Rol', 'İşlem', 'Tablo', 'Kayıt ID', 
               'Değişiklik', 'IP', 'URL', 'Başarılı']
    
    # Başlık satırını formatla
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Verileri ekle
    for row, log in enumerate(logs, 2):
        ws.cell(row=row, column=1, value=log.id)
        ws.cell(row=row, column=2, value=log.islem_tarihi.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=3, value=log.kullanici_adi)
        ws.cell(row=row, column=4, value=log.kullanici_rol)
        ws.cell(row=row, column=5, value=log.islem_tipi)
        ws.cell(row=row, column=6, value=log.tablo_adi)
        ws.cell(row=row, column=7, value=log.kayit_id)
        ws.cell(row=row, column=8, value=log.degisiklik_ozeti or '')
        ws.cell(row=row, column=9, value=log.ip_adresi or '')
        ws.cell(row=row, column=10, value=log.url or '')
        ws.cell(row=row, column=11, value='Evet' if log.basarili else 'Hayır')
    
    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 10
    
    # Excel dosyasını kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"audit_trail_{get_kktc_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Audit export işlemini logla
    from utils.audit import audit_export
    audit_export('audit_logs', f'Excel export: {len(logs)} kayıt')
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================================================
# SİSTEM SIFIRLAMA - ÖZEL ŞİFRE İLE KORUMALI
# ============================================================================

RESET_PASSWORD = "518518Erkan!"  # Özel sistem sıfırlama şifresi

@app.route('/resetsystem', methods=['GET', 'POST'])
@csrf.exempt  # CSRF korumasını kaldır (kendi validasyonumuz var)
def reset_system():
    """Sistem sıfırlama sayfası - Özel şifre ile korumalı"""
    
    if request.method == 'GET':
        # Şifre giriş sayfasını göster
        return render_template('reset_system.html', show_stats=False)
    
    # POST işlemi
    action = request.form.get('action')
    reset_password = request.form.get('reset_password', '')
    
    # Şifre kontrolü
    if reset_password != RESET_PASSWORD:
        flash('❌ Hatalı sistem sıfırlama şifresi!', 'error')
        return render_template('reset_system.html', show_stats=False)
    
    # İstatistikleri göster
    if action == 'check':
        try:
            stats = {
                'kullanici_sayisi': Kullanici.query.count(),
                'otel_sayisi': Otel.query.count(),
                'kat_sayisi': Kat.query.count(),
                'oda_sayisi': Oda.query.count(),
                'urun_grubu_sayisi': UrunGrup.query.count(),
                'urun_sayisi': Urun.query.count(),
                'stok_hareket_sayisi': StokHareket.query.count(),
                'zimmet_sayisi': PersonelZimmet.query.count(),
                'zimmet_detay_sayisi': PersonelZimmetDetay.query.count(),
                'minibar_islem_sayisi': MinibarIslem.query.count(),
                'minibar_detay_sayisi': MinibarIslemDetay.query.count(),
                'log_sayisi': SistemLog.query.count(),
                'hata_sayisi': HataLog.query.count(),
                'audit_sayisi': db.session.execute(db.text("SELECT COUNT(*) FROM audit_logs")).scalar() or 0
            }
            
            return render_template('reset_system.html', 
                                 show_stats=True, 
                                 stats=stats,
                                 password=reset_password)
        
        except Exception as e:
            flash(f'❌ İstatistikler alınırken hata: {str(e)}', 'error')
            return render_template('reset_system.html', show_stats=False)
    
    # Sistem sıfırlama işlemi
    elif action == 'reset':
        # Onay checkbox kontrolü
        if not request.form.get('confirm_reset'):
            flash('❌ Sıfırlama onayı verilmedi!', 'error')
            return redirect(url_for('reset_system'))
        
        try:
            # Tüm tabloları temizle (sıralama önemli - foreign key kısıtları)
            print("\n" + "="*60)
            print("🔴 SİSTEM SIFIRLAMA BAŞLADI")
            print("="*60)
            
            # 1. MinibarIslemDetay (foreign key: minibar_islemleri)
            count = db.session.execute(db.text("DELETE FROM minibar_islem_detay")).rowcount
            print(f"✓ MinibarIslemDetay silindi: {count} kayıt")
            
            # 2. MinibarIslem
            count = db.session.execute(db.text("DELETE FROM minibar_islemleri")).rowcount
            print(f"✓ MinibarIslem silindi: {count} kayıt")
            
            # 3. PersonelZimmetDetay (foreign key: personel_zimmet)
            count = db.session.execute(db.text("DELETE FROM personel_zimmet_detay")).rowcount
            print(f"✓ PersonelZimmetDetay silindi: {count} kayıt")
            
            # 4. PersonelZimmet
            count = db.session.execute(db.text("DELETE FROM personel_zimmet")).rowcount
            print(f"✓ PersonelZimmet silindi: {count} kayıt")
            
            # 5. StokHareket
            count = db.session.execute(db.text("DELETE FROM stok_hareketleri")).rowcount
            print(f"✓ StokHareket silindi: {count} kayıt")
            
            # 6. Urun (foreign key: urun_gruplari)
            count = db.session.execute(db.text("DELETE FROM urunler")).rowcount
            print(f"✓ Urun silindi: {count} kayıt")
            
            # 7. UrunGrup
            count = db.session.execute(db.text("DELETE FROM urun_gruplari")).rowcount
            print(f"✓ UrunGrup silindi: {count} kayıt")
            
            # 8. Oda (foreign key: katlar)
            count = db.session.execute(db.text("DELETE FROM odalar")).rowcount
            print(f"✓ Oda silindi: {count} kayıt")
            
            # 9. Kat (foreign key: oteller)
            count = db.session.execute(db.text("DELETE FROM katlar")).rowcount
            print(f"✓ Kat silindi: {count} kayıt")
            
            # 10. LOG VE AUDIT TABLOLARI ÖNCE SİLİNMELİ (foreign key: kullanicilar)
            # SistemLog
            count = db.session.execute(db.text("DELETE FROM sistem_loglari")).rowcount
            print(f"✓ SistemLog silindi: {count} kayıt")
            
            # HataLog
            count = db.session.execute(db.text("DELETE FROM hata_loglari")).rowcount
            print(f"✓ HataLog silindi: {count} kayıt")
            
            # AuditLog (kullanıcılara foreign key var!)
            count = db.session.execute(db.text("DELETE FROM audit_logs")).rowcount
            print(f"✓ AuditLog silindi: {count} kayıt")
            
            # 11. OtomatikRapor (kullanıcılara foreign key olabilir)
            count = db.session.execute(db.text("DELETE FROM otomatik_raporlar")).rowcount
            print(f"✓ OtomatikRapor silindi: {count} kayıt")
            
            # 12. ARTIK KULLANICILARı SİLEBİLİRİZ
            # Kullanici (foreign key: oteller)
            count = db.session.execute(db.text("DELETE FROM kullanicilar")).rowcount
            print(f"✓ Kullanici silindi: {count} kayıt")
            
            # 13. Otel
            count = db.session.execute(db.text("DELETE FROM oteller")).rowcount
            print(f"✓ Otel silindi: {count} kayıt")
            
            # 14. SistemAyar - setup_tamamlandi'yi sıfırla
            db.session.execute(db.text("DELETE FROM sistem_ayarlari WHERE anahtar = 'setup_tamamlandi'"))
            print(f"✓ Setup ayarı sıfırlandı")
            
            # Auto-increment değerlerini sıfırla
            tables = [
                'minibar_islem_detay', 'minibar_islemleri',
                'personel_zimmet_detay', 'personel_zimmet',
                'stok_hareketleri', 'urunler', 'urun_gruplari',
                'odalar', 'katlar', 'kullanicilar', 'oteller',
                'sistem_loglari', 'hata_loglari', 'audit_logs', 'otomatik_raporlar'
            ]
            
            for table in tables:
                try:
                    db.session.execute(db.text(f"ALTER TABLE {table} AUTO_INCREMENT = 1"))
                except:
                    pass  # Bazı tablolar primary key olmayabilir
            
            print(f"✓ Auto-increment değerleri sıfırlandı")
            
            # Commit
            db.session.commit()
            
            print("="*60)
            print("✅ SİSTEM SIFIRLAMA TAMAMLANDI")
            print("="*60)
            print()
            
            # Session'ı temizle
            session.clear()
            
            # Başarı mesajı ve yönlendirme
            flash('✅ Sistem başarıyla sıfırlandı! Tüm veriler silindi ve sistem ilk kurulum aşamasına döndü.', 'success')
            flash('🔄 Şimdi ilk kurulum sayfasına yönlendiriliyorsunuz...', 'info')
            
            return redirect(url_for('setup'))
        
        except Exception as e:
            db.session.rollback()
            print(f"\n❌ HATA: {str(e)}\n")
            flash(f'❌ Sistem sıfırlanırken hata oluştu: {str(e)}', 'error')
            return redirect(url_for('reset_system'))
    
    # Geçersiz action
    flash('❌ Geçersiz işlem!', 'error')
    return redirect(url_for('reset_system'))



# ============================================================================
# SYSTEM BACKUP - SUPER ADMIN ENDPOINT (GİZLİ)
# ============================================================================

@app.route('/systembackupsuperadmin', methods=['GET', 'POST'])
def system_backup_login():
    """Gizli super admin backup login sayfası - Sadece şifre ile giriş - LOGLANMAZ"""
    if request.method == 'POST':
        access_code = request.form.get('access_code', '').strip()
        
        # Sabit şifre kontrolü
        if access_code == '518518Erkan':
            session['super_admin_logged_in'] = True
            session['super_admin_login_time'] = get_kktc_now().isoformat()
            # LOG KAYDEDILMEZ - Gizli erişim
            return redirect(url_for('system_backup_panel'))
        else:
            flash('❌ Invalid access code!', 'error')
    
    return render_template('super_admin_login.html')


@app.route('/systembackupsuperadmin/panel')
def system_backup_panel():
    """Super admin backup panel - istatistikler ve backup özellikleri"""
    # Super admin kontrolü
    if not session.get('super_admin_logged_in'):
        return redirect(url_for('system_backup_login'))
    
    from models import (
        Otel, Kat, Oda, UrunGrup, Urun, Kullanici, 
        StokHareket, MinibarIslem, MinibarIslemDetay, PersonelZimmet, PersonelZimmetDetay
    )
    
    try:
        # Veritabanı istatistiklerini topla
        stats = {
            'otel_count': Otel.query.count(),
            'kat_count': Kat.query.count(),
            'oda_count': Oda.query.count(),
            'urun_grup_count': UrunGrup.query.count(),
            'urun_count': Urun.query.count(),
            'kullanici_count': Kullanici.query.count(),
            'stok_hareket_count': StokHareket.query.count(),
            'minibar_kontrol_count': MinibarIslem.query.count(),
            'database_name': app.config['SQLALCHEMY_DATABASE_URI'].split('/')[-1].split('?')[0],
            'current_time': get_kktc_now().strftime('%d.%m.%Y %H:%M:%S'),
            'last_backup': session.get('last_backup_time'),
        }
        
        # Tablo detayları
        stats['table_details'] = {
            'oteller': stats['otel_count'],
            'katlar': stats['kat_count'],
            'odalar': stats['oda_count'],
            'urun_gruplari': stats['urun_grup_count'],
            'urunler': stats['urun_count'],
            'kullanicilar': stats['kullanici_count'],
            'stok_hareketleri': stats['stok_hareket_count'],
            'minibar_islemleri': stats['minibar_kontrol_count'],
            'minibar_islem_detaylari': MinibarIslemDetay.query.count(),
            'personel_zimmetleri': PersonelZimmet.query.count(),
            'personel_zimmet_detaylari': PersonelZimmetDetay.query.count(),
        }
        
        stats['table_count'] = len(stats['table_details'])
        stats['total_records'] = sum(stats['table_details'].values())
        
        return render_template('system_backup.html', stats=stats)
        
    except Exception as e:
        flash(f'❌ İstatistikler yüklenirken hata: {str(e)}', 'error')
        return redirect(url_for('system_backup_login'))


@app.route('/systembackupsuperadmin/download', methods=['POST'])
def system_backup_download():
    """SQL backup dosyasını indir - Python ile direkt export"""
    # Super admin kontrolü
    if not session.get('super_admin_logged_in'):
        return redirect(url_for('system_backup_login'))
    
    backup_type = request.form.get('backup_type', 'full')
    
    try:
        from io import StringIO
        
        # SQL dump içeriği
        sql_dump = StringIO()
        
        # Header
        timestamp = get_kktc_now().strftime('%Y-%m-%d %H:%M:%S')
        sql_dump.write("-- Minibar Takip Sistemi Database Backup\n")
        sql_dump.write(f"-- Backup Date: {timestamp}\n")
        sql_dump.write(f"-- Backup Type: {backup_type}\n")
        sql_dump.write("-- Generated by: Super Admin Panel\n\n")
        sql_dump.write("SET FOREIGN_KEY_CHECKS=0;\n")
        sql_dump.write("SET SQL_MODE='NO_AUTO_VALUE_ON_ZERO';\n")
        sql_dump.write("SET NAMES utf8mb4;\n\n")
        
        # Tüm tabloları al
        from sqlalchemy import text
        tables_query = text("SHOW TABLES")
        tables = db.session.execute(tables_query).fetchall()
        
        for table_tuple in tables:
            table_name = table_tuple[0]
            sql_dump.write(f"-- Table: {table_name}\n")
            
            if backup_type == 'full':
                # Tablo yapısı
                create_query = text(f"SHOW CREATE TABLE `{table_name}`")
                create_result = db.session.execute(create_query).fetchone()
                sql_dump.write(f"DROP TABLE IF EXISTS `{table_name}`;\n")
                sql_dump.write(f"{create_result[1]};\n\n")
            
            # Veri export
            select_query = text(f"SELECT * FROM `{table_name}`")
            rows = db.session.execute(select_query).fetchall()
            
            if rows:
                # Column isimleri
                columns_query = text(f"SHOW COLUMNS FROM `{table_name}`")
                columns = db.session.execute(columns_query).fetchall()
                column_names = [col[0] for col in columns]
                
                sql_dump.write(f"INSERT INTO `{table_name}` (`{'`, `'.join(column_names)}`) VALUES\n")
                
                for i, row in enumerate(rows):
                    values = []
                    for val in row:
                        if val is None:
                            values.append('NULL')
                        elif isinstance(val, (int, float)):
                            values.append(str(val))
                        elif isinstance(val, datetime):
                            values.append(f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'")
                        else:
                            # String escape
                            escaped = str(val).replace("\\", "\\\\").replace("'", "\\'")
                            values.append(f"'{escaped}'")
                    
                    comma = ',' if i < len(rows) - 1 else ';'
                    sql_dump.write(f"({', '.join(values)}){comma}\n")
                
                sql_dump.write("\n")
        
        sql_dump.write("SET FOREIGN_KEY_CHECKS=1;\n")
        
        # StringIO'yu bytes'a çevir
        sql_content = sql_dump.getvalue()
        sql_bytes = io.BytesIO(sql_content.encode('utf-8'))
        
        # Son backup zamanını kaydet
        session['last_backup_time'] = get_kktc_now().strftime('%d.%m.%Y %H:%M:%S')
        
        # Dosya adı
        filename = f'minibar_backup_{backup_type}_{get_kktc_now().strftime("%Y%m%d_%H%M%S")}.sql'
        
        # Dosyayı gönder
        return send_file(
            sql_bytes,
            as_attachment=True,
            download_name=filename,
            mimetype='application/sql'
        )
        
    except Exception as e:
        flash(f'❌ Backup oluşturulurken hata: {str(e)}', 'error')
        # LOG KAYDEDILMEZ - Gizli operasyon
        return redirect(url_for('system_backup_panel'))

def init_database():
    """Veritabanı ve tabloları otomatik kontrol et - GÜVENLİ MOD"""
    try:
        with app.app_context():
            # Sadece bağlantı testi yap, tablo oluşturma!
            # Production'da mevcut verilere dokunmamak için
            inspector = inspect(db.engine)
            existing_tables = inspector.get_table_names()
            
            if existing_tables:
                print(f"✅ Veritabanı bağlantısı başarılı - {len(existing_tables)} tablo mevcut")
                return True
            else:
                print("⚠️  Henüz tablo yok!")
                print("🔧 Lütfen 'python init_db.py' komutunu çalıştırın.")
                return False
    except Exception as e:
        print(f"❌ Veritabanı hatası: {e}")
        print()
        print("🔧 Lütfen 'python init_db.py' komutunu çalıştırın.")
        return False


# ============================================
# KAT SORUMLUSU STOK YÖNETİMİ ROTALARI
# ============================================

@app.route('/kat-sorumlusu/zimmet-stoklarim')
@login_required
@role_required('kat_sorumlusu')
def kat_sorumlusu_zimmet_stoklarim():
    """Zimmet stok listesi sayfası"""
    from utils.helpers import get_kat_sorumlusu_zimmet_stoklari
    
    try:
        kullanici_id = session['kullanici_id']
        
        # Zimmet stoklarını getir
        zimmet_stoklari = get_kat_sorumlusu_zimmet_stoklari(kullanici_id)
        
        # Log kaydı
        log_islem('goruntuleme', 'zimmet_stoklari', {
            'kullanici_id': kullanici_id,
            'zimmet_sayisi': len(zimmet_stoklari)
        })
        
        return render_template('kat_sorumlusu/zimmet_stoklarim.html',
                             zimmet_stoklari=zimmet_stoklari)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok')
        flash('Zimmet stokları yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('kat_sorumlusu_dashboard'))


@app.route('/kat-sorumlusu/kritik-stoklar')
@login_required
@role_required('kat_sorumlusu')
def kat_sorumlusu_kritik_stoklar():
    """Kritik stoklar sayfası"""
    from utils.helpers import get_kat_sorumlusu_kritik_stoklar
    
    try:
        kullanici_id = session['kullanici_id']
        
        # Kritik stokları getir
        kritik_stoklar = get_kat_sorumlusu_kritik_stoklar(kullanici_id)
        
        # Log kaydı
        log_islem('goruntuleme', 'kritik_stoklar', {
            'kullanici_id': kullanici_id,
            'stokout_sayisi': kritik_stoklar['istatistik']['stokout_sayisi'],
            'kritik_sayisi': kritik_stoklar['istatistik']['kritik_sayisi']
        })
        
        return render_template('kat_sorumlusu/kritik_stoklar.html',
                             kritik_stoklar=kritik_stoklar)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok')
        flash('Kritik stoklar yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('kat_sorumlusu_dashboard'))


@app.route('/kat-sorumlusu/siparis-hazirla', methods=['GET', 'POST'])
@login_required
@role_required('kat_sorumlusu')
def kat_sorumlusu_siparis_hazirla():
    """Sipariş hazırlama sayfası"""
    from utils.helpers import olustur_otomatik_siparis, kaydet_siparis_talebi
    
    try:
        kullanici_id = session['kullanici_id']
        
        if request.method == 'POST':
            # Sipariş listesini al
            siparis_data = request.get_json()
            siparis_listesi = siparis_data.get('siparis_listesi', [])
            aciklama = siparis_data.get('aciklama', '')
            
            # Sipariş talebini kaydet
            sonuc = kaydet_siparis_talebi(kullanici_id, siparis_listesi, aciklama)
            
            return jsonify(sonuc)
        
        # GET request - Otomatik sipariş listesi oluştur
        siparis_bilgileri = olustur_otomatik_siparis(kullanici_id)
        
        # Log kaydı
        log_islem('goruntuleme', 'siparis_hazirla', {
            'kullanici_id': kullanici_id,
            'urun_sayisi': siparis_bilgileri['toplam_urun_sayisi']
        })
        
        return render_template('kat_sorumlusu/siparis_hazirla.html',
                             siparis_bilgileri=siparis_bilgileri)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok')
        if request.method == 'POST':
            return jsonify({'success': False, 'message': 'Bir hata oluştu'}), 500
        flash('Sipariş hazırlama sayfası yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('kat_sorumlusu_dashboard'))


@app.route('/kat-sorumlusu/urun-gecmisi/<int:urun_id>')
@login_required
@role_required('kat_sorumlusu')
def kat_sorumlusu_urun_gecmisi(urun_id):
    """Ürün kullanım geçmişi sayfası"""
    from utils.helpers import get_zimmet_urun_gecmisi
    
    try:
        kullanici_id = session['kullanici_id']
        
        # Tarih filtresi
        gun_sayisi = request.args.get('gun_sayisi', 30, type=int)
        
        # Ürün geçmişini getir
        gecmis = get_zimmet_urun_gecmisi(kullanici_id, urun_id, gun_sayisi)
        
        if not gecmis:
            flash('Ürün bulunamadı.', 'danger')
            return redirect(url_for('kat_sorumlusu_zimmet_stoklarim'))
        
        # Log kaydı
        log_islem('goruntuleme', 'urun_gecmisi', {
            'kullanici_id': kullanici_id,
            'urun_id': urun_id,
            'gun_sayisi': gun_sayisi
        })
        
        return render_template('kat_sorumlusu/urun_gecmisi.html',
                             gecmis=gecmis,
                             gun_sayisi=gun_sayisi)
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok', extra_info={'urun_id': urun_id})
        flash('Ürün geçmişi yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('kat_sorumlusu_zimmet_stoklarim'))


@app.route('/kat-sorumlusu/zimmet-export')
@login_required
@role_required('kat_sorumlusu')
def kat_sorumlusu_zimmet_export():
    """Zimmet stoklarını Excel'e export et"""
    from utils.helpers import export_zimmet_stok_excel
    
    try:
        kullanici_id = session['kullanici_id']
        
        # Excel dosyasını oluştur
        excel_buffer = export_zimmet_stok_excel(kullanici_id)
        
        if not excel_buffer:
            flash('Excel dosyası oluşturulamadı.', 'danger')
            return redirect(url_for('kat_sorumlusu_zimmet_stoklarim'))
        
        # Log kaydı
        log_islem('export', 'zimmet_stoklari', {
            'kullanici_id': kullanici_id,
            'format': 'excel'
        })
        
        filename = f'zimmet_stoklari_{get_kktc_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok')
        flash('Excel export işlemi başarısız oldu.', 'danger')
        return redirect(url_for('kat_sorumlusu_zimmet_stoklarim'))


@app.route('/api/kat-sorumlusu/kritik-seviye-guncelle', methods=['POST'])
@login_required
@role_required('kat_sorumlusu')
def api_kat_sorumlusu_kritik_seviye_guncelle():
    """AJAX - Kritik seviye güncelleme"""
    from utils.helpers import guncelle_kritik_seviye
    
    try:
        data = request.get_json()
        zimmet_detay_id = data.get('zimmet_detay_id')
        kritik_seviye = data.get('kritik_seviye')
        
        if not zimmet_detay_id or not kritik_seviye:
            return jsonify({
                'success': False,
                'message': 'Eksik parametreler'
            }), 400
        
        # Kritik seviyeyi güncelle
        sonuc = guncelle_kritik_seviye(zimmet_detay_id, int(kritik_seviye))
        
        if sonuc['success']:
            return jsonify(sonuc)
        else:
            return jsonify(sonuc), 400
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok')
        return jsonify({
            'success': False,
            'message': 'Bir hata oluştu'
        }), 500


@app.route('/api/kat-sorumlusu/siparis-kaydet', methods=['POST'])
@login_required
@role_required('kat_sorumlusu')
def api_kat_sorumlusu_siparis_kaydet():
    """AJAX - Sipariş talebini kaydet"""
    from utils.helpers import kaydet_siparis_talebi
    
    try:
        data = request.get_json()
        siparis_listesi = data.get('siparis_listesi', [])
        aciklama = data.get('aciklama', '')
        
        # Validasyon
        if not siparis_listesi or len(siparis_listesi) == 0:
            return jsonify({
                'success': False,
                'message': 'Sipariş listesi boş olamaz'
            }), 400
        
        # Sipariş talebini kaydet
        personel_id = session.get('kullanici_id')
        sonuc = kaydet_siparis_talebi(personel_id, siparis_listesi, aciklama)
        
        if sonuc['success']:
            # Log kaydı
            log_islem('ekleme', 'siparis_talebi', {
                'personel_id': personel_id,
                'urun_sayisi': len(siparis_listesi),
                'aciklama': aciklama
            })
            return jsonify(sonuc)
        else:
            return jsonify(sonuc), 400
        
    except Exception as e:
        log_hata(e, modul='kat_sorumlusu_stok', extra_info={
            'function': 'api_kat_sorumlusu_siparis_kaydet',
            'personel_id': session.get('kullanici_id')
        })
        return jsonify({
            'success': False,
            'message': 'Sipariş kaydedilirken bir hata oluştu'
        }), 500


# ============================================
# ODA KONTROL VE YENİDEN DOLUM ROTALARI
# ============================================

# İlk dolum route'u kaldırıldı - Setup bazlı sistem kullanılıyor
# Yeni route: /minibar-kontrol-setup

@app.route('/api/kat-sorumlusu/minibar-urunler', methods=['POST'])
@login_required
@role_required('kat_sorumlusu')
def api_minibar_urunler():
    """Odanın minibar ürünlerini getir - son işlemdeki stok durumu"""
    try:
        data = request.get_json()
        oda_id = data.get('oda_id')
        
        if not oda_id:
            return jsonify({'success': False, 'message': 'Oda ID gerekli'}), 400
        
        # Oda bilgilerini getir
        oda = Oda.query.get(oda_id)
        if not oda:
            return jsonify({'success': False, 'message': 'Oda bulunamadı'}), 404
        
        # Kullanıcının bu otele erişimi var mı kontrol et
        kullanici = get_current_user()
        if not kullanici or kullanici.otel_id != oda.kat.otel_id:
            return jsonify({'success': False, 'message': 'Bu odaya erişim yetkiniz yok'}), 403
        
        # Son minibar işlemini bul
        son_islem = MinibarIslem.query.filter_by(
            oda_id=oda_id
        ).order_by(MinibarIslem.islem_tarihi.desc()).first()
        
        urunler = []
        
        if son_islem:
            # Son işlemdeki ürünleri getir
            detaylar = MinibarIslemDetay.query.filter_by(
                islem_id=son_islem.id
            ).join(Urun).all()
            
            for detay in detaylar:
                urunler.append({
                    'id': detay.id,
                    'urun_id': detay.urun_id,
                    'urun_adi': detay.urun.urun_adi,
                    'mevcut_miktar': detay.bitis_stok,
                    'birim': detay.urun.birim,
                    'grup_adi': detay.urun.grup.grup_adi if detay.urun.grup else '-'
                })
        
        return jsonify({
            'success': True,
            'data': {
                'oda_no': oda.oda_no,
                'kat_adi': oda.kat.kat_adi,
                'urunler': urunler
            }
        })
        
    except Exception as e:
        log_hata(e, modul='api_minibar_urunler')
        return jsonify({'success': False, 'message': 'Ürünler yüklenirken hata oluştu'}), 500


@app.route('/api/kat-sorumlusu/yeniden-dolum', methods=['POST'])
@login_required
@role_required('kat_sorumlusu')
def api_yeniden_dolum():
    """Yeniden dolum işlemi"""
    try:
        data = request.get_json() or {}
        oda_id = data.get('oda_id')
        urun_id = data.get('urun_id')
        eklenecek_miktar = data.get('eklenecek_miktar')
        
        # Validasyon
        if not all([oda_id, urun_id, eklenecek_miktar]):
            return jsonify({
                'success': False,
                'message': 'Eksik parametre'
            }), 400
        
        # Miktar kontrolü
        try:
            eklenecek_miktar = float(eklenecek_miktar)
            if eklenecek_miktar <= 0:
                return jsonify({
                    'success': False,
                    'message': 'Lütfen geçerli bir miktar giriniz'
                }), 400
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'message': 'Geçersiz miktar formatı'
            }), 400
        
        # Oda ve ürün kontrolü
        oda = db.session.get(Oda, oda_id)
        urun = db.session.get(Urun, urun_id)
        
        if not oda:
            return jsonify({
                'success': False,
                'message': 'Oda bulunamadı'
            }), 404
        
        # Kat sorumlusunun otel kontrolü
        kullanici = get_current_user()
        if not kullanici or not kullanici.otel_id:
            return jsonify({
                'success': False,
                'message': 'Henüz bir otele atanmadınız'
            }), 403
        
        # Odanın katını kontrol et ve otele ait olup olmadığını doğrula
        if oda.kat.otel_id != kullanici.otel_id:
            return jsonify({
                'success': False,
                'message': 'Bu oda sizin atandığınız otele ait değil'
            }), 403
        
        if not urun:
            return jsonify({
                'success': False,
                'message': 'Ürün bulunamadı'
            }), 404
        
        # Kullanıcının zimmet stoğunu kontrol et
        kullanici_id = session['kullanici_id']
        
        # Aktif zimmetlerde bu ürünü ara
        zimmet_detaylar = db.session.query(PersonelZimmetDetay).join(
            PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id
        ).filter(
            PersonelZimmet.personel_id == kullanici_id,
            PersonelZimmet.durum == 'aktif',
            PersonelZimmetDetay.urun_id == urun_id
        ).all()
        
        if not zimmet_detaylar:
            return jsonify({
                'success': False,
                'message': f'Stoğunuzda {urun.urun_adi} bulunmamaktadır'
            }), 422
        
        # Toplam kalan miktarı hesapla
        toplam_kalan = sum(detay.kalan_miktar or 0 for detay in zimmet_detaylar)
        
        if toplam_kalan < eklenecek_miktar:
            return jsonify({
                'success': False,
                'message': f'Stoğunuzda yeterli {urun.urun_adi} bulunmamaktadır. Mevcut: {toplam_kalan}, İstenen: {eklenecek_miktar}'
            }), 422
        
        # Odanın son minibar işlemini bul
        son_islem = MinibarIslem.query.filter_by(
            oda_id=oda_id
        ).order_by(MinibarIslem.id.desc()).first()
        
        # Mevcut stok bilgisini al (şu anki gerçek miktar)
        mevcut_stok = 0
        baslangic_stok_ilk_dolum = 0
        if son_islem:
            for detay in son_islem.detaylar:
                if detay.urun_id == urun_id:
                    mevcut_stok = detay.bitis_stok if detay.bitis_stok is not None else 0
                    # İlk dolumdan bu yana toplam ne kadar eklendi
                    if son_islem.islem_tipi == 'ilk_dolum':
                        baslangic_stok_ilk_dolum = detay.bitis_stok
                    break
        
        # Transaction başlat
        try:
            # Zimmetlerden düş (FIFO mantığı)
            kalan_miktar = eklenecek_miktar
            for zimmet_detay in zimmet_detaylar:
                if kalan_miktar <= 0:
                    break
                
                detay_kalan = zimmet_detay.kalan_miktar or 0
                if detay_kalan > 0:
                    kullanilacak = min(detay_kalan, kalan_miktar)
                    zimmet_detay.kullanilan_miktar = (zimmet_detay.kullanilan_miktar or 0) + kullanilacak
                    zimmet_detay.kalan_miktar = (zimmet_detay.miktar or 0) - zimmet_detay.kullanilan_miktar
                    kalan_miktar -= kullanilacak
            
            # Yeni minibar işlemi oluştur
            yeni_islem = MinibarIslem(
                oda_id=oda_id,
                personel_id=kullanici_id,
                islem_tipi='yeniden_dolum',  # Veritabanındaki enum değeri
                aciklama=f'Yeniden dolum: {urun.urun_adi}'
            )
            db.session.add(yeni_islem)
            db.session.flush()  # ID'yi almak için
            
            # DOĞRU MANTIK:
            # Mevcut stok (DB'de kayıtlı) = 3 şişe
            # Eklenecek miktar (kat sorumlusu giriyor) = 1 şişe
            # Şu anki gerçek miktar = Mevcut - Eklenecek = 3 - 1 = 2 şişe (misafir 1 içmiş)
            # Yeni stok = Mevcut (değişmez, tekrar başlangıca dönüyor) = 3 şişe
            # Tüketim = Eklenecek miktar = 1 şişe
            
            su_anki_gercek_miktar = mevcut_stok - eklenecek_miktar
            yeni_stok = mevcut_stok  # Toplam değişmez, başlangıca dönüyor
            tuketim_miktari = eklenecek_miktar  # Eklenen miktar kadar tüketim olmuştur
            
            # İşlem detayı oluştur
            islem_detay = MinibarIslemDetay(
                islem_id=yeni_islem.id,
                urun_id=urun_id,
                baslangic_stok=su_anki_gercek_miktar,  # Şu anki gerçek miktar (dolum öncesi)
                bitis_stok=yeni_stok,  # Dolum sonrası miktar (başlangıca dönüyor)
                eklenen_miktar=eklenecek_miktar,
                tuketim=tuketim_miktari,  # Tüketim kaydı oluştur
                zimmet_detay_id=zimmet_detaylar[0].id if zimmet_detaylar else None
            )
            db.session.add(islem_detay)
            
            # Commit
            db.session.commit()
            
            # Audit Trail
            audit_create('minibar_islem', yeni_islem.id, yeni_islem)
            
            # Log kaydı
            log_islem('ekleme', 'yeniden_dolum', {
                'oda_id': oda_id,
                'oda_no': oda.oda_no,
                'urun_id': urun_id,
                'urun_adi': urun.urun_adi,
                'eklenecek_miktar': eklenecek_miktar,
                'yeni_stok': yeni_stok
            })
            
            # Kalan zimmet miktarını hesapla
            kalan_zimmet = sum(detay.kalan_miktar or 0 for detay in zimmet_detaylar)
            
            return jsonify({
                'success': True,
                'message': 'Dolum işlemi başarıyla tamamlandı',
                'data': {
                    'yeni_miktar': yeni_stok,
                    'kalan_zimmet': kalan_zimmet
                }
            })
            
        except Exception as e:
            db.session.rollback()
            raise
        
    except ValueError as e:
        log_hata(e, modul='yeniden_dolum', extra_info={'oda_id': oda_id, 'urun_id': urun_id})
        return jsonify({
            'success': False,
            'message': str(e)
        }), 400
    except Exception as e:
        log_hata(e, modul='yeniden_dolum')
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': 'İşlem sırasında bir hata oluştu. Lütfen tekrar deneyiniz'
        }), 500


# ============================================
# QR KOD SİSTEMİ ROTALARI
# ============================================

# ============================================
# Tüm route'lar merkezi olarak register edildi (routes/__init__.py)
# ============================================

# ============================================
# SCHEDULER - Otomatik Görevler
# ============================================
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from utils.file_management_service import FileManagementService

def start_scheduler():
    """Zamanlanmış görevleri başlat"""
    # Debug modunda sadece child process'te (gerçek uygulama) çalıştır
    # WERKZEUG_RUN_MAIN='true' sadece child process'te set edilir
    is_reloader_process = os.environ.get('WERKZEUG_RUN_MAIN') is None
    if is_reloader_process and app.debug:
        return  # Ana reloader process'inde scheduler başlatma
    
    scheduler = BackgroundScheduler()
    
    # Her gün saat 02:00'de eski dosyaları temizle
    scheduler.add_job(
        func=lambda: FileManagementService.cleanup_old_files(),
        trigger=CronTrigger(hour=2, minute=0),
        id='cleanup_old_files',
        name='Eski doluluk dosyalarını temizle',
        replace_existing=True
    )
    
    # ML SYSTEM JOBS - Sadece ML_ENABLED=true ise çalışır
    ml_enabled = os.getenv('ML_ENABLED', 'false').lower() == 'true'
    
    if ml_enabled:
        # Sabah 08:00 - Akşam 20:00 arası her saat başı veri toplama
        scheduler.add_job(
            func=lambda: collect_ml_data(),
            trigger='cron',
            hour='8-20',  # 08:00, 09:00, ..., 20:00
            minute=0,
            id='ml_data_collection',
            name='ML Veri Toplama',
            replace_existing=True
        )
        
        # Sabah 08:00 - Akşam 20:00 arası her saat başı anomali tespiti (30. dakikada)
        scheduler.add_job(
            func=lambda: detect_anomalies(),
            trigger='cron',
            hour='8-20',  # 08:00, 09:00, ..., 20:00
            minute=30,    # Veri toplamadan 30 dk sonra
            id='ml_anomaly_detection',
            name='ML Anomali Tespiti',
            replace_existing=True
        )
        
        # Her gece yarısı model eğitimi
        ml_training_schedule = os.getenv('ML_TRAINING_SCHEDULE', '0 0 * * *')  # Cron format
        scheduler.add_job(
            func=lambda: train_ml_models(),
            trigger=CronTrigger.from_crontab(ml_training_schedule),
            id='ml_model_training',
            name='ML Model Eğitimi',
            replace_existing=True
        )
        
        # Günde 2 kez stok bitiş kontrolü (sabah 9 ve akşam 6)
        scheduler.add_job(
            func=lambda: check_stock_depletion(),
            trigger='cron',
            hour='9,18',
            id='ml_stock_depletion_check',
            name='ML Stok Bitiş Kontrolü',
            replace_existing=True
        )
        
        # Her gece 03:00'te eski alertleri temizle
        scheduler.add_job(
            func=lambda: cleanup_old_alerts(),
            trigger='cron',
            hour=3,
            minute=0,
            id='ml_alert_cleanup',
            name='ML Alert Temizleme',
            replace_existing=True
        )
        
        # Her gece 04:00'te eski model versiyonlarını temizle
        scheduler.add_job(
            func=lambda: cleanup_old_models(),
            trigger='cron',
            hour=4,
            minute=0,
            id='ml_model_cleanup',
            name='ML Model Cleanup',
            replace_existing=True
        )
        
        print("✅ ML Scheduler başlatıldı")
        print("   - Veri toplama: 08:00-20:00 arası her saat başı")
        print("   - Anomali tespiti: 08:30-20:30 arası her saat")
        print(f"   - Model eğitimi: {ml_training_schedule}")
        print("   - Stok bitiş kontrolü: Günde 2 kez (09:00, 18:00)")
        print("   - Alert temizleme: Her gece 03:00")
        print("   - Model cleanup: Her gece 04:00")
    
    scheduler.start()
    print("✅ Scheduler başlatıldı (Günlük dosya temizleme: 02:00)")

def collect_ml_data():
    """ML veri toplama job'u"""
    try:
        from utils.ml.data_collector import DataCollector
        with app.app_context():
            collector = DataCollector(db)
            collector.collect_all_metrics()
            # Eski metrikleri temizle (90 günden eski)
            collector.cleanup_old_metrics(days=90)
    except Exception as e:
        logger.error(f"❌ ML veri toplama hatası: {str(e)}")

def detect_anomalies():
    """ML anomali tespiti job'u"""
    try:
        from utils.ml.anomaly_detector import AnomalyDetector
        with app.app_context():
            detector = AnomalyDetector(db)
            detector.detect_all_anomalies()
    except Exception as e:
        logger.error(f"❌ ML anomali tespiti hatası: {str(e)}")

def train_ml_models():
    """ML model eğitimi job'u"""
    try:
        from utils.ml.model_trainer import ModelTrainer
        with app.app_context():
            trainer = ModelTrainer(db)
            trainer.train_all_models()
    except Exception as e:
        logger.error(f"❌ ML model eğitimi hatası: {str(e)}")

def check_stock_depletion():
    """Stok bitiş kontrolü job'u"""
    try:
        from utils.ml.metrics_calculator import MetricsCalculator
        with app.app_context():
            calculator = MetricsCalculator(db)
            calculator.check_stock_depletion_alerts()
    except Exception as e:
        logger.error(f"❌ Stok bitiş kontrolü hatası: {str(e)}")

def cleanup_old_alerts():
    """Eski alertleri temizle job'u"""
    try:
        from utils.ml.alert_manager import AlertManager
        with app.app_context():
            alert_manager = AlertManager(db)
            alert_manager.cleanup_old_alerts(days=90)
    except Exception as e:
        logger.error(f"❌ Alert temizleme hatası: {str(e)}")

def cleanup_old_models():
    """Eski model versiyonlarını temizle job'u"""
    try:
        from utils.ml.model_manager import ModelManager
        with app.app_context():
            model_manager = ModelManager(db)
            
            # Eski model versiyonlarını temizle (son 3 versiyon sakla)
            result = model_manager.cleanup_old_models(keep_versions=3)
            
            # Disk kullanımını kontrol et
            disk_info = model_manager._check_disk_space()
            
            # Disk kullanımı %90'ı geçtiyse alert oluştur
            if disk_info['percent'] > 90:
                logger.warning(
                    f"⚠️  DISK KULLANIMI YÜKSEK: {disk_info['percent']:.1f}% "
                    f"({disk_info['used_gb']:.2f}GB / {disk_info['total_gb']:.2f}GB)"
                )
                
                # ML Alert oluştur
                from models import MLAlert
                from datetime import datetime, timezone
                
                alert = MLAlert(
                    alert_type='stok_anomali',  # En yakın tip
                    severity='kritik',
                    entity_id=0,  # Sistem seviyesi
                    metric_value=disk_info['percent'],
                    expected_value=80.0,
                    deviation_percent=(disk_info['percent'] - 80.0) / 80.0 * 100,
                    message=f"ML model dizini disk kullanımı kritik seviyede: {disk_info['percent']:.1f}%",
                    suggested_action="Eski model dosyalarını manuel olarak temizleyin veya disk alanını artırın",
                    created_at=get_kktc_now()
                )
                db.session.add(alert)
                db.session.commit()
            
            logger.info(
                f"✅ Model cleanup tamamlandı: "
                f"{result['deleted_count']} model silindi, "
                f"{result['freed_space_mb']:.2f}MB alan boşaltıldı"
            )
            
    except Exception as e:
        logger.error(f"❌ Model cleanup hatası: {str(e)}")

# Scheduler'ı başlat
try:
    start_scheduler()
except Exception as e:
    print(f"⚠️  Scheduler başlatılamadı: {str(e)}")

# ============================================

# Error Handlers - Session timeout ve diğer hatalar için
@app.errorhandler(500)
def internal_error(error):
    """500 hatası - Session timeout durumunda login'e yönlendir"""
    db.session.rollback()
    if 'kullanici_id' not in session:
        flash('Oturumunuz sona erdi. Lütfen tekrar giriş yapın.', 'warning')
        return redirect(url_for('login'))
    logger.error(f"500 Hatası: {error}")
    return render_template('errors/500.html'), 500

@app.errorhandler(401)
def unauthorized_error(error):
    """401 hatası - Yetkisiz erişim"""
    flash('Bu sayfaya erişim yetkiniz yok. Lütfen giriş yapın.', 'warning')
    return redirect(url_for('login'))

@app.errorhandler(403)
def forbidden_error(error):
    """403 hatası - Yasaklı erişim"""
    flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
    return redirect(url_for('dashboard'))

@app.errorhandler(404)
def not_found_error(error):
    """404 hatası - Sayfa bulunamadı"""
    if 'kullanici_id' not in session:
        return redirect(url_for('login'))
    return render_template('errors/404.html'), 404

# Session kontrolü - Her istekte
@app.before_request
def check_session_validity():
    """Her istekte session geçerliliğini kontrol et"""
    from datetime import datetime as dt
    # Static dosyalar ve login sayfası hariç
    if request.endpoint and not request.endpoint.startswith('static') and request.endpoint not in ['login', 'logout']:
        if 'kullanici_id' in session:
            # Session son aktivite zamanını kontrol et
            last_activity = session.get('last_activity')
            if last_activity:
                try:
                    last_time = dt.fromisoformat(last_activity)
                    timeout = app.config.get('PERMANENT_SESSION_LIFETIME', timedelta(hours=8))
                    if isinstance(timeout, int):
                        timeout = timedelta(seconds=timeout)
                    if dt.now() - last_time > timeout:
                        session.clear()
                        flash('Oturumunuz sona erdi. Lütfen tekrar giriş yapın.', 'warning')
                        return redirect(url_for('login'))
                except:
                    pass
            # Son aktivite zamanını güncelle
            session['last_activity'] = dt.now().isoformat()


if __name__ == '__main__':
    print()
    print("=" * 60)
    print("🏨 OTEL MİNİBAR TAKİP SİSTEMİ")
    print("=" * 60)
    print()

    # Veritabanını başlat
    if init_database():
        print()
        print("🚀 Uygulama başlatılıyor...")
        # Railway PORT environment variable desteği
        port = int(os.getenv('PORT', 5014))
        debug_mode = os.getenv('FLASK_ENV', 'development') == 'development'

        # HTTPS desteği için SSL context (mobil kamera erişimi için gerekli)
        ssl_context = None
        use_https = os.getenv('USE_HTTPS', 'false').lower() == 'true'

        if use_https:
            cert_file = os.path.join(os.path.dirname(__file__), 'cert.pem')
            key_file = os.path.join(os.path.dirname(__file__), 'key.pem')

            if os.path.exists(cert_file) and os.path.exists(key_file):
                ssl_context = (cert_file, key_file)
                print(f"🔒 HTTPS Aktif: https://0.0.0.0:{port}")
                print(f"📱 Mobil erişim: https://<IP-ADRESINIZ>:{port}")
                print("⚠️  Self-signed sertifika kullanıldığı için tarayıcıda güvenlik uyarısı alabilirsiniz.")
                print("   Mobilde 'Advanced' > 'Proceed to site' seçeneğini kullanın.")
            else:
                print("⚠️  SSL sertifikası bulunamadı. Sertifika oluşturmak için:")
                print("   python generate_ssl_cert.py")
                print("📍 HTTP Modu: http://0.0.0.0:{port}")
                print("⚠️  Mobil kamera erişimi için HTTPS gereklidir!")
        else:
            print(f"📍 HTTP Modu: http://0.0.0.0:{port}")
            print("⚠️  Mobil kamera erişimi için HTTPS gereklidir!")
            print("   HTTPS'i aktifleştirmek için .env dosyasına USE_HTTPS=true ekleyin")

        print("🌙 Dark/Light tema: Sağ üstten değiştirilebilir")
        print()
        print("Durdurmak için CTRL+C kullanın")
        print("=" * 60)
        print()

        try:
            app.run(
                debug=debug_mode,
                host='0.0.0.0',
                port=port,
                ssl_context=ssl_context,
                use_reloader=True
            )
        except Exception as e:
            print(f"❌ Flask başlatma hatası: {e}")
            import traceback
            traceback.print_exc()
    else:
        print()
        print("❌ Uygulama başlatılamadı. Lütfen veritabanı ayarlarını kontrol edin.")
        print()
        exit(1)


# ============================================================================
# API: SETUP YÖNETİMİ
# ============================================================================

@app.route('/api/setuplar', methods=['GET'])
# Setup CRUD API'leri routes/sistem_yoneticisi_routes.py'de tanımlı
# - GET /api/setuplar - Liste
# - POST /api/setuplar - Yeni ekle
# - PUT /api/setuplar/<id> - Güncelle
# - DELETE /api/setuplar/<id> - Sil
# NOT: Setup içerik, ürün grupları ve ürün listesi API'leri routes/sistem_yoneticisi_routes.py'de tanımlı
# NOT: Oda tipleri ve Setup atama API'leri routes/sistem_yoneticisi_routes.py'de tanımlı


# ============================================================================
# DOCKER HEALTH CHECK ENDPOINT
# ============================================================================

@app.route('/health', methods=['GET'])
def health_check():
    """
    Docker container health check endpoint
    Database bağlantısını kontrol eder
    """
    try:
        # Database bağlantısını test et
        db.session.execute(db.text('SELECT 1'))
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': get_kktc_now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'database': 'disconnected',
            'error': str(e),
            'timestamp': get_kktc_now().isoformat()
        }), 503
